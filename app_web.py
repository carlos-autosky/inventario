"""
Sistema de Inventario v4.10.1 — Interfaz Web (Streamlit)
"""
# ── Performance instrumentation (lo primero, para medir TODO el rerun) ──
import time as _ptime
_PERF_T0 = _ptime.perf_counter()
_PERF_CHECKPOINTS = []
def _perf(label):
    _PERF_CHECKPOINTS.append((label, _ptime.perf_counter() - _PERF_T0))
_perf("script_start")

import streamlit as st
import streamlit.components.v1 as _components
import tempfile, io, os, sys
import pandas as pd
from datetime import datetime, date, timedelta, timezone
from collections import defaultdict

# Zona horaria Ecuador (America/Guayaquil). Ecuador no observa DST, por lo que
# un offset fijo -05:00 es correcto y evita dependencias extra (pytz/zoneinfo).
_EC_TZ = timezone(timedelta(hours=-5))

def _now_ec():
    """Hora actual en Ecuador (GMT-5). Devuelve datetime naive para servir
    como drop-in replacement del antiguo datetime.now() en strftime y
    comparaciones contra fechas ya naive del resto del código."""
    return datetime.now(_EC_TZ).replace(tzinfo=None)

sys.path.insert(0, os.path.dirname(__file__))
from app.engine import InventoryEngine
from app.config import PRIMARY_WAREHOUSE
from app.toma_fisica_module import DEFAULT_LOCATIONS
_perf("imports_done")

# Fragment decorator: aisla reruns al bloque decorado (Streamlit 1.37+ estable,
# 1.33-1.36 experimental, anterior: sin efecto — funciona como rerun global)
try:
    _fragment = st.fragment
except AttributeError:
    try:
        _fragment = st.experimental_fragment
    except AttributeError:
        def _fragment(fn): return fn

def _rerun_frag():
    """Rerun sólo del fragment actual (Streamlit 1.37+). Fallback: rerun global."""
    try:
        st.rerun(scope="fragment")
    except Exception:
        st.rerun()

APP_VERSION = "v4.17.0"
BUILD_TIME  = "22/04/2026 GMT-5"

# ── Diagnóstico de inicio (log) ──────────────────────────────
import logging as _logging
_log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inicio.log")
try:
    with open(_log_path, "a", encoding="utf-8") as _lf:
        _lf.write(f"[STREAMLIT] app_web.py cargado correctamente\n")
        _lf.write(f"[STREAMLIT] __file__ = {os.path.abspath(__file__)}\n")
        _lf.write(f"[STREAMLIT] APP_VERSION = {APP_VERSION}\n")
        _lf.write(f"[STREAMLIT] BUILD_TIME  = {BUILD_TIME}\n")
except Exception: pass

# Forzar recarga: limpiar estado de sesión si la versión cambió
if st.session_state.get("_app_version") != "v4.17.0":
    st.session_state.clear()
    st.session_state["_app_version"] = "v4.17.0"

st.set_page_config(page_title="Inventario v4.10.1", page_icon="📦",
                   layout="wide", initial_sidebar_state="expanded")

# ── Estado compartido multi-sesión ──────────────────────────────
# El engine y la lista de archivos son UN SOLO objeto compartido por
# todas las sesiones (cache_resource = singleton). Así, cuando un cliente
# sube archivos, los demás ven los datos sin tener que recargarlos.
# Persistimos en disco (consolidado.xlsx / toma_fisica.xlsx) para que
# los datos sobrevivan reinicios del servidor.
import threading
_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONSOLIDADO_PATH = os.path.join(_BASE_DIR, "consolidado.xlsx")
TOMA_FISICA_PATH = os.path.join(_BASE_DIR, "toma_fisica.xlsx")
RAPIDA_PATH      = os.path.join(_BASE_DIR, "toma_fisica_rapida.xlsx")
_PERF_LOG_PATH   = os.path.join(_BASE_DIR, "perf.log")
_SHARED_WRITE_LOCK = threading.Lock()

@st.cache_resource
def _get_perf_history():
    return {"runs": []}

def _perf_flush():
    if not _PERF_CHECKPOINTS: return
    try:
        ts = _now_ec().strftime("%H:%M:%S.%f")[:-3]
        total_ms = _PERF_CHECKPOINTS[-1][1] * 1000
        ck_list = []
        prev = 0.0
        parts = []
        for label, t in _PERF_CHECKPOINTS:
            dt_ms = (t - prev) * 1000
            ck_list.append((label, round(dt_ms)))
            parts.append(f"{label}={dt_ms:.0f}")
            prev = t
        with open(_PERF_LOG_PATH, "a", encoding="utf-8") as f:
            f.write(f"{ts} TOTAL={total_ms:.0f}ms | {' '.join(parts)}\n")
        hist = _get_perf_history()
        hist["runs"].append({"ts": ts, "total_ms": round(total_ms), "ck": ck_list})
        hist["runs"] = hist["runs"][-30:]
    except Exception: pass
_RAPIDA_COLS = ["Fecha","Ubicación","Código Producto","Nombre Producto",
                "Cantidad Física","Observación"]

@st.cache_resource
def _get_shared_engine():
    e = InventoryEngine()
    if os.path.exists(CONSOLIDADO_PATH):
        try: e.load_inventory_file(CONSOLIDADO_PATH)
        except Exception: pass
    if os.path.exists(TOMA_FISICA_PATH):
        try: e.load_physical_count(TOMA_FISICA_PATH)
        except Exception: pass
    return e

@st.cache_resource
def _get_shared_files():
    state = {"files_loaded": [], "files_stats": []}
    if os.path.exists(CONSOLIDADO_PATH):
        state["files_loaded"].append("consolidado.xlsx (persistido)")
    return state

def _persist_raw(df):
    try:
        with _SHARED_WRITE_LOCK:
            df.to_excel(CONSOLIDADO_PATH, index=False, engine="openpyxl")
    except Exception as ex:
        log(f"⚠ No se pudo persistir consolidado: {ex}")

def _persist_physical(df):
    try:
        with _SHARED_WRITE_LOCK:
            df.to_excel(TOMA_FISICA_PATH, index=False, engine="openpyxl")
    except Exception as ex:
        log(f"⚠ No se pudo persistir toma física: {ex}")

@st.cache_resource
def _get_shared_rapid():
    state = {"df": pd.DataFrame(columns=_RAPIDA_COLS)}
    if os.path.exists(RAPIDA_PATH):
        try:
            df = pd.read_excel(RAPIDA_PATH)
            # Compat: versión previa usaba "Bodega" en lugar de "Ubicación"
            if "Bodega" in df.columns and "Ubicación" not in df.columns:
                df = df.rename(columns={"Bodega":"Ubicación"})
            for c in _RAPIDA_COLS:
                if c not in df.columns: df[c] = ""
            state["df"] = df[_RAPIDA_COLS]
        except Exception:
            pass
    return state

def _persist_rapid(df):
    try:
        with _SHARED_WRITE_LOCK:
            df.to_excel(RAPIDA_PATH, index=False, engine="openpyxl")
    except Exception as ex:
        log(f"⚠ No se pudo persistir toma rápida: {ex}")

# ── Importaciones (pedidos al proveedor en tránsito / por llegar) ──
IMPORTACIONES_PATH = os.path.join(_BASE_DIR, "importaciones.xlsx")
_IMP_COLS = [
    "ID", "Pedido", "Proveedor",
    "Código Producto", "Nombre Producto",
    "Cantidad", "Tipo Embarque", "Fecha Tentativa",
    "Estado", "Fecha Registro", "Fecha Llegada Real", "Observaciones",
]
_IMP_ESTADOS = ["INGRESADA", "LLEGADA", "PROCESADA"]
_IMP_TIPOS   = ["Marítimo", "Aéreo"]

def _imp_empty_df():
    return pd.DataFrame(columns=_IMP_COLS)

@st.cache_resource
def _get_shared_importaciones():
    state = {"df": _imp_empty_df()}
    if os.path.exists(IMPORTACIONES_PATH):
        try:
            df = pd.read_excel(IMPORTACIONES_PATH)
            for c in _IMP_COLS:
                if c not in df.columns: df[c] = ""
            # Normalizar tipos: Cantidad numérica; fechas como date si se puede
            df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors="coerce").fillna(0).astype(int)
            state["df"] = df[_IMP_COLS]
        except Exception:
            pass
    return state

def _persist_importaciones(df):
    try:
        with _SHARED_WRITE_LOCK:
            df.to_excel(IMPORTACIONES_PATH, index=False, engine="openpyxl")
    except Exception as ex:
        log(f"⚠ No se pudo persistir importaciones: {ex}")

def _next_import_id(df):
    """Genera el siguiente ID correlativo IMP-NNNN examinando los IDs existentes."""
    if df is None or df.empty or "ID" not in df.columns:
        return "IMP-0001"
    import re as _re
    max_n = 0
    for _v in df["ID"].astype(str):
        m = _re.match(r"IMP-(\d+)", _v.strip())
        if m:
            try: max_n = max(max_n, int(m.group(1)))
            except Exception: pass
    return f"IMP-{max_n+1:04d}"

def _imp_fecha_fmt(v):
    """Formatea una fecha tentativa/real a DD/MM/YYYY, robusto a strings/NaT/datetime."""
    if v is None or (isinstance(v, float) and pd.isna(v)): return ""
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts): return str(v) if v else ""
        return ts.strftime("%d/%m/%Y")
    except Exception:
        return str(v) if v else ""

def _imp_tipo_abbr(t):
    """Abreviación de tipo embarque para strings compactos (A=Aéreo, M=Marítimo)."""
    s = str(t or "").strip().lower()
    if s.startswith("a"): return "A"
    if s.startswith("m"): return "M"
    return "?"

def _aggregate_importaciones(imp_df):
    """Agrega el DataFrame de importaciones por SKU. Devuelve dict:
      sku -> {"en_transito": int, "llegadas": int, "fechas_txt": str}
    - en_transito: suma de Cantidad en estado INGRESADA (aún no llegó)
    - llegadas:    suma de Cantidad en estado LLEGADA  (en bodega, aún no en
                   Contifico — se suma al Stock Disponible)
    - fechas_txt:  string con fechas tentativas e.g. '05/05/2026 (A·40), 20/06/2026 (M·60)'
                   sólo para tránsito (INGRESADA). PROCESADA no aporta a nada."""
    out = {}
    if imp_df is None or imp_df.empty:
        return out
    df = imp_df.copy()
    df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors="coerce").fillna(0).astype(int)
    df["Estado"]   = df["Estado"].fillna("").astype(str).str.upper()

    # Agrupar por SKU
    for sku, grp in df.groupby(df["Código Producto"].astype(str)):
        if not sku or sku == "nan": continue
        ingr = grp[grp["Estado"] == "INGRESADA"]
        lleg = grp[grp["Estado"] == "LLEGADA"]
        en_tr = int(ingr["Cantidad"].sum())
        lleg_q = int(lleg["Cantidad"].sum())

        # Fechas de tránsito: ordenadas asc, formato DD/MM/YYYY (tipo·qty)
        _piezas = []
        _ingr_sorted = ingr.copy()
        _ingr_sorted["_d"] = pd.to_datetime(_ingr_sorted["Fecha Tentativa"],
                                              errors="coerce")
        _ingr_sorted = _ingr_sorted.sort_values("_d", na_position="last")
        for _, r in _ingr_sorted.iterrows():
            _f = _imp_fecha_fmt(r["Fecha Tentativa"])
            _t = _imp_tipo_abbr(r["Tipo Embarque"])
            _q = int(r["Cantidad"])
            _piezas.append(f"{_f} ({_t}·{_q})" if _f else f"({_t}·{_q})")
        fechas_txt = ", ".join(_piezas)

        out[sku] = {"en_transito": en_tr, "llegadas": lleg_q,
                    "fechas_txt": fechas_txt}
    return out

def _apply_importaciones_to_sku_base(base_df, imp_df, excluded_skus=None):
    """Enriquece el DataFrame base (con columnas Código Producto, Nombre Producto,
    Stock Disponible) sumando llegadas confirmadas al stock y agregando columnas
    'En Tránsito' y 'Fechas Tránsito'. SKUs sólo presentes en importaciones
    (nunca vistos en Contifico) se agregan como filas nuevas con Stock=llegadas.
    Respeta la lista de SKUs excluidos globalmente.

    **Nota**: Stock Disponible se clampa a 0 (nunca negativo) — los reportes
    ejecutivos de disponibilidad no deben mostrar cantidades negativas; un
    cuadre desfavorable en el motor se interpreta como 'sin stock'."""
    excluded_skus = set(excluded_skus or [])
    agg = _aggregate_importaciones(imp_df)

    df = base_df.copy()
    if "En Tránsito" not in df.columns:
        df["En Tránsito"] = 0
    if "Fechas Tránsito" not in df.columns:
        df["Fechas Tránsito"] = ""

    if not agg:
        df["En Tránsito"] = df["En Tránsito"].fillna(0).astype(int)
        df["Fechas Tránsito"] = df["Fechas Tránsito"].fillna("").astype(str)
        df["Stock Disponible"] = df["Stock Disponible"].fillna(0).astype(int).clip(lower=0)
        return df

    # Merge: para cada SKU en agg, ajustar fila existente o agregar nueva
    existing_skus = set(df["Código Producto"].astype(str))
    # Mapa SKU→Nombre desde importaciones (para SKUs nuevos)
    _nombres_imp = {}
    if imp_df is not None and not imp_df.empty:
        for _sku, _grp in imp_df.groupby(imp_df["Código Producto"].astype(str)):
            _nom = _grp["Nombre Producto"].astype(str).replace("nan","").str.strip()
            _nom = _nom[_nom != ""]
            _nombres_imp[_sku] = _nom.iloc[0] if len(_nom) else ""

    # Ajustar filas existentes
    for i, row in df.iterrows():
        sku = str(row["Código Producto"])
        if sku in agg and sku not in excluded_skus:
            a = agg[sku]
            df.at[i, "Stock Disponible"] = int(row["Stock Disponible"]) + a["llegadas"]
            df.at[i, "En Tránsito"]      = a["en_transito"]
            df.at[i, "Fechas Tránsito"]  = a["fechas_txt"]

    # Filas nuevas (SKUs sólo en importaciones)
    _new_rows = []
    for sku, a in agg.items():
        if sku in existing_skus: continue
        if sku in excluded_skus: continue
        if (a["en_transito"] + a["llegadas"]) == 0: continue
        _new_rows.append({
            "Código Producto": sku,
            "Nombre Producto": _nombres_imp.get(sku, ""),
            "Stock Disponible": a["llegadas"],
            "En Tránsito":      a["en_transito"],
            "Fechas Tránsito":  a["fechas_txt"],
        })
    if _new_rows:
        df = pd.concat([df, pd.DataFrame(_new_rows)], ignore_index=True)

    # Clamp a 0 — Stock Disponible en reportes ejecutivos no puede ser negativo
    df["Stock Disponible"] = df["Stock Disponible"].fillna(0).astype(int).clip(lower=0)
    df["En Tránsito"]      = df["En Tránsito"].fillna(0).astype(int)
    df["Fechas Tránsito"]  = df["Fechas Tránsito"].fillna("").astype(str)
    return df.sort_values("Código Producto").reset_index(drop=True)

# Ubicaciones personalizadas — archivo JSON global compartido entre sesiones
UBIC_CUSTOM_PATH = os.path.join(_BASE_DIR, "ubicaciones_custom.json")
LOGO_PATH        = os.path.join(_BASE_DIR, "logo.png")

def _get_logo_path():
    """Devuelve la ruta del logo si existe, sino None.
    Prioriza `logo.png` (subido por el usuario vía sidebar), luego
    `Logo AutoSky 500x300.png` (default del repo)."""
    if os.path.exists(LOGO_PATH):
        return LOGO_PATH
    _default = os.path.join(_BASE_DIR, "Logo AutoSky 500x300.png")
    if os.path.exists(_default):
        return _default
    return None

def _persist_logo(png_bytes):
    try:
        with _SHARED_WRITE_LOCK:
            with open(LOGO_PATH, "wb") as f:
                f.write(png_bytes)
    except Exception as ex:
        log(f"⚠ No se pudo persistir logo: {ex}")
FILTROS_PATH     = os.path.join(_BASE_DIR, "filtros_config.json")

# Filtros globales compartidos (SKUs y Bodegas excluidas del análisis)
# Se persisten a disco para que sobrevivan a refresh y sesión nueva.
@st.cache_resource
def _get_shared_filtros():
    state = {"excluded_skus": set(), "excluded_warehouses": set()}
    if os.path.exists(FILTROS_PATH):
        try:
            import json
            with open(FILTROS_PATH, encoding="utf-8") as f:
                data = json.load(f)
            state["excluded_skus"]        = set(data.get("excluded_skus", []))
            state["excluded_warehouses"]  = set(data.get("excluded_warehouses", []))
        except Exception:
            pass
    return state

def _persist_filtros(excluded_skus, excluded_warehouses):
    try:
        import json
        with _SHARED_WRITE_LOCK:
            with open(FILTROS_PATH, "w", encoding="utf-8") as f:
                json.dump({
                    "excluded_skus":       sorted(list(excluded_skus)),
                    "excluded_warehouses": sorted(list(excluded_warehouses)),
                }, f, ensure_ascii=False, indent=2)
    except Exception as ex:
        log(f"⚠ No se pudo persistir filtros: {ex}")

@st.cache_resource
def _get_custom_ubic():
    state = {"list": []}
    if os.path.exists(UBIC_CUSTOM_PATH):
        try:
            import json
            with open(UBIC_CUSTOM_PATH, encoding="utf-8") as f:
                state["list"] = json.load(f).get("ubicaciones", [])
        except Exception: pass
    return state

def _persist_custom_ubic(lst):
    try:
        import json
        with _SHARED_WRITE_LOCK:
            with open(UBIC_CUSTOM_PATH, "w", encoding="utf-8") as f:
                json.dump({"ubicaciones": list(lst)}, f, ensure_ascii=False, indent=2)
    except Exception as ex:
        log(f"⚠ No se pudo persistir ubicaciones: {ex}")

def _get_all_ubic():
    """Devuelve DEFAULT_LOCATIONS + custom, preservando orden y sin duplicados."""
    custom = _get_custom_ubic()["list"]
    seen = set(); result = []
    for u in list(DEFAULT_LOCATIONS) + list(custom):
        if u not in seen:
            result.append(u); seen.add(u)
    return result

VENTANA_OPTS = {
    "Últimos 30 días":  30,
    "Últimos 90 días":  90,
    "Últimos 180 días": 180,
    "Último año":       365,
    "Todo el período":  None,
    "Personalizado":    "custom",
}

def _compute_window_sales(r, ventana_label, custom_start=None, custom_end=None):
    """Calcula ventas por SKU y días de la ventana elegida.

    El consumo diario debe basarse en el ritmo RECIENTE (últimos N días),
    no en todo el histórico: si creciste 3x, las ventas antiguas diluyen
    el promedio y subestiman las compras.

    Devuelve: (days, ventas_by_sku, win_min, win_max)
    """
    df_f = r.filtered
    _data_max = df_f["Fecha"].max()
    _data_min = df_f["Fecha"].min()

    if ventana_label == "Personalizado":
        win_min = pd.Timestamp(custom_start) if custom_start else _data_min
        win_max = pd.Timestamp(custom_end)   if custom_end   else _data_max
    else:
        n_days = VENTANA_OPTS.get(ventana_label)
        if n_days is None:
            win_min = _data_min
            win_max = _data_max
        else:
            win_min = _data_max - pd.Timedelta(days=n_days)
            win_max = _data_max

    # Clamp dentro del rango de datos
    if win_min < _data_min: win_min = _data_min
    if win_max > _data_max: win_max = _data_max
    if win_min > win_max:   win_min = win_max

    df_win = df_f[(df_f["Fecha"] >= win_min) & (df_f["Fecha"] <= win_max)]
    typ = df_win["Tipo"].fillna("").astype(str).str.upper()
    ref = df_win["Referencia"].fillna("").astype(str).str.upper()
    mask = (typ == "EGR") & ref.str.startswith("FAC")
    ventas_win = df_win[mask]
    if ventas_win.empty:
        ventas_by_sku = {}
    else:
        ventas_by_sku = (ventas_win.groupby("Código Producto")["Cantidad"]
                         .sum().abs().to_dict())

    days = max(int((win_max - win_min).days) + 1, 1)
    return days, ventas_by_sku, win_min, win_max

def _ubic_sheet_name(ubic):
    """Reproduce la transformación que hace la plantilla al convertir
    un nombre de ubicación en nombre de hoja Excel (max 28 chars,
    ciertos caracteres reemplazados)."""
    s = str(ubic)[:28]
    for ch, rp in [("/","-"),("\\","-"),("?",""),("*",""),
                    ("[",""),("]",""),(":","")]:
        s = s.replace(ch, rp)
    return s

def _detect_historial_format(path):
    """True si el Excel es un historial exportado: 1 hoja con encabezados
    Fecha, Ubicación, Código Producto, Nombre Producto, Cantidad Física."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if len(wb.sheetnames) != 1:
            return False
        ws = wb[wb.sheetnames[0]]
        first_row = next(ws.iter_rows(values_only=True), None)
        if not first_row:
            return False
        expected = {"Fecha","Ubicación","Código Producto",
                    "Nombre Producto","Cantidad Física"}
        headers = {str(v).strip() for v in first_row if v is not None}
        return expected.issubset(headers)
    except Exception:
        return False

def _parse_historial(path):
    """Parsea el Excel de historial y devuelve filas listas para rap_df.
    Formato: 1 hoja con columnas Fecha, Ubicación, Código, Nombre, Cantidad, Obs."""
    df = pd.read_excel(path)
    req = ["Fecha","Ubicación","Código Producto","Nombre Producto","Cantidad Física"]
    if not all(c in df.columns for c in req):
        return [], set()
    rows = []
    ubicaciones = set()
    for _, r in df.iterrows():
        try:
            qty = float(r["Cantidad Física"])
        except (TypeError, ValueError):
            continue
        if qty <= 0:
            continue
        ubic = str(r.get("Ubicación") or "").strip()
        if not ubic:
            continue
        ubicaciones.add(ubic)
        fecha = r.get("Fecha")
        if pd.isna(fecha) or fecha is None or str(fecha).strip() == "":
            fecha_str = _now_ec().strftime("%Y-%m-%d %H:%M")
        elif isinstance(fecha, (datetime, date)):
            fecha_str = fecha.strftime("%Y-%m-%d %H:%M") if isinstance(fecha, datetime) else fecha.strftime("%Y-%m-%d")
        else:
            fecha_str = str(fecha)
        _obs_val = r.get("Observación","")
        _obs = str(_obs_val).strip() if pd.notna(_obs_val) else ""
        rows.append({
            "Fecha": fecha_str,
            "Ubicación": ubic,
            "Código Producto": str(r["Código Producto"] or "").strip(),
            "Nombre Producto": str(r["Nombre Producto"] or "").strip(),
            "Cantidad Física": qty,
            "Observación": _obs,
        })
    return rows, ubicaciones

def _parse_plantilla_toma(path, registered_ubic):
    """Parsea un Excel con el formato de la plantilla de toma física.

    Por cada hoja (excepto 'RESUMEN GENERAL'):
      - Valida que el nombre corresponda a una ubicación registrada.
      - Lee B2 como fecha de toma (opcional).
      - Lee filas desde la 4: A=Código, B=Nombre, C=Cantidad, D=Observación.
      - Ignora celdas vacías o cantidad ≤ 0 (significan "no contado").

    Devuelve:
      {
        "valid_sheets": {ubicacion: {"fecha": date|None, "rows": [...]}},
        "invalid_sheets": [nombres de hoja que no corresponden a ubicación registrada]
      }
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # Mapa: nombre_hoja (transformado) -> ubicación real
    reverse = {_ubic_sheet_name(u): u for u in registered_ubic}

    result = {"valid_sheets": {}, "invalid_sheets": []}

    for sname in wb.sheetnames:
        if sname.upper() == "RESUMEN GENERAL":
            continue
        if sname not in reverse:
            result["invalid_sheets"].append(sname)
            continue

        ubic = reverse[sname]
        ws = wb[sname]

        # Fecha desde B2
        b2 = ws["B2"].value
        fecha = None
        if b2 is not None and str(b2).strip():
            try:
                if isinstance(b2, datetime):
                    fecha = b2.date()
                elif isinstance(b2, date):
                    fecha = b2
                else:
                    fecha = pd.to_datetime(str(b2), errors="coerce")
                    fecha = fecha.date() if fecha is not pd.NaT and fecha is not None else None
            except Exception:
                fecha = None

        # Filas desde 4
        rows = []
        for r in range(4, (ws.max_row or 0) + 1):
            codigo = ws.cell(r, 1).value
            if codigo is None: continue
            _code = str(codigo).strip()
            if not _code: continue
            if _code.upper() in ("TOTAL", "TOTAL GENERAL"):
                break

            nombre = ws.cell(r, 2).value
            cantidad = ws.cell(r, 3).value
            obs = ws.cell(r, 4).value

            try:
                qty = float(cantidad) if cantidad is not None else 0.0
            except (ValueError, TypeError):
                qty = 0.0

            # Vacío o 0 = "no contado" → skip
            if qty <= 0:
                continue

            rows.append({
                "codigo":   _code,
                "nombre":   str(nombre or "").strip(),
                "cantidad": qty,
                "obs":      str(obs or "").strip(),
            })

        result["valid_sheets"][ubic] = {"fecha": fecha, "rows": rows}

    return result

# ── Session state ───────────────────────────────────────────────
def _init():
    shared_files   = _get_shared_files()
    shared_filtros = _get_shared_filtros()
    defs = {"engine": _get_shared_engine(), "result": None,
            "files_loaded": shared_files["files_loaded"],
            "files_stats":  shared_files["files_stats"],
            "log": [], "dark_mode": False,
            # Filtros globales persistidos (se cargan de filtros_config.json)
            "excluded_skus": set(shared_filtros["excluded_skus"]),
            "excl_wh":       set(shared_filtros["excluded_warehouses"])}
    for k,v in defs.items():
        if k not in st.session_state: st.session_state[k] = v
    # Propagar filtros al engine al primer init (para analyze correcto)
    _e = st.session_state.engine
    _e.excluded_skus        = set(st.session_state.excluded_skus)
    _e.excluded_warehouses  = set(st.session_state.excl_wh)
    # Si la sesión arranca con datos pre-cargados (el servidor ya los tenía)
    # pero sin cálculo, disparar auto-análisis para que los KPIs aparezcan
    # sin exigir al usuario pulsar "Calcular"
    if (_e.raw_df is not None and st.session_state.result is None):
        st.session_state["_recalc_pending"] = True
_init()
eng = st.session_state.engine
dark = st.session_state.dark_mode
_perf("session_init")

# ── Tema ────────────────────────────────────────────────────────
if dark:
    BG,PANEL,BORDER,TEXT,MUTED = "#0f172a","#1e293b","#334155","#f1f5f9","#94a3b8"
    TH,TDE,TDO,HOVER = "#0f172a","#1e293b","#162032","#243447"
    ACC,SUC,WRN,DNG = "#38bdf8","#4ade80","#fbbf24","#f87171"
else:
    BG,PANEL,BORDER,TEXT,MUTED = "#f8fafc","#ffffff","#e2e8f0","#0f172a","#64748b"
    TH,TDE,TDO,HOVER = "#f1f5f9","#ffffff","#f8fafc","#e0f2fe"
    ACC,SUC,WRN,DNG = "#0284c7","#16a34a","#d97706","#dc2626"

# ── Autosky Design System CSS ─────────────────────────────
_CSS = '''
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;600&display=swap');
:root {
  --sky:#0ea5e9; --sky-d:#0284c7; --sky-dd:#0369a1;
  --sky-l:#e0f2fe; --sky-ll:#f0f9ff;
  --bg:#f0f9ff; --surface:#ffffff; --surface2:#f8fafc;
  --border:#e2e8f0; --border2:#cbd5e1;
  --text:#0f172a; --text2:#475569; --text3:#94a3b8;
  --green:#059669; --green-l:#ecfdf5;
  --red:#dc2626; --red-l:#fef2f2;
  --amber:#d97706; --amber-l:#fffbeb;
  --purple:#7c3aed; --purple-l:#f3e8ff;
  --radius:10px; --radius-sm:6px;
  --shadow:0 1px 3px rgba(0,0,0,.08);
}
html,body,[data-testid='stAppViewContainer']{
  background:var(--bg)!important;
  font-family:'Inter','Segoe UI',system-ui,sans-serif;
  font-size:13px; color:var(--text);
}
.as-banner{
  background:linear-gradient(135deg,#0ea5e9,#38bdf8,#7dd3fc);
  padding:12px 22px; border-radius:10px; margin-bottom:16px;
  display:flex; align-items:center; justify-content:space-between;
  box-shadow:0 2px 8px rgba(14,165,233,.35);
}
.as-logo{font-size:18px;font-weight:800;color:#fff;letter-spacing:.06em;}
.as-logo span{font-weight:300;opacity:.85;}
.as-build{text-align:right;color:rgba(255,255,255,.9);
  font-size:10px;font-family:'JetBrains Mono',monospace;}
.as-build .v{font-size:13px;font-weight:700;}
section[data-testid='stSidebar']{
  background:var(--surface)!important;
  border-right:1px solid var(--border)!important;
}
section[data-testid='stSidebar'] *{color:var(--text)!important;}
section[data-testid='stSidebar'] .stSelectbox label,
section[data-testid='stSidebar'] .stMultiSelect label,
section[data-testid='stSidebar'] .stCheckbox label{
  font-size:10px!important; font-weight:600!important;
  color:var(--text3)!important; text-transform:uppercase; letter-spacing:.06em;
}
.kpi-row{display:flex;gap:10px;margin-bottom:10px;flex-wrap:wrap;}
.kpi-card{
  background:var(--surface); border:1px solid var(--border);
  border-radius:var(--radius); padding:14px 16px;
  flex:1 1 140px; min-width:0;
  position:relative; overflow:hidden; box-shadow:var(--shadow);
}
.kpi-card::before{
  content:''; position:absolute; top:0; left:0; right:0; height:3px;
  background:var(--accent,var(--sky));
}
.kpi-label{font-size:10px;font-weight:600;color:var(--text3);
  text-transform:uppercase;letter-spacing:.06em;margin-bottom:5px;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.kpi-value{font-size:21px;font-weight:700;color:var(--text);
  font-family:'JetBrains Mono',monospace;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.kpi-sub{font-size:10px;color:var(--text3);margin-top:3px;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
@media (max-width: 640px){
  .kpi-row{gap:6px;}
  .kpi-card{flex:1 1 calc(50% - 3px); padding:10px 12px;}
  .kpi-label{font-size:9px;letter-spacing:.02em;margin-bottom:3px;}
  .kpi-value{font-size:17px;}
  .kpi-sub{font-size:9px;}
}
@media (max-width: 420px){
  .kpi-value{font-size:15px;}
}
.kpi-card.a{--accent:var(--sky);}
.kpi-card.s{--accent:var(--green);}
.kpi-card.w{--accent:var(--amber);}
.kpi-card.d{--accent:var(--red);}
.kpi-card.p{--accent:var(--purple);}
.tc{overflow:auto;max-height:540px;border-radius:var(--radius);
  border:1px solid var(--border);background:var(--surface);
  box-shadow:var(--shadow);
  /* CRÍTICO: position relative para que sticky funcione dentro */
  position:relative;}
.tc.piv{max-height:680px;}
.it{width:100%;border-collapse:separate;border-spacing:0;font-size:12px;
  color:var(--text);font-family:'Inter',sans-serif;}

/* ── Header sticky vertical (top:0 relativo al .tc) ── */
.it thead th{
  background:var(--surface2);color:var(--text3);
  font-size:10px;font-weight:600;letter-spacing:.06em;text-transform:uppercase;
  padding:8px 12px;
  position:sticky;top:0;z-index:3;
  border-bottom:2px solid var(--border2);
  white-space:nowrap;cursor:pointer;user-select:none;}

/* ── Col 1 sticky horizontal ── */
.it thead th:first-child,
.it tbody  td:first-child,
.it tfoot  td:first-child{
  position:sticky;left:0;z-index:2;
  background:var(--surface2);
  border-right:1px solid var(--border2);}

/* ── Col 2 sticky horizontal (para pivot con SKU + Nombre) ── */
.it thead th.sc2,
.it tbody  td.sc2,
.it tfoot  td.sc2{
  position:sticky;z-index:2;
  background:var(--surface2);
  border-right:2px solid var(--border2);}

/* ── Esquina (col1+header): z-index máximo ── */
.it thead th:first-child{z-index:5;}
.it thead th.sc2{z-index:5;}

/* ── Zebra ── */
.it tbody tr:nth-child(even) td            {background:var(--surface2);}
.it tbody tr:nth-child(odd)  td            {background:var(--surface);}
.it tbody tr:nth-child(even) td:first-child{background:var(--surface2);}
.it tbody tr:nth-child(odd)  td:first-child{background:var(--surface);}
.it tbody tr:nth-child(even) td.sc2        {background:var(--surface2);}
.it tbody tr:nth-child(odd)  td.sc2        {background:var(--surface);}

/* ── Hover sobre zebra ── */
.it tbody tr:hover td                      {background:var(--sky-ll)!important;}

.it thead th:hover{color:var(--sky);}
.it tbody td{padding:7px 12px;border-bottom:1px solid var(--surface2);color:var(--text);}
.it .n{text-align:right;font-family:'JetBrains Mono',monospace;}
/* Variante compacta para Rotación: filas densas, una por SKU (sin wrap) */
.it.it-rot thead th{padding:4px 6px!important;font-size:9px!important;
  letter-spacing:.04em;white-space:nowrap;
  /* Sticky top reforzado para que no desaparezca al scroll vertical */
  position:sticky!important;top:0!important;
  background:var(--surface2)!important;
  z-index:10!important;}
/* Esquina (Código): sticky top Y left al tiempo, z-index máximo */
.it.it-rot thead th:first-child{
  left:0!important;z-index:12!important;}
/* Nombre: segundo sticky top (sin left), z-index alto */
.it.it-rot thead th:nth-child(2){z-index:10!important;}
.it.it-rot tbody td{padding:2px 6px!important;font-size:10px;line-height:1.2;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
/* Código (col 1) sticky left, fondo opaco, z-index por debajo del header */
.it.it-rot tbody td:first-child{
  font-family:'JetBrains Mono',monospace;font-size:10px;
  position:sticky!important;left:0!important;z-index:4!important;}
/* Columna Nombre (2da) con ancho flexible y truncado */
.it.it-rot tbody td.nom{max-width:320px;}
.it.it-rot tfoot td{padding:4px 6px!important;font-size:10px;}
.it.it-rot tfoot td:first-child{position:sticky!important;left:0!important;z-index:4!important;}
.tc.rot-tc{max-height:640px;}
.it tfoot tr.tot td{background:var(--sky-ll)!important;font-weight:700;
  border-top:2px solid var(--sky-l);color:var(--sky-d);}
.it tfoot tr.tot td{background:var(--sky-ll)!important;font-weight:700;
  border-top:2px solid var(--sky-l);color:var(--sky-d);}
.zb{display:flex;align-items:center;gap:6px;margin-bottom:6px;}
.zb button{background:var(--surface);border:1px solid var(--border);
  color:var(--text2);border-radius:var(--radius-sm);
  padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;}
.zb button:hover{background:var(--sky);color:#fff;border-color:var(--sky);}
.zb .zb-info{font-size:11px;color:var(--text3);}
.stTabs [data-baseweb='tab-list']{gap:2px;background:var(--surface2);
  border-radius:var(--radius-sm);padding:4px;border:1px solid var(--border);}
.stTabs [data-baseweb='tab']{border-radius:5px;font-size:12px;
  font-weight:600;color:var(--text2);padding:6px 14px;}
.stTabs [aria-selected='true']{background:var(--sky)!important;color:#fff!important;}
.stDataFrame td,.stDataFrame th{color:var(--text)!important;}
.stButton button[kind='primary']{background:var(--sky)!important;
  border-color:var(--sky)!important;border-radius:var(--radius-sm)!important;
  font-weight:600!important;}
.stButton button[kind='primary']:hover{background:var(--sky-d)!important;}
.stButton button{border-radius:var(--radius-sm)!important;font-weight:600!important;}
.stAlert{border-radius:var(--radius)!important;}
.stSelectbox>div>div,.stMultiSelect>div>div{border-radius:var(--radius-sm)!important;}
.stTextInput>div>div>input{border-radius:var(--radius-sm)!important;}

/* ═══════════════════════════════════════════════
   FORZAR TEMA CLARO EN TODOS LOS COMPONENTES
   ═══════════════════════════════════════════════ */

/* App background completo */
.stApp, .stApp > div,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
[data-testid="block-container"],
.main, .main .block-container {
  background-color: #f0f9ff !important;
  color: #0f172a !important;
}

/* Sidebar fondo blanco */
[data-testid="stSidebar"],
[data-testid="stSidebar"] > div,
[data-testid="stSidebar"] .sidebar-content {
  background-color: #ffffff !important;
}
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] div {
  color: #0f172a !important;
}

/* Todos los textos */
p, span, div, label, h1, h2, h3, h4 {
  color: #0f172a;
}

/* File uploader */
[data-testid="stFileUploader"],
[data-testid="stFileUploadDropzone"] {
  background-color: #ffffff !important;
  border: 2px dashed #bae6fd !important;
  border-radius: 10px !important;
  color: #0f172a !important;
}
[data-testid="stFileUploadDropzone"] * { color: #0f172a !important; }
[data-testid="stFileUploadDropzone"]:hover {
  border-color: #0ea5e9 !important;
  background-color: #f0f9ff !important;
}

/* Upload button dentro del dropzone */
[data-testid="stFileUploader"] button {
  background: #0ea5e9 !important;
  color: #fff !important;
  border-radius: 6px !important;
  border: none !important;
}

/* Uploaded file item */
[data-testid="stFileUploaderFile"],
[data-testid="uploadedFileData"] {
  background: #f0f9ff !important;
  border: 1px solid #bae6fd !important;
  border-radius: 8px !important;
}
[data-testid="stFileUploaderFile"] * { color: #0f172a !important; }

/* Inputs, selects */
.stTextInput input, .stNumberInput input,
.stDateInput input, .stTimeInput input {
  background: #ffffff !important;
  color: #0f172a !important;
  border: 1px solid #e2e8f0 !important;
  border-radius: 6px !important;
}
.stSelectbox > div > div > div,
.stMultiSelect > div > div > div {
  background: #ffffff !important;
  color: #0f172a !important;
}

/* Dropdown options */
[data-baseweb="popover"], [data-baseweb="menu"],
[role="listbox"], [role="option"] {
  background: #ffffff !important;
  color: #0f172a !important;
}
[role="option"]:hover { background: #f0f9ff !important; }

/* Number input */
[data-testid="stNumberInput"] input {
  background: #ffffff !important;
  color: #0f172a !important;
}
[data-testid="stNumberInput"] button {
  background: #f1f5f9 !important;
  color: #0f172a !important;
  border: 1px solid #e2e8f0 !important;
}

/* Checkbox, toggle */
[data-testid="stCheckbox"] label,
[data-testid="stToggle"] label { color: #0f172a !important; }

/* Metrics */
[data-testid="stMetric"] * { color: #0f172a !important; }
[data-testid="stMetricValue"] {
  color: #0f172a !important;
  font-family: 'JetBrains Mono', monospace !important;
}

/* Success/Info/Warning/Error alerts */
.stAlert > div { border-radius: 8px !important; }
[data-testid="stNotification"] { border-radius: 8px !important; }

/* Success boxes (like "1 archivo cargado") */
.element-container .stAlert [data-baseweb="notification"] {
  background: #ecfdf5 !important;
  color: #065f46 !important;
  border: 1px solid #6ee7b7 !important;
}

/* Dividers */
hr { border-color: #e2e8f0 !important; }

/* Spinner */
[data-testid="stSpinner"] { color: #0ea5e9 !important; }

/* Caption text */
.stCaption, [data-testid="stCaptionContainer"] {
  color: #64748b !important;
}

/* Markdown text */
.stMarkdown p, .stMarkdown span { color: #0f172a !important; }

/* Tabs area background */
[data-testid="stHorizontalBlock"],
[data-testid="stVerticalBlock"] {
  background: transparent !important;
}

/* Column backgrounds */
[data-testid="column"] { background: transparent !important; }

/* Expander */
[data-testid="stExpander"] {
  background: #ffffff !important;
  border: 1px solid #e2e8f0 !important;
  border-radius: 10px !important;
}
[data-testid="stExpander"] summary { color: #0f172a !important; }

/* Download button */
.stDownloadButton button {
  background: #f8fafc !important;
  color: #0284c7 !important;
  border: 1px solid #bae6fd !important;
  border-radius: 6px !important;
  font-weight: 600 !important;
}
.stDownloadButton button:hover {
  background: #e0f2fe !important;
  border-color: #0ea5e9 !important;
}
'''
# ── CSS adicional tema oscuro (se inyecta sobre el base claro) ──
_DARK_CSS = '''
html,body,[data-testid='stAppViewContainer'],
.stApp, .stApp > div,
[data-testid="stAppViewContainer"],
[data-testid="stAppViewBlockContainer"],
[data-testid="block-container"],
.main, .main .block-container {
  background-color: #0f172a !important;
  color: #f1f5f9 !important;
}
:root {
  --bg:#0f172a; --surface:#1e293b; --surface2:#162032;
  --border:#334155; --text:#f1f5f9; --text2:#cbd5e1; --text3:#94a3b8;
  --sky-ll:#162032; --sky-l:#1e3a5f;
  --green-l:#052e16; --red-l:#450a0a; --amber-l:#451a03; --purple-l:#2e1065;
}
section[data-testid='stSidebar'],
[data-testid="stSidebar"],
[data-testid="stSidebar"] > div {
  background-color: #1e293b !important;
}
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] div { color: #f1f5f9 !important; }
p, span, div, label, h1, h2, h3, h4 { color: #f1f5f9; }
.stTextInput input, .stNumberInput input,
.stDateInput input, .stTimeInput input {
  background: #1e293b !important; color: #f1f5f9 !important;
  border: 1px solid #334155 !important;
}
.stSelectbox > div > div > div,
.stMultiSelect > div > div > div {
  background: #1e293b !important; color: #f1f5f9 !important;
}
[data-baseweb="popover"], [data-baseweb="menu"],
[role="listbox"], [role="option"] {
  background: #1e293b !important; color: #f1f5f9 !important;
}
[role="option"]:hover { background: #0f172a !important; }
[data-testid="stFileUploader"],
[data-testid="stFileUploadDropzone"] {
  background-color: #1e293b !important;
  border: 2px dashed #334155 !important; color: #f1f5f9 !important;
}
[data-testid="stFileUploadDropzone"] * { color: #f1f5f9 !important; }
[data-testid="stFileUploaderFile"],
[data-testid="uploadedFileData"] {
  background: #162032 !important;
  border: 1px solid #334155 !important;
}
[data-testid="stFileUploaderFile"] * { color: #f1f5f9 !important; }
[data-testid="stFileUploader"] button {
  background: #0ea5e9 !important; color: #fff !important;
}
[data-testid="stExpander"] {
  background: #1e293b !important; border: 1px solid #334155 !important;
}
[data-testid="stExpander"] summary { color: #f1f5f9 !important; }
[data-testid="stMetric"] * { color: #f1f5f9 !important; }
.stMarkdown p, .stMarkdown span { color: #f1f5f9 !important; }
hr { border-color: #334155 !important; }
.stDownloadButton button {
  background: #1e293b !important; color: #38bdf8 !important;
  border: 1px solid #334155 !important;
}
.stDownloadButton button:hover {
  background: #0f172a !important; border-color: #38bdf8 !important;
}
.it tbody td { color: #f1f5f9 !important; }
.it thead th { background: #162032 !important; color: #94a3b8 !important; }
.it tbody tr:hover td { background: #162032 !important; }
.it tfoot tr.tot td { background: #1e3a5f !important; color: #38bdf8 !important; }
'''

_ACTIVE_CSS = _CSS + (_DARK_CSS if dark else "")
st.markdown(f'''<style>{_ACTIVE_CSS}</style>
<script>
// ── Zoom ──────────────────────────────────────────────────────
function asZoom(uid, delta) {{
  var t = document.getElementById("tbl_" + uid);
  if (!t) return;
  var sz = parseFloat(window.getComputedStyle(t).fontSize) || 12;
  t.style.fontSize = Math.min(20, Math.max(8, sz + delta)) + "px";
}}
function asZoomReset(uid) {{
  var t = document.getElementById("tbl_" + uid);
  if (t) t.style.fontSize = "12px";
}}

</script>''', unsafe_allow_html=True)

# ── Resize de columnas via window.parent (accede al DOM real) ───
_components.html("""<script>
(function() {
  var doc = window.parent.document;

  // ── Col resize ───────────────────────────────────────────────
  function initResize(table) {
    table.style.tableLayout = 'auto';
    table.querySelectorAll('thead th').forEach(function(th) {
      if (th.dataset.resizeInit) return;
      th.dataset.resizeInit = '1';
      th.style.position   = 'relative';
      th.style.overflow   = 'visible';
      th.style.userSelect = 'none';
      var handle = doc.createElement('div');
      handle.style.cssText = 'position:absolute;top:0;right:0;bottom:0;width:6px;cursor:col-resize;z-index:10;background:transparent';
      handle.addEventListener('mouseenter', function() { handle.style.background = 'rgba(14,165,233,.45)'; });
      handle.addEventListener('mouseleave', function() { if (!handle._drag) handle.style.background = 'transparent'; });
      var startX, startW, drag = false;
      handle.addEventListener('mousedown', function(e) {
        e.stopPropagation(); e.preventDefault();
        drag = true; handle._drag = true;
        startX = e.pageX; startW = th.offsetWidth;
        handle.style.background   = 'rgba(14,165,233,.7)';
        doc.body.style.cursor     = 'col-resize';
        doc.body.style.userSelect = 'none';
      });
      doc.addEventListener('mousemove', function(e) {
        if (!drag) return;
        var w = Math.max(40, startW + (e.pageX - startX));
        th.style.width = th.style.minWidth = th.style.maxWidth = w + 'px';
      });
      doc.addEventListener('mouseup', function() {
        if (!drag) return;
        drag = false; handle._drag = false;
        handle.style.background   = 'transparent';
        doc.body.style.cursor     = '';
        doc.body.style.userSelect = '';
      });
      th.appendChild(handle);
    });
  }

  // ── Two-finger / trackpad pan en contenedor .tc ──────────────
  // Detecta touchstart con 2 dedos O wheel con deltaX (trackpad horizontal)
  // y los traduce en scrollLeft/scrollTop del contenedor .tc
  function initPan(tc) {
    if (tc.dataset.panInit) return;
    tc.dataset.panInit = '1';

    // Trackpad horizontal (wheel event con deltaX)
    tc.addEventListener('wheel', function(e) {
      // Si hay desplazamiento horizontal real (trackpad/dos dedos) lo aplicamos
      if (Math.abs(e.deltaX) > Math.abs(e.deltaY)) {
        e.preventDefault();
        tc.scrollLeft += e.deltaX;
      }
      // Si hay desplazamiento vertical con shift → scroll horizontal
      else if (e.shiftKey) {
        e.preventDefault();
        tc.scrollLeft += e.deltaY;
      }
      // Vertical normal: dejar comportamiento por defecto (no interceptar)
    }, { passive: false });

    // Touch (móvil / tablet): dos dedos → pan libre
    var t0x, t0y, t0sl, t0st;
    tc.addEventListener('touchstart', function(e) {
      if (e.touches.length !== 2) return;
      t0x  = (e.touches[0].pageX + e.touches[1].pageX) / 2;
      t0y  = (e.touches[0].pageY + e.touches[1].pageY) / 2;
      t0sl = tc.scrollLeft;
      t0st = tc.scrollTop;
    }, { passive: true });

    tc.addEventListener('touchmove', function(e) {
      if (e.touches.length !== 2) return;
      e.preventDefault();
      var cx = (e.touches[0].pageX + e.touches[1].pageX) / 2;
      var cy = (e.touches[0].pageY + e.touches[1].pageY) / 2;
      tc.scrollLeft = t0sl - (cx - t0x);
      tc.scrollTop  = t0st - (cy - t0y);
    }, { passive: false });
  }

  // ── Scan: inicializar resize y pan en todas las tablas/contenedores ──
  function scanTables() {
    doc.querySelectorAll('table.it').forEach(initResize);
    doc.querySelectorAll('div.tc').forEach(initPan);
  }

  var obs = new MutationObserver(function(muts) {
    muts.forEach(function(m) {
      m.addedNodes.forEach(function(n) {
        if (n.nodeType !== 1) return;
        if (n.matches) {
          if (n.matches('table.it')) initResize(n);
          if (n.matches('div.tc'))   initPan(n);
        }
        if (n.querySelectorAll) {
          n.querySelectorAll('table.it').forEach(initResize);
          n.querySelectorAll('div.tc').forEach(initPan);
        }
      });
    });
  });
  obs.observe(doc.body, { childList: true, subtree: true });

  scanTables();
  setTimeout(scanTables, 800);
  setTimeout(scanTables, 2000);
})();
</script>""", height=0, scrolling=False)
_perf("css_js_emitted")

# ── Helpers ─────────────────────────────────────────────────────
def log(m):
    ts = _now_ec().strftime("%H:%M:%S")
    st.session_state.log.insert(0, f"[{ts}] {m}")
    st.session_state.log = st.session_state.log[:300]

def fmt(v, t="n"):
    try:
        f = float(v)
        if t=="i": return f"{int(f):,}"
        if t=="p": return f"{f:.1f}%"
        return f"{f:,.2f}"
    except: return "—"

def kc(label, val, cls="", sub=""):
    sub_html = f'<div class="kpi-sub">{sub}</div>' if sub else ''
    return f'<div class="kpi-card {cls}"><div class="kpi-label">{label}</div><div class="kpi-value">{val}</div>{sub_html}</div>'

def to_xl(df):
    b = io.BytesIO()
    with pd.ExcelWriter(b, engine="openpyxl") as w: df.to_excel(w, index=False)
    return b.getvalue()

def _col_fmt_mode(col_name, is_num):
    """Decide el formato de una columna numérica por semántica del nombre:
    - 'money': columnas monetarias (valor, costo, precio, monto, usd, $) →
      2 decimales fijos siempre (ej. 17.00, 4,882.82).
    - 'int':   columnas de cantidad/stock (stock, cantidad, compras, ventas,
      muestras, dev, baja, cuadre, unidades) → entero sin decimales.
    - 'auto':  numérica sin match claro → int si valor redondo, sino 2 dec.
    - 'text':  no numérica (respeta str(v) con escape del caller).

    El orden importa: 'valor' se evalúa antes que 'compras'/'ventas' para
    que 'Valor Compras' caiga en money, no en int."""
    if not is_num:
        return "text"
    _lc = str(col_name).lower()
    if any(k in _lc for k in ("valor", "costo", "precio", "monto",
                               "importe", "usd", "$")):
        return "money"
    if any(k in _lc for k in ("stock", "cantidad", "compras", "ventas",
                               "muestras", "dev", "baja", "cuadre",
                               "unidades", "piezas", "registros",
                               "inventario")):
        return "int"
    return "auto"


def _fmt_cell(v, mode):
    """Formatea un valor según el modo devuelto por _col_fmt_mode.
    Devuelve siempre str; el caller se encarga de escapar HTML si aplica."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    if mode == "money":
        try:
            return f"{float(v):,.2f}"
        except (ValueError, TypeError):
            s = str(v)
            return "" if s in ("nan", "None", "NaN") else s
    if mode == "int":
        try:
            return f"{int(round(float(v))):,}"
        except (ValueError, TypeError):
            s = str(v)
            return "" if s in ("nan", "None", "NaN") else s
    if mode == "auto":
        try:
            f = float(v)
            if f == int(f):
                return f"{int(f):,}"
            return f"{f:,.2f}"
        except (ValueError, TypeError):
            s = str(v)
            return "" if s in ("nan", "None", "NaN") else s
    s = str(v)
    return "" if s in ("nan", "None", "NaN") else s


def _logo_data_uri():
    """Logo como data URI base64 para embeber en HTML standalone (sin
    dependencia de archivos externos al abrir el .html en otra máquina)."""
    _p = _get_logo_path()
    if not _p or not os.path.exists(_p):
        return ""
    try:
        import base64
        with open(_p, "rb") as f:
            _b64 = base64.b64encode(f.read()).decode("ascii")
        _ext = os.path.splitext(_p)[1].lower().lstrip(".") or "png"
        _mt = "image/jpeg" if _ext in ("jpg", "jpeg") else f"image/{_ext}"
        return f"data:{_mt};base64,{_b64}"
    except Exception:
        return ""


def to_html(df, title="Reporte"):
    """Export HTML standalone con logo embebido, header corporativo, fecha
    GMT-5, anchos proporcionales, zebra, wrap automático y estilos aptos
    para impresión (page-break-inside:avoid en filas, header repetido)."""
    import html as _h

    cols = list(df.columns)
    _is_num = {c: pd.api.types.is_numeric_dtype(df[c]) for c in cols}

    # Mismo heurístico de ancho que el PDF: Nombre/Descripción anchas,
    # Código/SKU/Fecha/Estado medias, numéricas angostas.
    _weights = []
    for c in cols:
        _lc = str(c).lower()
        if _is_num[c]:
            _weights.append(1.0)
        elif any(k in _lc for k in ("código", "codigo", "sku", "fecha",
                                      "estado", "tipo")):
            _weights.append(1.5)
        elif any(k in _lc for k in ("nombre", "descrip", "detalle",
                                      "observ", "producto")):
            _weights.append(3.5)
        else:
            _weights.append(2.0)
    _tw = sum(_weights) or 1
    _col_widths = [f"{w/_tw*100:.2f}%" for w in _weights]

    def _esc(x):
        s = str(x)
        return "" if s in ("nan", "None", "NaN") else _h.escape(s)

    _fmt_modes = {c: _col_fmt_mode(c, _is_num[c]) for c in cols}

    _colgroup = "".join(f'<col style="width:{w}"/>' for w in _col_widths)
    _hdrs = "".join(f"<th>{_esc(c)}</th>" for c in cols)

    _rows = []
    for _, r in df.iterrows():
        _cells = []
        for c in cols:
            mode = _fmt_modes[c]
            cls = "num" if mode in ("money", "int", "auto") else "txt"
            _cells.append(f'<td class="{cls}">{_h.escape(_fmt_cell(r[c], mode))}</td>')
        _rows.append(f"<tr>{''.join(_cells)}</tr>")

    _logo = _logo_data_uri()
    _logo_html = (f'<img src="{_logo}" alt="AutoSky" class="logo"/>'
                  if _logo else "")
    _fecha = _now_ec().strftime("%d/%m/%Y %H:%M")
    _summary = f"{len(df):,} registros &nbsp;·&nbsp; {len(cols)} columnas"
    _tesc = _esc(title)

    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"><title>{_tesc}</title>
<style>
*{{box-sizing:border-box}}
body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
     padding:24px 32px;background:#fff;color:#111827;margin:0}}
.hdr{{display:flex;align-items:center;border-bottom:2px solid #1E3A5F;
     padding-bottom:12px;margin-bottom:4px}}
.hdr .logo{{height:44px;margin-right:16px}}
.hdr .ttl{{flex:1}}
.hdr h1{{color:#1E3A5F;font-size:22px;margin:0 0 2px 0;font-weight:700}}
.hdr .sub{{color:#64748B;font-size:12px}}
.hdr .meta{{color:#64748B;font-size:11px;text-align:right;white-space:nowrap;line-height:1.4}}
table{{border-collapse:collapse;width:100%;table-layout:fixed;
      font-size:11px;margin-top:14px}}
thead th{{background:#1E3A5F;color:#fff;padding:8px 10px;text-align:center;
         font-weight:700;border:1px solid #1E3A5F;font-size:11px}}
tbody td{{padding:6px 10px;border:1px solid #E2E8F0;vertical-align:top;
         word-wrap:break-word;overflow-wrap:break-word;color:#111827}}
tbody tr:nth-child(even) td{{background:#F8FAFC}}
tbody tr:nth-child(odd) td{{background:#FFFFFF}}
tbody td.num{{text-align:right;font-variant-numeric:tabular-nums}}
tbody td.txt{{text-align:left}}
.ftr{{margin-top:16px;padding-top:10px;border-top:1px solid #E2E8F0;
     color:#64748B;font-size:10px;text-align:center}}
@media print{{
  body{{padding:12mm}}
  .hdr{{margin-bottom:2px}}
  table{{page-break-inside:auto}} tr{{page-break-inside:avoid}}
  thead{{display:table-header-group}}
}}
</style></head><body>
<div class="hdr">
  {_logo_html}
  <div class="ttl"><h1>{_tesc}</h1><div class="sub">{_summary}</div></div>
  <div class="meta">{_fecha}<br/>AutoSky Inventario</div>
</div>
<table><colgroup>{_colgroup}</colgroup>
<thead><tr>{_hdrs}</tr></thead>
<tbody>{''.join(_rows)}</tbody></table>
<div class="ftr">AutoSky Inventario · Generado {_fecha}</div>
</body></html>""".encode()

def _export_resumen_excel(pivot_df, ubic_cols, title="Resumen Toma Física"):
    """Export ejecutivo en Excel con formato listo para imprimir:
    - Título y metadatos en header
    - Tabla con colores corporativos, zebra, bordes
    - Fila TOTAL con fórmulas SUM por columna
    - Freeze panes en SKU+Producto (izquierda) y 5 filas arriba
    - Anchos adaptados para impresión en paisaje"""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen"

    # Colores corporativos
    C_HDR   = PatternFill("solid", fgColor="1E3A5F")  # azul oscuro
    C_SUB   = PatternFill("solid", fgColor="0EA5E9")  # azul claro
    C_ZEB1  = PatternFill("solid", fgColor="FFFFFF")
    C_ZEB2  = PatternFill("solid", fgColor="F8FAFC")
    C_TOT   = PatternFill("solid", fgColor="DBEAFE")  # azul muy claro total
    C_TOTB  = PatternFill("solid", fgColor="1E40AF")  # azul oscuro total fila
    F_HDR   = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    F_TIT   = Font(bold=True, size=16, color="1E3A5F", name="Calibri")
    F_SUB   = Font(bold=False, size=10, color="64748B", name="Calibri")
    F_CELL  = Font(size=10, color="111827", name="Calibri")
    F_NUM   = Font(size=10, color="111827", name="Calibri")
    F_TOT   = Font(bold=True, size=10, color="1E40AF", name="Calibri")
    F_TOTG  = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    _side   = Side(style="thin", color="CBD5E1")
    BRD     = Border(_side, _side, _side, _side)
    _tside  = Side(style="medium", color="1E40AF")
    TOT_BRD = Border(_tside, _tside, _tside, _tside)

    cols = list(pivot_df.columns)
    ncols = len(cols)
    last_col_letter = get_column_letter(ncols)

    # Logo en A1 (si existe). Ocupa 4 filas (rango A1:B4). Texto a partir de C1.
    _logo_path = _get_logo_path()
    _text_start_col = "A"
    _has_logo = False
    if _logo_path and os.path.exists(_logo_path):
        try:
            from openpyxl.drawing.image import Image as _XLImage
            _img = _XLImage(_logo_path)
            # Tamaño para ocupar aprox 2 columnas × 4 filas
            _img.width  = 150   # pixels
            _img.height = 90
            _img.anchor = "A1"
            ws.add_image(_img)
            _has_logo = True
            _text_start_col = "C"
            # Ensanchar A y B para contener el logo
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 14
        except Exception:
            _has_logo = False

    # Altura de filas 1-3 si hay logo (para que quepa)
    if _has_logo:
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 22

    # Fila 1: Título (empieza en C si hay logo, sino en A)
    _tit_range = f"{_text_start_col}1:{last_col_letter}1"
    ws.merge_cells(_tit_range)
    ws[f"{_text_start_col}1"] = title
    ws[f"{_text_start_col}1"].font = F_TIT
    ws[f"{_text_start_col}1"].alignment = Alignment(horizontal="left", vertical="center")
    if not _has_logo:
        ws.row_dimensions[1].height = 26

    # Fila 2: fecha
    _fecha_range = f"{_text_start_col}2:{last_col_letter}2"
    ws.merge_cells(_fecha_range)
    ws[f"{_text_start_col}2"] = f"Generado: {_now_ec().strftime('%d/%m/%Y %H:%M')}"
    ws[f"{_text_start_col}2"].font = F_SUB
    ws[f"{_text_start_col}2"].alignment = Alignment(horizontal="left", vertical="center")

    # Fila 3: estadísticas
    _n_skus = len(pivot_df)
    _n_ub   = len(ubic_cols)
    _total_u = int(pivot_df["Total"].sum()) if "Total" in pivot_df.columns else 0
    _stats_range = f"{_text_start_col}3:{last_col_letter}3"
    ws.merge_cells(_stats_range)
    ws[f"{_text_start_col}3"] = (f"{_n_skus:,} SKUs · {_n_ub} ubicaciones · "
                                  f"{_total_u:,} unidades totales contadas")
    ws[f"{_text_start_col}3"].font = F_SUB
    ws[f"{_text_start_col}3"].alignment = Alignment(horizontal="left", vertical="center")

    # Fila 4 vacía (separador — más espacio si hay logo que termina aquí)
    ws.row_dimensions[4].height = 8 if _has_logo else 6

    # Fila 5: encabezados
    HDR_ROW = 5
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(HDR_ROW, ci, str(c))
        cell.fill = C_HDR
        cell.font = F_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BRD
    ws.row_dimensions[HDR_ROW].height = 28

    # Filas de datos (zebra)
    DATA_START = HDR_ROW + 1
    for ri, (_, row) in enumerate(pivot_df.iterrows(), DATA_START):
        _fill = C_ZEB1 if ri % 2 == 1 else C_ZEB2
        for ci, c in enumerate(cols, 1):
            val = row[c]
            cell = ws.cell(ri, ci, val if pd.notna(val) else None)
            cell.fill = _fill
            cell.border = BRD
            if c in ("Código Producto", "Nombre Producto"):
                cell.font = F_CELL
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c == "Total":
                cell.font = F_TOT
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            else:
                cell.font = F_NUM
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"

    # Fila TOTAL con fórmulas SUM por columna numérica
    TOT_ROW = DATA_START + len(pivot_df)
    for ci, c in enumerate(cols, 1):
        cell = ws.cell(TOT_ROW, ci)
        cell.fill = C_TOTB
        cell.font = F_TOTG
        cell.border = TOT_BRD
        if ci == 1:  # Primera columna: etiqueta
            cell.value = "TOTAL GENERAL"
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif c == "Nombre Producto":
            cell.value = ""
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            letter = get_column_letter(ci)
            cell.value = f"=SUM({letter}{DATA_START}:{letter}{TOT_ROW-1})"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
    ws.row_dimensions[TOT_ROW].height = 24

    # Anchos de columnas
    for ci, c in enumerate(cols, 1):
        letter = get_column_letter(ci)
        if c == "Código Producto":
            ws.column_dimensions[letter].width = 14
        elif c == "Nombre Producto":
            ws.column_dimensions[letter].width = 42
        elif c == "Total":
            ws.column_dimensions[letter].width = 14
        else:
            ws.column_dimensions[letter].width = max(13, min(22, len(str(c)) + 4))

    # Freeze panes: tras las 2 primeras columnas y tras encabezados
    _freeze_col = "C"  # tras Nombre Producto
    if "Nombre Producto" not in cols:
        _freeze_col = "B"  # solo tras SKU
    ws.freeze_panes = f"{_freeze_col}{DATA_START}"

    # Configuración de impresión: paisaje + ajustar a una página de ancho
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.5
    ws.page_margins.top = ws.page_margins.bottom = 0.7
    ws.print_title_rows = f"{HDR_ROW}:{HDR_ROW}"  # repetir header en cada página

    # Encabezado y pie de impresión
    ws.oddHeader.left.text = title
    ws.oddHeader.left.size = 10
    ws.oddHeader.left.color = "1E3A5F"
    ws.oddHeader.right.text = _now_ec().strftime("%d/%m/%Y %H:%M")
    ws.oddHeader.right.size = 9
    ws.oddHeader.right.color = "64748B"
    ws.oddFooter.center.text = "Página &P de &N"
    ws.oddFooter.center.size = 9
    ws.oddFooter.center.color = "64748B"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_resumen_pdf(pivot_df, ubic_cols, title="Resumen Toma Física"):
    """Export ejecutivo en PDF: landscape, tabla con totales al final,
    header corporativo y pie con paginación."""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer)
    except ImportError:
        return None

    buf = io.BytesIO()
    C_HDR   = colors.HexColor("#1E3A5F")
    C_TXT   = colors.HexColor("#111827")
    C_MUT   = colors.HexColor("#64748B")
    C_SUB   = colors.HexColor("#F1F5F9")
    C_ZEB   = colors.HexColor("#F8FAFC")
    C_TOTBG = colors.HexColor("#1E40AF")
    C_BRD   = colors.HexColor("#CBD5E1")

    _logo_path = _get_logo_path()

    def _fp(canvas, doc):
        canvas.saveState()
        _page_w, _page_h = landscape(A4)
        # Logo en la esquina superior izquierda (si existe)
        _text_x = 1.5*cm
        if _logo_path and os.path.exists(_logo_path):
            try:
                from reportlab.lib.utils import ImageReader
                _img_reader = ImageReader(_logo_path)
                _iw, _ih = _img_reader.getSize()
                _target_h = 1.1*cm
                _target_w = _iw * (_target_h / _ih)
                canvas.drawImage(_logo_path, 1.5*cm, _page_h - _target_h - 0.3*cm,
                                  width=_target_w, height=_target_h,
                                  preserveAspectRatio=True, mask="auto")
                _text_x = 1.5*cm + _target_w + 0.4*cm
            except Exception:
                pass
        # Header: título + fecha a la derecha
        canvas.setFont("Helvetica-Bold", 10); canvas.setFillColor(C_HDR)
        canvas.drawString(_text_x, _page_h-0.8*cm, title)
        canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
        canvas.drawRightString(_page_w-1.5*cm, _page_h-0.8*cm,
                                _now_ec().strftime("%d/%m/%Y %H:%M"))
        # Línea horizontal
        canvas.setStrokeColor(C_HDR); canvas.setLineWidth(1.2)
        canvas.line(1.5*cm, _page_h-1.5*cm, _page_w-1.5*cm, _page_h-1.5*cm)
        # Footer: paginación
        canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
        canvas.drawCentredString(_page_w/2, 0.7*cm,
                                  f"Página {doc.page}  ·  AutoSky Inventario")
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                             topMargin=2.0*cm, bottomMargin=1.2*cm,
                             leftMargin=1.5*cm, rightMargin=1.5*cm,
                             title=title)
    styles = getSampleStyleSheet()
    _sty_h  = ParagraphStyle("h", fontSize=16, textColor=C_HDR,
                              fontName="Helvetica-Bold", spaceAfter=4)
    _sty_s  = ParagraphStyle("s", fontSize=9, textColor=C_MUT,
                              fontName="Helvetica", spaceAfter=12)

    _n_skus = len(pivot_df)
    _n_ub   = len(ubic_cols)
    _total_u = int(pivot_df["Total"].sum()) if "Total" in pivot_df.columns else 0
    sub = (f"{_n_skus:,} SKUs &nbsp;·&nbsp; {_n_ub} ubicaciones &nbsp;·&nbsp; "
           f"<b>{_total_u:,} unidades totales</b>")

    elements = [
        Paragraph(title, _sty_h),
        Paragraph(sub, _sty_s),
    ]

    cols = list(pivot_df.columns)
    # Anchos proporcionales para A4 landscape (se dedica más espacio a las
    # ubicaciones que antes, y el nombre producto reduce un poco)
    PW = landscape(A4)[0] - 3*cm
    _wide_cols = {"Código Producto": 0.10, "Nombre Producto": 0.22, "Total": 0.07}
    _n_ubic_cols = sum(1 for c in cols if c not in _wide_cols)
    _remaining = 1 - sum(_wide_cols.get(c, 0) for c in cols if c in _wide_cols)
    _each = _remaining / _n_ubic_cols if _n_ubic_cols else 0
    cws = [PW * _wide_cols.get(c, _each) for c in cols]

    # Estilos Paragraph para wrap de texto
    _hdr_style = ParagraphStyle(
        "pdf_hdr", fontSize=8, textColor=colors.white,
        fontName="Helvetica-Bold", alignment=1,  # CENTER
        leading=10, wordWrap="CJK",  # wordWrap CJK permite cortar entre letras
    )
    _sku_style = ParagraphStyle(
        "pdf_sku", fontSize=8, textColor=C_TXT,
        fontName="Helvetica", alignment=0,  # LEFT
        leading=10, wordWrap="LTR",
    )
    _name_style = ParagraphStyle(
        "pdf_name", fontSize=7.5, textColor=C_TXT,
        fontName="Helvetica", alignment=0,  # LEFT
        leading=9, wordWrap="LTR",
    )

    # Construir data con Paragraphs donde hace falta wrap
    def _fmt_cell(v, c):
        if pd.isna(v) or v is None: return ""
        if c in ("Código Producto", "Nombre Producto"):
            return str(v)
        try:
            return f"{int(round(float(v))):,}"
        except (ValueError, TypeError):
            return str(v)

    # Header row: Paragraph permite wrap del texto
    _header_row = [Paragraph(str(c), _hdr_style) for c in cols]
    data = [_header_row]

    for _, row in pivot_df.iterrows():
        _row_cells = []
        for c in cols:
            val = _fmt_cell(row[c], c)
            if c == "Código Producto":
                _row_cells.append(Paragraph(val, _sku_style))
            elif c == "Nombre Producto":
                _row_cells.append(Paragraph(val, _name_style))
            else:
                # Números: str simple, alineados a la derecha por TableStyle
                _row_cells.append(val)
        data.append(_row_cells)

    # Fila TOTAL: sumar columnas numéricas
    _tot = []
    for c in cols:
        if c == cols[0]:
            _tot.append("TOTAL GENERAL")
        elif c == "Nombre Producto":
            _tot.append("")
        else:
            try:
                s = pd.to_numeric(pivot_df[c], errors="coerce").fillna(0).sum()
                _tot.append(f"{int(round(s)):,}")
            except Exception:
                _tot.append("")
    data.append(_tot)

    ts = TableStyle([
        # Encabezado — más alto para acomodar texto envuelto
        ("BACKGROUND", (0,0), (-1,0), C_HDR),
        ("VALIGN",     (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,0), 8),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
        ("LEFTPADDING",(0,0), (-1,0), 4),
        ("RIGHTPADDING",(0,0), (-1,0), 4),
        # Body general
        ("FONTSIZE",   (0,0), (-1,-1), 8),
        ("VALIGN",     (0,1), (-1,-1), "MIDDLE"),
        ("FONTNAME",   (0,1), (-1,-2), "Helvetica"),
        # Alineación numérica (columnas a partir de la 3: Total + Ubicaciones)
        ("ALIGN",      (2,1), (-1,-2), "RIGHT"),
        ("LEFTPADDING",(0,1), (-1,-2), 4),
        ("RIGHTPADDING",(0,1), (-1,-2), 4),
        ("TOPPADDING", (0,1), (-1,-2), 4),
        ("BOTTOMPADDING", (0,1), (-1,-2), 4),
        ("GRID",       (0,0), (-1,-1), 0.4, C_BRD),
        # Zebra
        ("ROWBACKGROUNDS", (0,1), (-1,-2), [colors.white, C_ZEB]),
        # Fila TOTAL
        ("BACKGROUND", (0,-1), (-1,-1), C_TOTBG),
        ("TEXTCOLOR",  (0,-1), (-1,-1), colors.white),
        ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
        ("FONTSIZE",   (0,-1), (-1,-1), 9),
        ("ALIGN",      (2,-1), (-1,-1), "RIGHT"),
        ("ALIGN",      (0,-1), (0,-1),  "LEFT"),
        ("TOPPADDING", (0,-1), (-1,-1), 8),
        ("BOTTOMPADDING", (0,-1), (-1,-1), 8),
    ])

    t = Table(data, colWidths=cws, repeatRows=1)
    t.setStyle(ts)
    elements.append(t)

    doc.build(elements, onFirstPage=_fp, onLaterPages=_fp)
    return buf.getvalue()


def to_html_mobile(df, fecha_corte_txt):
    """HTML standalone responsive optimizado para visualización móvil:
    una card por SKU con Código/Nombre/Stock, buscador client-side (JS),
    logo embebido en base64. Columnas esperadas: SKU, Nombre Producto,
    Stock Disponible. Opcional: 'En Tránsito' y 'Fechas Tránsito' — si
    existen, se muestra una línea extra con las cantidades en camino y
    las fechas tentativas. Devuelve bytes UTF-8."""
    from html import escape as _esc
    _logo = _logo_data_uri()
    _now_txt = _now_ec().strftime("%d/%m/%Y %H:%M GMT-5")
    _n_total = len(df)
    _n_pos   = int((df["Stock Disponible"] > 0).sum())
    _n_zero  = int((df["Stock Disponible"] == 0).sum())
    _n_neg   = int((df["Stock Disponible"] < 0).sum())
    _has_transito = "En Tránsito" in df.columns
    _n_transito = int((df["En Tránsito"] > 0).sum()) if _has_transito else 0
    _u_transito = int(df["En Tránsito"].sum()) if _has_transito else 0

    _cards = []
    for _, row in df.iterrows():
        _sku  = _esc(str(row["SKU"]))
        _name = _esc(str(row["Nombre Producto"]))
        _stk  = int(row["Stock Disponible"])
        _cls  = "pos" if _stk > 0 else ("neg" if _stk < 0 else "zero")
        _tr_qty = int(row["En Tránsito"]) if _has_transito else 0
        _tr_fechas = _esc(str(row.get("Fechas Tránsito",""))) if _has_transito else ""
        _srch_extra = (" " + str(row.get("Fechas Tránsito",""))).lower() if _has_transito else ""
        _srch = _esc((str(row["SKU"]) + " " + str(row["Nombre Producto"]) + _srch_extra)
                     .lower())
        _transito_html = ""
        if _tr_qty > 0:
            _transito_html = (
                f'<div class="transito">🚢 En tránsito: '
                f'<b>{_tr_qty:,}</b> u'
                f'{(" · " + _tr_fechas) if _tr_fechas else ""}'
                f'</div>'
            )
        _cards.append(
            f'<div class="sku" data-search="{_srch}">'
            f'<div class="sku-info">'
            f'<div class="sku-code">{_sku}</div>'
            f'<div class="sku-name">{_name}</div>'
            f'{_transito_html}'
            f'</div>'
            f'<div class="stock {_cls}">{_stk:,}</div>'
            f'</div>'
        )
    _logo_tag = (f'<img src="{_logo}" alt="AutoSky">' if _logo else '')
    _transito_summary = ""
    if _has_transito and _u_transito > 0:
        _transito_summary = (
            f'<div class="card tr"><strong>{_u_transito:,}</strong>'
            f'<span>En tránsito ({_n_transito} SKU)</span></div>'
        )

    html = f"""<!DOCTYPE html>
<html lang="es"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5">
<title>Stock Ejecutivo · AutoSky</title>
<style>
  *{{box-sizing:border-box}}
  body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
       background:#f0f9ff;color:#0f172a;margin:0;padding:0;
       -webkit-font-smoothing:antialiased}}
  .header{{background:linear-gradient(135deg,#0ea5e9,#1e3a5f);color:#fff;
          padding:22px 16px 18px;text-align:center;
          box-shadow:0 4px 14px rgba(30,58,95,.18)}}
  .header img{{max-height:44px;background:#fff;padding:5px 9px;border-radius:8px}}
  .header h1{{margin:12px 0 4px;font-size:22px;font-weight:700}}
  .header p{{margin:0;font-size:13px;opacity:.92}}
  .wrap{{max-width:720px;margin:0 auto;padding:16px}}
  .summary{{display:flex;gap:8px;margin:4px 0 16px}}
  .summary .card{{flex:1;background:#fff;border-radius:12px;padding:12px 8px;
                  text-align:center;box-shadow:0 2px 6px rgba(0,0,0,.05)}}
  .summary strong{{font-size:22px;font-weight:700;display:block;line-height:1.1}}
  .summary .card.pos strong{{color:#059669}}
  .summary .card.zero strong{{color:#94a3b8}}
  .summary .card.neg strong{{color:#dc2626}}
  .summary .card.tot strong{{color:#0ea5e9}}
  .summary .card.tr strong{{color:#b45309}}
  .summary span{{font-size:11px;color:#64748b;text-transform:uppercase;
                 letter-spacing:.5px;margin-top:4px;display:block}}
  .search{{position:sticky;top:0;z-index:10;background:#f0f9ff;
           padding:8px 0 6px}}
  .search input{{width:100%;padding:12px 14px;border:1px solid #cbd5e1;
                 border-radius:10px;font-size:15px;outline:none;
                 background:#fff;transition:border-color .15s}}
  .search input:focus{{border-color:#0ea5e9;
                       box-shadow:0 0 0 3px rgba(14,165,233,.15)}}
  .sku{{background:#fff;border-radius:12px;padding:13px 14px;margin-bottom:7px;
        box-shadow:0 1px 3px rgba(0,0,0,.05);display:flex;
        justify-content:space-between;align-items:center;gap:10px}}
  .sku-info{{flex:1;min-width:0}}
  .sku-code{{font-weight:600;color:#0ea5e9;font-size:14px;letter-spacing:.3px}}
  .sku-name{{color:#475569;font-size:13px;margin-top:2px;
             overflow:hidden;text-overflow:ellipsis;
             display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical}}
  .transito{{color:#b45309;font-size:11.5px;margin-top:4px;
             background:#fef3c7;padding:3px 7px;border-radius:6px;
             display:inline-block;font-weight:500}}
  .transito b{{font-weight:700}}
  .stock{{font-size:22px;font-weight:700;padding:0 4px;min-width:60px;
          text-align:right;line-height:1}}
  .stock.pos{{color:#059669}}
  .stock.zero{{color:#94a3b8}}
  .stock.neg{{color:#dc2626}}
  .empty{{text-align:center;color:#94a3b8;padding:30px 0;font-size:14px}}
  .footer{{text-align:center;color:#64748b;font-size:11px;padding:20px 10px 26px}}
  @media (min-width:700px){{
    .sku-name{{font-size:14px}}
    .stock{{font-size:24px}}
  }}
</style></head>
<body>
  <div class="header">
    {_logo_tag}
    <h1>Stock Ejecutivo</h1>
    <p>Actualizado hasta {fecha_corte_txt} · {_n_total} SKUs</p>
  </div>
  <div class="wrap">
    <div class="summary">
      <div class="card pos"><strong>{_n_pos}</strong><span>Con stock</span></div>
      <div class="card zero"><strong>{_n_zero}</strong><span>En cero</span></div>
      <div class="card neg"><strong>{_n_neg}</strong><span>Negativo</span></div>
      <div class="card tot"><strong>{_n_total}</strong><span>Total</span></div>
      {_transito_summary}
    </div>
    <div class="search">
      <input type="text" id="q" placeholder="🔍 Buscar SKU o producto…"
             oninput="filterSku()" autocomplete="off">
    </div>
    <div id="list">
      {"".join(_cards)}
    </div>
    <div id="empty" class="empty" style="display:none">
      Sin resultados para tu búsqueda
    </div>
    <div class="footer">AutoSky Inventario · Generado {_now_txt}</div>
  </div>
<script>
function filterSku(){{
  var q=document.getElementById('q').value.toLowerCase().trim();
  var items=document.querySelectorAll('.sku');
  var shown=0;
  items.forEach(function(el){{
    var match=!q || el.dataset.search.indexOf(q)!==-1;
    el.style.display=match?'':'none';
    if(match) shown++;
  }});
  document.getElementById('empty').style.display=shown?'none':'block';
}}
</script>
</body></html>"""
    return html.encode("utf-8")


def to_pdf(df, title="Reporte"):
    """Export genérico a PDF (landscape A4) con logo corporativo, header con
    fecha GMT-5, footer con paginación, wrap de texto en columnas no numéricas
    y anchos proporcionales (más espacio para Nombre/Descripción)."""
    try:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph)
    except ImportError:
        return None

    try:
        buf = io.BytesIO()
        C_HDR = colors.HexColor("#1E3A5F")
        C_TXT = colors.HexColor("#111827")
        C_MUT = colors.HexColor("#64748B")
        C_ZEB = colors.HexColor("#F8FAFC")
        C_BRD = colors.HexColor("#CBD5E1")

        _logo_path = _get_logo_path()

        def _fp(canvas, doc):
            canvas.saveState()
            _pw, _ph = landscape(A4)
            _text_x = 1.5*cm
            if _logo_path and os.path.exists(_logo_path):
                try:
                    from reportlab.lib.utils import ImageReader
                    _ir = ImageReader(_logo_path)
                    _iw, _ih = _ir.getSize()
                    _target_h = 1.1*cm
                    _target_w = _iw * (_target_h / _ih)
                    canvas.drawImage(_logo_path, 1.5*cm,
                                      _ph - _target_h - 0.3*cm,
                                      width=_target_w, height=_target_h,
                                      preserveAspectRatio=True, mask="auto")
                    _text_x = 1.5*cm + _target_w + 0.4*cm
                except Exception:
                    pass
            canvas.setFont("Helvetica-Bold", 10); canvas.setFillColor(C_HDR)
            canvas.drawString(_text_x, _ph - 0.8*cm, title)
            canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
            canvas.drawRightString(_pw - 1.5*cm, _ph - 0.8*cm,
                                    _now_ec().strftime("%d/%m/%Y %H:%M"))
            canvas.setStrokeColor(C_HDR); canvas.setLineWidth(1.2)
            canvas.line(1.5*cm, _ph - 1.5*cm, _pw - 1.5*cm, _ph - 1.5*cm)
            canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
            canvas.drawCentredString(_pw/2, 0.7*cm,
                                      f"Página {doc.page}  ·  AutoSky Inventario")
            canvas.restoreState()

        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                 topMargin=2.0*cm, bottomMargin=1.2*cm,
                                 leftMargin=1.5*cm, rightMargin=1.5*cm,
                                 title=title)

        cols = list(df.columns)
        _is_num = {c: pd.api.types.is_numeric_dtype(df[c]) for c in cols}

        # Pesos proporcionales de ancho: columnas de nombre/descripción reciben
        # más espacio, SKU/fecha/numéricas menos. Así `Nombre Producto` ya no
        # se solapa. Orden importa: código/sku antes que "producto".
        _weights = []
        for c in cols:
            _lc = str(c).lower()
            if _is_num[c]:
                _weights.append(1.0)
            elif any(k in _lc for k in ("código", "codigo", "sku", "fecha",
                                          "estado", "tipo")):
                _weights.append(1.5)
            elif any(k in _lc for k in ("nombre", "descrip", "detalle",
                                          "observ", "producto")):
                _weights.append(3.5)
            else:
                _weights.append(2.0)
        _tw = sum(_weights) or 1
        PW = landscape(A4)[0] - 3*cm
        cws = [PW * (w / _tw) for w in _weights]

        _hdr_style = ParagraphStyle("pdf_hdr", fontSize=8, textColor=colors.white,
                                     fontName="Helvetica-Bold", alignment=1,
                                     leading=10, wordWrap="CJK")
        _txt_style = ParagraphStyle("pdf_txt", fontSize=7.5, textColor=C_TXT,
                                     fontName="Helvetica", alignment=0,
                                     leading=9, wordWrap="LTR")

        _fmt_modes = {c: _col_fmt_mode(c, _is_num[c]) for c in cols}

        _header_row = [Paragraph(str(c), _hdr_style) for c in cols]
        data = [_header_row]
        for _, row in df.iterrows():
            _rc = []
            for c in cols:
                mode = _fmt_modes[c]
                s = _fmt_cell(row[c], mode)
                if mode == "text":
                    _rc.append(Paragraph(s, _txt_style) if s else "")
                else:
                    _rc.append(s)
            data.append(_rc)

        ts = [
            ("BACKGROUND",    (0,0), (-1,0),  C_HDR),
            ("VALIGN",        (0,0), (-1,0),  "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,0),  6),
            ("BOTTOMPADDING", (0,0), (-1,0),  6),
            ("LEFTPADDING",   (0,0), (-1,0),  3),
            ("RIGHTPADDING",  (0,0), (-1,0),  3),
            ("FONTSIZE",      (0,1), (-1,-1), 7.5),
            ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("LEFTPADDING",   (0,1), (-1,-1), 3),
            ("RIGHTPADDING",  (0,1), (-1,-1), 3),
            ("TOPPADDING",    (0,1), (-1,-1), 3),
            ("BOTTOMPADDING", (0,1), (-1,-1), 3),
            ("GRID",          (0,0), (-1,-1), 0.3, C_BRD),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, C_ZEB]),
        ]
        for i, c in enumerate(cols):
            if _is_num[c]:
                ts.append(("ALIGN", (i,1), (i,-1), "RIGHT"))

        t = Table(data, colWidths=cws, repeatRows=1)
        t.setStyle(TableStyle(ts))

        doc.build([t], onFirstPage=_fp, onLaterPages=_fp)
        return buf.getvalue()
    except Exception:
        return None

def to_pdf_ejecutivo(df, title="Stock Ejecutivo", subtitle=""):
    """PDF ejecutivo vertical A4, tabla compacta pensada para caber en una
    sola hoja. Usado por la pestaña Exportar Portal: 3 columnas
    (SKU / Nombre Producto / Stock Disponible). El tamaño de fuente se
    reduce automáticamente si el total de filas excede lo que cabe al
    tamaño base, para maximizar la probabilidad de un solo page."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph)
    except ImportError:
        return None

    try:
        buf = io.BytesIO()
        C_HDR = colors.HexColor("#1E3A5F")
        C_TXT = colors.HexColor("#111827")
        C_MUT = colors.HexColor("#64748B")
        C_ZEB = colors.HexColor("#F8FAFC")
        C_BRD = colors.HexColor("#CBD5E1")

        _logo_path = _get_logo_path()

        def _fp(canvas, doc):
            canvas.saveState()
            _pw, _ph = A4
            _text_x = 1.5*cm
            if _logo_path and os.path.exists(_logo_path):
                try:
                    from reportlab.lib.utils import ImageReader
                    _ir = ImageReader(_logo_path)
                    _iw, _ih = _ir.getSize()
                    _target_h = 1.1*cm
                    _target_w = _iw * (_target_h / _ih)
                    canvas.drawImage(_logo_path, 1.2*cm,
                                      _ph - _target_h - 0.3*cm,
                                      width=_target_w, height=_target_h,
                                      preserveAspectRatio=True, mask="auto")
                    _text_x = 1.2*cm + _target_w + 0.4*cm
                except Exception:
                    pass
            canvas.setFont("Helvetica-Bold", 10); canvas.setFillColor(C_HDR)
            canvas.drawString(_text_x, _ph - 0.7*cm, title)
            if subtitle:
                canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
                canvas.drawString(_text_x, _ph - 1.15*cm, subtitle)
            canvas.setFont("Helvetica", 8); canvas.setFillColor(C_MUT)
            canvas.drawRightString(_pw - 1.2*cm, _ph - 0.7*cm,
                                    _now_ec().strftime("%d/%m/%Y %H:%M"))
            canvas.setStrokeColor(C_HDR); canvas.setLineWidth(1.0)
            canvas.line(1.2*cm, _ph - 1.5*cm, _pw - 1.2*cm, _ph - 1.5*cm)
            canvas.setFont("Helvetica", 7); canvas.setFillColor(C_MUT)
            canvas.drawCentredString(_pw/2, 0.6*cm,
                                      f"Página {doc.page}  ·  AutoSky Inventario")
            canvas.restoreState()

        # Escalar fuente según número de filas para maximizar chance de 1 page.
        # Los nombres largos envuelven a 2 líneas: de ahí la agresividad a n>=55.
        n = len(df)
        if   n <= 28:  fs_body, row_pad = 9.0, 2.0
        elif n <= 40:  fs_body, row_pad = 7.8, 1.4
        elif n <= 55:  fs_body, row_pad = 5.8, 0.6
        elif n <= 75:  fs_body, row_pad = 5.0, 0.4
        elif n <= 95:  fs_body, row_pad = 4.4, 0.3
        else:          fs_body, row_pad = 4.0, 0.2
        fs_head = fs_body + 0.5

        doc = SimpleDocTemplate(buf, pagesize=A4,
                                 topMargin=2.0*cm, bottomMargin=1.1*cm,
                                 leftMargin=1.2*cm, rightMargin=1.2*cm,
                                 title=title)

        cols = list(df.columns)
        _is_num = {c: pd.api.types.is_numeric_dtype(df[c]) for c in cols}

        # Pesos: el Nombre Producto se lleva la mayor parte del ancho
        _weights = []
        for c in cols:
            _lc = str(c).lower()
            if _is_num[c]:
                _weights.append(1.2)
            elif any(k in _lc for k in ("sku", "código", "codigo")):
                _weights.append(1.3)
            elif any(k in _lc for k in ("nombre", "producto", "descrip")):
                _weights.append(4.2)
            else:
                _weights.append(1.8)
        _tw = sum(_weights) or 1
        PW = A4[0] - 2.4*cm
        cws = [PW * (w / _tw) for w in _weights]

        _hdr_style = ParagraphStyle("pdf_hdr", fontSize=fs_head,
                                     textColor=colors.white,
                                     fontName="Helvetica-Bold", alignment=1,
                                     leading=fs_head+1, wordWrap="CJK")
        _txt_style = ParagraphStyle("pdf_txt", fontSize=fs_body,
                                     textColor=C_TXT, fontName="Helvetica",
                                     alignment=0, leading=fs_body+1,
                                     wordWrap="LTR")

        _fmt_modes = {c: _col_fmt_mode(c, _is_num[c]) for c in cols}
        _header_row = [Paragraph(str(c), _hdr_style) for c in cols]
        data = [_header_row]
        for _, row in df.iterrows():
            _rc = []
            for c in cols:
                mode = _fmt_modes[c]
                s = _fmt_cell(row[c], mode)
                if mode == "text":
                    _rc.append(Paragraph(s, _txt_style) if s else "")
                else:
                    _rc.append(s)
            data.append(_rc)

        ts = [
            ("BACKGROUND",    (0,0), (-1,0),  C_HDR),
            ("VALIGN",        (0,0), (-1,0),  "MIDDLE"),
            ("TOPPADDING",    (0,0), (-1,0),  3),
            ("BOTTOMPADDING", (0,0), (-1,0),  3),
            ("LEFTPADDING",   (0,0), (-1,0),  2),
            ("RIGHTPADDING",  (0,0), (-1,0),  2),
            ("FONTSIZE",      (0,1), (-1,-1), fs_body),
            ("VALIGN",        (0,1), (-1,-1), "MIDDLE"),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("LEFTPADDING",   (0,1), (-1,-1), 2),
            ("RIGHTPADDING",  (0,1), (-1,-1), 2),
            ("TOPPADDING",    (0,1), (-1,-1), row_pad),
            ("BOTTOMPADDING", (0,1), (-1,-1), row_pad),
            ("GRID",          (0,0), (-1,-1), 0.25, C_BRD),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, C_ZEB]),
        ]
        for i, c in enumerate(cols):
            if _is_num[c]:
                ts.append(("ALIGN", (i,1), (i,-1), "RIGHT"))

        t = Table(data, colWidths=cws, repeatRows=1)
        t.setStyle(TableStyle(ts))

        doc.build([t], onFirstPage=_fp, onLaterPages=_fp)
        return buf.getvalue()
    except Exception:
        return None


def dl3(df, name, key):
    c1,c2,c3=st.columns(3)
    with c1: st.download_button("📊 Excel",to_xl(df),f"{name}.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=f"{key}_xl",width='stretch')
    with c2: st.download_button("🌐 HTML",to_html(df,name),f"{name}.html","text/html",key=f"{key}_htm",width='stretch')
    with c3:
        pdf=to_pdf(df,name)
        if pdf: st.download_button("📄 PDF",pdf,f"{name}.pdf","application/pdf",key=f"{key}_pdf",width='stretch')
        else: st.button("📄 PDF",disabled=True,key=f"{key}_pdfx",width='stretch')

def _strict_opts(options, query, format_func=None, keep_selected=None):
    """Filtra opciones por substring estricto (no fuzzy). Streamlit usa fuzzy
    por subsecuencia en el dropdown interno (ej: 'AKE' matchea 'FRANK KEMPIN').

    Para evitar esto, si no hay query devuelve [] (más keep_selected) — fuerza
    al usuario a escribir en el text_input EXTERNO, cuyo filtrado es estricto.
    Si ya hay query, devuelve los items que realmente contienen ese texto."""
    fx = format_func if format_func else str
    if query:
        ql = query.lower()
        filt = [o for o in options if ql in fx(o).lower()]
    else:
        filt = []  # Sin búsqueda, sin opciones — se obliga a usar el buscador
    if keep_selected:
        for c in keep_selected:
            if c not in filt: filt.append(c)
    return filt

def pfilt(df, txt, cols=("Código Producto","Nombre Producto")):
    if not txt: return df
    s=txt.lower()
    m=None
    for c in cols:
        if c in df.columns:
            mk=df[c].fillna("").astype(str).str.lower().str.contains(s,regex=False)
            m=mk if m is None else m|mk
    return df[m] if m is not None else df


# Lista BLANCA de columnas monetarias (2 decimales).
# Todo lo demás en columnas numéricas se trata como entero (unidades).
_MONEY_COLS = {
    # valores financieros
    "valor inventario","valor compras","valor ventas","valor stock",
    "valor unitario","valor unitario promedio","valor total","v.unit",
    "v.total","valor inv.",
    # costos
    "costo","costo prom.","costo promedio","costo ponderado",
    # precios
    "pvp","pvp promedio","precio","precio promedio","p. unitario",
    # márgenes y rentabilidad
    "margen","margen%","rentabilidad","utilidad",
    # ratios con decimales
    "rotación","días inv.","cons/día","p.reorden","sug.compra",
    "marítimo","aéreo",
}

def _is_int_col(col_name):
    """Retorna True si la columna debe mostrarse como entero (sin decimales)."""
    return col_name.lower().strip() not in _MONEY_COLS

def tbl(df, nc=None, uid="t"):
    """Tabla HTML con zoom y totales. Stock=entero, Valores=2dec."""
    if df is None or df.empty: return "<p>Sin datos</p>"
    nc=nc or []
    hdrs="".join(f"<th class='{'n' if c in nc else ''}'>{c}</th>" for c in df.columns)
    rows=""; tots=defaultdict(float); first_col=df.columns[0] if len(df.columns)>0 else ""
    for _,row in df.iterrows():
        cells=""
        for c in df.columns:
            v=row[c]; disp=str(v) if str(v) not in("nan","None","NaN") else ""
            if c in nc:
                try:
                    fv=float(v); tots[c]+=fv
                    disp=f"{int(round(fv)):,}" if _is_int_col(c) else f"{fv:,.2f}"
                except: pass
                cells+=f"<td class='n'>{disp}</td>"
            else: cells+=f"<td>{disp}</td>"
        rows+=f"<tr>{cells}</tr>"
    def _tot_fmt(c):
        v=tots[c]
        return f"{int(round(v)):,}" if _is_int_col(c) else f"{v:,.2f}"
    tcells="".join(
        (f"<td class='n'>{_tot_fmt(c)}</td>" if c in nc and tots[c]!=0 else
         (f"<td><b>TOTAL</b></td>" if c==first_col else "<td></td>"))
        for c in df.columns)
    uid_safe=uid.replace("-","_")
    return f"""<div class="zb">
  <span style="color:var(--text3);font-size:11px;font-weight:700">ZOOM</span>
  <button onclick="asZoom('{uid_safe}',-1)">−</button>
  <button onclick="asZoom('{uid_safe}',1)">+</button>
  <button onclick="asZoomReset('{uid_safe}')">↺</button>
  <span style="color:var(--text3);font-size:10px">{len(df):,} registros</span>
</div>
<div class="tc"><table class="it" id="tbl_{uid_safe}">
<thead><tr>{hdrs}</tr></thead><tbody>{rows}</tbody>
<tfoot><tr class="tot">{tcells}</tr></tfoot>
</table></div>"""

# ══ Función universal: tabla modo componente (sticky header + cols fijas) ══
def _comp_tbl(df, nc, uid, freeze_cols=2, height=600, title="", groups=None, legend=""):
    """
    Renderiza df como st.components.html con:
    - Header siempre visible (sticky top)
    - freeze_cols columnas izquierda fijas (sticky left)
    - Zebra, ceros vacíos en nc, zoom, two-finger pan
    groups: lista de dicts {label, color_bg, color_text, rows_df} para separadores
    legend: HTML extra debajo de la tabla
    """
    import json
    if df is None or df.empty:
        _components.html("<p style='color:#94a3b8;padding:12px'>Sin datos.</p>", height=60)
        return

    all_cols = list(df.columns)
    # Anchos fijos para cols congeladas
    FW = [110, 240, 120, 120]   # col0, col1, col2, col3

    def col_left(i):
        return sum(FW[:i]) if i < len(FW) else 0

    def fmt_val(v, col):
        s = str(v)
        if s in ("nan","None","NaN",""): return ""
        if col in nc:
            try:
                fv = float(v)
                return f"{int(round(fv)):,}" if _is_int_col(col) else f"{fv:,.2f}"
            except: return s
        return s

    # ── Encabezados ───────────────────────────────────────────
    hdrs = ""
    for i,c in enumerate(all_cols):
        left  = f"left:{col_left(i)}px;" if i < freeze_cols else ""
        zidx  = "5" if i < freeze_cols else "3"
        bdr_r = "border-right:2px solid #94a3b8;" if i == freeze_cols-1 else ("border-right:1px solid #e2e8f0;" if i < freeze_cols else "")
        align = "text-align:right;" if c in nc else ""
        hdrs += (f'<th style="position:sticky;top:0;{left}z-index:{zidx};'
                 f'background:#f1f5f9;{bdr_r}border-bottom:2px solid #cbd5e1;'
                 f'padding:7px 10px;font-size:10px;font-weight:700;'
                 f'text-transform:uppercase;color:#64748b;white-space:nowrap;{align}">{c}</th>')

    # ── Filas ─────────────────────────────────────────────────
    rows = ""; tots = {c:0.0 for c in nc}
    ri = 0
    if groups:
        # Tabla agrupada con separadores de sección (T_INV)
        for g in groups:
            g_df  = g["df"]
            g_lbl = g["label"]
            g_bg  = g.get("bg","#e0f2fe")
            g_col = g.get("col","#0369a1")
            n_sub = {c:0.0 for c in nc}
            # Fila separadora de grupo
            rows += (f'<tr><td colspan="{len(all_cols)}" '
                     f'style="background:{g_bg};padding:5px 12px;font-weight:700;'
                     f'font-size:11px;color:{g_col};letter-spacing:.04em;'
                     f'position:sticky;left:0">'
                     f'🏪 {g_lbl}</td></tr>')
            for _,row in g_df.iterrows():
                bg = "#f8fafc" if ri%2==0 else "#ffffff"; ri+=1
                cells = ""
                for i,c in enumerate(all_cols):
                    v = row[c]; disp = fmt_val(v, c)
                    left  = f"left:{col_left(i)}px;" if i < freeze_cols else ""
                    zidx  = "2" if i < freeze_cols else "0"
                    bdr_r = "border-right:2px solid #94a3b8;" if i==freeze_cols-1 else ("border-right:1px solid #e2e8f0;" if i<freeze_cols else "")
                    align = "text-align:right;font-family:monospace;" if c in nc else ""
                    cells += (f'<td style="position:sticky;{left}z-index:{zidx};'
                              f'background:{bg};{bdr_r}border-bottom:1px solid #f1f5f9;'
                              f'padding:6px 10px;font-size:12px;{align}">{disp}</td>')
                    if c in nc:
                        try: n_sub[c]+=float(v); tots[c]+=float(v)
                        except: pass
                rows += f"<tr>{cells}</tr>"
            # Subtotal grupo
            sc = ""
            for i,c in enumerate(all_cols):
                left  = f"left:{col_left(i)}px;" if i < freeze_cols else ""
                zidx  = "2" if i < freeze_cols else "0"
                bdr_r = "border-right:2px solid #94a3b8;" if i==freeze_cols-1 else ("border-right:1px solid #e2e8f0;" if i<freeze_cols else "")
                if i==0:
                    sc += (f'<td style="position:sticky;{left}z-index:{zidx};'
                           f'background:#dbeafe;{bdr_r}border-top:1px solid #bfdbfe;'
                           f'padding:5px 10px;font-size:10px;font-weight:700;color:#1d4ed8">'
                           f'{g_lbl} — subtotal</td>')
                elif c in nc and n_sub[c]!=0:
                    d = f"{int(round(n_sub[c])):,}" if _is_int_col(c) else f"{n_sub[c]:,.2f}"
                    sc += (f'<td style="position:sticky;{left}z-index:{zidx};'
                           f'background:#dbeafe;{bdr_r}border-top:1px solid #bfdbfe;'
                           f'text-align:right;font-family:monospace;padding:5px 10px;'
                           f'font-weight:700;color:#1d4ed8">{d}</td>')
                else:
                    sc += (f'<td style="position:sticky;{left}z-index:{zidx};'
                           f'background:#dbeafe;{bdr_r}border-top:1px solid #bfdbfe;'
                           f'padding:5px 10px"></td>')
            rows += f"<tr>{sc}</tr>"
    else:
        # Tabla plana (T_SKU unidades, T_SAM)
        for _,row in df.iterrows():
            bg = "#f8fafc" if ri%2==0 else "#ffffff"; ri+=1
            cells = ""
            for i,c in enumerate(all_cols):
                v = row[c]; disp = fmt_val(v, c)
                left  = f"left:{col_left(i)}px;" if i < freeze_cols else ""
                zidx  = "2" if i < freeze_cols else "0"
                bdr_r = "border-right:2px solid #94a3b8;" if i==freeze_cols-1 else ("border-right:1px solid #e2e8f0;" if i<freeze_cols else "")
                align = "text-align:right;font-family:monospace;" if c in nc else ""
                cells += (f'<td style="position:sticky;{left}z-index:{zidx};'
                          f'background:{bg};{bdr_r}border-bottom:1px solid #f1f5f9;'
                          f'padding:6px 10px;font-size:12px;{align}">{disp}</td>')
                if c in nc:
                    try: tots[c]+=float(row[c])
                    except: pass
            rows += f"<tr>{cells}</tr>"

    # ── Fila TOTAL ────────────────────────────────────────────
    tfooter = ""
    for i,c in enumerate(all_cols):
        left  = f"left:{col_left(i)}px;" if i < freeze_cols else ""
        zidx  = "2" if i < freeze_cols else "0"
        bdr_r = "border-right:2px solid #94a3b8;" if i==freeze_cols-1 else ("border-right:1px solid #e2e8f0;" if i<freeze_cols else "")
        S = (f"position:sticky;{left}z-index:{zidx};{bdr_r}"
             f"background:#e0f2fe;font-weight:700;padding:7px 10px;"
             f"border-top:2px solid #7dd3fc;color:#0369a1")
        if i==0:
            tfooter += f'<td style="{S}">TOTAL GENERAL</td>'
        elif c in nc and tots[c]!=0:
            d = f"{int(round(tots[c])):,}" if _is_int_col(c) else f"{tots[c]:,.2f}"
            tfooter += f'<td style="{S};text-align:right;font-family:monospace">{d}</td>'
        else:
            tfooter += f'<td style="{S}"></td>'

    n_rows = len(df) if not groups else sum(len(g["df"]) for g in groups)
    n_info = title if title else f"{n_rows:,} registros"

    legend_html = f"<div style='font-size:10px;color:#64748b;margin-top:6px;padding:6px 10px;background:#f8fafc;border-radius:6px;border-left:3px solid #0ea5e9'>{legend}</div>" if legend else ""

    html = (
        "<!DOCTYPE html><html><head><meta charset=\"UTF-8\">"
        "<style>"
        "*{box-sizing:border-box;margin:0;padding:0}"
        "body{font-family:Inter,Segoe UI,sans-serif;background:#f0f9ff}"
        ".wrap{border:1px solid #e2e8f0;border-radius:8px;overflow:auto;"
        f"max-height:{height}px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08)}}"
        ".zb{display:flex;align-items:center;gap:6px;margin-bottom:6px}"
        ".zb span{font-size:11px;font-weight:700;color:#94a3b8}"
        ".zb .inf{font-size:10px;font-weight:400}"
        ".zb button{background:#fff;border:1px solid #e2e8f0;border-radius:5px;"
        "padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;color:#475569}"
        ".zb button:hover{background:#0ea5e9;color:#fff;border-color:#0ea5e9}"
        "table{border-collapse:separate;border-spacing:0;width:max-content}"
        "tr:hover td{background:#e0f2fe!important}"
        "</style></head><body>"
        f'<div class=\"zb\">'
        "<span>ZOOM</span>"
        "<button onclick=\"var t=document.getElementById('_t');var s=parseFloat(t.style.fontSize||'12');t.style.fontSize=(s-1)+'px'\">−</button>"
        "<button onclick=\"var t=document.getElementById('_t');var s=parseFloat(t.style.fontSize||'12');t.style.fontSize=(s+1)+'px'\">+</button>"
        "<button onclick=\"document.getElementById('_t').style.fontSize='12px'\">↺</button>"
        f'<span class=\"inf\">{n_info}</span>'
        "</div>"
        '<div class=\"wrap\">'
        f'<table id=\"_t\"><thead><tr>{hdrs}</tr></thead>'
        f'<tbody>{rows}</tbody>'
        f'<tfoot><tr>{tfooter}</tr></tfoot>'
        "</table></div>"
        f"{legend_html}"
        "<script>"
        "var w=document.querySelector('.wrap');"
        "var t0x,t0y,t0sl,t0st;"
        "w.addEventListener('touchstart',function(e){if(e.touches.length!==2)return;"
        "t0x=(e.touches[0].pageX+e.touches[1].pageX)/2;"
        "t0y=(e.touches[0].pageY+e.touches[1].pageY)/2;"
        "t0sl=w.scrollLeft;t0st=w.scrollTop;},{passive:true});"
        "w.addEventListener('touchmove',function(e){if(e.touches.length!==2)return;"
        "e.preventDefault();"
        "var cx=(e.touches[0].pageX+e.touches[1].pageX)/2;"
        "var cy=(e.touches[0].pageY+e.touches[1].pageY)/2;"
        "w.scrollLeft=t0sl-(cx-t0x);w.scrollTop=t0st-(cy-t0y);"
        "},{passive:false});"
        "w.addEventListener('wheel',function(e){"
        "if(Math.abs(e.deltaX)>Math.abs(e.deltaY)){e.preventDefault();w.scrollLeft+=e.deltaX;}"
        "else if(e.shiftKey){e.preventDefault();w.scrollLeft+=e.deltaY;}"
        "},{passive:false});"
        "</script></body></html>"
    )
    _components.html(html, height=height+60, scrolling=False)


# ── Sidebar ─────────────────────────────────────────────────────
_perf("before_sidebar")
with st.sidebar:
    ca,cb=st.columns([4,1])
    with ca:
        st.markdown(
            f'''<div style="font-size:16px;font-weight:800;color:#0ea5e9;
            font-family:Inter,sans-serif;letter-spacing:.04em">
            AUTO<span style="font-weight:300;opacity:.8">SKY</span>
            </div>
            <div style="font-size:10px;font-family:monospace;color:#475569;margin-top:2px">
            <b style="color:#0ea5e9">{APP_VERSION}</b> &nbsp;|&nbsp; {BUILD_TIME}
            </div>''',
            unsafe_allow_html=True
        )
    with cb:
        st.markdown("")
        nd=st.toggle("🌙",value=dark,help="Tema oscuro/claro")
        if nd!=dark: st.session_state.dark_mode=nd; st.rerun()
    st.divider()

    st.markdown("### 📂 Cargar Excel")
    st.caption("XLS / XLSX del sistema contable — acumulativo")
    uploaded=st.file_uploader("Archivos",type=["xlsx","xls"],
                               accept_multiple_files=True,label_visibility="collapsed")
    if uploaded:
        # Al subir archivos reales, remover la entrada sintética del consolidado persistido
        st.session_state.files_loaded[:] = [
            f for f in st.session_state.files_loaded if "(persistido)" not in f
        ]
        new=[f for f in uploaded if f.name not in st.session_state.files_loaded]
        if new:
            for uf in new:
                with st.spinner(f"Procesando {uf.name}..."):
                    try:
                        suf=os.path.splitext(uf.name)[1]
                        with tempfile.NamedTemporaryFile(delete=False,suffix=suf) as tmp:
                            tmp.write(uf.getvalue()); tp=tmp.name
                        if eng.raw_df is None:
                            eng.load_inventory_file(tp)
                            _nuevos=len(eng.raw_df); _dupes=0
                        else:
                            from app.engine import InventoryEngine as _IE
                            te=_IE(); te.load_inventory_file(tp)
                            # Deduplicar por columna "Código" (N° de movimiento único)
                            if "Código" in eng.raw_df.columns and "Código" in te.raw_df.columns:
                                _exist=set(eng.raw_df["Código"].astype(str).str.strip())
                                _mask=~te.raw_df["Código"].astype(str).str.strip().isin(_exist)
                                _dupes=int((~_mask).sum()); _nuevos=int(_mask.sum())
                                if _nuevos>0:
                                    eng.raw_df=pd.concat([eng.raw_df,te.raw_df[_mask]],ignore_index=True)
                            else:
                                # Fallback: deduplicar por todas las columnas
                                _prev=len(eng.raw_df)
                                eng.raw_df=pd.concat([eng.raw_df,te.raw_df],ignore_index=True).drop_duplicates()
                                _nuevos=len(eng.raw_df)-_prev; _dupes=len(te.raw_df)-_nuevos
                        try: os.unlink(tp)
                        except: pass
                        st.session_state.files_loaded.append(uf.name)
                        # Guardar stats por archivo
                        st.session_state.setdefault("files_stats",[]).append({
                            "nombre": uf.name,
                            "nuevos": _nuevos,
                            "dupes":  _dupes,
                        })
                        st.session_state.result=None
                        _msg=f"✓ {uf.name} | {_nuevos:,} nuevos"
                        if _dupes: _msg+=f" | {_dupes} duplicados omitidos"
                        log(_msg)
                        # Marcar para recálculo automático
                        st.session_state["_recalc_pending"]=True
                    except Exception as e: st.error(f"{uf.name}: {e}")
            # Persistir consolidado para que todas las sesiones/clientes
            # vean los mismos datos y sobrevivan al reinicio del servidor
            if eng.raw_df is not None:
                _persist_raw(eng.raw_df)
            st.rerun()

    if st.session_state.files_loaded:
        # ── Rango de fechas consolidado ────────────────────────
        if eng.raw_df is not None and "Fecha" in eng.raw_df.columns:
            try:
                _dates=pd.to_datetime(eng.raw_df["Fecha"],errors="coerce").dropna()
                if not _dates.empty:
                    _d1=_dates.min().strftime("%d/%m/%Y")
                    _d2=_dates.max().strftime("%d/%m/%Y")
                    _pc_bg    = "#1e3a5f" if dark else "#f0f9ff"
                    _pc_bdr   = "#2d5a8e" if dark else "#bae6fd"
                    _pc_title = "#7dd3fc" if dark else "#0284c7"
                    _pc_date  = "#f1f5f9" if dark else "#0f172a"
                    _pc_muted = "#94a3b8" if dark else "#64748b"
                    st.markdown(
                        f"<div style='background:{_pc_bg};border:1px solid {_pc_bdr};"
                        f"border-radius:8px;padding:8px 12px;font-size:11px;margin-bottom:4px'>"
                        f"<b style='color:{_pc_title}'>📅 Período cargado</b><br>"
                        f"<span style='font-size:13px;font-weight:700;color:{_pc_date}'>"
                        f"{_d1} → {_d2}</span><br>"
                        f"<span style='color:{_pc_muted}'>{len(eng.raw_df):,} movimientos · "
                        f"{len(st.session_state.files_loaded)} archivo(s)</span>"
                        f"</div>",
                        unsafe_allow_html=True)
            except: pass

        # ── Detalle por archivo ────────────────────────────────
        for _st_item in st.session_state.get("files_stats",[]):
            _dup_txt=f" · <span style='color:#f59e0b'>{_st_item['dupes']} dup.</span>" if _st_item.get("dupes") else ""
            st.markdown(
                f"<div style='font-size:11px;color:{MUTED};padding:2px 4px'>"
                f"📄 {_st_item['nombre']} — "
                f"<b style='color:#059669'>{_st_item['nuevos']:,}</b> reg.{_dup_txt}"
                f"</div>",
                unsafe_allow_html=True)

        c1,c2=st.columns(2)
        with c1:
            if st.button("🗑 Limpiar",width='stretch'):
                # Mutar en sitio para afectar a todas las sesiones
                eng.raw_df=None; eng.physical_df=None
                st.session_state.files_loaded.clear()
                st.session_state.files_stats.clear()
                st.session_state.result=None
                for _p in (CONSOLIDADO_PATH, TOMA_FISICA_PATH):
                    try:
                        if os.path.exists(_p): os.unlink(_p)
                    except Exception: pass
                st.rerun()
        with c2:
            if eng.raw_df is not None:
                st.download_button("📥 Exportar",to_xl(eng.raw_df),
                    "consolidado.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width='stretch',key="exp_cons")
    st.divider()

    if eng.raw_df is not None:
        st.markdown("### 🚫 Exclusiones globales")
        # Mapa código → nombre para búsqueda predictiva por ambos
        _excl_map = (eng.raw_df[["Código Producto","Nombre Producto"]]
                     .drop_duplicates().sort_values("Nombre Producto"))
        _excl_map = dict(zip(_excl_map["Código Producto"].astype(str),
                             _excl_map["Nombre Producto"].astype(str)))
        all_skus = list(_excl_map.keys())

        _fmt_sku = lambda s: f"{s} — {_excl_map.get(s, '')}"
        # Inicializar widget state desde filtros persistidos, sin usar `default`
        # para evitar warning "default + key" de Streamlit.
        if "es_ms" not in st.session_state:
            st.session_state["es_ms"] = list(st.session_state.excluded_skus)
        excl_s=st.multiselect("Excluir SKUs", all_skus,
                               key="es_ms",
                               format_func=_fmt_sku,
                               placeholder="Escribe código o nombre…")
        _prev_excl_s = set(st.session_state.excluded_skus)
        st.session_state.excluded_skus=set(excl_s); eng.excluded_skus=set(excl_s)
        _sku_changed = _prev_excl_s != set(excl_s)
        if _sku_changed:
            st.session_state["_recalc_pending"] = True

        all_wh=eng.get_warehouses()
        if "ew_ms" not in st.session_state:
            st.session_state["ew_ms"] = list(st.session_state.excl_wh)
        excl_w=st.multiselect("Excluir Bodegas", all_wh,
                               key="ew_ms",
                               placeholder="Escribe para filtrar…")
        _prev_excl_w = set(st.session_state.excl_wh)
        st.session_state.excl_wh=set(excl_w)
        # Exclusión GLOBAL: descarta movimientos (origen o destino) en esas bodegas.
        # Afecta KPIs, rotación, kardex, compras — todo el pipeline.
        eng.excluded_warehouses = set(excl_w)
        _wh_changed = _prev_excl_w != set(excl_w)
        if _wh_changed:
            st.session_state["_recalc_pending"] = True

        # Persistir filtros a disco + cache_resource compartido si cambió algo.
        # Así sobreviven refresh, nueva sesión y son iguales para todos los clientes.
        if _sku_changed or _wh_changed:
            _shared_f = _get_shared_filtros()
            _shared_f["excluded_skus"]       = set(excl_s)
            _shared_f["excluded_warehouses"] = set(excl_w)
            _persist_filtros(set(excl_s), set(excl_w))

        st.divider()

        st.markdown("### ⚙️ Calcular")
        cutoff=st.date_input("Fecha de corte",date.today(),format="DD/MM/YYYY")
        wh_mode=st.selectbox("Bodegas",["Todas","Solo principal","Selección manual"])
        sel_wh=[]
        if wh_mode=="Selección manual": sel_wh=st.multiselect("Bodegas",all_wh)
        if st.button("▶ Calcular",type="primary",width='stretch'):
            with st.spinner("Calculando..."):
                try:
                    r=eng.analyze(str(cutoff),wh_mode,sel_wh)
                    st.session_state.result=r; log(f"OK | {cutoff} | {len(r.filtered):,} mov.")
                    st.success("✓")
                except Exception as e: st.error(str(e)); log(f"Error: {e}")

        # Auto-recálculo cuando se carga un archivo nuevo
        if st.session_state.pop("_recalc_pending", False) and eng.raw_df is not None:
            try:
                _ar=eng.analyze(str(cutoff),wh_mode,sel_wh)
                st.session_state.result=_ar
                log(f"Auto-calculado | {cutoff} | {len(_ar.filtered):,} mov.")
            except: pass

r=st.session_state.result

# ── Sin datos ────────────────────────────────────────────────────
if eng.raw_df is None:
    st.markdown("## 📦 Sistema de Inventario")
    _card_bg  = "#1e293b" if dark else "#f0f9ff"
    _card_bdr = "#334155" if dark else "#bae6fd"
    _card_ver = "#38bdf8" if dark else "#0284c7"
    _card_sep = "#475569" if dark else "#94a3b8"
    _card_txt = "#94a3b8" if dark else "#475569"
    st.markdown(f"""
<div style='display:inline-flex;align-items:center;gap:12px;
  background:{_card_bg};border:1px solid {_card_bdr};border-radius:8px;
  padding:8px 16px;margin-bottom:16px;'>
  <span style='color:{_card_ver};font-weight:700;font-family:monospace;font-size:14px'>{APP_VERSION}</span>
  <span style='color:{_card_sep}'>|</span>
  <span style='color:{_card_txt};font-family:monospace;font-size:12px'>{BUILD_TIME}</span>
</div>""", unsafe_allow_html=True)
    st.info("👈 Cargue uno o más archivos Excel. Puede cargar archivos adicionales y los datos se acumularán.")
    st.stop()

# ── Banner Autosky ───────────────────────────────────────────────
st.markdown(f"""<div class="as-banner">
  <div>
    <div class="as-logo">AUTO<span>SKY</span>&nbsp;
      <span style="font-size:12px;font-weight:400;opacity:.85">Sistema de Inventario</span>
    </div>
    <div style="font-size:10px;opacity:.8;color:#fff">Gestión · Análisis · Control de Inventario</div>
  </div>
  <div class="as-build">
    <div class="v">{APP_VERSION}</div>
    <div>{BUILD_TIME}</div>
  </div>
</div>""", unsafe_allow_html=True)

# ── Período de datos cargados ───────────────────────────────────
if eng.raw_df is not None and "Fecha" in eng.raw_df.columns:
    try:
        _all_d=pd.to_datetime(eng.raw_df["Fecha"],errors="coerce").dropna()
        if not _all_d.empty:
            _pd1=_all_d.min().strftime("%d/%m/%Y")
            _pd2=_all_d.max().strftime("%d/%m/%Y")
            _pb_bg  = "#1e3a5f" if dark else "#f0f9ff"
            _pb_bdr = "#2d5a8e" if dark else "#bae6fd"
            _pb_lbl = "#7dd3fc" if dark else "#475569"
            _pb_dt  = "#f1f5f9" if dark else "#0f172a"
            _pb_arr = "#7dd3fc" if dark else "#94a3b8"
            _pb_mut = "#94a3b8" if dark else "#64748b"
            st.markdown(
                f"<div class='as-periodo-banner' style='display:inline-flex;align-items:center;gap:16px;"
                f"background:{_pb_bg};border:1px solid {_pb_bdr};border-radius:8px;"
                f"padding:7px 16px;margin-bottom:8px;font-size:12px'>"
                f"<span style='color:{_pb_lbl};font-weight:600'>📅 PERÍODO DE ANÁLISIS</span>"
                f"<span style='color:{_pb_dt};font-weight:700;font-size:14px'>{_pd1}</span>"
                f"<span style='color:{_pb_arr}'>→</span>"
                f"<span style='color:{_pb_dt};font-weight:700;font-size:14px'>{_pd2}</span>"
                f"<span style='color:{_pb_mut}'>{len(eng.raw_df):,} movimientos</span>"
                f"</div>",
                unsafe_allow_html=True)
    except: pass

# ── Aviso de exclusiones globales activas ──────────────────────
_act_excl_sku = list(st.session_state.get("excluded_skus", set()))
_act_excl_wh  = list(st.session_state.get("excl_wh", set()))
if _act_excl_sku or _act_excl_wh:
    _ex_lines = []
    if _act_excl_wh:
        _ex_lines.append(
            f"🏪 <b>{len(_act_excl_wh)} bodega(s) excluida(s):</b> "
            + ", ".join(f"<code>{b}</code>" for b in _act_excl_wh[:5])
            + (f" <i>+{len(_act_excl_wh)-5} más</i>" if len(_act_excl_wh) > 5 else "")
        )
    if _act_excl_sku:
        _ex_lines.append(
            f"🏷 <b>{len(_act_excl_sku)} SKU(s) excluido(s)</b>"
        )
    st.markdown(
        f"<div class='as-excl-banner' style='background:#fef3c7;border:1px solid #f59e0b;border-radius:8px;"
        f"padding:8px 14px;margin-bottom:10px;font-size:12px;color:#92400e;"
        f"line-height:1.5'>"
        f"<b>⚠ Exclusiones globales activas</b> — los movimientos relacionados se "
        f"descartan del pipeline (KPIs, Rotación, Kardex, Compras, todo).<br>"
        + "<br>".join(_ex_lines) +
        f"</div>",
        unsafe_allow_html=True
    )

# ── KPIs ────────────────────────────────────────────────────────
if r is not None:
    kpis=r.kpis
    r1="".join([kc("Stock Total",fmt(kpis.get("Stock total",0),"i"),"a"),
                kc("Disponible",fmt(kpis.get("Stock disponible",0),"i"),"s"),
                kc("Muestras",fmt(kpis.get("Stock en muestras",0),"i"),"w"),
                kc("Valor Inv.",f'${fmt(kpis.get("Valor inventario",0))}', "a"),
                kc("Compras Acum.",f'${fmt(kpis.get("Compras acumuladas",0))}', "d")])
    r2="".join([kc("Rotación",fmt(kpis.get("Rotación",0),"p")),
                kc("Días Inv.",f'{kpis.get("Días de inventario",0):.0f} d'),
                kc("Consumo/día",f'{kpis.get("Consumo promedio",0):.2f} u'),
                kc("Exactitud",fmt(kpis.get("Exactitud inventario",0),"p"))])
    st.markdown(f'<div class="kpi-row">{r1}</div>',unsafe_allow_html=True)
    st.markdown(f'<div class="kpi-row">{r2}</div>',unsafe_allow_html=True)
    st.markdown("---")


_perf("main_kpis_done")

# ── Pestañas ─────────────────────────────────────────────────────
tabs=st.tabs(["🏪 Inv×Bodega","🔍 Detalle SKU","📊 SKU×Bodega","👥 Muestras",
              "📈 Período","🔄 Rotación","📐 Cálculos","🧾 Compras","📋 Kardex","🏭 Toma Física","📦 Importación","🔍 Auditoría","📤 Exportar Portal"])
(T_INV,T_SKU,T_PIV,T_SAM,T_ANA,T_ROT,T_CAL,T_PUR,T_KDX,T_PHY,T_IMP,T_AUD,T_EXP)=tabs
_perf("tabs_defined")

excl_s=list(st.session_state.excluded_skus)
excl_w=list(st.session_state.excl_wh)

# ══ TAB 1 INVENTARIO ════════════════════════════════════════════
@_fragment
def _render_tab_inv():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None: st.info("Ejecute el análisis.")
    else:
        df_inv=r.inventory_by_warehouse.copy()
        if excl_w: df_inv=df_inv[~df_inv["Bodega"].isin(excl_w)]
        for _drop in ("Grupo Visual","GRUPO VISUAL","Grupo Movimiento","Categoría Producto"):
            if _drop in df_inv.columns: df_inv=df_inv.drop(columns=[_drop])

        # ── Filtros ───────────────────────────────────────────────
        _all_skus=sorted(df_inv["Código Producto"].dropna().unique().tolist()) if "Código Producto" in df_inv.columns else []
        _all_bods=sorted(df_inv["Bodega"].dropna().unique().tolist()) if "Bodega" in df_inv.columns else []

        # Mapa SKU → nombre para búsqueda rápida
        _inv_name_map = dict(zip(df_inv["Código Producto"].astype(str),
                                 df_inv["Nombre Producto"].astype(str))) if "Código Producto" in df_inv.columns else {}
        _inv_fmt = lambda s: f"{s} — {_inv_name_map.get(s, '')}" if _inv_name_map.get(s) else s

        fc1,fc2,fc3,fc4=st.columns([4,4,1,1])
        with fc1:
            sel_skus=st.multiselect("🔍 SKU / Producto", _all_skus,
                key="i_skus", placeholder="Escribe código o nombre…",
                format_func=_inv_fmt)
        with fc2:
            sel_bods=st.multiselect("🏪 Bodega", _all_bods,
                key="i_bods", placeholder="Escribe para filtrar…")
        with fc3:
            st.markdown("")
            stk_only=st.checkbox("Solo con stock",True,key="i_s")
        with fc4:
            st.markdown("")
            st.button("✔ Aplicar",key="i_apply",width='stretch',
                      help="Cerrar selector y aplicar filtros")

        # Aplicar filtros
        df=df_inv.copy()
        if sel_skus: df=df[df["Código Producto"].isin(sel_skus)]
        if sel_bods: df=df[df["Bodega"].isin(sel_bods)]
        if stk_only and "Stock" in df.columns: df=df[df["Stock"]>0]

        # Columnas numéricas — excluir Bodega de la columna visible
        _skip_cols=("Código Producto","Nombre Producto","Bodega")
        nc=[c for c in df.columns if c not in _skip_cols]

        # ── Tabla agrupada por Bodega (sin columna Bodega) ────────
        def _inv_grouped(df, nc, uid):
            if df is None or df.empty: return "<p style='color:var(--text3);padding:12px'>Sin datos para los filtros seleccionados.</p>"
            uid_safe=uid.replace("-","_")
            bodegas=sorted(df["Bodega"].dropna().unique().tolist()) if "Bodega" in df.columns else []
            # Columnas a mostrar: excluir "Bodega" de la tabla
            show_cols=[c for c in df.columns if c!="Bodega"]
            nc_show=[c for c in nc if c in show_cols]
            hdrs="".join(f"<th class='{'n' if c in nc_show else ''}'>{c}</th>" for c in show_cols)
            body=""
            grand=defaultdict(float)
            for bod_name in bodegas:
                sub=df[df["Bodega"]==bod_name] if "Bodega" in df.columns else df
                if sub.empty: continue
                n_cols=len(show_cols)
                body+=f"<tr style='background:var(--sky-l)'><td colspan='{n_cols}' style='padding:5px 12px;font-weight:700;font-size:11px;color:var(--sky-d);letter-spacing:.04em'>🏪 {bod_name}</td></tr>"
                sub_tots=defaultdict(float)
                for _,row in sub.iterrows():
                    cells=""
                    for c in show_cols:
                        v=row[c]; raw=str(v)
                        disp="" if raw in("nan","None","NaN") else raw
                        if c in nc_show:
                            try:
                                fv=float(v); sub_tots[c]+=fv; grand[c]+=fv
                                disp=f"{int(round(fv)):,}" if _is_int_col(c) else f"{fv:,.2f}"
                            except: pass
                            cells+=f"<td class='n'>{disp}</td>"
                        else:
                            cells+=f"<td>{disp}</td>"
                    body+=f"<tr>{cells}</tr>"
                # Subtotal bodega
                sc=""
                for i,c in enumerate(show_cols):
                    if i==0:
                        sc+=f"<td style='font-weight:700;font-size:10px;color:var(--sky-d)'>{bod_name} — subtotal</td>"
                    elif c in nc_show and sub_tots[c]!=0:
                        v=sub_tots[c]
                        d=f"{int(round(v)):,}" if _is_int_col(c) else f"{v:,.2f}"
                        sc+=f"<td class='n' style='font-weight:700;color:var(--sky-d)'>{d}</td>"
                    else:
                        sc+="<td></td>"
                body+=f"<tr style='background:var(--sky-ll);border-top:1px solid var(--border2)'>{sc}</tr>"
            # TOTAL GENERAL
            gc=""
            for i,c in enumerate(show_cols):
                if i==0:
                    gc+=f"<td style='font-weight:800;color:var(--sky-d)'>TOTAL GENERAL</td>"
                elif c in nc_show and grand[c]!=0:
                    v=grand[c]
                    d=f"{int(round(v)):,}" if _is_int_col(c) else f"{v:,.2f}"
                    gc+=f"<td class='n' style='font-weight:800;color:var(--sky-d)'>{d}</td>"
                else:
                    gc+="<td></td>"
            n_shown=sum(1 for _,r in df.iterrows() for _ in [None])
            return f"""<div class="zb">
  <span style="color:var(--text3);font-size:11px;font-weight:700">ZOOM</span>
  <button onclick="asZoom('{uid_safe}',-1)">−</button>
  <button onclick="asZoom('{uid_safe}',1)">+</button>
  <button onclick="asZoomReset('{uid_safe}')">↺</button>
  <span style="color:var(--text3);font-size:10px">{len(df):,} registros · {len(bodegas)} bodega(s)</span>
</div>
<div class="tc"><table class="it" id="tbl_{uid_safe}">
<thead><tr>{hdrs}</tr></thead><tbody>{body}</tbody>
<tfoot><tr class="tot">{gc}</tr></tfoot>
</table></div>"""
        # Construir groups para _comp_tbl
        _bodegas = sorted(df["Bodega"].dropna().unique().tolist()) if "Bodega" in df.columns else []
        _show_cols = [c for c in df.columns if c != "Bodega"]
        _nc_show   = [c for c in nc if c in _show_cols]
        _groups = []
        for _bn in _bodegas:
            _sub = df[df["Bodega"]==_bn][_show_cols].copy() if "Bodega" in df.columns else df[_show_cols].copy()
            if not _sub.empty:
                _groups.append({"label":_bn, "df":_sub,
                                 "bg":"#e0f2fe", "col":"#0369a1"})
        _comp_tbl(df[_show_cols], _nc_show, "inv",
                  freeze_cols=2, height=580,
                  title=f"{len(df):,} registros · {len(_bodegas)} bodega(s)",
                  groups=_groups if len(_bodegas)>1 else None)
        dl3(df,"inventario_bodega","inv")

with T_INV:
    _render_tab_inv()

# ══ TAB 2 DETALLE SKU ═══════════════════════════════════════════
@_fragment
def _render_tab_sku():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None: st.info("Ejecute el análisis.")
    else:
        df=r.sku_summary.copy()

        # Dropear columnas no relevantes para el reporte (gana espacio
        # horizontal y evita solapamiento con ~19 columnas)
        for _drop in ("Categoría Producto",):
            if _drop in df.columns: df=df.drop(columns=[_drop])

        # ── Renombrar columnas del engine a labels legibles ───────
        _ren = {
            "Dev_Proveedor":     "Dev. Proveedor",
            "Dev_Cliente":       "Dev. Cliente",
            "Baja_Inventario":   "Baja Inv.",
            "Muestras_Enviadas": "Muestras Env.",
            "Muestras_Devueltas":"Muestras Dev.",
            "Valor_Compras":     "Valor Compras",
            "Valor_Ventas":      "Valor Ventas",
        }
        df = df.rename(columns={k:v for k,v in _ren.items() if k in df.columns})

        # ── Forzar enteros en todas las columnas de unidades ──────
        _unit_cols = ["Compras","Dev. Proveedor","Ventas","Dev. Cliente",
                      "Baja Inv.",
                      "Muestras Env.","Muestras Dev.",
                      "Stock Disponible","Stock Muestras","Stock Total",
                      "En Tránsito"]
        for _uc in _unit_cols:
            if _uc in df.columns:
                df[_uc] = df[_uc].fillna(0).astype(int)

        # ── Cuadre: solo movimientos reales (muestras son internas)
        #   Compras − Dev.Proveedor − Ventas + Dev.Cliente = Stock Total
        #   Muestras Env./Dev. son transferencias internas — no afectan Stock Total
        # ── Costo Promedio Ponderado = Valor_Compras / Compras ───
        if "Valor Compras" in df.columns and "Compras" in df.columns:
            df["Costo Prom."] = (
                df["Valor Compras"] / df["Compras"].replace(0, float("nan"))
            ).round(2).fillna(0.0)
        elif "Valor_Compras" in df.columns and "Compras" in df.columns:
            df["Costo Prom."] = (
                df["Valor_Compras"] / df["Compras"].replace(0, float("nan"))
            ).round(2).fillna(0.0)

        # Si la columna falta (sku_summary vacío), devolver Series de ceros
        # indexado como el df: evita que el `.astype(int)` del resultado caiga
        # sobre un int escalar (0) cuando TODAS las columnas están ausentes.
        def _safe(d, col):
            if col in d.columns: return d[col]
            return pd.Series(0, index=d.index, dtype=int)
        df["✓ Cuadre"] = (
              _safe(df,"Compras")
            - _safe(df,"Dev. Proveedor")
            - _safe(df,"Baja Inv.")
            - _safe(df,"Ventas")
            + _safe(df,"Dev. Cliente")
        ).astype(int)
        df["Δ vs Stock"] = (df["✓ Cuadre"] - _safe(df,"Stock Total")).astype(int)

        # ── Enriquecer con importaciones (columnas informativas) ─
        # Añade "En Tránsito" (u en estado INGRESADA) y "Fechas Tránsito"
        # (fechas tentativas formateadas). NO modifica Stock Disponible aquí
        # — ese se mantiene puramente contable para preservar el cuadre.
        _imp_agg_det = _aggregate_importaciones(_get_shared_importaciones()["df"])
        if _imp_agg_det:
            df["En Tránsito"]     = df["Código Producto"].astype(str).map(
                lambda s: _imp_agg_det.get(s, {}).get("en_transito", 0)).astype(int)
            df["Fechas Tránsito"] = df["Código Producto"].astype(str).map(
                lambda s: _imp_agg_det.get(s, {}).get("fechas_txt", ""))
        else:
            df["En Tránsito"]     = 0
            df["Fechas Tránsito"] = ""

        # Multiselect predictivo con filtro estricto de substring
        _sk_opts_df = df[["Código Producto","Nombre Producto"]].drop_duplicates()
        _sk_opts_df = _sk_opts_df.sort_values("Código Producto")
        _sk_labels = [f"{c} — {n}" for c,n in zip(
            _sk_opts_df["Código Producto"].astype(str),
            _sk_opts_df["Nombre Producto"].astype(str))]

        sk_f1, sk_f2 = st.columns([3, 2])
        with sk_f1:
            _sk_sel = st.multiselect("🔍 SKU / Producto", _sk_labels, key="sk_f",
                                      placeholder="Escribe código o nombre…")
        with sk_f2:
            _sort_opts = {
                "Código ↑":            ("Código Producto",      True),
                "Código ↓":            ("Código Producto",      False),
                "Nombre ↑":            ("Nombre Producto",      True),
                "Nombre ↓":            ("Nombre Producto",      False),
                "Stock Total ↓":       ("Stock Total",          False),
                "Ventas ↓":            ("Ventas",               False),
                "Valor Inventario ↓":  ("Valor Inventario",     False),
                "Δ vs Stock ↓":        ("Δ vs Stock",           False),
            }
            _sort_label = st.selectbox("Ordenar por",
                                       list(_sort_opts.keys()),
                                       index=0, key="sk_sort",
                                       help="Default: Código ascendente (útil para búsqueda por SKU).")
        _sort_col, _sort_asc = _sort_opts[_sort_label]

        if _sk_sel:
            _sk_codes_sel = {s.split(" — ")[0] for s in _sk_sel}
            df = df[df["Código Producto"].astype(str).isin(_sk_codes_sel)]

        # Ordenamiento aplicado a las tablas que se muestran abajo
        if _sort_col in df.columns:
            df = df.sort_values(_sort_col, ascending=_sort_asc,
                                na_position="last").reset_index(drop=True)

        # ── Tabla Unidades ────────────────────────────────────────
        # Orden: flujo de entrada/salida | resultado | muestras (info) | cuadre
        st.markdown("##### 📦 Movimiento de Unidades")
        _u_cols = [
            "Código Producto","Nombre Producto",
            "Costo Prom.",               # último costo promedio ponderado
            "Compras","Dev. Proveedor","Baja Inv.",
            "Ventas","Dev. Cliente",
            "Stock Disponible","Stock Total",
            "En Tránsito","Fechas Tránsito",
            "✓ Cuadre","Δ vs Stock",
            "Stock Muestras",
        ]
        mu=[c for c in _u_cols if c in df.columns]
        nu=[c for c in mu if c not in("Código Producto","Nombre Producto")]
        _comp_tbl(df[mu], nu, "su", freeze_cols=2, height=520,
                  title=f"{len(df):,} SKUs",
                  legend="<b>✓ Cuadre</b> = Compras − Dev.Proveedor − <b>Baja Inv.</b> − Ventas + Dev.Cliente &nbsp;·&nbsp; "
                         "<b>Δ vs Stock</b> = Cuadre − Stock Total (0 = correcto) &nbsp;·&nbsp; "
                         "<b>Baja Inv.</b>: EGR con Descripción «BAJA DE INVENTARIO» (merma, deterioro, obsoleto) &nbsp;·&nbsp; "
                         "<b>Stock Muestras</b>: informativo, transferencias internas &nbsp;·&nbsp; "
                         "<b>En Tránsito</b>: importaciones en estado INGRESADA (informativo — no afecta Stock ni Cuadre)")

        st.markdown("")
        # ── Tabla Valores Financieros ─────────────────────────────
        st.markdown("##### 💰 Valores Financieros")
        _f_cols = [
            "Código Producto","Nombre Producto",
            "Valor Compras","Valor Ventas","Valor Inventario",
        ]
        mf=[c for c in _f_cols if c in df.columns]
        nf=[c for c in mf if c not in("Código Producto","Nombre Producto")]
        st.markdown(tbl(df[mf],nf,"sf"),unsafe_allow_html=True)

        dl3(df,"detalle_sku","sku")

with T_SKU:
    _render_tab_sku()

# ══ TAB 3 SKU×BODEGA ════════════════════════════════════════════
@_fragment
def _render_tab_piv():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None: st.info("Ejecute el análisis.")
    else:
        df=r.inventory_by_warehouse.copy()
        if excl_w: df=df[~df["Bodega"].isin(excl_w)]

        # Filtros
        _pv_skus=sorted(df["Código Producto"].dropna().unique().tolist()) if "Código Producto" in df.columns else []
        # Mapa para búsqueda predictiva
        _pv_name_map = dict(zip(df["Código Producto"].astype(str),
                                df["Nombre Producto"].astype(str))) if "Código Producto" in df.columns else {}
        _pv_fmt = lambda s: f"{s} — {_pv_name_map.get(s, '')}" if _pv_name_map.get(s) else s

        pc1,pc2,pc3=st.columns([5,1,1])
        with pc1:
            sel_pv=st.multiselect("🔍 SKU", _pv_skus, key="pv_skus",
                placeholder="Escribe código o nombre…", format_func=_pv_fmt)
        with pc2:
            st.markdown("")
            excl_bp=st.checkbox("Excluir Bodega Ppal",True,key="pv_e")
        with pc3:
            st.markdown("")
            # Botón Aplicar: cierra el dropdown y fuerza rerun
            st.button("✔ Aplicar",key="pv_apply",
                      help="Cerrar selector y aplicar filtros",
                      width='stretch')

        if excl_bp: df=df[df["Bodega"]!=PRIMARY_WAREHOUSE]
        if sel_pv:  df=df[df["Código Producto"].isin(sel_pv)]
        try:
            pv=df.pivot_table(index=["Código Producto","Nombre Producto"],
                              columns="Bodega",values="Stock",
                              aggfunc="sum",fill_value=0).reset_index()
            bc=[c for c in pv.columns if c not in("Código Producto","Nombre Producto")]
            rn={c:c.replace("Bodega ","").replace("BODEGA ","") for c in bc}
            pv=pv.rename(columns=rn); bcr=list(rn.values())
            pv=pv[pv[bcr].sum(axis=1)>0]
            for _bc in bcr: pv[_bc]=pv[_bc].fillna(0).astype(int)
            pv["TOTAL"]=pv[bcr].sum(axis=1).astype(int)

            # ── Tabla pivot via st.components (sticky real garantizado) ──
            C1W, C2W  = 110, 240
            num_cols  = bcr + ["TOTAL"]
            all_cols  = list(pv.columns)

            def _td(style, content=""):
                return f'<td style="{style}">{content}</td>'
            def _th(style, content=""):
                return f'<th style="{style}">{content}</th>'

            S_STICKY_TOP   = "position:sticky;top:0;z-index:{z};background:#f1f5f9;border-bottom:2px solid #cbd5e1;padding:7px 10px;font-size:10px;font-weight:700;text-transform:uppercase;color:#64748b;white-space:nowrap"
            S_STICKY_LEFT0 = "position:sticky;left:0;z-index:{z};background:{bg};border-right:1px solid #e2e8f0;border-bottom:1px solid #f1f5f9;padding:6px 10px;font-size:12px;white-space:nowrap;font-weight:600;font-family:monospace"
            S_STICKY_LEFT1 = f"position:sticky;left:{C1W}px;z-index:{{z}};background:{{bg}};border-right:2px solid #94a3b8;border-bottom:1px solid #f1f5f9;padding:6px 10px;font-size:12px"
            S_NUM          = "text-align:right;padding:6px 10px;border-bottom:1px solid #f1f5f9;font-family:monospace;font-size:12px;background:{bg}"
            S_TXT          = "padding:6px 10px;border-bottom:1px solid #f1f5f9;font-size:12px;background:{bg}"

            # ── Encabezados ────────────────────────────────────────
            hdrs = ""
            for i,c in enumerate(all_cols):
                if i == 0:
                    hdrs += _th(S_STICKY_TOP.format(z=5) + f";left:0;border-right:1px solid #cbd5e1;width:{C1W}px;min-width:{C1W}px", c)
                elif i == 1:
                    hdrs += _th(S_STICKY_TOP.format(z=5) + f";left:{C1W}px;border-right:2px solid #94a3b8;width:{C2W}px;min-width:{C2W}px", c)
                else:
                    align = "text-align:right;" if c in num_cols else ""
                    hdrs += _th(S_STICKY_TOP.format(z=3) + ";" + align, c)

            # ── Filas de datos ─────────────────────────────────────
            rows = ""; tots = {c:0 for c in num_cols}
            for ri,(_,row) in enumerate(pv.iterrows()):
                bg = "#f8fafc" if ri%2==0 else "#ffffff"
                cells = ""
                for i,c in enumerate(all_cols):
                    v = row[c]
                    if i == 0:
                        txt = str(v) if str(v) not in("nan","None","NaN") else ""
                        cells += _td(S_STICKY_LEFT0.format(z=2, bg=bg), txt)
                    elif i == 1:
                        txt = str(v) if str(v) not in("nan","None","NaN") else ""
                        cells += _td(S_STICKY_LEFT1.format(z=2, bg=bg), txt)
                    elif c in num_cols:
                        try:
                            iv = int(round(float(v))); tots[c] += iv
                            disp = f"{iv:,}" if iv != 0 else ""
                        except:
                            disp = ""
                        cells += _td(S_NUM.format(bg=bg), disp)
                    else:
                        cells += _td(S_TXT.format(bg=bg), str(v) if str(v) not in("nan","None","NaN") else "")
                rows += f"<tr>{cells}</tr>"

            # ── Fila TOTAL ─────────────────────────────────────────
            S_TOT = "background:#e0f2fe;font-weight:700;padding:7px 10px;border-top:2px solid #7dd3fc;color:#0369a1"
            tfooter = ""
            for i,c in enumerate(all_cols):
                if i == 0:
                    tfooter += _td(S_TOT + f";position:sticky;left:0;z-index:2", "TOTAL")
                elif i == 1:
                    tfooter += _td(S_TOT + f";position:sticky;left:{C1W}px;z-index:2;border-right:2px solid #94a3b8", "")
                elif c in num_cols:
                    tfooter += _td(S_TOT + ";text-align:right;font-family:monospace", f"{tots[c]:,}")
                else:
                    tfooter += _td(S_TOT, "")

            html_piv = (
                "<!DOCTYPE html><html><head><meta charset=\"UTF-8\">"
                "<style>"
                "*{box-sizing:border-box;margin:0;padding:0}"
                "body{font-family:Inter,Segoe UI,sans-serif;background:#f0f9ff}"
                ".wrap{border:1px solid #e2e8f0;border-radius:8px;overflow:auto;"
                "max-height:620px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08)}"
                ".zb{display:flex;align-items:center;gap:6px;margin-bottom:6px}"
                ".zb span{font-size:11px;font-weight:700;color:#94a3b8}"
                ".zb .inf{font-size:10px;font-weight:400}"
                ".zb button{background:#fff;border:1px solid #e2e8f0;border-radius:5px;"
                "padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;color:#475569}"
                ".zb button:hover{background:#0ea5e9;color:#fff;border-color:#0ea5e9}"
                "table{border-collapse:separate;border-spacing:0;width:max-content}"
                "tr:hover td{background:#e0f2fe!important}"
                "</style></head><body>"
                "<div class=\"zb\">"
                "<span>ZOOM</span>"
                "<button onclick=\"var t=document.getElementById('pt');var s=parseFloat(t.style.fontSize||'12');t.style.fontSize=(s-1)+'px'\">−</button>"
                "<button onclick=\"var t=document.getElementById('pt');var s=parseFloat(t.style.fontSize||'12');t.style.fontSize=(s+1)+'px'\">+</button>"
                "<button onclick=\"document.getElementById('pt').style.fontSize='12px'\">↺</button>"
                f"<span class=\"inf\">{len(pv):,} SKUs &middot; {len(bcr)} bodegas</span>"
                "</div>"
                "<div class=\"wrap\">"
                f"<table id=\"pt\"><thead><tr>{hdrs}</tr></thead>"
                f"<tbody>{rows}</tbody>"
                f"<tfoot><tr>{tfooter}</tr></tfoot>"
                "</table></div>"
                "<script>"
                "var w=document.querySelector('.wrap');"
                "var t0x,t0y,t0sl,t0st;"
                "w.addEventListener('touchstart',function(e){if(e.touches.length!==2)return;"
                "t0x=(e.touches[0].pageX+e.touches[1].pageX)/2;"
                "t0y=(e.touches[0].pageY+e.touches[1].pageY)/2;"
                "t0sl=w.scrollLeft;t0st=w.scrollTop;},{passive:true});"
                "w.addEventListener('touchmove',function(e){if(e.touches.length!==2)return;"
                "e.preventDefault();"
                "var cx=(e.touches[0].pageX+e.touches[1].pageX)/2;"
                "var cy=(e.touches[0].pageY+e.touches[1].pageY)/2;"
                "w.scrollLeft=t0sl-(cx-t0x);w.scrollTop=t0st-(cy-t0y);"
                "},{passive:false});"
                "w.addEventListener('wheel',function(e){"
                "if(Math.abs(e.deltaX)>Math.abs(e.deltaY)){e.preventDefault();w.scrollLeft+=e.deltaX;}"
                "else if(e.shiftKey){e.preventDefault();w.scrollLeft+=e.deltaY;}"
                "},{passive:false});"
                "</script></body></html>"
            )

            _components.html(html_piv, height=700, scrolling=False)
            dl3(pv,"sku_x_bodega","piv")
        except Exception as e: st.error(str(e))

with T_PIV:
    _render_tab_piv()

# ══ TAB 4 MUESTRAS ══════════════════════════════════════════════
@_fragment
def _render_tab_sam():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None: st.info("Ejecute el análisis.")
    else:
        _s1,_s2=st.tabs(["👥 Resumen por Cliente","📋 Movimientos TRA"])

        # ── Sub-tab 1: Resumen por Cliente ────────────────────────
        with _s1:
            df=r.samples_by_client.copy()
            if df.empty:
                st.warning("Sin muestras.")
            else:
                c1,c2=st.columns([3,1])
                flt=c1.text_input("🔍 Cliente","",key="sa_f",placeholder="Nombre...")
                pend=c2.checkbox("Solo pendientes",True,key="sa_p")
                df=pfilt(df,flt,cols=("Cliente",))
                if pend and "Stock en Cliente" in df.columns:
                    df=df[df["Stock en Cliente"]>0]
                m1,m2,m3=st.columns(3)
                m1.metric("Enviadas",  int(df.get("Entregadas",    pd.Series([0])).sum()))
                m2.metric("Devueltas", int(df.get("Devueltas",     pd.Series([0])).sum()))
                m3.metric("Saldo",     int(df.get("Stock en Cliente",pd.Series([0])).sum()))
                nc=[c for c in df.columns if c not in("Código Producto","Nombre Producto","Última Devolución")]
                _freeze=2 if "Código Producto" in df.columns else 1
                _comp_tbl(df,nc,"sam",freeze_cols=_freeze,height=500,title=f"{len(df):,} clientes")
                dl3(df,"muestras","sam")

        # ── Sub-tab 2: Movimientos TRA detallados ─────────────────
        with _s2:
            if eng.raw_df is None:
                st.info("Cargue un archivo.")
            else:
                # Filtrar solo TRA del raw_df
                _raw=eng.raw_df.copy()
                if excl_s: _raw=_raw[~_raw["Código Producto"].isin(excl_s)]
                _tra=_raw[_raw["Tipo"].fillna("").str.upper()=="TRA"].copy()

                if _tra.empty:
                    st.warning("No hay movimientos de transferencia (TRA).")
                else:
                    # Normalizar fecha
                    _tra["Fecha"]=pd.to_datetime(_tra["Fecha"],errors="coerce")
                    _tra["Fecha_str"]=_tra["Fecha"].dt.strftime("%d/%m/%Y").fillna("")

                    # Tipo de movimiento — SIN emoji para comparaciones limpias
                    from app.config import PRIMARY_WAREHOUSE as _PW
                    _tra["Mov"]=_tra.apply(
                        lambda r: "Enviada"  if r["Bodega Origen"]==_PW else
                                  "Devuelta" if r["Bodega Destino"]==_PW else
                                  "Interna", axis=1)
                    _tra["Bodega"]=_tra.apply(
                        lambda r: r["Bodega Destino"] if r["Bodega Origen"]==_PW
                                  else r["Bodega Origen"], axis=1)

                    # Filtros
                    _fc1,_fc2,_fc3=st.columns([3,3,2])
                    _bods_tra=["Todas"]+sorted(_tra["Bodega"].dropna().unique().tolist())
                    _flt_cli =_fc1.text_input("🔍 Bodega/Cliente","",key="tra_f",placeholder="Nombre...")
                    _flt_sku =_fc2.text_input("🔍 SKU","",key="tra_s",placeholder="Código o nombre...")
                    _mov_flt =_fc3.selectbox("Movimiento",["Todos","📤 Enviada","📥 Devuelta","↔ Interna"],key="tra_m")

                    _t=_tra.copy()
                    if _flt_cli:
                        _t=_t[_t["Bodega"].fillna("").str.upper().str.contains(_flt_cli.upper())]
                    if _flt_sku:
                        _mask=(_t["Código Producto"].fillna("").str.upper().str.contains(_flt_sku.upper()) |
                               _t["Nombre Producto"].fillna("").str.upper().str.contains(_flt_sku.upper()))
                        _t=_t[_mask]
                    if _mov_flt!="Todos":
                        _t=_t[_t["Mov"]==_mov_flt]

                    # Orden: Bodega → SKU → Fecha
                    _t=_t.sort_values(["Bodega","Código Producto","Fecha"],na_position="last")

                    # Columnas: SKU | Nombre | N°Reg | MOV | Fecha | Desc | Cant
                    _COLS=["Código Producto","Nombre Producto","N° Registro","Mov","Fecha","Descripción","Cantidad"]
                    # Anchos MÍNIMOS — el usuario puede expandir arrastrando
                    _CW={"Código Producto":90,"Nombre Producto":200,
                         "N° Registro":138,"Mov":90,"Fecha":82,
                         "Descripción":280,"Cantidad":62}
                    _BOD_BG="#1e3a5f"; _BOD_FG="#7dd3fc"
                    _SKU_BG="#dbeafe"; _SKU_FG="#1d4ed8"
                    _TOT_BG="#1e40af"; _TOT_FG="#ffffff"
                    _DAT_BG=["#ffffff","#f0f9ff"]

                    # Helper <td> — siempre nowrap + overflow ellipsis para textos largos
                    def _td_tra(val, bg, extra_style=""):
                        v = str(val) if str(val) not in ("nan","None","NaN","") else ""
                        return ("<td style='background:"+bg
                                +";border-bottom:1px solid #f1f5f9;padding:4px 8px;"
                                "font-size:12px;white-space:nowrap;overflow:hidden;"
                                "text-overflow:ellipsis;max-width:400px;"
                                +extra_style+"' title='"+v.replace("'","")+"'>"+v+"</td>")

                    # Iconos MOV — flecha verde ↑ Enviada, flecha azul ↓ Devuelta
                    _MOV_ICON={"Enviada":"<span style='color:#059669;font-weight:900'>&#8593;</span> Enviada",
                               "Devuelta":"<span style='color:#0284c7;font-weight:900'>&#8595;</span> Devuelta",
                               "Interna":"&#8596; Interna"}

                    # Headers con ancho mínimo + position:relative para resize handle
                    TH_BASE=("position:sticky;top:0;z-index:3;background:#1e3a5f;"
                             "border-bottom:2px solid #0ea5e9;padding:6px 8px;font-size:10px;"
                             "font-weight:700;text-transform:uppercase;color:#e0f2fe;"
                             "white-space:nowrap;position:relative;overflow:visible")
                    _hdrs_tra="".join(
                        "<th style='"+TH_BASE
                        +";min-width:"+str(_CW.get(c,80))+"px"
                        +(";text-align:right" if c=="Cantidad" else "")
                        +"'>"+c
                        # Handle de resize (div invisible en borde derecho)
                        +"<div style='position:absolute;top:0;right:0;bottom:0;width:5px;"
                         "cursor:col-resize;z-index:10' "
                         "onmousedown='startResize(event,this.parentElement)'></div>"
                        +"</th>"
                        for c in _COLS)

                    _rows_html=""; _grand_env=0; _grand_dev=0; _ri=0

                    for _bod in _t["Bodega"].dropna().unique():
                        _df_b=_t[_t["Bodega"]==_bod]
                        _rows_html+=("<tr><td colspan='"+str(len(_COLS))
                                     +"' style='background:"+_BOD_BG
                                     +";padding:6px 12px;font-size:12px;font-weight:800;color:"+_BOD_FG
                                     +";border-top:3px solid #0ea5e9;border-bottom:2px solid #0ea5e9'>"
                                     +"🏪 "+str(_bod)+"</td></tr>")
                        _bod_env=0; _bod_dev=0

                        for _sku in _df_b["Código Producto"].dropna().unique():
                            _df_s=_df_b[_df_b["Código Producto"]==_sku]
                            _sku_nom=str(_df_s["Nombre Producto"].iloc[0]) if len(_df_s)>0 else ""
                            _sku_env=0; _sku_dev=0

                            for __,_row in _df_s.iterrows():
                                _bg=_DAT_BG[_ri%2]; _ri+=1
                                _qty=int(float(_row.get("Cantidad",0))) if str(_row.get("Cantidad","")) not in("","nan","None") else 0
                                _mov=str(_row.get("Mov",""))
                                # Enviada = positivo (+), Devuelta = negativo (−)
                                if _mov=="Enviada":
                                    _sku_env+=_qty
                                    _qty_disp="+"+str(_qty)
                                    _qty_col="#065f46"  # verde oscuro
                                elif _mov=="Devuelta":
                                    _sku_dev+=_qty
                                    _qty_disp="-"+str(_qty)
                                    _qty_col="#b91c1c"  # rojo oscuro
                                else:
                                    _qty_disp=str(_qty)
                                    _qty_col="#374151"
                                _cells=(
                                    _td_tra(_row.get("Código Producto",""), _bg)
                                    +_td_tra(_sku_nom, _bg)
                                    +_td_tra(_row.get("Código",""), _bg, "font-family:monospace;font-size:11px")
                                    # MOV con icono de flecha coloreada
                                    +"<td style='background:"+_bg+";border-bottom:1px solid #f1f5f9;"
                                     "padding:4px 8px;font-size:12px;white-space:nowrap'>"
                                     +_MOV_ICON.get(_mov,_mov)+"</td>"
                                    +_td_tra(_row.get("Fecha_str",""), _bg, "white-space:nowrap")
                                    +_td_tra(str(_row.get("Descripción",""))[:80], _bg)
                                    +"<td style='background:"+_bg+";border-bottom:1px solid #f1f5f9;"
                                     "padding:4px 8px;text-align:right;font-family:monospace;"
                                     "font-weight:700;color:"+_qty_col+";white-space:nowrap'>"
                                     +_qty_disp+"</td>"
                                )
                                _rows_html+="<tr>"+_cells+"</tr>"

                            # Subtotal SKU: total = Enviadas - Devueltas (suma simple con signo)
                            _sku_neto=_sku_env-_sku_dev
                            _bod_env+=_sku_env; _bod_dev+=_sku_dev
                            _neto_col="#065f46" if _sku_neto>0 else ("#b91c1c" if _sku_neto<0 else "#374151")
                            _ss=("background:"+_SKU_BG+";border-top:1px solid #93c5fd;"
                                 "border-bottom:1px solid #93c5fd;padding:5px 8px;"
                                 "font-size:11px;font-weight:700;color:"+_SKU_FG+";")
                            _s_cells=(
                                "<td style='"+_ss+"'>"+_sku+"</td>"
                                +"<td style='"+_ss+"'>"+_sku_nom+"</td>"
                                +"<td style='"+_ss+"'>Subtotal</td>"
                                +"<td style='"+_ss+"'></td>"
                                +"<td style='"+_ss+"'></td>"
                                +"<td style='"+_ss+"'>+"+str(_sku_env)+" / -"+str(_sku_dev)+"</td>"
                                +"<td style='"+_ss+"text-align:right;font-family:monospace;color:"+_neto_col+"'>"
                                +str(_sku_neto)+"</td>"
                            )
                            _rows_html+="<tr>"+_s_cells+"</tr>"

                        # Total Bodega
                        _bod_neto=_bod_env-_bod_dev
                        _grand_env+=_bod_env; _grand_dev+=_bod_dev
                        _bneto_col="#065f46" if _bod_neto>0 else ("#b91c1c" if _bod_neto<0 else "#374151")
                        _bs=("background:"+_TOT_BG+";border-top:2px solid #3b82f6;"
                             "border-bottom:2px solid #3b82f6;padding:6px 8px;"
                             "font-size:12px;font-weight:800;color:"+_TOT_FG+";")
                        _b_cells=(
                            "<td style='"+_bs+"'>TOTAL "+str(_bod)+"</td>"
                            +"<td style='"+_bs+"'></td>"
                            +"<td style='"+_bs+"'></td>"
                            +"<td style='"+_bs+"'></td>"
                            +"<td style='"+_bs+"'></td>"
                            +"<td style='"+_bs+"'>+"+str(_bod_env)+" / -"+str(_bod_dev)+"</td>"
                            +"<td style='"+_bs+"text-align:right;font-family:monospace;color:"
                            +_bneto_col+"'>"+str(_bod_neto)+"</td>"
                        )
                        _rows_html+="<tr>"+_b_cells+"</tr>"

                    # Gran total
                    _grand_neto=_grand_env-_grand_dev
                    _gneto_col="#065f46" if _grand_neto>0 else ("#b91c1c" if _grand_neto<0 else "#374151")
                    _gs=("background:#0ea5e9;color:#fff;font-weight:800;"
                         "border-top:3px solid #0284c7;padding:7px 8px;")
                    _gt_tra=(
                        "<td style='"+_gs+"'>TOTAL GENERAL</td>"
                        +"<td style='"+_gs+"'></td>"
                        +"<td style='"+_gs+"'></td>"
                        +"<td style='"+_gs+"'></td>"
                        +"<td style='"+_gs+"'></td>"
                        +"<td style='"+_gs+"'>+"+str(_grand_env)+" / -"+str(_grand_dev)+"</td>"
                        +"<td style='"+_gs+"text-align:right;font-family:monospace;color:"
                        +_gneto_col+"'>"+str(_grand_neto)+"</td>"
                    )
                    _n_mov=str(len(_t))
                    _html_tra=(
                        "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
                        "<style>*{box-sizing:border-box;margin:0;padding:0}"
                        "body{font-family:Inter,sans-serif;background:#f0f9ff;padding:2px}"
                        ".zb{display:flex;align-items:center;gap:6px;margin-bottom:6px}"
                        ".zb span{font-size:11px;font-weight:700;color:#94a3b8}"
                        ".zb .i{font-size:10px;font-weight:400}"
                        ".zb button{background:#fff;border:1px solid #e2e8f0;border-radius:5px;"
                        "padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;color:#475569}"
                        ".zb button:hover{background:#0ea5e9;color:#fff}"
                        ".wt{overflow:auto;max-height:620px;border:1px solid #e2e8f0;"
                        "border-radius:8px;background:#fff}"
                        "table{border-collapse:separate;border-spacing:0;"
                        "width:max-content;font-size:12px;table-layout:auto}"
                        "td{white-space:nowrap}"
                        "tr:hover td{filter:brightness(.95)}"
                        "</style></head><body>"
                        "<div class='zb'><span>ZOOM</span>"
                        "<button onclick='tZ(-1)'>&#8722;</button>"
                        "<button onclick='tZ(1)'>+</button>"
                        "<button onclick='tR()'>&#8635;</button>"
                        "<span class='i'>"+_n_mov+" movimientos</span></div>"
                        "<div class='wt'><table id='trt'>"
                        "<thead><tr>"+_hdrs_tra+"</tr></thead>"
                        "<tbody>"+_rows_html+"</tbody>"
                        "<tfoot><tr>"+_gt_tra+"</tr></tfoot>"
                        "</table></div>"
                        "<script>"
                        "function tZ(d){var t=document.getElementById('trt');"
                        "var s=parseFloat(t.style.fontSize||'12');"
                        "t.style.fontSize=Math.min(20,Math.max(8,s+d))+'px';}"
                        "function tR(){document.getElementById('trt').style.fontSize='12px';}"
                        "var _rTh=null,_rX=0,_rW=0;"
                        "function startResize(e,th){"
                        "e.preventDefault();e.stopPropagation();"
                        "_rTh=th;_rX=e.pageX;_rW=th.offsetWidth;"
                        "document.body.style.cursor='col-resize';"
                        "document.body.style.userSelect='none';}"
                        "document.addEventListener('mousemove',function(e){"
                        "if(!_rTh)return;"
                        "var w=Math.max(40,_rW+(e.pageX-_rX));"
                        "_rTh.style.width=w+'px';_rTh.style.minWidth=w+'px';});"
                        "document.addEventListener('mouseup',function(){"
                        "if(!_rTh)return;_rTh=null;"
                        "document.body.style.cursor='';"
                        "document.body.style.userSelect='';});"
                        "var w=document.querySelector('.wt');"
                        "w.addEventListener('wheel',function(e){"
                        "if(Math.abs(e.deltaX)>Math.abs(e.deltaY)){e.preventDefault();w.scrollLeft+=e.deltaX;}"
                        "else if(e.shiftKey){e.preventDefault();w.scrollLeft+=e.deltaY;}"
                        "},{passive:false});"
                        "</script></body></html>"
                    )
                    st.markdown(f"**{len(_t):,} movimientos** encontrados")
                    _components.html(_html_tra, height=730, scrolling=False)
                    _exp_c=["Bodega","Código Producto","Nombre Producto","Código","Fecha_str","Descripción","Mov","Cantidad"]
                    _exp_df=_t[[c for c in _exp_c if c in _t.columns]].rename(columns={"Fecha_str":"Fecha","Código":"N° Registro"})
                    dl3(_exp_df,"muestras_tra","tra_exp")

with T_SAM:
    _render_tab_sam()

# ══ TAB 5 ANÁLISIS PERÍODO ══════════════════════════════════════
@_fragment
def _render_tab_ana():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if eng.raw_df is None: st.info("Cargue un archivo.")
    else:
        c1,c2,c3=st.columns([2,2,1])
        df=date.today()-timedelta(days=365)
        d_f=c1.date_input("Desde",df,format="DD/MM/YYYY",key="an_f")
        d_t=c2.date_input("Hasta",date.today(),format="DD/MM/YYYY",key="an_t")
        with c3:
            st.markdown("")
            go=st.button("▶ Calcular",type="primary",key="an_btn")
        if go:
            if d_f>d_t: st.error("Fecha inicial > final.")
            else:
                with st.spinner("Calculando..."):
                    dfa=eng.raw_df.copy()
                    if excl_s: dfa=dfa[~dfa["Código Producto"].isin(excl_s)]
                    dfa=dfa[(dfa["Fecha"]>=pd.Timestamp(d_f))&(dfa["Fecha"]<=pd.Timestamp(d_t))]
                    ref=dfa["Referencia"].fillna("").astype(str).str.upper()
                    typ=dfa["Tipo"].fillna("").astype(str).str.upper()
                    vdf=dfa[(typ=="EGR")&ref.str.startswith("FAC")].copy()
                    cdf=dfa[(typ=="ING")&ref.str.startswith("FAC")].copy()
                    cpq,cpv,cpm=defaultdict(float),defaultdict(float),{}
                    for _,row in cdf.sort_values(["Código Producto","Fecha"]).iterrows():
                        sku=row["Código Producto"]; qty=float(row["Cantidad"]) or 1
                        vt=float(row["Valor Total"]); cu=vt/qty if qty>0 and vt>0 else 0
                        if cu>0:
                            nq=cpq[sku]+qty; nv=cpv[sku]+vt; cpq[sku]=nq; cpv[sku]=nv; cpm[sku]=nv/nq
                    st.session_state["an_v"]=vdf; st.session_state["an_cpm"]=cpm
        if "an_v" in st.session_state and not st.session_state["an_v"].empty:
            vdf=st.session_state["an_v"]; cpm=st.session_state["an_cpm"]
            s1,s2=st.tabs(["📊 Top 10 Vendidos","💰 Top 10 Rentabilidad"])
            with s1:
                t10v=(vdf.groupby(["Código Producto","Nombre Producto"])
                      .agg(Unidades=("Cantidad","sum"),Ventas=("Valor Total","sum"))
                      .reset_index().nlargest(10,"Ventas")
                      .sort_values("Ventas",ascending=False))
                if not t10v.empty:
                    a,b=st.columns([1,1])
                    with a: st.markdown(tbl(t10v,["Unidades","Ventas"],"av"),unsafe_allow_html=True)
                    # Gráfica también ordenada de mayor a menor
                    with b: st.bar_chart(t10v.sort_values("Ventas",ascending=False)
                                            .set_index("Código Producto")["Ventas"])
                    dl3(t10v,"top10_vendidos","av")
            with s2:
                rows_r=[]
                for (sku,nom),grp in vdf.groupby(["Código Producto","Nombre Producto"]):
                    v=float(grp["Valor Total"].sum())
                    ct=sum(float(rv["Cantidad"])*cpm.get(sku,0) for _,rv in grp.iterrows())
                    rent=v-ct
                    rows_r.append({"Código":sku,"Nombre":nom,"Ventas":round(v,2),"Costo":round(ct,2),"Rentabilidad":round(rent,2),"Margen%":round(rent/v*100,2) if v else 0})
                t10r=pd.DataFrame(rows_r).nlargest(10,"Rentabilidad")
                if not t10r.empty:
                    a,b=st.columns([1,1])
                    with a: st.markdown(tbl(t10r,["Ventas","Costo","Rentabilidad","Margen%"],"ar"),unsafe_allow_html=True)
                    with b: st.bar_chart(t10r.set_index("Código")["Rentabilidad"])
                    dl3(t10r,"top10_rentabilidad","ar")

with T_ANA:
    _render_tab_ana()

# ══ TAB 6 ROTACIÓN ══════════════════════════════════════════════
@_fragment
def _render_tab_rot():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None: st.info("Ejecute el análisis.")
    else:
        c1,c2,c3,c4,c5=st.columns([2,2,2,2,3])
        lt_m=c1.number_input("🚢 Marítimo(d)",1,365,45,key="rm")
        lt_a=c2.number_input("✈ Aéreo(d)",1,180,15,key="ra")
        saf=c3.number_input("Stock Seg.(d)",0,180,15,key="rs")
        cob=c4.number_input("🎯 Cobertura(d)",1,730,60,key="rcob",
                            help="Días totales de cobertura deseados. Sug.Cob. "
                                 "= Consumo_diario × Cobertura − Stock_actual.")
        flt_r=c5.text_input("🔍 Filtrar SKU","",key="rf",placeholder="Código o nombre...")

        # ── Ventana para el cálculo del consumo ─────────────────────
        vc1, vc2, vc3 = st.columns([2, 2, 2])
        with vc1:
            ventana = st.selectbox(
                "📅 Ventana consumo",
                list(VENTANA_OPTS.keys()), index=1,
                key="rc_ventana",
                help="Período para calcular el consumo promedio. Si la empresa "
                     "ha crecido, usar Últimos 90d es más realista que Todo el período.",
            )
        _cus_s = _cus_e = None
        if ventana == "Personalizado":
            _dmin = r.filtered["Fecha"].min()
            _dmax = r.filtered["Fecha"].max()
            with vc2:
                _cus_s = st.date_input("Desde", value=(_dmax - pd.Timedelta(days=90)).date(),
                                        format="DD/MM/YYYY", key="rc_win_s")
            with vc3:
                _cus_e = st.date_input("Hasta", value=_dmax.date(),
                                        format="DD/MM/YYYY", key="rc_win_e")

        calc_r=st.button("▶ Calcular Rotación",type="primary",key="rc")
        if calc_r:
            sku=r.sku_summary.copy()
            if excl_s: sku=sku[~sku["Código Producto"].isin(excl_s)]

            # Ventas por SKU SOLO en la ventana seleccionada
            days, _ventas_win, _win_min, _win_max = _compute_window_sales(
                r, ventana, custom_start=_cus_s, custom_end=_cus_e)
            st.session_state["_rc_win_info"] = {
                "label": ventana, "days": days,
                "min": _win_min, "max": _win_max,
                "n_skus_con_ventas": len(_ventas_win),
            }
            def make_rot(lt):
                rows=[]
                for _,row in sku.iterrows():
                    cod=row["Código Producto"]; nom=row["Nombre Producto"]
                    stk=max(0.0,float(row.get("Stock Disponible",0)))
                    vtas=float(_ventas_win.get(cod, 0))
                    # Costo promedio y valores $
                    _comp_u = float(row.get("Compras", 0) or 0)
                    _val_comp = float(row.get("Valor Compras", row.get("Valor_Compras", 0)) or 0)
                    costo_prom = (_val_comp / _comp_u) if _comp_u > 0 else 0.0
                    val_stock  = stk * costo_prom
                    cogs_aprox = vtas * costo_prom
                    cons=vtas/days
                    rot_u=vtas/stk if stk>0 else(999 if vtas>0 else 0)
                    rot_d=cogs_aprox/val_stock if val_stock>0 else 0.0
                    dinv=stk/cons if cons>0 else(0 if stk==0 else 9999)
                    sug_lt=max(0.0,cons*(lt+saf)-stk)
                    sug_cob=max(0.0,cons*cob-stk)
                    pr=cons*lt
                    if cons>0 and stk==0: alrt="SIN STOCK"
                    elif cons>0 and rot_u>2 and dinv<lt: alrt="BEST-SELLER"
                    elif cons>0 and dinv<lt: alrt="CRÍTICO"
                    elif cons>0 and dinv<lt+saf: alrt="BAJO"
                    elif cons==0: alrt="SIN VENTA"
                    else: alrt="OK"
                    rows.append({"Código":cod,"Nombre":nom,"Stock":int(stk),"Ventas(u)":int(vtas),
                                 "Cons/día":round(cons,3),"P.Reorden":round(pr,1),
                                 "Días Inv.":round(dinv,1),
                                 "Rotación":round(rot_u,2),
                                 "Rotación $":round(rot_d,2),
                                 "Sug.LT+SS":int(round(sug_lt)),
                                 "Sug.Cob.":int(round(sug_cob)),
                                 "Estado":alrt})
                order={"SIN STOCK":0,"BEST-SELLER":1,"CRÍTICO":2,"BAJO":3,"OK":4,"SIN VENTA":5}
                rows.sort(key=lambda x:(order.get(x["Estado"],6),-x["Ventas(u)"]))
                return pd.DataFrame(rows)
            st.session_state["rot_m"]=make_rot(lt_m)
            st.session_state["rot_a"]=make_rot(lt_a)

        CK={"SIN STOCK":"#fca5a5","BEST-SELLER":"#fed7aa","CRÍTICO":"#fca5a5",
            "BAJO":"#fde68a","OK":"#bbf7d0","SIN VENTA":"#e2e8f0"}
        nr=["Stock","Ventas(u)","Cons/día","P.Reorden","Días Inv.",
            "Rotación","Rotación $","Sug.LT+SS","Sug.Cob."]

        def rot_table(df_r, uid):
            hdrs="".join(f"<th class='{'n' if c in nr else ''}'>{c}</th>" for c in df_r.columns)
            rows=""; tot_lt=0; tot_cob=0
            for _,row in df_r.iterrows():
                cells=""
                for c in df_r.columns:
                    v=row[c]; disp=str(v) if str(v) not in("nan","None","NaN") else ""
                    if c=="Estado":
                        bg=CK.get(v,"")
                        cells+=f"<td style='background:{bg};color:#111;font-weight:700'>{disp}</td>"
                    elif c in nr:
                        cells+=f"<td class='n'>{disp}</td>"
                        if c=="Sug.LT+SS":
                            try: tot_lt+=int(v)
                            except: pass
                        elif c=="Sug.Cob.":
                            try: tot_cob+=int(v)
                            except: pass
                    elif c=="Nombre":
                        # Truncar nombre largo con tooltip (ver completo al hover)
                        _full = disp.replace('"','&quot;').replace('<','&lt;').replace('>','&gt;')
                        _disp = disp if len(disp) <= 50 else disp[:47] + "…"
                        cells+=f'<td class="nom" title="{_full}">{_disp}</td>'
                    else: cells+=f"<td>{disp}</td>"
                rows+=f"<tr>{cells}</tr>"
            def _tot_cell(c, i):
                if c=="Sug.LT+SS": return f"<td class='n'>{tot_lt:,}</td>"
                if c=="Sug.Cob.":  return f"<td class='n'>{tot_cob:,}</td>"
                if i==0: return "<td><b>TOTAL</b></td>"
                return "<td></td>"
            tcells="".join(_tot_cell(c, i) for i,c in enumerate(df_r.columns))
            return f"""<div class="zb">
  <button onclick="var t=document.getElementById('t_{uid}');t.style.fontSize=Math.max(8,parseFloat(getComputedStyle(t).fontSize)-1)+'px'">−</button>
  <button onclick="var t=document.getElementById('t_{uid}');t.style.fontSize=Math.min(20,parseFloat(getComputedStyle(t).fontSize)+1)+'px'">+</button>
  <button onclick="document.getElementById('t_{uid}').style.fontSize='12px'">↺</button>
  <span style="color:{MUTED};font-size:10px">{len(df_r):,} SKUs | ⚠️ {len(df_r[df_r['Estado'].isin(['SIN STOCK','CRÍTICO','BEST-SELLER'])])} críticos</span>
</div>
<div class="tc rot-tc"><table class="it it-rot" id="t_{uid}">
<thead><tr>{hdrs}</tr></thead><tbody>{rows}</tbody>
<tfoot><tr class="tot">{tcells}</tr></tfoot>
</table></div>"""

        if "rot_m" in st.session_state:
            _wi = st.session_state.get("_rc_win_info", {})
            if _wi:
                st.info(
                    f"📅 **Ventana de consumo**: {_wi['label']} · "
                    f"{_wi['min'].strftime('%d/%m/%Y')} → {_wi['max'].strftime('%d/%m/%Y')} "
                    f"({_wi['days']} días) · "
                    f"{_wi['n_skus_con_ventas']:,} SKUs con ventas en la ventana."
                )
            s1,s2=st.tabs(["🚢 Marítimo","✈ Aéreo"])
            for tab_r,key,lbl in [(s1,"rot_m","Marítimo"),(s2,"rot_a","Aéreo")]:
                with tab_r:
                    df_r=pfilt(st.session_state[key],flt_r,cols=("Código","Nombre"))
                    st.markdown(rot_table(df_r,f"rot_{lbl}"),unsafe_allow_html=True)
                    with st.expander("ℹ Leyenda — qué significa cada columna y cada estado"):
                        st.markdown("""
**Columnas de la tabla**

| Columna | Qué es | Cómo se calcula |
|---|---|---|
| **Código / Nombre** | Identificación del producto | — |
| **Stock** | Unidades disponibles hoy | Del análisis global |
| **Ventas (u)** | Unidades vendidas en la ventana elegida | Σ EGR FAC en la ventana |
| **Cons/día** | Ritmo actual de consumo | Ventas ÷ Días de la ventana |
| **P. Reorden** | Nivel de stock al que debes disparar una compra para que llegue justo antes de agotarse | Cons/día × Lead time |
| **Días Inv.** | Cuántos días te alcanza el stock al ritmo actual | Stock ÷ Cons/día |
| **Rotación** | Veces que rotas el stock en la ventana | Ventas ÷ Stock |
| **Rotación $** | Lo mismo pero en dinero (indicador financiero) | COGS ÷ Valor_stock |
| **Sug. LT+SS** | Compra para cubrir Lead time + Stock Seguridad | max(0, Cons/día × (LT+SS) − Stock) |
| **Sug. Cob.** | Compra para alcanzar la Cobertura objetivo (días) | max(0, Cons/día × Cobertura − Stock) |
| **Estado** | Clasificación por urgencia (ver abajo) | — |

---

**Estados — por qué se asigna cada uno**

| Estado | Significado | Regla que lo activa |
|---|---|---|
| 🔴 **SIN STOCK** | Hubo ventas pero el stock quedó en 0. Pierdes ventas ahora mismo. | `Ventas > 0` **AND** `Stock = 0` |
| 🔥 **BEST-SELLER** | Producto de alta demanda en riesgo de quiebre. Urgente reponer. | `Rotación > 2×` **AND** `Días Inv. < Lead time` |
| 🔴 **CRÍTICO** | No alcanzas a sobrevivir el tiempo de entrega si pides hoy. | `Días Inv. < Lead time` |
| 🟡 **BAJO** | Cubres el lead time pero no el stock de seguridad. Pide pronto. | `Días Inv. < Lead time + Stock Seguridad` |
| 🟢 **OK** | Stock suficiente para cubrir lead time + seguridad. Tranquilo. | `Días Inv. ≥ Lead time + Stock Seguridad` |
| ⚪ **SIN VENTA** | No hubo ventas en la ventana. Puede ser estacional o descontinuado. | `Ventas = 0` en la ventana |

**Tip**: si un producto aparece como 🔥 BEST-SELLER, **prioriza el aéreo** aunque sea más caro — el costo del quiebre supera al flete extra.
                        """)
                    dl3(df_r,f"rotacion_{lbl.lower()}",f"rot_{lbl}")

with T_ROT:
    _render_tab_rot()

# ══ TAB: CÁLCULOS PASO A PASO ═══════════════════════════════════
@_fragment
def _render_tab_cal():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None or eng.raw_df is None:
        st.info("Ejecute el análisis primero (sidebar → Calcular).")
        return

    st.markdown("### 📐 Cálculos de Rotación y Compras — paso a paso")
    st.caption(
        "Elige un SKU y ajusta los parámetros. La tabla muestra cómo se "
        "obtiene cada valor, fórmula por fórmula."
    )

    # Selector de SKU (predictivo con format_func nativo).
    # Excluye los SKUs globalmente excluidos (sidebar).
    _opts = (eng.raw_df[["Código Producto","Nombre Producto"]]
             .drop_duplicates().sort_values("Código Producto"))
    _opts["Código Producto"] = _opts["Código Producto"].astype(str)
    if eng.excluded_skus:
        _opts = _opts[~_opts["Código Producto"].isin(eng.excluded_skus)]
    _codes = _opts["Código Producto"].astype(str).tolist()
    _names = _opts["Nombre Producto"].astype(str).tolist()
    _labels = [f"{c} — {n}" for c,n in zip(_codes, _names)]
    if not _labels:
        st.warning("No hay SKUs disponibles.")
        return

    cs1, cs2 = st.columns([3, 2])
    with cs1:
        sel_label = st.selectbox("🔍 SKU a analizar", _labels, key="cal_sku",
                                  placeholder="Escribe código o nombre…")
    sel_idx = _labels.index(sel_label) if sel_label in _labels else 0
    sel_code = _codes[sel_idx]; sel_name = _names[sel_idx]

    p1, p2, p3 = st.columns(3)
    with p1:
        lt = st.number_input("🚢/✈ Lead time (d)", 1, 365, value=45, key="cal_lt",
                              help="Tiempo en días desde que pides hasta que recibes.")
    with p2:
        saf = st.number_input("🛡 Stock seguridad (d)", 0, 180, value=15, key="cal_saf",
                               help="Buffer en días contra variabilidad o demora.")
    with p3:
        cob = st.number_input("🎯 Cobertura objetivo (d)", 1, 730, value=60, key="cal_cob",
                               help="Cuántos días de demanda quieres tener cubiertos.")

    # ── Ventana para el consumo ──
    vc1, vc2, vc3 = st.columns([2, 2, 2])
    with vc1:
        ventana = st.selectbox(
            "📅 Ventana consumo",
            list(VENTANA_OPTS.keys()), index=1,
            key="cal_ventana",
            help="Período para consumo promedio. 'Últimos 90 días' es estándar "
                 "en retail; evita diluir el ritmo actual con ventas antiguas.",
        )
    _cus_s = _cus_e = None
    if ventana == "Personalizado":
        _dmin = r.filtered["Fecha"].min()
        _dmax = r.filtered["Fecha"].max()
        with vc2:
            _cus_s = st.date_input("Desde",
                value=(_dmax - pd.Timedelta(days=90)).date(),
                format="DD/MM/YYYY", key="cal_win_s")
        with vc3:
            _cus_e = st.date_input("Hasta", value=_dmax.date(),
                format="DD/MM/YYYY", key="cal_win_e")

    # ── Datos del SKU ──
    sku_sel = r.sku_summary[r.sku_summary["Código Producto"] == sel_code]
    if sku_sel.empty:
        st.warning(
            f"El SKU **{sel_code}** no aparece en el análisis actual "
            "(puede estar excluido o sin movimientos en el período)."
        )
        return
    row = sku_sel.iloc[0]

    # Ventas y días calculados desde la ventana (NO del histórico completo)
    days, _ventas_win, _win_min, _win_max = _compute_window_sales(
        r, ventana, custom_start=_cus_s, custom_end=_cus_e)

    stk   = max(0.0, float(row.get("Stock Disponible", 0)))
    vtas  = float(_ventas_win.get(sel_code, 0))
    comp_u= float(row.get("Compras", 0) or 0)
    val_c = float(row.get("Valor Compras", row.get("Valor_Compras", 0)) or 0)
    costo = (val_c / comp_u) if comp_u > 0 else 0.0
    val_stk  = stk * costo
    cogs_apx = vtas * costo

    cons    = vtas / days if days else 0
    rot_u   = (vtas / stk) if stk > 0 else (999 if vtas > 0 else 0)
    rot_d   = (cogs_apx / val_stk) if val_stk > 0 else 0.0
    dinv_p  = (stk / cons) if cons > 0 else (0 if stk == 0 else 9999)
    dinv_a  = (365 / rot_u) if (0 < rot_u < 999) else (9999 if rot_u == 0 else 0)
    dem_lt  = cons * lt
    ss_u    = cons * saf
    reord   = dem_lt + ss_u
    nec_cob = cons * cob
    sug_cob = max(0.0, nec_cob - stk)
    sug_lt  = max(0.0, cons * (lt + saf) - stk)

    # Alerta
    if cons > 0 and stk == 0:
        alerta = "🔴 SIN STOCK"
    elif cons > 0 and rot_u > 2 and dinv_p < lt:
        alerta = "🔥 BEST-SELLER EN RIESGO"
    elif cons > 0 and dinv_p < lt:
        alerta = "🔴 CRÍTICO"
    elif cons > 0 and dinv_p < lt + saf:
        alerta = "🟡 BAJO"
    elif cons == 0:
        alerta = "⚪ SIN VENTA"
    else:
        alerta = "🟢 OK"

    st.markdown(f"#### {sel_name}  ·  `{sel_code}`")
    st.caption(
        f"📅 **Ventana de consumo**: {ventana} · "
        f"{_win_min.strftime('%d/%m/%Y')} → {_win_max.strftime('%d/%m/%Y')} ({days} días)"
    )

    # ── Tabla paso a paso ──
    pasos = [
        ("1",  "Días de la ventana",      "fecha_fin − fecha_inicio + 1",     f"{days:,} d"),
        ("2",  "Stock actual",             "del análisis (hoy)",              f"{int(stk):,} u"),
        ("3",  "Ventas en la ventana",     "Σ unidades EGR FAC en ventana",   f"{int(vtas):,} u"),
        ("4",  "Compras del período",      "Σ unidades ING FAC",              f"{int(comp_u):,} u"),
        ("5",  "Costo promedio unit.",     "Valor_Compras / Compras",         f"${costo:,.2f}"),
        ("6",  "Valor del stock",          "Stock × Costo_prom",              f"${val_stk:,.2f}"),
        ("7",  "COGS aprox.",              "Ventas_u × Costo_prom",           f"${cogs_apx:,.2f}"),
        ("8",  "Consumo diario",           "Ventas / Días_período",           f"{cons:.2f} u/día"),
        ("9",  "Rotación (unidades)",      "Ventas / Stock",                  f"{rot_u:.2f}×"),
        ("10", "Rotación ($)",             "COGS / Valor_stock",              f"{rot_d:.2f}×"),
        ("11", "Días inv. (período)",      "Stock / Consumo_diario",          f"{dinv_p:.1f} d"),
        ("12", "Días inv. (anualizado)",   "365 / Rotación_u",                f"{dinv_a:.1f} d"),
        ("13", "Demanda en Lead time",     f"Cons × LT ({lt}d)",              f"{dem_lt:.1f} u"),
        ("14", "Stock de seguridad",       f"Cons × SS ({saf}d)",             f"{ss_u:.1f} u"),
        ("15", "Punto de reorden",         "Demanda_LT + SS",                 f"{reord:.1f} u"),
        ("16", f"Necesidad para {cob}d",   "Cons × Cobertura",                f"{nec_cob:.1f} u"),
        ("17", "Sug. compra (cobertura)",  "max(0, Necesidad − Stock)",       f"{int(round(sug_cob)):,} u"),
        ("18", "Sug. compra (LT+SS)",      "max(0, Cons×(LT+SS) − Stock)",    f"{int(round(sug_lt)):,} u"),
        ("19", "Estado / Alerta",          "evaluación de reglas",            alerta),
    ]
    cal_df = pd.DataFrame(pasos, columns=["#", "Variable", "Fórmula", "Valor"])
    st.dataframe(cal_df, width='stretch', hide_index=True, height=560,
                 column_config={
                     "#":        st.column_config.TextColumn("Paso", width="small"),
                     "Variable": st.column_config.TextColumn("Variable"),
                     "Fórmula":  st.column_config.TextColumn("Fórmula"),
                     "Valor":    st.column_config.TextColumn("Valor", width="medium"),
                 })

    # ── Interpretación narrativa ──
    st.markdown("#### 📝 Interpretación")
    lines = []
    if cons == 0:
        lines.append("Este SKU **no tuvo ventas** en el período. Revisa si está "
                     "descontinuado, fuera de temporada o si el filtro de fechas excluye ventas.")
    else:
        lines.append(
            f"- Vendiste **{int(vtas):,} u** en **{days} días** → consumo diario "
            f"promedio **{cons:.2f} u/día**."
        )
        lines.append(
            f"- Stock actual: **{int(stk):,} u** (valor **${val_stk:,.2f}** al "
            f"costo promedio **${costo:,.2f}**)."
        )
        lines.append(
            f"- Al ritmo actual te alcanza para **{dinv_p:.0f} días**. "
            f"Rotación del período: **{rot_u:.2f}×** "
            f"(anualizado: 1 rotación cada {dinv_a:.0f}d)."
        )
        lines.append(
            f"- Demanda durante los **{lt} días de lead time**: {dem_lt:.0f} u. "
            f"Punto de reorden (incluye {saf}d de seguridad): **{reord:.0f} u**."
        )
        if sug_cob > 0:
            lines.append(
                f"- Para cubrir **{cob} días** necesitas **{int(round(sug_cob)):,} u** "
                f"(ya tienes {int(stk):,})."
            )
        else:
            lines.append(
                f"- Tu stock actual cubre **más de {cob} días**. No hace falta comprar."
            )
        if dinv_p < lt:
            lines.append(
                f"- ⚠ Tu lead time ({lt}d) es **mayor** a los días de stock que te quedan "
                f"({dinv_p:.0f}d). Si pides hoy **te quedarás sin stock** antes de que llegue."
            )
        elif dinv_p < lt + saf:
            lines.append(
                f"- ⚠ Estás apenas cubriendo lead time + seguridad. Pide **pronto**."
            )
    st.markdown("\n".join(lines))

    # Export del cálculo
    st.download_button("📥 Exportar cálculo a Excel",
        to_xl(cal_df),
        f"calculo_{sel_code}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="cal_xl", width='stretch')

with T_CAL:
    _render_tab_cal()

# ══ TAB 7 COMPRAS ═══════════════════════════════════════════════
@_fragment
def _render_tab_pur():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if eng.raw_df is None: st.info("Cargue un archivo.")
    else:
        # Multiselect predictivo con filtro estricto.
        # Excluye SKUs globalmente excluidos (sidebar).
        _pu_opts = eng.raw_df[["Código Producto","Nombre Producto"]].drop_duplicates()
        _pu_opts = _pu_opts.sort_values("Código Producto")
        _pu_opts["Código Producto"] = _pu_opts["Código Producto"].astype(str)
        if eng.excluded_skus:
            _pu_opts = _pu_opts[~_pu_opts["Código Producto"].isin(eng.excluded_skus)]
        _pu_labels = [f"{c} — {n}" for c,n in zip(
            _pu_opts["Código Producto"].astype(str), _pu_opts["Nombre Producto"].astype(str))]
        _pu_sel = st.multiselect("🔍 SKU / Producto", _pu_labels, key="pu_f",
                                  placeholder="Escribe código o nombre…")
        df=eng.raw_df.copy()
        if excl_s: df=df[~df["Código Producto"].isin(excl_s)]
        ref=df["Referencia"].fillna("").astype(str).str.upper()
        typ=df["Tipo"].fillna("").astype(str).str.upper()
        cdf=df[(typ=="ING")&ref.str.startswith("FAC")].copy()
        if _pu_sel:
            _pu_codes_sel = {s.split(" — ")[0] for s in _pu_sel}
            cdf = cdf[cdf["Código Producto"].astype(str).isin(_pu_codes_sel)]
        if cdf.empty: st.warning("Sin compras.")
        else:
            cdf=cdf.sort_values(["Código Producto","Fecha"])
            rows=[]
            cpq,cpv=defaultdict(float),defaultdict(float)
            for _,row in cdf.iterrows():
                sku=row["Código Producto"]; qty=float(row["Cantidad"]) or 1
                vt=float(row["Valor Total"]); vu=vt/qty
                nq=cpq[sku]+qty; nv=cpv[sku]+vt; cpq[sku]=nq; cpv[sku]=nv; cp=nv/nq
                rows.append({"Fecha":row["Fecha"].strftime("%d/%m/%Y") if pd.notna(row["Fecha"]) else "",
                             "Factura":str(row.get("Referencia","")).strip(),"Código":sku,
                             "Nombre":str(row.get("Nombre Producto",""))[:50],
                             "Desc.":str(row.get("Descripción",""))[:40],
                             "Cant.":int(qty),"V.Total":round(vt,2),
                             "V.Unit":round(vu,4),"Costo Prom.":round(cp,4),"_g":sku})
            all_r=[]
            for sku,grp in pd.DataFrame(rows).groupby("_g",sort=False):
                for _,row in grp.iterrows(): all_r.append(row.drop("_g"))
                s=grp.iloc[-1].copy(); s["Fecha"]="SUBTOTAL"; s["Factura"]=""
                s["Cant."]=int(grp["Cant."].sum()); s["V.Total"]=round(grp["V.Total"].sum(),2)
                s["V.Unit"]=""; s["Desc."]=""; all_r.append(s.drop("_g"))
            out=pd.DataFrame(all_r)

            # ── Tabla compras con color alternado por grupo SKU ──
            _nc_pur=["Cant.","V.Total","V.Unit","Costo Prom."]
            # Paleta de 2 fondos alternados por SKU
            _pal=["#f0f9ff","#ffffff"]   # celeste muy suave / blanco
            _sub_bg="#dbeafe"            # azul pastel para filas SUBTOTAL

            def _pur_table(df, nc, uid):
                if df is None or df.empty:
                    return "<p>Sin datos</p>"
                uid_s = uid.replace("-", "_")
                all_cols = list(df.columns)
                PAL = ["#f0f9ff", "#ffffff"]
                SUB_BG = "#dbeafe"
                TH_BASE = (
                    "position:sticky;top:0;z-index:3;background:#f1f5f9;"
                    "border-bottom:2px solid #cbd5e1;padding:7px 10px;font-size:10px;"
                    "font-weight:700;text-transform:uppercase;color:#64748b;white-space:nowrap"
                )
                hdrs = "".join(
                    "<th style='" + TH_BASE + (";text-align:right" if c in nc else "") + "'>" + c + "</th>"
                    for c in all_cols
                )
                rows = ""
                tots = {c: 0.0 for c in nc}
                gi = 0
                last_sku = ""
                for _, row in df.iterrows():
                    sku = str(row.get("Código", ""))
                    is_sub = (str(row.get("Fecha", "")) == "SUBTOTAL")
                    if not is_sub and sku != last_sku and sku:
                        if last_sku:
                            gi += 1
                        last_sku = sku
                    bg = SUB_BG if is_sub else PAL[gi % 2]
                    cells = ""
                    for c in all_cols:
                        v = row[c]
                        raw = str(v)
                        disp = "" if raw in ("nan", "None", "NaN") else raw
                        if c in nc:
                            try:
                                fv = float(v)
                                if not is_sub:
                                    tots[c] += fv
                                disp = ("{:,}".format(int(round(fv))) if _is_int_col(c)
                                        else "{:,.2f}".format(fv))
                            except:
                                pass
                        fw  = "font-weight:700;" if is_sub else ""
                        aln = "text-align:right;font-family:monospace;" if c in nc else ""
                        bdr = "border-top:1px solid #93c5fd;" if is_sub else "border-bottom:1px solid #f1f5f9;"
                        cells += (
                            "<td style='background:" + bg + ";padding:6px 10px;font-size:12px;"
                            + aln + fw + bdr + "'>" + disp + "</td>"
                        )
                    rows += "<tr>" + cells + "</tr>"

                def _tf(c):
                    v = tots[c]
                    return ("{:,}".format(int(round(v))) if _is_int_col(c)
                            else "{:,.2f}".format(v))

                first = all_cols[0] if all_cols else ""
                TOT = (
                    "background:#e0f2fe;font-weight:800;padding:7px 10px;"
                    "border-top:2px solid #7dd3fc;color:#0369a1"
                )
                tcells = "".join(
                    (
                        "<td style='" + TOT + ";text-align:right;font-family:monospace'>" + _tf(c) + "</td>"
                        if (c in nc and tots[c] != 0) else
                        ("<td style='" + TOT + "'>TOTAL</td>" if c == first
                         else "<td style='" + TOT + "'></td>")
                    )
                    for c in all_cols
                )
                # Barra ZOOM — onclick usa comillas simples para JS
                on_m = "asZoom('" + uid_s + "',-1)"
                on_p = "asZoom('" + uid_s + "',1)"
                on_r = "asZoomReset('" + uid_s + "')"
                zoom = (
                    "<div class='zb'>"
                    "<span style='color:var(--text3);font-size:11px;font-weight:700'>ZOOM</span>"
                    "<button onclick='" + on_m + "'>&#8722;</button>"
                    "<button onclick='" + on_p + "'>+</button>"
                    "<button onclick='" + on_r + "'>&#8635;</button>"
                    "<span style='color:var(--text3);font-size:10px'>"
                    + str(len(df)) + " registros</span></div>"
                )
                return (
                    zoom
                    + "<div class='tc'><table class='it' id='tbl_" + uid_s + "'>"
                    + "<thead><tr>" + hdrs + "</tr></thead>"
                    + "<tbody>" + rows + "</tbody>"
                    + "<tfoot><tr>" + tcells + "</tr></tfoot>"
                    + "</table></div>"
                )

            # Wrapper HTML completo para sticky header en st.components
            _pur_html = _pur_table(out, _nc_pur, "pur")
            _pur_doc = (
                "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
                "<link rel='preconnect' href='https://fonts.googleapis.com'>"
                "<style>"
                "*{box-sizing:border-box;margin:0;padding:0}"
                "body{font-family:Inter,Segoe UI,sans-serif;background:#f0f9ff;padding:2px}"
                ".tc{overflow:auto;max-height:560px;border:1px solid #e2e8f0;"
                "border-radius:8px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08);"
                "position:relative}"
                ".it{width:100%;border-collapse:separate;border-spacing:0;font-size:12px}"
                ".it thead th{position:sticky!important;top:0!important;z-index:3!important;"
                "background:#f1f5f9;border-bottom:2px solid #cbd5e1;padding:7px 10px;"
                "font-size:10px;font-weight:700;text-transform:uppercase;color:#64748b;"
                "white-space:nowrap}"
                ".it .n{text-align:right;font-family:monospace}"
                ".zb{display:flex;align-items:center;gap:6px;margin-bottom:6px;font-family:Inter,sans-serif}"
                ".zb span{font-size:11px;font-weight:700;color:#94a3b8}"
                ".zb button{background:#fff;border:1px solid #e2e8f0;border-radius:5px;"
                "padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;color:#475569}"
                ".zb button:hover{background:#0ea5e9;color:#fff}"
                "tr:hover td{background:#e0f2fe!important}"
                "</style></head><body>"
                + _pur_html
                + "<script>"
                "var w=document.querySelector('.tc');"
                "if(w){w.addEventListener('wheel',function(e){"
                "if(Math.abs(e.deltaX)>Math.abs(e.deltaY)){e.preventDefault();w.scrollLeft+=e.deltaX;}"
                "else if(e.shiftKey){e.preventDefault();w.scrollLeft+=e.deltaY;}"
                "},{passive:false});}"
                "</script>"
                "</body></html>"
            )
            _components.html(_pur_doc, height=640, scrolling=False)
            dl3(out, "historico_compras", "pur")

with T_PUR:
    _render_tab_pur()

# ══ TAB 8 KARDEX ════════════════════════════════════════════════
@_fragment
def _render_tab_kdx():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if eng.raw_df is None: st.info("Cargue un archivo.")
    else:
        c1,c2,c3,c4=st.columns([2,2,3,1])
        kdf=date.today()-timedelta(days=365)
        kd_f=c1.date_input("Desde",kdf,format="DD/MM/YYYY",key="kf")
        kd_t=c2.date_input("Hasta",date.today(),format="DD/MM/YYYY",key="kt")
        # Selector predictivo con filtro estricto de substring.
        # Excluye SKUs globalmente excluidos (sidebar).
        _kdx_opts = (eng.raw_df[["Código Producto","Nombre Producto"]]
                     .drop_duplicates().sort_values("Código Producto"))
        _kdx_opts["Código Producto"] = _kdx_opts["Código Producto"].astype(str)
        if eng.excluded_skus:
            _kdx_opts = _kdx_opts[~_kdx_opts["Código Producto"].isin(eng.excluded_skus)]
        _kdx_codes  = _kdx_opts["Código Producto"].astype(str).tolist()
        _kdx_names  = _kdx_opts["Nombre Producto"].astype(str).tolist()
        _kdx_labels = [f"{c} — {n}" for c,n in zip(_kdx_codes, _kdx_names)]
        _opts_kx = ["— Todos —"] + _kdx_labels
        kd_s_label = c3.selectbox("SKU", _opts_kx, key="ks")
        kd_s = "— Todos —" if kd_s_label == "— Todos —" else kd_s_label.split(" — ")[0]
        with c4:
            st.markdown("")
            gen=st.button("▶ Generar",type="primary",key="kg")
        if gen:
            with st.spinner("Calculando..."):
                raw=eng.raw_df.copy()
                sf="" if kd_s=="— Todos —" else kd_s
                if sf: raw=raw[raw["Código Producto"]==sf]
                df_ts=pd.Timestamp(kd_f); dt_ts=pd.Timestamp(kd_t)
                ref=raw["Referencia"].fillna("").astype(str).str.upper()
                typ=raw["Tipo"].fillna("").astype(str).str.upper()
                desc=raw["Descripción"].fillna("").astype(str).str.upper()
                # Baja de inventario (EGR + descripción contiene "BAJA DE INVENTARIO")
                raw["_ib"]=(typ=="EGR") & desc.str.contains("BAJA DE INVENTARIO", regex=False, na=False)
                # Dev. Cliente: ref NCT o "BAJA DE FACTURA" en desc
                raw["_ic"]=(typ=="ING") & (
                    ref.str.startswith("NCT") |
                    desc.str.contains("BAJA DE FACTURA", regex=False, na=False)
                )
                # Compras: ref FAC/NVE o "FACTURA DE COMPRA" / "COMPRA INVENTARIO" en desc
                raw["_ip"]=(typ=="ING") & ~raw["_ic"] & (
                    ref.str.startswith("FAC") |
                    ref.str.startswith("NVE") |
                    desc.str.contains("FACTURA DE COMPRA", regex=False, na=False) |
                    desc.str.contains("COMPRA INVENTARIO", regex=False, na=False) |
                    desc.str.contains("COMPRA DE INVENTARIO", regex=False, na=False)
                )
                # Excluir bajas de venta y dev. proveedor (prioridad)
                raw["_is"]=(typ=="EGR")&ref.str.startswith("FAC") & ~raw["_ib"]
                raw["_ir"]=(typ=="EGR")&ref.str.startswith("NCT") & ~raw["_ib"]
                raw["_it"]=typ=="TRA"
                raw=raw.sort_values(["Código Producto","Fecha"]).reset_index(drop=True)
                hist=raw[raw["Fecha"]<df_ts]
                cp={};cpq={};cpv={};stk={}
                for _,row in hist.iterrows():
                    sku=row["Código Producto"]; qty=float(row["Cantidad"])
                    stk[sku]=stk.get(sku,0)
                    if row["_ip"] or row["_ic"]: stk[sku]+=qty
                    elif row["_is"] or row["_ir"] or row["_ib"]: stk[sku]-=qty
                    if row["_ip"] and qty>0:
                        vt=float(row["Valor Total"]); cu=vt/qty if qty>0 else 0
                        if cu>0:
                            nq=cpq.get(sku,0)+qty; nv=cpv.get(sku,0)+vt
                            cpq[sku]=nq; cpv[sku]=nv; cp[sku]=nv/nq
                rows=[]
                for sku,grp in raw.groupby("Código Producto"):
                    nom=str(grp["Nombre Producto"].iloc[0])
                    s0=stk.get(sku,0); c0=cp.get(sku,0)
                    rows.append({"Fecha":kd_f.strftime("%d/%m/%Y"),"Código":sku,"Nombre":nom,
                                 "N°Reg":"","Referencia":"—","Descripción":"SALDO INICIAL","Tipo":"INICIO",
                                 "Cantidad":round(s0,2),"V.Unit":round(c0,4),
                                 "Costo Prom.":round(c0,4),"Saldo":round(s0,2),"Valor Inv.":round(s0*c0,2)})
                    cc=cp.get(sku,0); cq=cpq.get(sku,0); cv=cpv.get(sku,0); saldo=s0
                    for _,row in grp[(grp["Fecha"]>=df_ts)&(grp["Fecha"]<=dt_ts)].iterrows():
                        qty=float(row["Cantidad"]); vt=float(row["Valor Total"]); vu=vt/qty if qty>0 else 0
                        fd=row["Fecha"].strftime("%d/%m/%Y") if pd.notna(row["Fecha"]) else ""
                        if row["_ip"]: tipo="INGRESO"; ef=+qty; nq=cq+qty; nv=cv+vt; cc=nv/nq; cq=nq; cv=nv
                        elif row["_ic"]: tipo="ING DEV.CLI"; ef=+qty
                        elif row["_is"]: tipo="EGRESO"; ef=-qty
                        elif row["_ir"]: tipo="EGR DEV.PROV"; ef=-qty
                        elif row["_ib"]: tipo="BAJA INV."; ef=-qty
                        elif row["_it"]: tipo="TRANSFERENCIA"; ef=0
                        else: tipo="OTRO"; ef=0
                        saldo+=ef
                        # N° registro del movimiento (col "Código" del Excel)
                        n_reg=str(row.get("Código","")).strip()
                        ref_val=str(row.get("Referencia","")).strip()
                        # Para transferencias: mostrar N°Reg prominente en Referencia
                        if tipo=="TRANSFERENCIA" and n_reg:
                            ref_disp="[Reg:" + n_reg + "]"
                        else:
                            ref_disp=ref_val
                        rows.append({"Fecha":fd,"Código":sku,"Nombre":nom,
                                     "N°Reg":n_reg,
                                     "Referencia":ref_disp,
                                     "Descripción":str(row.get("Descripción","")).strip(),
                                     "Tipo":tipo,"Cantidad":round(abs(qty),2),
                                     "V.Unit":round(vu,4),"Costo Prom.":round(cc,4),
                                     "Saldo":round(saldo,2),"Valor Inv.":round(saldo*cc,2)})
                dk=pd.DataFrame(rows)
                st.session_state["kdx"]=dk; st.session_state["kdx_s"]=sf
                log(f"Kardex {sf or 'todos'}: {len(dk):,} filas")

        if "kdx" in st.session_state and not st.session_state["kdx"].empty:
            dk=st.session_state["kdx"]
            CK2={
                "INICIO":     ("#dbeafe","#1e40af"),
                "INGRESO":    ("#d1fae5","#065f46"),
                "ING DEV.CLI":("#fef9c3","#854d0e"),
                "EGRESO":     ("#f3f4f6","#374151"),
                "EGR DEV.PROV":("#f3f4f6","#374151"),
                "BAJA INV.":  ("#fee2e2","#991b1b"),
                "TRANSFERENCIA":("#ede9fe","#5b21b6"),
            }
            nk=["Cantidad","V.Unit","Costo Prom.","Saldo","Valor Inv."]
            all_cols=list(dk.columns)

            TH_CSS = (
                "position:sticky;top:0;z-index:3;background:#1e3a5f;"
                "border-bottom:2px solid #0ea5e9;padding:7px 10px;font-size:10px;"
                "font-weight:700;text-transform:uppercase;color:#e0f2fe;white-space:nowrap"
            )
            hdrs_parts = []
            for c in all_cols:
                align = ";text-align:right" if c in nk else ""
                hdrs_parts.append("<th style='" + TH_CSS + align + "'>" + c + "</th>")
            hdrs = "".join(hdrs_parts)

            # ── Filas con separadores entre SKUs ───────────────────
            rows_parts = []
            last_sku = ""
            sku_col = "Código" if "Código" in all_cols else ("Código Producto" if "Código Producto" in all_cols else "")
            for _, row in dk.iterrows():
                sku = str(row.get(sku_col, "")) if sku_col else ""
                tp  = str(row.get("Tipo", ""))
                bg, fg = CK2.get(tp, ("#ffffff","#111827"))
                if sku and sku != last_sku and last_sku:
                    # Fila separadora pronunciada al cambiar de SKU
                    rows_parts.append(
                        "<tr><td colspan='" + str(len(all_cols)) + "' "
                        "style='background:#1e3a5f;padding:5px 12px;"
                        "font-size:11px;font-weight:700;color:#7dd3fc;"
                        "letter-spacing:.06em;border-top:3px solid #0ea5e9;"
                        "border-bottom:2px solid #0ea5e9'>"
                        "&#9644;&#9644; " + sku + " &#9644;&#9644;"
                        "</td></tr>"
                    )
                last_sku = sku
                cells_parts = []
                for c in all_cols:
                    v    = row[c]
                    raw  = str(v)
                    disp = "" if raw in ("nan","None","NaN") else raw
                    if c in nk:
                        try:
                            fv = float(v)
                            disp = ("{:,.4f}".format(fv) if c in ("V.Unit","Costo Prom.")
                                    else "{:,.2f}".format(fv))
                        except:
                            pass
                        cells_parts.append(
                            "<td style='background:" + bg + ";color:" + fg + ";"
                            "text-align:right;font-family:monospace;padding:6px 10px;"
                            "border-bottom:1px solid #f1f5f9'>" + disp + "</td>"
                        )
                    else:
                        cells_parts.append(
                            "<td style='background:" + bg + ";color:" + fg + ";"
                            "padding:6px 10px;border-bottom:1px solid #f1f5f9'>" + disp + "</td>"
                        )
                rows_parts.append("<tr>" + "".join(cells_parts) + "</tr>")
            rows = "".join(rows_parts)

            # ── Leyenda ────────────────────────────────────────────
            leg_parts = []
            for t,(b,f) in CK2.items():
                leg_parts.append(
                    "<span style='background:" + b + ";color:" + f + ";"
                    "padding:2px 8px;border-radius:4px;font-size:10px;"
                    "border:1px solid #d1d5db;margin-right:4px'>"
                    "&#9632; " + t + "</span>"
                )
            leg = "".join(leg_parts)

            n_mov = str(len(dk))
            kdx_html = (
                "<!DOCTYPE html><html><head><meta charset='UTF-8'>"
                "<style>"
                "*{box-sizing:border-box;margin:0;padding:0}"
                "body{font-family:Inter,Segoe UI,sans-serif;background:#f0f9ff;padding:2px}"
                ".leg{display:flex;flex-wrap:wrap;gap:4px;margin-bottom:8px}"
                ".zb{display:flex;align-items:center;gap:6px;margin-bottom:6px}"
                ".zb span{font-size:11px;font-weight:700;color:#94a3b8}"
                ".zb .inf{font-size:10px;font-weight:400}"
                ".zb button{background:#fff;border:1px solid #e2e8f0;border-radius:5px;"
                "padding:2px 10px;cursor:pointer;font-weight:700;font-size:13px;color:#475569}"
                ".zb button:hover{background:#0ea5e9;color:#fff}"
                ".wrap{overflow:auto;max-height:580px;border:1px solid #e2e8f0;"
                "border-radius:8px;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,.08)}"
                "table{border-collapse:separate;border-spacing:0;"
                "width:max-content;font-size:12px}"
                "tr:hover td{filter:brightness(.93)}"
                "</style></head><body>"
                "<div class='leg'>" + leg + "</div>"
                "<div class='zb'>"
                "<span>ZOOM</span>"
                "<button onclick=\"var t=document.getElementById('kdt');"
                "var s=parseFloat(t.style.fontSize||'12');"
                "t.style.fontSize=(s-1)+'px'\">&#8722;</button>"
                "<button onclick=\"var t=document.getElementById('kdt');"
                "var s=parseFloat(t.style.fontSize||'12');"
                "t.style.fontSize=(s+1)+'px'\">+</button>"
                "<button onclick=\"document.getElementById('kdt').style.fontSize='12px'\">&#8635;</button>"
                "<span class='inf'>" + n_mov + " movimientos</span>"
                "</div>"
                "<div class='wrap'>"
                "<table id='kdt'>"
                "<thead><tr>" + hdrs + "</tr></thead>"
                "<tbody>" + rows + "</tbody>"
                "</table></div>"
                "<script>"
                "var w=document.querySelector('.wrap');"
                "w.addEventListener('wheel',function(e){"
                "if(Math.abs(e.deltaX)>Math.abs(e.deltaY)){"
                "e.preventDefault();w.scrollLeft+=e.deltaX;}"
                "else if(e.shiftKey){e.preventDefault();w.scrollLeft+=e.deltaY;}"
                "},{passive:false});"
                "</script>"
                "</body></html>"
            )
            _components.html(kdx_html, height=720, scrolling=False)
            sfn=st.session_state["kdx_s"].replace("/","_") or "todos"
            dl3(dk,f"kardex_{sfn}","kdx")

with T_KDX:
    _render_tab_kdx()

# ── Fragment: Toma Física por Ubicación (tabla tipo planilla) ───
# Aislado en @st.fragment: las ediciones de celdas solo re-ejecutan esta
# función, no toda la app.
def _safe_eval(expr):
    """Evalúa una expresión aritmética segura (+ − × ÷ paréntesis y decimales).
    NO permite nombres, atributos, funciones, imports. Lanza ValueError si falla."""
    import ast, operator
    _OPS = {
        ast.Add: operator.add, ast.Sub: operator.sub,
        ast.Mult: operator.mul, ast.Div: operator.truediv,
        ast.USub: operator.neg, ast.UAdd: operator.pos,
        ast.Pow: operator.pow,
    }
    clean = str(expr).replace("×","*").replace("÷","/").replace(",",".").strip()
    if not clean:
        raise ValueError("expresión vacía")
    try:
        tree = ast.parse(clean, mode="eval")
    except SyntaxError as ex:
        raise ValueError(f"sintaxis inválida: {ex}")
    def _ev(n):
        if isinstance(n, ast.Expression): return _ev(n.body)
        if isinstance(n, ast.Constant):
            if isinstance(n.value, (int, float)): return n.value
            raise ValueError(f"valor no permitido: {n.value!r}")
        if isinstance(n, ast.BinOp):
            op = _OPS.get(type(n.op))
            if op is None: raise ValueError(f"operador no permitido: {type(n.op).__name__}")
            return op(_ev(n.left), _ev(n.right))
        if isinstance(n, ast.UnaryOp):
            op = _OPS.get(type(n.op))
            if op is None: raise ValueError(f"operador unario no permitido")
            return op(_ev(n.operand))
        raise ValueError(f"elemento no permitido: {type(n).__name__}")
    return _ev(tree)

def _build_toma_table(eng, rap_df, ubicacion):
    """Construye DataFrame con TODOS los SKUs + su cantidad anterior para la ubicación dada.
    Respeta la exclusión global de SKUs: los SKUs en eng.excluded_skus no aparecen
    porque no interesan en el análisis."""
    sk = (eng.raw_df[["Código Producto","Nombre Producto"]]
          .drop_duplicates()
          .sort_values("Código Producto")
          .reset_index(drop=True)).copy()
    sk["Código Producto"] = sk["Código Producto"].astype(str)
    sk["Nombre Producto"]  = sk["Nombre Producto"].astype(str)
    # Aplicar exclusión global de SKUs (sidebar) — no aparecen en la toma
    if eng.excluded_skus:
        sk = sk[~sk["Código Producto"].isin(eng.excluded_skus)].reset_index(drop=True)
    prev_map = {}
    if rap_df is not None and not rap_df.empty:
        pv = rap_df[rap_df["Ubicación"].astype(str) == str(ubicacion)].copy()
        if not pv.empty:
            pv = pv.sort_values("Fecha").drop_duplicates(
                subset=["Código Producto"], keep="last")
            prev_map = dict(zip(pv["Código Producto"].astype(str),
                                pd.to_numeric(pv["Cantidad Física"], errors="coerce")))
    sk["Anterior"]    = sk["Código Producto"].map(prev_map)
    sk["Nueva"]       = pd.NA
    sk["Observación"] = ""
    # Orden: Producto a la IZQUIERDA (nombre), SKU a la derecha, luego datos editables
    return sk[["Nombre Producto","Código Producto","Anterior","Nueva","Observación"]]

@_fragment
def _render_toma_fragment():
    _frag_t0 = _ptime.perf_counter()
    _eng = st.session_state.engine
    if _eng.raw_df is None:
        st.info("Primero carga los Excel de movimientos en el sidebar.")
        return
    rap_state = _get_shared_rapid()
    rap_df    = rap_state["df"]

    all_ubic = _get_all_ubic()

    # ── Modo pantalla completa (oculta sidebar + banner + KPIs + barra tabs) ──
    if "toma_fullscreen" not in st.session_state:
        st.session_state["toma_fullscreen"] = False

    if st.session_state["toma_fullscreen"]:
        # CSS modo minimalista: solo lo esencial para contar.
        st.markdown("""
        <style>
        /* Ocultar sidebar */
        section[data-testid="stSidebar"] { display: none !important; }
        /* Ocultar banner AutoSky, período, KPIs y aviso de exclusiones */
        .as-banner,
        .as-periodo-banner,
        .as-excl-banner,
        .kpi-row { display: none !important; }
        /* Ocultar el título "Toma Física" (h3 #...) */
        h3.as-toma-title { display: none !important; }
        /* Ocultar TODAS las barras de tabs (principal y sub-tabs) */
        [data-baseweb="tab-list"] { display: none !important; }
        /* Ocultar el caption instructivo y otros elementos opcionales */
        .toma-hide-in-fs { display: none !important; }
        /* Ocultar el separador horizontal hr previo al título Toma */
        hr { display: none !important; }
        /* Expandir el main area a full width */
        section[data-testid="stMain"] .block-container,
        .main .block-container { max-width: 100% !important;
            padding-top: .5rem !important;
            padding-left: .5rem !important; padding-right: .5rem !important; }
        /* Panel de tab sin padding extra */
        [data-baseweb="tab-panel"] { padding-top: 0 !important; }
        /* Header/toolbar superior de Streamlit */
        header[data-testid="stHeader"] { display: none !important; }
        </style>
        """, unsafe_allow_html=True)

    # Toggle fullscreen arriba de todo (esquina superior derecha)
    _fs_c1, _fs_c2 = st.columns([5, 1])
    with _fs_c2:
        _fs_label = "🔙 Salir modo toma" if st.session_state["toma_fullscreen"] else "🖥 Pantalla completa"
        if st.button(_fs_label, key="toma_fs_btn", width='stretch',
                     help="Oculta sidebar, banner y otras pestañas para conteo sin distracciones. "
                          "Ideal en celular o tablet."):
            st.session_state["toma_fullscreen"] = not st.session_state["toma_fullscreen"]
            _rerun_frag()

    # Selector + agregar ubicación
    col_u, col_add = st.columns([3, 1])
    with col_u:
        sel_ubic = st.selectbox("📍 Ubicación donde se hará la toma",
                                all_ubic, key="tu_ubic")
    with col_add:
        st.markdown("")
        with st.popover("➕ Nueva ubicación", width='stretch'):
            _new_ub = st.text_input("Nombre de la nueva ubicación",
                                     key="tu_new_ubic_name",
                                     placeholder="Ej: Bodega Norte")
            if st.button("Agregar", key="tu_add_ubic", type="primary"):
                _n = (_new_ub or "").strip()
                if not _n:
                    st.error("Escribe un nombre.")
                elif _n in all_ubic:
                    st.warning(f"«Ya existe la ubicación {_n}».")
                else:
                    custom = _get_custom_ubic()
                    custom["list"].append(_n)
                    _persist_custom_ubic(custom["list"])
                    log(f"Nueva ubicación agregada: {_n}")
                    st.success(f"✓ Agregada: {_n}")
                    _rerun_frag()

    # Construir la tabla base para esta ubicación
    table_df = _build_toma_table(_eng, rap_df, sel_ubic)

    # Instrucciones — solo visible en modo normal (ocultas en fullscreen)
    if not st.session_state.get("toma_fullscreen", False):
        st.markdown(
            "<div class='toma-hide-in-fs' style='color:#64748b;font-size:13px;"
            "margin-bottom:6px'>"
            "Escribe la cantidad en la columna <b>Nueva</b>. Tab avanza a Observación; "
            "Tab otra vez baja a la siguiente fila. Flechas para navegar libremente. "
            "Columna <b>Anterior</b> (gris) es la última toma registrada y no es editable."
            "</div>",
            unsafe_allow_html=True
        )

    # Controles compactos: toggle para ocultar columna Producto
    _compact_key = f"toma_compact_{sel_ubic}"
    if _compact_key not in st.session_state:
        st.session_state[_compact_key] = False
    _compact = st.toggle(
        "📏 Vista compacta (ocultar columna Producto)",
        key=_compact_key,
        help="Oculta la columna de nombre del producto para ver más filas por pantalla. "
             "Ideal en celular."
    )

    # Key dinámica por ubicación: al cambiar de ubicación, el editor reinicia
    editor_key = f"tu_ed_{sel_ubic}"

    # ── 🧮 Calculadora (expresiones aritméticas para cajas × unidades) ──
    with st.expander("🧮 Calculadora — aplicar expresión a un SKU"):
        st.caption(
            "Útil para cajas × unidades. Ej: `5*33*3+2` = 497. "
            "Operadores: `+ − × ÷ ( )` y decimales. "
            "El resultado se aplica al SKU seleccionado en la tabla de arriba."
        )
        _calc_codes = table_df["Código Producto"].astype(str).tolist()
        _calc_names = table_df["Nombre Producto"].astype(str).tolist()
        _calc_labels = [f"{c} — {n}" for c,n in zip(_calc_codes, _calc_names)]
        cc1, cc2 = st.columns([3, 2])
        with cc1:
            _sel_sku_calc = st.selectbox(
                "SKU al que aplicar el resultado", _calc_labels,
                key=f"calc_sku_{sel_ubic}",
                placeholder="Escribe código o nombre…",
            )
        with cc2:
            _expr = st.text_input(
                "Expresión (ej: 5*33*3+2)",
                key=f"calc_expr_{sel_ubic}",
                placeholder="5*33*3+2",
            )
        # Evaluación en vivo + aplicar
        _result = None; _err = None
        if _expr and _expr.strip():
            try:
                _v = _safe_eval(_expr)
                # Validar que sea un número finito no negativo
                import math
                if math.isnan(_v) or math.isinf(_v):
                    _err = "resultado no válido (infinito/NaN)"
                elif _v < 0:
                    _err = "el resultado debe ser ≥ 0"
                else:
                    _result = _v
            except ValueError as _ex:
                _err = str(_ex)
        rc1, rc2 = st.columns([3, 2])
        with rc1:
            if _result is not None:
                st.markdown(
                    f"<div style='background:#d1fae5;border:1px solid #059669;"
                    f"border-radius:8px;padding:10px 14px;font-weight:700;"
                    f"color:#065f46;font-size:18px'>= {_result:g} u</div>",
                    unsafe_allow_html=True,
                )
            elif _err:
                st.error(f"Error: {_err}")
            else:
                st.caption("Escribe una expresión para ver el resultado.")
        with rc2:
            _disabled = (_result is None or _sel_sku_calc is None)
            if st.button("✔ Aplicar al SKU", type="primary",
                         width='stretch',
                         key=f"calc_apply_{sel_ubic}",
                         disabled=_disabled):
                # Obtener el índice del SKU en la tabla del editor
                _sel_code = _sel_sku_calc.split(" — ")[0]
                try:
                    _row_idx = _calc_codes.index(_sel_code)
                except ValueError:
                    _row_idx = None
                if _row_idx is not None:
                    # Modificar el estado del data_editor inyectando edited_rows
                    _ed_state = st.session_state.get(editor_key, {}) or {}
                    _edited_rows = _ed_state.get("edited_rows", {}) if isinstance(_ed_state, dict) else {}
                    _edited_rows[_row_idx] = {
                        **(_edited_rows.get(_row_idx, {})),
                        "Nueva": int(round(_result)),
                    }
                    _ed_state["edited_rows"] = _edited_rows
                    st.session_state[editor_key] = _ed_state
                    st.success(f"✓ Aplicado {int(round(_result))} u al SKU {_sel_code}")
                    _rerun_frag()


    # Vista compacta: dropear columna Producto si está activada
    table_df_display = table_df
    if _compact and "Nombre Producto" in table_df_display.columns:
        table_df_display = table_df_display.drop(columns=["Nombre Producto"])

    # En modo compacto dar más ancho a SKU para que no se trunque
    _sku_width = "medium" if _compact else "small"
    # Altura mayor en fullscreen
    _editor_height = 720 if st.session_state.get("toma_fullscreen", False) else 520

    edited = st.data_editor(
        table_df_display,
        width='stretch',
        num_rows="fixed",
        hide_index=True,
        column_config={
            "Nombre Producto":  st.column_config.TextColumn("Producto", disabled=True),
            "Código Producto":  st.column_config.TextColumn("SKU", width=_sku_width, disabled=True),
            "Anterior":         st.column_config.NumberColumn("Anterior", format="%d",
                                    disabled=True,
                                    help="Última toma registrada para este SKU en esta ubicación."),
            "Nueva":            st.column_config.NumberColumn("Nueva cantidad",
                                    format="%d", min_value=0,
                                    help="Escribe la cantidad contada. Vacío = pendiente."),
            "Observación":      st.column_config.TextColumn("Observación",
                                    help="Opcional. Ej: dañado, en caja master..."),
        },
        key=editor_key,
        height=_editor_height,
    )

    # Progreso
    total    = len(edited)
    contados = int(edited["Nueva"].notna().sum())
    faltan   = total - contados
    pct      = (contados / total) if total else 0.0
    st.progress(pct,
        text=f"✅ Contados: {contados:,}  |  ⏳ Pendientes: {faltan:,}  "
             f"|  {pct*100:.0f}% de {total:,}")

    # Guardar / Descartar / Borrar
    cg1, cg2, cg3 = st.columns([2,1,1])
    with cg1:
        if st.button(f"💾 Guardar toma de «{sel_ubic}»",
                     type="primary", width='stretch',
                     key=f"tu_save_{sel_ubic}"):
            to_save = edited[edited["Nueva"].notna()].copy()
            if to_save.empty:
                st.warning("No hay cantidades registradas para guardar.")
            else:
                _now = _now_ec().strftime("%Y-%m-%d %H:%M")
                # En modo compacto la columna "Nombre Producto" no está en `edited`;
                # lo busco del table_df original (que siempre la tiene).
                _name_map = dict(zip(
                    table_df["Código Producto"].astype(str),
                    table_df["Nombre Producto"].astype(str),
                )) if "Nombre Producto" in table_df.columns else {}
                _rows = []
                for _, row in to_save.iterrows():
                    _obs = row.get("Observación","")
                    _cod = str(row["Código Producto"])
                    _nom = str(row["Nombre Producto"]) if "Nombre Producto" in row.index \
                           else _name_map.get(_cod, "")
                    _rows.append({
                        "Fecha": _now,
                        "Ubicación": sel_ubic,
                        "Código Producto": _cod,
                        "Nombre Producto":  _nom,
                        "Cantidad Física": float(row["Nueva"]),
                        "Observación": _obs if pd.notna(_obs) else "",
                    })
                new_df = pd.concat([rap_df, pd.DataFrame(_rows)], ignore_index=True)
                rap_state["df"] = new_df
                _persist_rapid(new_df)
                log(f"Toma «{sel_ubic}»: {len(_rows)} items guardados")
                # Limpiar estado del editor para esta ubicación
                st.session_state.pop(editor_key, None)
                st.success(f"✓ Guardados {len(_rows)} items en «{sel_ubic}». "
                           f"Cambia de ubicación arriba para continuar.")
                _rerun_frag()
    with cg2:
        if st.button("↺ Descartar edición", width='stretch',
                     key=f"tu_reset_{sel_ubic}",
                     help="Limpia los cambios sin guardar en la tabla de arriba."):
            st.session_state.pop(editor_key, None)
            _rerun_frag()
    with cg3:
        # Eliminar TODA la toma guardada de esta ubicación — acción destructiva
        _n_saved = int((rap_df["Ubicación"].astype(str) == str(sel_ubic)).sum()) \
                   if not rap_df.empty else 0
        with st.popover(f"🔥 Borrar toma ({_n_saved})",
                        width='stretch',
                        disabled=(_n_saved == 0),
                        help=("No hay tomas guardadas en esta ubicación"
                              if _n_saved == 0 else
                              f"Elimina todas las {_n_saved} filas guardadas en «{sel_ubic}»")):
            st.warning(
                f"⚠ Vas a borrar **TODAS** las tomas guardadas en "
                f"**«{sel_ubic}»** ({_n_saved} registro(s)). Acción **irreversible**. "
                f"El resto de ubicaciones no se tocará."
            )
            if st.button("✔ Confirmar eliminación",
                         type="primary",
                         key=f"tu_del_confirm_{sel_ubic}"):
                new_df = rap_df[rap_df["Ubicación"].astype(str) != str(sel_ubic)].copy()
                rap_state["df"] = new_df
                _persist_rapid(new_df)
                log(f"Toma «{sel_ubic}» eliminada: {_n_saved} registro(s) borrados")
                st.session_state.pop(editor_key, None)
                st.success(f"✓ Toma de «{sel_ubic}» eliminada ({_n_saved} registros).")
                _rerun_frag()

    # Log fragment timing
    _frag_ms = (_ptime.perf_counter() - _frag_t0) * 1000
    try:
        _fts = _now_ec().strftime("%H:%M:%S.%f")[:-3]
        with open(_PERF_LOG_PATH, "a", encoding="utf-8") as _ff:
            _ff.write(f"{_fts} FRAGMENT=toma TOTAL={_frag_ms:.0f}ms\n")
        _fh = _get_perf_history()
        _fh["runs"].append({"ts": _fts, "total_ms": round(_frag_ms),
                            "ck": [("fragment_toma", round(_frag_ms))]})
        _fh["runs"] = _fh["runs"][-30:]
    except Exception: pass


# ── Fragment: Resumen de Tomas Físicas (pivot SKU × Ubicación) ──
@_fragment
def _render_resumen_fragment():
    rap_state = _get_shared_rapid()
    rap_df    = rap_state["df"]

    if rap_df is None or rap_df.empty:
        st.info("Aún no hay tomas registradas. Ve a «⚡ Toma» para empezar.")
        # Permitir restaurar un backup previo aun sin tomas registradas:
        # es el único momento donde la pantalla se bloqueaba (post-reboot),
        # dejando al usuario sin forma de subir el ZIP de respaldo.
        with st.expander("♻ Restaurar desde backup ZIP (post-reboot)",
                          expanded=True):
            st.caption(
                "Si tienes un backup ZIP descargado antes de un reboot, "
                "súbelo aquí para restaurar historial de tomas, filtros "
                "de exclusión y ubicaciones custom."
            )
            import json, zipfile
            _up_zip = st.file_uploader("Subir backup ZIP",
                type=["zip"], key="res_restore_zip_empty",
                label_visibility="collapsed")
            if _up_zip is not None and st.button(
                "♻ Restaurar desde ZIP", type="primary",
                width='stretch', key="res_restore_btn_empty"
            ):
                try:
                    _in = zipfile.ZipFile(io.BytesIO(_up_zip.getvalue()))
                    _names = _in.namelist()
                    _summary = []
                    if "toma_fisica_rapida.xlsx" in _names:
                        with _in.open("toma_fisica_rapida.xlsx") as _fh:
                            _df_imp = pd.read_excel(_fh)
                        rap_state["df"] = _df_imp
                        _persist_rapid(_df_imp)
                        _summary.append(
                            f"📋 Historial: {len(_df_imp):,} filas restauradas")
                    if "filtros_config.json" in _names:
                        _raw = _in.read("filtros_config.json").decode("utf-8")
                        _data = json.loads(_raw)
                        _sku_set = set(_data.get("excluded_skus", []))
                        _wh_set  = set(_data.get("excluded_warehouses", []))
                        _shared_f = _get_shared_filtros()
                        _shared_f["excluded_skus"]       = _sku_set
                        _shared_f["excluded_warehouses"] = _wh_set
                        _persist_filtros(_sku_set, _wh_set)
                        st.session_state["excluded_skus"] = _sku_set
                        st.session_state["excl_wh"]       = _wh_set
                        _summary.append(
                            f"🚫 Filtros: {len(_sku_set)} SKU + "
                            f"{len(_wh_set)} bodegas")
                    if "ubicaciones_custom.json" in _names:
                        _raw = _in.read("ubicaciones_custom.json").decode("utf-8")
                        _data = json.loads(_raw)
                        _lst = list(_data.get("ubicaciones", []))
                        _c = _get_custom_ubic()
                        _c["list"] = _lst
                        _persist_custom_ubic(_lst)
                        _summary.append(
                            f"📍 Ubicaciones custom: {len(_lst)} restauradas")
                    if "importaciones.xlsx" in _names:
                        with _in.open("importaciones.xlsx") as _fh:
                            _df_imp_in = pd.read_excel(_fh)
                        for _c in _IMP_COLS:
                            if _c not in _df_imp_in.columns: _df_imp_in[_c] = ""
                        _df_imp_in = _df_imp_in[_IMP_COLS]
                        _imp_st = _get_shared_importaciones()
                        _imp_st["df"] = _df_imp_in
                        _persist_importaciones(_df_imp_in)
                        _summary.append(
                            f"📦 Importaciones: {len(_df_imp_in):,} registros restaurados")
                    if _summary:
                        st.success("✅ Restaurado:\n\n" +
                                    "\n".join(f"- {s}" for s in _summary))
                        log(f"Backup restaurado desde {_up_zip.name}: "
                            f"{'; '.join(_summary)}")
                        _rerun_frag()
                    else:
                        st.warning(
                            "El ZIP no contenía ninguno de los archivos "
                            "esperados (toma_fisica_rapida.xlsx, "
                            "filtros_config.json, ubicaciones_custom.json, "
                            "importaciones.xlsx).")
                except Exception as _ex:
                    st.error(f"Error al restaurar: {_ex}")
        return

    # Para cada (SKU, Ubicación) conservar sólo la última toma
    latest = rap_df.sort_values("Fecha").drop_duplicates(
        subset=["Código Producto","Ubicación"], keep="last")

    # Pivot: filas=SKU, columnas=Ubicación, valores=Cantidad
    pivot = latest.pivot_table(
        index=["Código Producto","Nombre Producto"],
        columns="Ubicación",
        values="Cantidad Física",
        aggfunc="sum",
    )
    # Ordenar columnas: primero DEFAULT_LOCATIONS (las que existan), luego resto alfabético
    _cols_used = list(pivot.columns)
    _ordered = [u for u in DEFAULT_LOCATIONS if u in _cols_used]
    _ordered += sorted([u for u in _cols_used if u not in _ordered])
    pivot = pivot[_ordered]

    # Total por fila (tratando NaN como 0 para la suma, sin alterar display)
    pivot_total = pivot.fillna(0).sum(axis=1).astype(int)
    pivot_display = pivot.reset_index()
    pivot_display["Total"] = pivot_total.values

    # KPIs arriba
    k1,k2,k3,k4 = st.columns(4)
    k1.metric("SKUs contados", f"{len(pivot_display):,}")
    k2.metric("Ubicaciones", f"{len(_ordered)}")
    k3.metric("Unidades totales", f"{int(pivot_total.sum()):,}")
    k4.metric("Registros", f"{len(rap_df):,}")

    # Configuración de columnas para display (números con formato)
    _col_cfg = {
        "Código Producto": st.column_config.TextColumn("SKU", width="small"),
        "Nombre Producto":  st.column_config.TextColumn("Producto"),
        "Total":            st.column_config.NumberColumn("Σ Total", format="%d"),
    }
    for _c in _ordered:
        _col_cfg[_c] = st.column_config.NumberColumn(_c, format="%d",
                         help=f"Cantidad contada en {_c}")

    st.dataframe(pivot_display, width='stretch', hide_index=True,
                 column_config=_col_cfg, height=560)

    # Export ejecutivo (listo para imprimir, con totales por columna)
    st.markdown("#### 📥 Exportar resumen ejecutivo")
    st.caption(
        "Los exports llevan formato corporativo: título, fecha, encabezados "
        "destacados, fila de TOTALES al final (sumas por columna), paisaje con "
        "repetición de header entre páginas. Listos para imprimir o enviar."
    )
    _fecha_nombre = _now_ec().strftime("%Y%m%d_%H%M")
    ec1, ec2 = st.columns(2)
    with ec1:
        st.download_button("📊 Excel ejecutivo",
            _export_resumen_excel(pivot_display, _ordered, "Resumen Toma Física"),
            f"resumen_toma_fisica_{_fecha_nombre}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch', key="res_xl")
    with ec2:
        _pdf = _export_resumen_pdf(pivot_display, _ordered, "Resumen Toma Física")
        if _pdf:
            st.download_button("📄 PDF ejecutivo",
                _pdf, f"resumen_toma_fisica_{_fecha_nombre}.pdf",
                "application/pdf",
                width='stretch', key="res_pdf")
        else:
            st.button("📄 PDF", disabled=True, width='stretch',
                     key="res_pdf_disabled",
                     help="Instalar reportlab: `pip install reportlab`")

    with st.expander("🧾 Historial completo de movimientos"):
        _hist = rap_df.sort_values("Fecha", ascending=False).reset_index(drop=True)
        # Tipos consistentes
        for _tc in ["Fecha","Ubicación","Código Producto","Nombre Producto","Observación"]:
            if _tc in _hist.columns:
                _hist[_tc] = _hist[_tc].fillna("").astype(str)
        st.dataframe(_hist, width='stretch', hide_index=True, height=320)
        st.download_button("📥 Exportar historial",
            to_xl(_hist), "historial_toma_fisica.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch', key="res_hist_xl")

    # ── Backup completo (todo en un ZIP) ────────────────────────
    with st.expander("💾 Backup completo — todo en un ZIP (recomendado antes de reboot)"):
        st.caption(
            "Descarga un ZIP con **historial de toma física + filtros de "
            "exclusión + ubicaciones custom**. Tras un reboot de Streamlit Cloud "
            "súbelo aquí para restaurar todo de una vez."
        )
        import json, zipfile

        # Construir ZIP en memoria
        _zip_buf = io.BytesIO()
        with zipfile.ZipFile(_zip_buf, "w", zipfile.ZIP_DEFLATED) as _zf:
            # Historial de toma física (si hay datos)
            if not rap_df.empty:
                _hist_bytes = to_xl(rap_df.sort_values("Fecha", ascending=False))
                _zf.writestr("toma_fisica_rapida.xlsx", _hist_bytes)
            # Filtros de exclusión
            _f = _get_shared_filtros()
            _zf.writestr("filtros_config.json", json.dumps({
                "excluded_skus":       sorted(list(_f["excluded_skus"])),
                "excluded_warehouses": sorted(list(_f["excluded_warehouses"])),
            }, ensure_ascii=False, indent=2))
            # Ubicaciones custom
            _c = _get_custom_ubic()
            _zf.writestr("ubicaciones_custom.json", json.dumps({
                "ubicaciones": list(_c["list"]),
            }, ensure_ascii=False, indent=2))
            # Importaciones (pedidos al proveedor — todos los estados)
            _imp_back = _get_shared_importaciones()["df"]
            if not _imp_back.empty:
                _zf.writestr("importaciones.xlsx", to_xl(_imp_back))
        _zip_bytes = _zip_buf.getvalue()

        _n_skus_excl = len(_get_shared_filtros()["excluded_skus"])
        _n_wh_excl   = len(_get_shared_filtros()["excluded_warehouses"])
        _n_custom    = len(_get_custom_ubic()["list"])
        _n_hist      = len(rap_df)
        _n_imp       = len(_get_shared_importaciones()["df"])

        st.markdown(
            f"**Contenido del backup:**\n\n"
            f"- 📋 **Historial de toma física**: {_n_hist:,} filas\n"
            f"- 🚫 **SKUs excluidos**: {_n_skus_excl}\n"
            f"- 🏪 **Bodegas excluidas**: {_n_wh_excl}\n"
            f"- 📍 **Ubicaciones custom**: {_n_custom}\n"
            f"- 📦 **Importaciones**: {_n_imp:,} registros"
        )

        _fecha_zip = _now_ec().strftime("%Y%m%d_%H%M")
        cz1, cz2 = st.columns(2)
        with cz1:
            st.download_button(
                "📥 Descargar backup completo (ZIP)",
                _zip_bytes,
                f"backup_autosky_{_fecha_zip}.zip",
                "application/zip",
                width='stretch', key="res_backup_zip")
        with cz2:
            _up_zip = st.file_uploader("Restaurar desde ZIP",
                                        type=["zip"], key="res_restore_zip",
                                        label_visibility="collapsed")
            if _up_zip is not None and st.button(
                "♻ Restaurar desde ZIP", type="primary",
                width='stretch', key="res_restore_btn"
            ):
                try:
                    _in = zipfile.ZipFile(io.BytesIO(_up_zip.getvalue()))
                    _names = _in.namelist()
                    _summary = []

                    # Historial
                    if "toma_fisica_rapida.xlsx" in _names:
                        with _in.open("toma_fisica_rapida.xlsx") as _fh:
                            _df_imp = pd.read_excel(_fh)
                        rap_state["df"] = _df_imp
                        _persist_rapid(_df_imp)
                        _summary.append(f"📋 Historial: {len(_df_imp):,} filas restauradas")

                    # Filtros
                    if "filtros_config.json" in _names:
                        _raw = _in.read("filtros_config.json").decode("utf-8")
                        _data = json.loads(_raw)
                        _sku_set = set(_data.get("excluded_skus", []))
                        _wh_set  = set(_data.get("excluded_warehouses", []))
                        _shared_f = _get_shared_filtros()
                        _shared_f["excluded_skus"]       = _sku_set
                        _shared_f["excluded_warehouses"] = _wh_set
                        _persist_filtros(_sku_set, _wh_set)
                        st.session_state["excluded_skus"] = _sku_set
                        st.session_state["excl_wh"]       = _wh_set
                        _summary.append(f"🚫 Filtros: {len(_sku_set)} SKU + {len(_wh_set)} bodegas")

                    # Ubicaciones custom
                    if "ubicaciones_custom.json" in _names:
                        _raw = _in.read("ubicaciones_custom.json").decode("utf-8")
                        _data = json.loads(_raw)
                        _lst = list(_data.get("ubicaciones", []))
                        _c = _get_custom_ubic()
                        _c["list"] = _lst
                        _persist_custom_ubic(_lst)
                        _summary.append(f"📍 Ubicaciones custom: {len(_lst)} restauradas")

                    # Importaciones
                    if "importaciones.xlsx" in _names:
                        with _in.open("importaciones.xlsx") as _fh:
                            _df_imp_in = pd.read_excel(_fh)
                        for _c in _IMP_COLS:
                            if _c not in _df_imp_in.columns: _df_imp_in[_c] = ""
                        _df_imp_in = _df_imp_in[_IMP_COLS]
                        _imp_st = _get_shared_importaciones()
                        _imp_st["df"] = _df_imp_in
                        _persist_importaciones(_df_imp_in)
                        _summary.append(f"📦 Importaciones: {len(_df_imp_in):,} registros restaurados")

                    if _summary:
                        st.success("✅ Restaurado:\n\n" + "\n".join(f"- {s}" for s in _summary))
                        log(f"Backup restaurado desde {_up_zip.name}: {'; '.join(_summary)}")
                        _rerun_frag()
                    else:
                        st.warning("El ZIP no contenía ninguno de los archivos esperados.")
                except Exception as _ex:
                    st.error(f"Error al restaurar: {_ex}")

    # ── Administración: borrar tomas guardadas ──────────────────
    with st.expander("🧹 Administrar datos guardados — borrar tomas"):
        st.caption("Para borrar registros guardados (ej: pruebas). Esta acción **NO** se puede deshacer.")

        _all_ubic_saved = sorted(rap_df["Ubicación"].astype(str).unique().tolist())

        # A) Borrar por ubicación(es)
        st.markdown("##### 🗂 Borrar por ubicación(es)")
        _to_del = st.multiselect(
            "Ubicaciones a borrar",
            _all_ubic_saved,
            key="res_del_ubic",
            placeholder="Elige una o más ubicaciones…",
        )
        if _to_del:
            _n_afect = int(rap_df["Ubicación"].astype(str).isin(_to_del).sum())
            with st.popover(f"🔥 Borrar {len(_to_del)} ubicación(es) ({_n_afect} registro(s))",
                            width='stretch'):
                st.warning(
                    f"⚠ Se borrarán **{_n_afect}** registro(s) de "
                    f"**{len(_to_del)}** ubicación(es): "
                    + ", ".join(f"«{u}»" for u in _to_del)
                )
                if st.button("✔ Confirmar", type="primary", key="res_del_sel_confirm"):
                    new_df = rap_df[~rap_df["Ubicación"].astype(str).isin(_to_del)].copy()
                    rap_state["df"] = new_df
                    _persist_rapid(new_df)
                    log(f"Resumen: borrados {_n_afect} registro(s) "
                        f"de {len(_to_del)} ubicación(es)")
                    # Limpiar editores activos de estas ubicaciones
                    for _u in _to_del:
                        st.session_state.pop(f"tu_ed_{_u}", None)
                    st.success(f"✓ {_n_afect} registro(s) eliminados.")
                    _rerun_frag()

        st.divider()

        # B) Borrar TODO
        st.markdown("##### 🔥 Borrar TODO el historial")
        st.caption(f"Actualmente hay **{len(rap_df):,}** registro(s) en **{len(_all_ubic_saved)}** ubicación(es).")
        with st.popover("🔥 Borrar TODAS las tomas guardadas",
                        width='stretch'):
            st.error(
                f"⚠ Esto eliminará **TODOS** los {len(rap_df):,} registro(s) de "
                f"toma física guardados, de **TODAS** las ubicaciones. "
                f"El archivo `toma_fisica_rapida.xlsx` quedará vacío. "
                f"Los archivos de inventario (consolidado.xlsx) NO se afectan."
            )
            _typed = st.text_input(
                "Escribe **BORRAR** para confirmar",
                key="res_del_all_confirm_txt",
                placeholder="BORRAR",
            )
            if st.button("✔ Confirmar borrado total",
                         type="primary",
                         disabled=(_typed.strip().upper() != "BORRAR"),
                         key="res_del_all_confirm_btn"):
                empty = pd.DataFrame(columns=_RAPIDA_COLS)
                rap_state["df"] = empty
                _persist_rapid(empty)
                log(f"Resumen: borrado TOTAL ({len(rap_df):,} registros eliminados)")
                # Limpiar TODOS los editores de tomas activos
                for _k in [k for k in list(st.session_state.keys())
                           if k.startswith("tu_ed_")]:
                    st.session_state.pop(_k, None)
                st.success("✓ Todo el historial de tomas físicas fue eliminado.")
                _rerun_frag()

# -- Fragment: sub-pestaña Importar --
@_fragment
def _render_importar_fragment():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    st.markdown("#### 📥 Importar toma física desde Excel")
    st.caption(
        "Detecta automáticamente dos formatos:\n\n"
        "• **Plantilla** (una hoja por ubicación, con columna Cantidad) — generada en «📋 Plantilla».\n"
        "• **Historial** (una sola hoja con columnas Fecha, Ubicación, SKU, Nombre, Cantidad, Observación) "
        "— exportada desde «📊 Resumen → Exportar historial». Útil para restaurar backups tras un reboot."
    )

    pf = st.file_uploader("Archivo (.xlsx / .xls)",
                          type=["xlsx","xls"], key="ph_up")

    if pf is not None:
        # Parsear en memoria
        _tp_ext = os.path.splitext(pf.name)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=_tp_ext) as _tmp:
            _tmp.write(pf.getvalue())
            _tp = _tmp.name

        # Detección de formato
        _is_historial = False
        try:
            _is_historial = _detect_historial_format(_tp)
        except Exception:
            pass

        if _is_historial:
            # ── Flujo HISTORIAL (restauración de backup) ──────────
            try:
                _hist_rows, _hist_ubic = _parse_historial(_tp)
            except Exception as _ex:
                st.error(f"Error al leer el historial: {_ex}")
                _hist_rows, _hist_ubic = [], set()
            try: os.unlink(_tp)
            except: pass

            if not _hist_rows:
                st.warning("El archivo no contiene filas válidas para importar.")
                return

            st.info(
                f"📘 Detectado formato **Historial** (restauración de backup). "
                f"Contiene **{len(_hist_rows):,}** filas en "
                f"**{len(_hist_ubic)}** ubicación(es). Las fechas originales se preservan."
            )

            # Validar que las ubicaciones existan (si no, ofrecer crearlas)
            _registered = set(_get_all_ubic())
            _missing = sorted([u for u in _hist_ubic if u not in _registered])

            if _missing:
                st.warning(
                    f"⚠ El historial incluye {len(_missing)} ubicación(es) no "
                    f"registradas en el sistema:\n\n"
                    + "\n".join(f"- `{u}`" for u in _missing)
                )
                _to_create_h = st.multiselect(
                    "Ubicaciones a crear automáticamente",
                    _missing, default=_missing,
                    key="imp_hist_create",
                    help="Las no seleccionadas omitirán sus filas al importar."
                )
                _btn_label = "✔ Crear ubicación(es) e importar historial"
            else:
                _to_create_h = []
                _btn_label = f"✔ Importar historial ({len(_hist_rows):,} filas)"

            # Preview resumen por ubicación
            _resumen = {}
            for _r in _hist_rows:
                _u = _r["Ubicación"]
                if _u not in _resumen: _resumen[_u] = {"rows":0,"units":0}
                _resumen[_u]["rows"]  += 1
                _resumen[_u]["units"] += _r["Cantidad Física"]
            _prev_df = pd.DataFrame([
                {"Ubicación": u, "Registros": v["rows"], "Unidades": int(v["units"]),
                 "Estado": "✓ Registrada" if u in _registered else
                           ("➕ Se creará" if u in _to_create_h else "⏭ Se omitirá")}
                for u, v in _resumen.items()
            ])
            st.dataframe(_prev_df, width='stretch', hide_index=True)

            if st.button(_btn_label, type="primary", width='stretch',
                         key="imp_hist_confirm"):
                # Crear ubicaciones nuevas seleccionadas
                if _to_create_h:
                    _custom = _get_custom_ubic()
                    _existing = set(_custom["list"])
                    for _n in _to_create_h:
                        if _n not in _existing:
                            _custom["list"].append(_n)
                    _persist_custom_ubic(_custom["list"])

                # Filtrar filas: solo ubicaciones registradas o recién creadas
                _valid_ubic = _registered | set(_to_create_h)
                _to_save = [r for r in _hist_rows if r["Ubicación"] in _valid_ubic]
                _skipped = len(_hist_rows) - len(_to_save)

                rap_state = _get_shared_rapid()
                rap_df    = rap_state["df"]
                if _to_save:
                    _merged = pd.concat([rap_df, pd.DataFrame(_to_save)],
                                        ignore_index=True)
                    rap_state["df"] = _merged
                    _persist_rapid(_merged)
                log(f"Import historial ({pf.name}): {len(_to_save)} filas restauradas"
                    + (f" ({_skipped} omitidas)" if _skipped else ""))
                _msg = f"✅ Restauradas {len(_to_save):,} filas del historial."
                if _skipped:
                    _msg += f" Se omitieron {_skipped} por ubicaciones descartadas."
                st.success(_msg)
                _rerun_frag()
            return   # Terminar aquí el flujo historial

        try:
            parsed = _parse_plantilla_toma(_tp, _get_all_ubic())
        except Exception as _ex:
            st.error(f"Error al leer el Excel: {_ex}")
            parsed = None
        finally:
            try: os.unlink(_tp)
            except: pass

        if parsed is not None:
            # Reset de lista de "ignorados" si cambió el archivo
            _fhash = hash(pf.getvalue()[:50000])
            if st.session_state.get("_imp_fhash") != _fhash:
                st.session_state.pop("_imp_ignore", None)
                st.session_state["_imp_fhash"] = _fhash

            _ignored = set(st.session_state.get("_imp_ignore", []))
            # Hojas inválidas aún visibles (no marcadas como ignoradas)
            _invalid_vis = [s for s in parsed["invalid_sheets"]
                            if s not in _ignored]

            if _invalid_vis:
                st.warning(
                    "⚠ El Excel contiene hojas que **no corresponden a "
                    "ubicaciones registradas**:\n\n"
                    + "\n".join(f"- `{s}`" for s in _invalid_vis)
                    + "\n\nPara las que marques, se **crearán nuevas ubicaciones**. "
                      "Las que **desmarques**, se **omitirán** de la importación."
                )
                _to_create = st.multiselect(
                    "Ubicaciones a crear en el sistema (desmarcar = omitir)",
                    _invalid_vis,
                    default=_invalid_vis,
                    key="imp_create_ubic",
                )
                _to_skip = [s for s in _invalid_vis if s not in _to_create]
                if _to_skip:
                    st.info(
                        "Se **omitirán** estas hojas (no se importan): "
                        + ", ".join(f"`{s}`" for s in _to_skip)
                    )

                _btn_label = "✔ Continuar con la importación"
                if _to_create:
                    _btn_label = (f"➕ Crear {len(_to_create)} y continuar"
                                   + (f" (omitir {len(_to_skip)})" if _to_skip else ""))
                if st.button(_btn_label, type="primary", width='stretch',
                             key="imp_create_and_continue"):
                    if _to_create:
                        _custom = _get_custom_ubic()
                        _already = set(_custom["list"])
                        _added = 0
                        for _name in _to_create:
                            if _name not in _already:
                                _custom["list"].append(_name)
                                _added += 1
                        _persist_custom_ubic(_custom["list"])
                        log(f"Importar: creadas {_added} ubicación(es) desde Excel")
                    if _to_skip:
                        st.session_state["_imp_ignore"] = list(_ignored | set(_to_skip))
                        log(f"Importar: omitidas {len(_to_skip)} hoja(s) — {', '.join(_to_skip)}")
                    _rerun_frag()

                with st.expander("ℹ Ubicaciones registradas actualmente"):
                    st.caption(", ".join(f"`{u}`" for u in _get_all_ubic()))
            elif not parsed["valid_sheets"] or all(
                    len(info["rows"]) == 0
                    for info in parsed["valid_sheets"].values()):
                st.warning(
                    "El archivo no contiene cantidades para importar. "
                    "Recuerda: celdas vacías o con valor 0 se consideran «no contadas» y se omiten."
                )
            else:
                # Preview
                st.markdown("##### 👁 Vista previa")
                _prev = []
                _need_date = []
                for ubic, info in parsed["valid_sheets"].items():
                    _n = len(info["rows"])
                    if _n == 0: continue
                    _prev.append({
                        "Ubicación": ubic,
                        "Items a importar": _n,
                        "Fecha toma (B2)": info["fecha"].strftime("%d/%m/%Y")
                                           if info["fecha"] else "— sin fecha —",
                    })
                    if info["fecha"] is None:
                        _need_date.append(ubic)
                _prev_df = pd.DataFrame(_prev)
                st.dataframe(_prev_df, width='stretch', hide_index=True)

                # Fecha fallback para hojas sin B2
                _fallback_date = None
                if _need_date:
                    st.warning(
                        "⚠ Estas hojas no traen fecha en la celda B2: "
                        + ", ".join(f"«{u}»" for u in _need_date)
                    )
                    _fallback_date = st.date_input(
                        "Fecha para estas hojas (si la dejas vacía, se usa la de importación)",
                        value=date.today(),
                        format="DD/MM/YYYY",
                        key="imp_fallback_date",
                    )

                _total_items = sum(len(info["rows"])
                                   for info in parsed["valid_sheets"].values())

                if st.button(f"✔ Confirmar importación ({_total_items} items en "
                             f"{len(_prev)} ubicación(es))",
                             type="primary", key="imp_confirm",
                             width='stretch'):
                    rap_state = _get_shared_rapid()
                    rap_df = rap_state["df"]
                    _now_str = _now_ec().strftime("%Y-%m-%d %H:%M")
                    _new_rows = []
                    for ubic, info in parsed["valid_sheets"].items():
                        if not info["rows"]: continue
                        if info["fecha"] is not None:
                            _fecha_str = info["fecha"].strftime("%Y-%m-%d") + " 00:00"
                        elif _fallback_date is not None:
                            _fecha_str = _fallback_date.strftime("%Y-%m-%d") + " 00:00"
                        else:
                            _fecha_str = _now_str
                        for _row in info["rows"]:
                            _new_rows.append({
                                "Fecha": _fecha_str,
                                "Ubicación": ubic,
                                "Código Producto": _row["codigo"],
                                "Nombre Producto": _row["nombre"],
                                "Cantidad Física": float(_row["cantidad"]),
                                "Observación": _row["obs"],
                            })
                    if _new_rows:
                        _merged = pd.concat([rap_df, pd.DataFrame(_new_rows)],
                                            ignore_index=True)
                        rap_state["df"] = _merged
                        _persist_rapid(_merged)
                        log(f"Import Excel ({pf.name}): {len(_new_rows)} items "
                            f"en {len(_prev)} ubicaciones")
                        st.success(
                            f"✅ Importados {len(_new_rows)} items en "
                            f"{len(_prev)} ubicación(es). Revisa «📊 Resumen» "
                            f"y «📊 Comparación»."
                        )
                        _rerun_frag()

# -- Fragment: sub-pestaña Plantilla --
@_fragment
def _render_plantilla_fragment():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if eng.raw_df is not None and st.button("📄 Generar plantilla",key="ph_t"):
        try:
            import openpyxl
            from openpyxl.styles import PatternFill,Font,Alignment,Border,Side
            from openpyxl.utils import get_column_letter,quote_sheetname
            # Plantilla respeta la exclusión global de SKUs
            sd=(eng.raw_df[["Código Producto","Nombre Producto"]].drop_duplicates()
                .sort_values("Código Producto").reset_index(drop=True))
            sd["Código Producto"] = sd["Código Producto"].astype(str)
            if eng.excluded_skus:
                sd = sd[~sd["Código Producto"].isin(eng.excluded_skus)].reset_index(drop=True)
            n=len(sd); DR=4
            H=PatternFill("solid",fgColor="1E3A5F"); Y=PatternFill("solid",fgColor="FEF9C3")
            G=PatternFill("solid",fgColor="D1FAE5"); Q=PatternFill("solid",fgColor="DBEAFE")
            F=PatternFill("solid",fgColor="F0FDF4"); E=PatternFill("solid",fgColor="F8FAFC")
            O=PatternFill("solid",fgColor="FFFFFF"); t=Side(style="thin",color="CBD5E1")
            brd=Border(t,t,t,t); tg=Side(style="thin",color="BBF7D0"); bg=Border(tg,tg,tg,tg)
            wb=openpyxl.Workbook(); ls={}
            _tpl_ubic = _get_all_ubic()
            for loc in _tpl_ubic:
                sn=loc[:28].replace("/","-").replace("\\","-").replace("?","").replace("*","").replace("[","").replace("]","").replace(":","")
                ws2=wb.create_sheet(sn); ls[loc]=(sn,ws2)
                ws2.merge_cells("A1:D1"); ws2["A1"]=f"TOMA FISICA — {loc.upper()}"
                ws2["A1"].font=Font(bold=True,size=12,color="FFFFFF"); ws2["A1"].fill=H
                ws2["A1"].alignment=Alignment(horizontal="center"); ws2.row_dimensions[1].height=24
                ws2["A2"]="FECHA TOMA:"; ws2["A2"].font=Font(bold=True,size=10,color="FFFFFF"); ws2["A2"].fill=H
                ws2["B2"].fill=PatternFill("solid",fgColor="EFF6FF"); ws2["B2"].font=Font(size=10,color="1E3A5F")
                for ci,h in enumerate(["Código","Nombre","Cantidad","Observación"],1):
                    ce=ws2.cell(3,ci,h); ce.font=Font(bold=True,size=9,color="1E3A5F")
                    ce.fill=Y; ce.alignment=Alignment(horizontal="center",wrap_text=True); ce.border=brd
                for ri,(_,row) in enumerate(sd.iterrows(),DR):
                    fill=E if ri%2==0 else O
                    ws2.cell(ri,1,str(row["Código Produto"] if "Código Produto" in row else row.get("Código Producto",""))).fill=fill
                    ws2.cell(ri,1).font=Font(size=9,color="1E40AF")
                    ws2.cell(ri,2,str(row["Nombre Producto"])).fill=fill; ws2.cell(ri,2).font=Font(size=9,color="111827")
                    qc=ws2.cell(ri,3,""); qc.fill=Q; qc.alignment=Alignment(horizontal="right"); qc.font=Font(size=10,bold=True,color="1E3A5F")
                    ws2.cell(ri,4,"").fill=fill
                    for ci in range(1,5): ws2.cell(ri,ci).border=brd
                    ws2.row_dimensions[ri].height=16
                tr2=n+DR; ws2.merge_cells(f"A{tr2}:B{tr2}")
                ws2.cell(tr2,1,"TOTAL").font=Font(bold=True,size=9,color="FFFFFF"); ws2.cell(tr2,1).fill=H
                tc2=ws2.cell(tr2,3,f"=SUM(C{DR}:C{tr2-1})")
                tc2.font=Font(bold=True,size=9,color="065F46"); tc2.fill=G
                tc2.alignment=Alignment(horizontal="right"); tc2.border=brd
                ws2.column_dimensions["A"].width=13; ws2.column_dimensions["B"].width=40
                ws2.column_dimensions["C"].width=11; ws2.column_dimensions["D"].width=30; ws2.freeze_panes="C4"
            ac=["Código","Nombre"]+_tpl_ubic+["TOTAL"]
            wr=wb.create_sheet("RESUMEN GENERAL",0); nc=len(ac)
            wr.merge_cells(f"A1:{get_column_letter(nc)}1")
            wr["A1"]="RESUMEN — Actualiza automáticamente"; wr["A1"].font=Font(bold=True,size=12,color="FFFFFF"); wr["A1"].fill=H
            wr["A1"].alignment=Alignment(horizontal="center"); wr.row_dimensions[1].height=24
            wr.merge_cells(f"A2:{get_column_letter(nc)}2")
            wr["A2"]="⚠ NO editar — calculado automáticamente desde cada hoja"
            wr["A2"].font=Font(bold=True,size=9,color="92400E"); wr["A2"].fill=PatternFill("solid",fgColor="FEF3C7")
            wr["A2"].alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); wr.row_dimensions[2].height=28
            for ci,col in enumerate(ac,1):
                ce=wr.cell(3,ci,col); ce.font=Font(bold=True,size=9,color="1E3A5F"); ce.fill=Y
                ce.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); ce.border=brd
            wr.row_dimensions[3].height=30
            for ri,(_,row) in enumerate(sd.iterrows(),DR):
                fill=E if ri%2==0 else O
                wr.cell(ri,1,str(row.get("Código Producto",""))).fill=fill; wr.cell(ri,1).font=Font(size=9,color="1E40AF")
                wr.cell(ri,2,str(row["Nombre Producto"])).fill=fill; wr.cell(ri,2).font=Font(size=9,color="111827")
                for li,loc in enumerate(_tpl_ubic):
                    ci=3+li; sn,_=ls[loc]
                    ce=wr.cell(ri,ci,f"=IFERROR({quote_sheetname(sn)}!C{ri},0)")
                    ce.fill=F; ce.font=Font(size=9,color="065F46"); ce.alignment=Alignment(horizontal="right"); ce.border=bg
                cf=get_column_letter(3); cl=get_column_letter(3+len(_tpl_ubic)-1)
                tc=wr.cell(ri,nc,f"=SUM({cf}{ri}:{cl}{ri})")
                tc.fill=G; tc.font=Font(bold=True,size=9,color="065F46"); tc.alignment=Alignment(horizontal="right"); tc.border=brd
                for ci in range(1,3): wr.cell(ri,ci).border=brd
                wr.row_dimensions[ri].height=16
            tr=n+DR; wr.merge_cells(f"A{tr}:B{tr}")
            wr.cell(tr,1,"TOTAL GENERAL").font=Font(bold=True,size=9,color="FFFFFF"); wr.cell(tr,1).fill=H
            for ci in range(3,nc+1):
                cl=get_column_letter(ci); ce=wr.cell(tr,ci,f"=SUM({cl}{DR}:{cl}{tr-1})")
                ce.font=Font(bold=True,size=9,color="FFFFFF"); ce.fill=H
                ce.alignment=Alignment(horizontal="right"); ce.border=brd
            wr.column_dimensions["A"].width=13; wr.column_dimensions["B"].width=38
            for ci in range(3,nc+1): wr.column_dimensions[get_column_letter(ci)].width=13
            wr.freeze_panes="C4"
            buf=io.BytesIO(); wb.save(buf)
            st.download_button("📥 Descargar plantilla",buf.getvalue(),"plantilla_toma_fisica.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.success("✓ Plantilla lista")
        except Exception as e: st.error(str(e))

# -- Fragment: sub-pestaña Comparacion --
@_fragment
def _render_comparacion_fragment():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    # Comparación unificada: lee de toma_fisica_rapida.xlsx
    # (mismo archivo que usan «⚡ Toma» e «📥 Importar»).
    _rap_state = _get_shared_rapid()
    _rap_df    = _rap_state["df"]
    if r is None:
        st.info("Ejecute el análisis primero.")
        return
    if _rap_df is None or _rap_df.empty:
        st.info("Aún no hay tomas registradas. Usa «⚡ Toma» o «📥 Importar».")
        return

    _inv = r.inventory_by_warehouse.copy()
    if excl_w: _inv = _inv[~_inv["Bodega"].isin(excl_w)]
    _all_bod = sorted(_inv["Bodega"].dropna().unique().tolist())

    # Selector de "bodegas contables" (físicamente contables vs consignadas en clientes).
    # Default: solo Bodega Principal. Se persiste en session_state.
    # Importante: inicializar session_state ANTES del widget; NO pasar `default=`
    # al widget cuando ya existe la key — Streamlit advierte y puede causar
    # comportamiento inestable.
    if "_cmp_contables" not in st.session_state:
        _default_contables = [b for b in _all_bod if b == PRIMARY_WAREHOUSE]
        if not _default_contables and _all_bod:
            _default_contables = [_all_bod[0]]
        st.session_state["_cmp_contables"] = _default_contables

    sc1, sc2 = st.columns([3, 2])
    with sc1:
        contables = st.multiselect(
            "🏢 Bodegas contables (físicamente contables — se comparan contra la toma)",
            _all_bod,
            key="_cmp_contables",
            help="Bodegas donde el inventario está en tu posesión y se puede contar. "
                 "El resto se considera 'consignado en clientes' y se reporta aparte.",
        )
    with sc2:
        _tol = st.number_input(
            "Tolerancia (± unidades para 'coincide')",
            min_value=0.0, max_value=100.0, value=0.5, step=0.5,
            key="_cmp_tol",
        )

    if not contables:
        st.warning("Selecciona al menos una bodega contable.")
        return

    # ── FÍSICO: suma por SKU de la última toma en cada ubicación
    _latest = _rap_df.sort_values("Fecha").drop_duplicates(
        subset=["Código Producto","Ubicación"], keep="last")
    _phys_sku = (_latest.groupby(["Código Producto","Nombre Producto"])
                 ["Cantidad Física"].sum().reset_index())

    # ── CALCULADO CONTABLE: stock SOLO en bodegas contables
    _inv_ct  = _inv[_inv["Bodega"].isin(contables)]
    _calc_ct = (_inv_ct.groupby(["Código Producto","Nombre Producto"])
                ["Stock"].sum().reset_index()
                .rename(columns={"Stock":"Cantidad Calculada"}))

    # ── CONSIGNADO EN CLIENTES: stock en bodegas NO contables (separado)
    _inv_cl  = _inv[~_inv["Bodega"].isin(contables)]
    _cons_cl = (_inv_cl.groupby(["Código Producto","Nombre Producto"])
                ["Stock"].sum().reset_index()
                .rename(columns={"Stock":"En Clientes"}))

    # Merge outer para la comparación
    _cmp = _phys_sku.merge(_calc_ct,
                           on=["Código Producto","Nombre Producto"],
                           how="outer").fillna(0)
    _cmp = _cmp.merge(_cons_cl,
                      on=["Código Producto","Nombre Producto"],
                      how="left").fillna(0)
    # Diferencia = Sistema − Físico
    #   > 0 → FALTANTE (sistema dice que debería haber más de lo contado)
    #   < 0 → SOBRANTE (hay más físicamente de lo que el sistema dice)
    _cmp["Diferencia"] = _cmp["Cantidad Calculada"] - _cmp["Cantidad Física"]
    _cmp["Coincide"]   = _cmp["Diferencia"].abs() <= _tol
    _cmp = _cmp.sort_values("Diferencia",
                            key=lambda s: s.abs(), ascending=False)

    # Banner explicativo
    st.markdown(
        f"<div style='background:#e0f2fe;border:1px solid #7dd3fc;border-radius:8px;"
        f"padding:8px 14px;margin:8px 0;font-size:12px;color:#075985;line-height:1.5'>"
        f"<b>📘 Lógica de la comparación</b><br>"
        f"• <b>Físico</b> = suma de la última cantidad contada por SKU en todas las "
        f"ubicaciones de la toma.<br>"
        f"• <b>Calculado</b> = stock del sistema SOLO en las bodegas marcadas como "
        f"«contables» ({', '.join(f'<code>{b}</code>' for b in contables)}).<br>"
        f"• Las bodegas <b>no contables</b> (consignadas en clientes) se listan "
        f"aparte abajo — su stock es tuyo pero no se puede contar físicamente.<br>"
        f"• <b>Diferencia</b> = Sistema − Físico. "
        f"<b>+</b> = faltante (sistema dice que debería haber más · investigar). "
        f"<b>−</b> = sobrante (hay más físicamente). "
        f"Coincide si |Diff| ≤ {_tol:g}."
        f"</div>",
        unsafe_allow_html=True
    )

    # ── Estado por fila: FALTANTE / SOBRANTE / OK ──
    # Diferencia > tol  → FALTANTE (sistema dice que debería haber más)
    # Diferencia < -tol → SOBRANTE (hay más físicamente)
    # |Diff| ≤ tol      → OK (coincide)
    def _estado_row(d):
        if abs(d) <= _tol:
            return "✅ OK"
        return "🟡 FALTANTE" if d > 0 else "🔵 SOBRANTE"
    _cmp["Estado"] = _cmp["Diferencia"].apply(_estado_row)

    # Ordenamiento inteligente: primero FALTANTE (mayor magnitud),
    # luego SOBRANTE, luego OK al final.
    _orden_est = {"🟡 FALTANTE": 0, "🔵 SOBRANTE": 1, "✅ OK": 2}
    _cmp["_ord"] = _cmp["Estado"].map(_orden_est)
    _cmp = _cmp.sort_values(["_ord", "Diferencia"],
                            key=lambda s: s.abs() if s.name == "Diferencia" else s,
                            ascending=[True, False])
    _cmp = _cmp.drop(columns=["_ord"])

    # KPIs
    _ex          = (_cmp["Coincide"].mean()*100) if len(_cmp) else 0
    _n_faltante  = int((_cmp["Estado"] == "🟡 FALTANTE").sum())
    _n_sobrante  = int((_cmp["Estado"] == "🔵 SOBRANTE").sum())
    _n_ok        = int((_cmp["Estado"] == "✅ OK").sum())
    _total_fis   = int(_cmp["Cantidad Física"].sum())
    _total_ct    = int(_cmp["Cantidad Calculada"].sum())
    _total_cl    = int(_cmp["En Clientes"].sum())
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Exactitud", f"{_ex:.1f}%")
    m2.metric("🟡 Faltantes", _n_faltante,
              help="SKUs donde el sistema dice que debería haber más de lo contado.")
    m3.metric("🔵 Sobrantes", _n_sobrante,
              help="SKUs donde hay más físicamente del que el sistema reporta.")
    m4.metric("✅ Coinciden", _n_ok)

    n1,n2,n3 = st.columns(3)
    n1.metric("Σ Físico contado", f"{_total_fis:,} u")
    n2.metric("Σ Calculado contable", f"{_total_ct:,} u")
    n3.metric("Σ En clientes", f"{_total_cl:,} u",
              help="Stock del sistema en bodegas consignadas a clientes (no contables)")

    # Filtro de vista — 5 opciones
    _show_mode = st.radio("Mostrar",
        ["Con diferencia","🟡 Solo faltantes","🔵 Solo sobrantes",
         "✅ Coinciden","Todo"],
        horizontal=True, key="cmp_mode")
    if _show_mode == "Con diferencia":
        _cmp_show = _cmp[_cmp["Estado"] != "✅ OK"].copy()
    elif _show_mode == "🟡 Solo faltantes":
        _cmp_show = _cmp[_cmp["Estado"] == "🟡 FALTANTE"].copy()
    elif _show_mode == "🔵 Solo sobrantes":
        _cmp_show = _cmp[_cmp["Estado"] == "🔵 SOBRANTE"].copy()
    elif _show_mode == "✅ Coinciden":
        _cmp_show = _cmp[_cmp["Estado"] == "✅ OK"].copy()
    else:
        _cmp_show = _cmp.copy()

    # Formatear diferencia con flecha/signo
    def _fmt_dif(d):
        d = int(d) if pd.notna(d) else 0
        if d > 0:  return f"+{d:,} ⬆"
        if d < 0:  return f"{d:,} ⬇"
        return "0 ="

    _cmp_show = _cmp_show[["Código Producto","Nombre Producto",
                           "Cantidad Calculada","Cantidad Física",
                           "Diferencia","En Clientes","Estado"]].reset_index(drop=True)
    for _c in ("Cantidad Calculada","Cantidad Física","En Clientes"):
        _cmp_show[_c] = pd.to_numeric(_cmp_show[_c], errors="coerce").fillna(0).astype(int)
    # Guardar numérico para export, mostrar con flecha
    _cmp_show["Δ"] = pd.to_numeric(_cmp_show["Diferencia"], errors="coerce").fillna(0).astype(int).apply(_fmt_dif)
    _cmp_show = _cmp_show.drop(columns=["Diferencia"])
    # Reordenar: Estado primero para que sea lo primero que se vea
    _cmp_show = _cmp_show[["Estado","Código Producto","Nombre Producto",
                           "Cantidad Calculada","Cantidad Física","Δ","En Clientes"]]

    st.dataframe(_cmp_show, width='stretch', hide_index=True, height=520,
                 column_config={
                     "Estado":          st.column_config.TextColumn("Estado", width="small"),
                     "Código Producto": st.column_config.TextColumn("SKU", width="small"),
                     "Nombre Producto": st.column_config.TextColumn("Producto"),
                     "Cantidad Calculada": st.column_config.NumberColumn(
                         "Sistema (contable)", format="%d",
                         help=f"Stock en: {', '.join(contables)}"),
                     "Cantidad Física":    st.column_config.NumberColumn("Físico", format="%d"),
                     "Δ":                  st.column_config.TextColumn("Diferencia", width="small",
                         help="Sistema − Físico. ⬆ = faltante, ⬇ = sobrante."),
                     "En Clientes":        st.column_config.NumberColumn(
                         "En clientes", format="%d",
                         help="Stock en bodegas consignadas (informativo)"),
                 })
    dl3(_cmp_show, "comparacion_toma", "ph")

    # ── Detalle de Consignado en Clientes ──────────────────────
    st.markdown("---")
    with st.expander(f"🏪 Stock consignado en clientes ({len(_inv_cl['Bodega'].unique())} bodegas · "
                     f"{_total_cl:,} u totales) — no se cuenta físicamente"):
        if _inv_cl.empty:
            st.caption("No hay stock en bodegas consignadas con los filtros actuales.")
        else:
            # Pivot SKU × Bodega de clientes
            _pv_cl = _inv_cl.pivot_table(
                index=["Código Producto","Nombre Producto"],
                columns="Bodega", values="Stock", aggfunc="sum"
            ).fillna(0).astype(int).reset_index()
            # Total por SKU
            _cl_cols = [c for c in _pv_cl.columns
                        if c not in ("Código Producto","Nombre Producto")]
            _pv_cl["Σ En clientes"] = _pv_cl[_cl_cols].sum(axis=1)
            _pv_cl = _pv_cl.sort_values("Σ En clientes", ascending=False)
            st.dataframe(_pv_cl, width='stretch', hide_index=True, height=400,
                         column_config={
                             "Código Producto": st.column_config.TextColumn("SKU", width="small"),
                             "Nombre Producto": st.column_config.TextColumn("Producto"),
                             "Σ En clientes":   st.column_config.NumberColumn("Σ Total", format="%d"),
                             **{c: st.column_config.NumberColumn(c, format="%d") for c in _cl_cols},
                         })
            st.download_button("📥 Exportar consignado a Excel",
                to_xl(_pv_cl), "consignado_clientes.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', key="cons_cl_xl")


_perf("before_tab_phy")
# ══ TAB 9 TOMA FÍSICA ═══════════════════════════════════════════
with T_PHY:
    st.markdown("### 🏭 Toma Física")
    p0,p_res,p1,p2,p3=st.tabs([
        "⚡ Toma", "📊 Resumen",
        "📥 Importar", "📋 Plantilla", "📊 Comparación"
    ])

    # ── Sub-pestaña: TOMA POR UBICACIÓN (aislada en @st.fragment) ──
    with p0:
        _render_toma_fragment()

    # ── Sub-pestaña: RESUMEN (aislado en @st.fragment) ────────────
    with p_res:
        _render_resumen_fragment()


    with p1:
        _render_importar_fragment()
    with p2:
        _render_plantilla_fragment()
    with p3:
        _render_comparacion_fragment()

# ══ TAB IMPORTACIÓN ════════════════════════════════════════════
@_fragment
def _render_tab_imp():
    """Gestión de pedidos al proveedor (importaciones).

    Tres estados del ciclo de vida del embarque:
      INGRESADA → el pedido aún no llega (aparece en reportes como "EN TRÁNSITO")
      LLEGADA   → ya está en bodega pero Contifico aún no lo registra
                  (se suma al Stock Disponible para que los reportes ejecutivos
                  reflejen el stock real, no el contable)
      PROCESADA → Contifico ya lo tiene en el Excel consolidado; el registro
                  deja de influir en cualquier cálculo (archivo histórico)

    Cada embarque es una fila independiente; las entregas parciales se modelan
    creando varias filas con el mismo `Pedido` (PO)."""
    eng = st.session_state.engine
    imp_state = _get_shared_importaciones()
    imp_df    = imp_state["df"]

    st.markdown("### 📦 Importación de mercadería")
    st.caption(
        "Pedidos al proveedor: en tránsito, llegados a bodega (aún no en "
        "Contifico) y procesados. Los estados INGRESADA y LLEGADA se reflejan "
        "en los reportes ejecutivos (pestaña **Exportar Portal**). Las "
        "entregas parciales se registran como filas separadas con el mismo "
        "número de Pedido."
    )

    # ── KPIs superiores ─────────────────────────────────────────
    _n_ingr = int((imp_df["Estado"].astype(str).str.upper() == "INGRESADA").sum())
    _n_lleg = int((imp_df["Estado"].astype(str).str.upper() == "LLEGADA").sum())
    _n_proc = int((imp_df["Estado"].astype(str).str.upper() == "PROCESADA").sum())
    _u_ingr = int(pd.to_numeric(imp_df.loc[imp_df["Estado"].astype(str).str.upper()=="INGRESADA","Cantidad"], errors="coerce").fillna(0).sum())
    _u_lleg = int(pd.to_numeric(imp_df.loc[imp_df["Estado"].astype(str).str.upper()=="LLEGADA","Cantidad"], errors="coerce").fillna(0).sum())

    k1,k2,k3,k4 = st.columns(4)
    k1.metric("🚢 En tránsito", f"{_n_ingr:,}", f"{_u_ingr:,} u")
    k2.metric("📦 En bodega (sin Contifico)", f"{_n_lleg:,}", f"{_u_lleg:,} u")
    k3.metric("✅ Procesadas", f"{_n_proc:,}")
    k4.metric("Registros totales", f"{len(imp_df):,}")

    # Listas para autocompletado / dropdowns
    _known_skus   = []
    _sku_name_map = {}
    if eng is not None and eng.raw_df is not None and not eng.raw_df.empty:
        _known_skus = sorted(set(eng.raw_df["Código Producto"].dropna().astype(str)))
        # Mapa SKU→Nombre desde raw_df (última ocurrencia gana)
        _raw_valid = eng.raw_df.dropna(subset=["Código Producto"])
        for _s, _grp in _raw_valid.groupby(_raw_valid["Código Producto"].astype(str)):
            _nm = _grp["Nombre Producto"].dropna().astype(str)
            _sku_name_map[_s] = _nm.iloc[0] if len(_nm) else ""

    _prev_proveedores = sorted(set(
        imp_df["Proveedor"].dropna().astype(str).str.strip().replace("", pd.NA).dropna()
    ))
    _prev_pedidos = sorted(set(
        imp_df["Pedido"].dropna().astype(str).str.strip().replace("", pd.NA).dropna()
    ))

    # ── Formulario: Nueva importación ───────────────────────────
    # Diseño: cabecera (Pedido/OC + Proveedor) aplica a TODAS las líneas
    # de la tabla. Cada línea se guarda como un registro independiente con
    # su propio ID IMP-XXXX, pero compartiendo Pedido + Proveedor. Así un
    # mismo SKU puede repetirse en 2 líneas con distinto Tipo Embarque
    # (ej: 40 aéreo + 60 marítimo) o fecha.
    with st.expander("➕ Nueva importación", expanded=(imp_df.empty)):
        st.caption(
            "El **Pedido (OC) + Proveedor** son el contenedor del pedido. "
            "Debajo agregás una fila por cada SKU + tipo de embarque — un "
            "mismo SKU puede tener 2 filas si viene parte aéreo y parte "
            "marítimo (o en fechas distintas). Cada línea se registra como "
            "embarque independiente para permitir llegadas parciales."
        )

        with st.form("imp_new_form", clear_on_submit=True):
            # ── Cabecera del pedido ─────────────────────────────
            ch1, ch2 = st.columns(2)
            with ch1:
                _f_pedido = st.text_input(
                    "📦 Pedido / OC",
                    placeholder="Ej: PO-2026-042",
                    help="Número de orden de compra al proveedor. Este valor "
                         "se aplica a todas las líneas registradas abajo.")
            with ch2:
                _f_proveedor = st.text_input(
                    "🏭 Proveedor",
                    placeholder="Nombre del proveedor")
            if _prev_proveedores:
                st.caption(f"💡 Proveedores anteriores: "
                           f"{', '.join(_prev_proveedores[:8])}"
                           + (" …" if len(_prev_proveedores) > 8 else ""))

            _f_obs_cab = st.text_area(
                "Observaciones generales del pedido (opcional)",
                placeholder="Notas que aplican a todo el pedido — ej: "
                            "condiciones, lead time, etc.",
                height=60)

            # ── Líneas del pedido (data_editor dinámico) ────────
            st.markdown("**Líneas del pedido**")
            _hint_skus = (f"{len(_known_skus)} SKUs registrados en el "
                          "consolidado — al guardar, si el SKU coincide, "
                          "el Nombre se autocompleta.") if _known_skus else (
                          "Aún no hay SKUs cargados del consolidado — "
                          "completa SKU y Nombre manualmente.")
            st.caption(f"Usa el botón **➕** de la tabla para agregar filas "
                       f"(una por línea del pedido). {_hint_skus}")

            # Template vacío: una fila placeholder que el usuario llena o
            # borra. Con Cantidad=0 la validación filtra la fila vacía al
            # submit sin requerir que el usuario la elimine manualmente.
            _default_fecha = _now_ec().date() + timedelta(days=30)
            _empty_lines = pd.DataFrame([
                {"SKU": "", "Nombre": "", "Cantidad": 0,
                 "Tipo Embarque": _IMP_TIPOS[0],
                 "Fecha Tentativa": _default_fecha,
                 "Observaciones": ""}
            ])

            _edit_lines = st.data_editor(
                _empty_lines,
                key="imp_new_lines_editor",
                num_rows="dynamic",
                width='stretch',
                hide_index=True,
                column_config={
                    "SKU": st.column_config.TextColumn(
                        "SKU", required=False, width="small",
                        help="Código del producto. Se autocompleta el "
                             "Nombre al guardar si está en el consolidado."),
                    "Nombre": st.column_config.TextColumn(
                        "Nombre (opcional)", width="medium",
                        help="Déjalo vacío para auto-rellenar desde el "
                             "consolidado; complétalo manualmente sólo "
                             "si el SKU es nuevo."),
                    "Cantidad": st.column_config.NumberColumn(
                        "Cantidad", min_value=0, step=1, format="%d",
                        help="Debe ser mayor a 0 para que la línea se "
                             "registre. Filas con 0 se ignoran."),
                    "Tipo Embarque": st.column_config.SelectboxColumn(
                        "Tipo Embarque", options=_IMP_TIPOS,
                        help="Marítimo o Aéreo. Para llegadas parciales "
                             "del mismo SKU, crea 2 filas distintas."),
                    "Fecha Tentativa": st.column_config.DateColumn(
                        "Fecha Tentativa", format="DD/MM/YYYY"),
                    "Observaciones": st.column_config.TextColumn(
                        "Obs. de línea (opcional)", width="medium"),
                },
            )

            _submit = st.form_submit_button(
                "💾 Registrar importación", type="primary",
                width='stretch')

            if _submit:
                # Validación
                _errs = []
                if not (_f_pedido or "").strip():
                    _errs.append("Falta **Pedido / OC**.")
                if not (_f_proveedor or "").strip():
                    _errs.append("Falta **Proveedor**.")

                _valid = _edit_lines.copy()
                _valid["SKU"] = _valid["SKU"].fillna("").astype(str).str.strip()
                _valid["Cantidad"] = pd.to_numeric(
                    _valid["Cantidad"], errors="coerce").fillna(0).astype(int)
                _valid = _valid[(_valid["SKU"] != "") & (_valid["Cantidad"] > 0)]
                if len(_valid) == 0:
                    _errs.append("Agrega al menos una línea con **SKU** "
                                 "y **Cantidad** > 0.")

                if _errs:
                    for _e in _errs: st.error(_e)
                else:
                    # Crear un registro por línea válida, compartiendo
                    # Pedido + Proveedor de la cabecera. ID correlativo
                    # consecutivo dentro del mismo submit.
                    _base_next = _next_import_id(imp_df)
                    _n = int(_base_next.split("-")[1])
                    _pedido = _f_pedido.strip()
                    _proveedor = _f_proveedor.strip()
                    _obs_cab = (_f_obs_cab or "").strip()
                    _now_ts = _now_ec().strftime("%Y-%m-%d %H:%M")

                    _new_rows = []
                    for _, _row in _valid.iterrows():
                        _sku = str(_row["SKU"]).strip()
                        _nombre = str(_row.get("Nombre","") or "").strip()
                        if not _nombre:
                            _nombre = _sku_name_map.get(_sku, "")
                        _qty = int(_row["Cantidad"])
                        _tipo = _row.get("Tipo Embarque") or _IMP_TIPOS[0]
                        _fecha = _row.get("Fecha Tentativa")
                        if _fecha is None or (
                            isinstance(_fecha, float) and pd.isna(_fecha)):
                            _fecha_str = ""
                        elif hasattr(_fecha, "strftime"):
                            _fecha_str = _fecha.strftime("%Y-%m-%d")
                        else:
                            _fecha_str = str(_fecha)
                        _obs_line = str(_row.get("Observaciones","") or "").strip()
                        # Observaciones final: combinar cabecera + línea
                        if _obs_cab and _obs_line:
                            _obs_final = f"{_obs_cab} · {_obs_line}"
                        else:
                            _obs_final = _obs_cab or _obs_line

                        _new_rows.append({
                            "ID":                f"IMP-{_n:04d}",
                            "Pedido":            _pedido,
                            "Proveedor":         _proveedor,
                            "Código Producto":   _sku,
                            "Nombre Producto":   _nombre,
                            "Cantidad":          _qty,
                            "Tipo Embarque":     _tipo,
                            "Fecha Tentativa":   _fecha_str,
                            "Estado":            "INGRESADA",
                            "Fecha Registro":    _now_ts,
                            "Fecha Llegada Real": "",
                            "Observaciones":     _obs_final,
                        })
                        _n += 1

                    _new_df = pd.concat(
                        [imp_df, pd.DataFrame(_new_rows, columns=_IMP_COLS)],
                        ignore_index=True)
                    imp_state["df"] = _new_df
                    _persist_importaciones(_new_df)
                    log(f"Importación creada: Pedido={_pedido} "
                        f"Proveedor={_proveedor} — {len(_new_rows)} líneas")
                    st.success(
                        f"✅ Registrado pedido **{_pedido}** ({_proveedor}): "
                        f"{len(_new_rows)} línea(s) en estado INGRESADA.")
                    _rerun_frag()

    st.divider()

    # Helper local: renderiza una sección de estado con data_editor editable,
    # selección de filas vía checkbox y botones de acción masiva.
    def _render_section(estado, titulo, accion_siguiente, help_txt,
                          editable_cantidad=True, editable_otros=True):
        _mask = imp_df["Estado"].astype(str).str.upper() == estado
        _sub  = imp_df[_mask].copy()
        st.markdown(f"#### {titulo}  ·  {len(_sub)} registro(s)")
        if help_txt:
            st.caption(help_txt)
        if _sub.empty:
            st.info("Sin registros en este estado.")
            return

        # Normalizar para display
        _sub["Cantidad"] = pd.to_numeric(_sub["Cantidad"], errors="coerce").fillna(0).astype(int)
        for _c in ("Pedido","Proveedor","Código Producto","Nombre Producto",
                    "Tipo Embarque","Fecha Tentativa","Observaciones",
                    "Fecha Llegada Real","Fecha Registro"):
            _sub[_c] = _sub[_c].fillna("").astype(str)
        _sub.insert(0, "Sel", False)

        # Columnas visibles por estado (Fecha Llegada Real sólo relevante en LLEGADA/PROCESADA)
        _show_cols = ["Sel","ID","Pedido","Proveedor","Código Producto",
                      "Nombre Producto","Cantidad","Tipo Embarque",
                      "Fecha Tentativa"]
        if estado in ("LLEGADA","PROCESADA"):
            _show_cols.append("Fecha Llegada Real")
        _show_cols += ["Observaciones","Fecha Registro"]

        # Configuración de columnas: editables sólo donde aplique
        _col_cfg = {
            "Sel": st.column_config.CheckboxColumn("✓", width="small",
                     help="Marca las filas a las que aplicar la acción masiva de abajo"),
            "ID": st.column_config.TextColumn("ID", disabled=True, width="small"),
            "Pedido":          st.column_config.TextColumn("Pedido / PO",
                                disabled=not editable_otros),
            "Proveedor":       st.column_config.TextColumn("Proveedor",
                                disabled=not editable_otros),
            "Código Producto": st.column_config.TextColumn("SKU",
                                disabled=not editable_otros),
            "Nombre Producto": st.column_config.TextColumn("Nombre",
                                disabled=not editable_otros),
            "Cantidad":        st.column_config.NumberColumn("Cantidad",
                                min_value=0, step=1, format="%d",
                                disabled=not editable_cantidad),
            "Tipo Embarque":   st.column_config.SelectboxColumn("Tipo",
                                options=_IMP_TIPOS, disabled=not editable_otros),
            "Fecha Tentativa": st.column_config.TextColumn("F. Tentativa",
                                disabled=not editable_otros),
            "Fecha Llegada Real": st.column_config.TextColumn("F. Llegada",
                                disabled=(estado == "PROCESADA")),
            "Observaciones":   st.column_config.TextColumn("Observaciones",
                                disabled=(estado == "PROCESADA"), width="medium"),
            "Fecha Registro":  st.column_config.TextColumn("F. Registro",
                                disabled=True, width="small"),
        }

        _edited = st.data_editor(
            _sub[_show_cols], key=f"imp_ed_{estado}",
            width='stretch', hide_index=True,
            column_config=_col_cfg, num_rows="fixed",
            disabled=(estado == "PROCESADA"))

        _ids_sel = [str(r["ID"]) for _, r in _edited.iterrows() if bool(r.get("Sel"))]
        _n_sel = len(_ids_sel)

        # Guardar cambios editados — compara vs original y actualiza imp_df global
        _changed = False
        _orig_by_id = _sub.set_index("ID")
        _edit_by_id = _edited.set_index("ID")
        _editable_fields = []
        if editable_otros:
            _editable_fields += ["Pedido","Proveedor","Código Producto",
                                   "Nombre Producto","Tipo Embarque",
                                   "Fecha Tentativa"]
        if editable_cantidad:
            _editable_fields.append("Cantidad")
        if estado != "PROCESADA":
            _editable_fields += ["Fecha Llegada Real","Observaciones"]

        for _id in _edit_by_id.index:
            if _id not in _orig_by_id.index: continue
            for _col in _editable_fields:
                if _col not in _edit_by_id.columns: continue
                _v_new = _edit_by_id.at[_id, _col]
                _v_old = _orig_by_id.at[_id, _col]
                if str(_v_new) != str(_v_old):
                    # Aplicar cambio al DataFrame global
                    _row_idx = imp_df.index[imp_df["ID"] == _id]
                    if len(_row_idx) > 0:
                        imp_df.at[_row_idx[0], _col] = (
                            int(_v_new) if _col == "Cantidad" else _v_new)
                    _changed = True

        # Botones de acción
        b1, b2, b3 = st.columns(3)
        with b1:
            _btn_save = st.button(
                f"💾 Guardar cambios editados", key=f"imp_save_{estado}",
                width='stretch', disabled=not _changed,
                type=("primary" if _changed else "secondary"))
        with b2:
            _btn_adv = st.button(
                f"{accion_siguiente[1]}  ({_n_sel} sel.)" if accion_siguiente else "—",
                key=f"imp_adv_{estado}", width='stretch',
                disabled=(accion_siguiente is None or _n_sel == 0),
                type=("primary" if _n_sel else "secondary"),
                help=(accion_siguiente[2] if accion_siguiente else None))
        with b3:
            _btn_del = st.button(
                f"🗑 Eliminar  ({_n_sel} sel.)", key=f"imp_del_{estado}",
                width='stretch', disabled=(_n_sel == 0))

        if _btn_save and _changed:
            imp_state["df"] = imp_df
            _persist_importaciones(imp_df)
            log(f"Importaciones: {estado} — guardados cambios inline")
            st.success("✅ Cambios guardados")
            _rerun_frag()

        if _btn_adv and accion_siguiente and _ids_sel:
            _next_state = accion_siguiente[0]
            _fecha_real_default = _now_ec().strftime("%Y-%m-%d")
            _mask2 = imp_df["ID"].isin(_ids_sel)
            imp_df.loc[_mask2, "Estado"] = _next_state
            if _next_state == "LLEGADA":
                # Asignar Fecha Llegada Real si está vacía
                _empty = _mask2 & (imp_df["Fecha Llegada Real"].fillna("").astype(str) == "")
                imp_df.loc[_empty, "Fecha Llegada Real"] = _fecha_real_default
            imp_state["df"] = imp_df
            _persist_importaciones(imp_df)
            log(f"Importaciones: {len(_ids_sel)} registros → {_next_state}")
            st.success(f"✅ {len(_ids_sel)} registro(s) pasaron a {_next_state}")
            _rerun_frag()

        if _btn_del and _ids_sel:
            # Usar nombre distinto para no hacer local el `imp_df` del scope
            # exterior — Python detecta la asignación y bloquea el read
            # anterior con UnboundLocalError si se llama `imp_df = ...`.
            _before = len(imp_df)
            _new_df = imp_df[~imp_df["ID"].isin(_ids_sel)].reset_index(drop=True)
            imp_state["df"] = _new_df
            _persist_importaciones(_new_df)
            log(f"Importaciones: eliminados {_before - len(_new_df)} registro(s)")
            st.success(f"🗑 Eliminados {_before - len(_new_df)} registro(s)")
            _rerun_frag()

    # ── Sección A: INGRESADAS ───────────────────────────────────
    _render_section(
        estado="INGRESADA",
        titulo="🚢 INGRESADAS — en tránsito (suman a EN TRÁNSITO en reportes)",
        accion_siguiente=("LLEGADA", "📦 Marcar como LLEGADA",
                           "Las filas seleccionadas pasan a 'en bodega'; "
                           "Fecha Llegada Real se auto-fija en hoy (editable)."),
        help_txt="Pedidos aún no recibidos. Todos los campos son editables "
                 "(incluye SKU y cantidad, por si se reajusta o divide en "
                 "parciales). Para dividir una entrega parcial: edita la "
                 "cantidad aquí y crea un nuevo registro con el mismo Pedido.",
        editable_cantidad=True, editable_otros=True)

    st.divider()

    # ── Sección B: LLEGADAS ─────────────────────────────────────
    _render_section(
        estado="LLEGADA",
        titulo="📦 LLEGADAS — en bodega (aún no en Contifico, suman a Stock Disponible)",
        accion_siguiente=("PROCESADA", "✅ Marcar como PROCESADA",
                           "Las filas seleccionadas pasan a histórico y "
                           "dejan de influir en los reportes. Úsalo cuando "
                           "Contifico ya haya registrado el ingreso."),
        help_txt="Mercadería ya física en bodega. Sólo la cantidad es "
                 "editable (por faltantes/robos/errores). Márcala como "
                 "PROCESADA cuando el ingreso aparezca en el consolidado "
                 "de Contifico — a partir de ese punto deja de sumar al stock.",
        editable_cantidad=True, editable_otros=False)

    st.divider()

    # ── Sección C: PROCESADAS (archivo) ─────────────────────────
    _render_section(
        estado="PROCESADA",
        titulo="✅ PROCESADAS — archivo histórico (no influyen en reportes)",
        accion_siguiente=None,
        help_txt="Registros ya reflejados en Contifico. Se mantienen para "
                 "trazabilidad pero no afectan Stock Disponible ni En Tránsito. "
                 "Sólo puedes eliminarlos.",
        editable_cantidad=False, editable_otros=False)

    st.divider()

    # ── Export historial ────────────────────────────────────────
    with st.expander("📥 Exportar historial completo"):
        st.caption("Descarga el registro de todas las importaciones (todos los estados).")
        if not imp_df.empty:
            _stamp = _now_ec().strftime("%Y%m%d_%H%M")
            st.download_button(
                "📥 importaciones.xlsx",
                to_xl(imp_df.sort_values(["Estado","Fecha Registro"],
                                           ascending=[True, False])),
                f"importaciones_{_stamp}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', key="imp_export_all")
        else:
            st.info("Aún no hay registros para exportar.")

with T_IMP:
    _render_tab_imp()

# ══ TAB 10 AUDITORÍA ═══════════════════════════════════════════
@_fragment
def _render_tab_aud():
    eng = st.session_state.engine
    if eng.raw_df is None:
        st.info("Cargue un archivo para ver la auditoría.")
        return

    st.markdown("### 🔍 Auditoría de datos")
    st.caption(
        "Revisa la cobertura del pipeline: cuántos movimientos se clasifican, "
        "cuáles quedan fuera y por qué. Útil para detectar registros con "
        "formato no estándar que no entran en los KPIs ni cuadres."
    )

    df = eng.raw_df.copy()
    # Aplicar las mismas reglas que engine.analyze() para este reporte.
    ref  = df["Referencia"].fillna("").astype(str).str.upper()
    typ  = df["Tipo"].fillna("").astype(str).str.upper()
    desc = df["Descripción"].fillna("").astype(str).str.upper()

    df["_ib"] = (typ == "EGR") & desc.str.contains("BAJA DE INVENTARIO", regex=False, na=False)
    df["_ic"] = (typ == "ING") & (
        ref.str.startswith("NCT") |
        desc.str.contains("BAJA DE FACTURA", regex=False, na=False)
    )
    df["_ip"] = (typ == "ING") & ~df["_ic"] & (
        ref.str.startswith("FAC") |
        ref.str.startswith("NVE") |
        desc.str.contains("FACTURA DE COMPRA", regex=False, na=False) |
        desc.str.contains("COMPRA INVENTARIO", regex=False, na=False) |
        desc.str.contains("COMPRA DE INVENTARIO", regex=False, na=False)
    )
    df["_is"] = (typ == "EGR") & ref.str.startswith("FAC") & ~df["_ib"]
    df["_ir"] = (typ == "EGR") & ref.str.startswith("NCT") & ~df["_ib"]
    df["_it"] = typ == "TRA"
    df["_any"] = df["_ip"] | df["_ic"] | df["_ib"] | df["_is"] | df["_ir"] | df["_it"]

    total = len(df)
    cls   = int(df["_any"].sum())
    noc   = total - cls

    # ── KPIs superiores ──
    k1,k2,k3,k4,k5 = st.columns(5)
    k1.metric("Movimientos totales", f"{total:,}")
    k2.metric("Clasificados", f"{cls:,}", f"{(cls/total*100) if total else 0:.1f}%")
    k3.metric("Sin clasificar", f"{noc:,}",
              f"−{(noc/total*100) if total else 0:.1f}%" if noc else "0%",
              delta_color="inverse")
    _rng = (df["Fecha"].max() - df["Fecha"].min()).days if total else 0
    k4.metric("Rango (días)", f"{_rng:,}")
    k5.metric("SKUs únicos", f"{df['Código Producto'].nunique():,}")

    # ── Resumen por categoría ──
    st.markdown("#### 📊 Desglose por clasificación")
    _cats = [
        ("💰 Compras",           "_ip",  "ING + FAC/NVE o descripción de compra"),
        ("↩ Dev. Cliente",       "_ic",  "ING + NCT o descripción 'BAJA DE FACTURA'"),
        ("🛒 Ventas",             "_is",  "EGR + FAC (y no es baja)"),
        ("📤 Dev. Proveedor",    "_ir",  "EGR + NCT (y no es baja)"),
        ("🔥 Baja de Inventario","_ib",  "EGR + descripción 'BAJA DE INVENTARIO'"),
        ("🔄 Transferencias",    "_it",  "Tipo TRA"),
    ]
    rows = []
    for nom, col, regla in _cats:
        n = int(df[col].sum())
        u = float(df.loc[df[col], "Cantidad"].sum())
        v = float(df.loc[df[col], "Valor Total"].sum())
        rows.append({
            "Categoría": nom,
            "Movimientos": n,
            "Unidades": int(u),
            "Valor total": f"${v:,.2f}",
            "% del total": f"{(n/total*100) if total else 0:.1f}%",
            "Regla": regla,
        })
    rows.append({
        "Categoría": "⚠ Sin clasificar",
        "Movimientos": noc,
        "Unidades": int(df.loc[~df["_any"], "Cantidad"].sum()),
        "Valor total": f"${float(df.loc[~df['_any'], 'Valor Total'].sum()):,.2f}",
        "% del total": f"{(noc/total*100) if total else 0:.1f}%",
        "Regla": "ninguna — no entra en el análisis",
    })
    _cat_df = pd.DataFrame(rows)
    st.dataframe(_cat_df, width='stretch', hide_index=True)

    # ── Filas sin clasificar — detalle ──
    st.markdown("#### ⚠ Filas sin clasificar")
    no_cls = df[~df["_any"]].copy()

    if no_cls.empty:
        st.success("✅ Cobertura 100%. Todos los movimientos entran en alguna categoría.")
    else:
        # Separar ajustes con cantidad 0 (inofensivos) del resto
        _zero = (pd.to_numeric(no_cls["Cantidad"], errors="coerce").fillna(0) == 0)
        no_cls_real = no_cls[~_zero]
        no_cls_zero = no_cls[_zero]

        if not no_cls_real.empty:
            _u = float(no_cls_real["Cantidad"].sum())
            _v = float(no_cls_real["Valor Total"].sum())
            st.error(
                f"❌ Hay **{len(no_cls_real):,}** filas con cantidad > 0 que NO se "
                f"están considerando en el análisis (total: **{int(_u):,} u · ${_v:,.2f}**). "
                f"Estas filas impactan el cuadre de stock. Revísalas:"
            )
            _show_cols = ["Fecha","Código Producto","Nombre Producto","Tipo",
                          "Referencia","Descripción","Cantidad","Valor Total",
                          "Bodega Origen","Bodega Destino"]
            _show_cols = [c for c in _show_cols if c in no_cls_real.columns]
            _show = no_cls_real[_show_cols].copy()
            # Tipos consistentes para el data editor
            for c in ["Fecha","Código Producto","Nombre Producto","Tipo",
                      "Referencia","Descripción","Bodega Origen","Bodega Destino"]:
                if c in _show.columns:
                    _show[c] = _show[c].fillna("").astype(str)
            st.dataframe(_show, width='stretch', hide_index=True, height=320)

            # Patrones de las filas sin clasificar
            st.markdown("##### 🔎 Patrones detectados")
            _pat_rows = []
            _combo = no_cls_real.groupby([
                no_cls_real["Tipo"].fillna("").astype(str).str.upper(),
                no_cls_real["Referencia"].fillna("").astype(str).str.upper().str[:3].rename("Ref[:3]"),
            ]).size().reset_index(name="Cantidad")
            st.dataframe(_combo, width='stretch', hide_index=True)

            st.caption(
                "💡 Si hay un patrón recurrente (ej: todas las filas con un tipo + prefijo "
                "específico), puede valer la pena ampliar las reglas del engine para "
                "incluirlas en el análisis."
            )

        if not no_cls_zero.empty:
            with st.expander(f"ℹ Filas sin clasificar con Cantidad=0 ({len(no_cls_zero)}) — no afectan stock"):
                _zc = no_cls_zero[["Fecha","Código Producto","Tipo","Referencia","Descripción"]].copy()
                for c in _zc.columns:
                    _zc[c] = _zc[c].fillna("").astype(str)
                st.dataframe(_zc, width='stretch', hide_index=True)

    # ── Calidad de datos: campos clave vacíos ──
    st.markdown("#### 📋 Calidad de datos")
    issues = []
    _n_fecha_nan  = int(df["Fecha"].isna().sum())
    _n_sku_nan    = int(df["Código Producto"].fillna("").astype(str).str.strip().eq("").sum())
    _n_tipo_nan   = int(df["Tipo"].fillna("").astype(str).str.strip().eq("").sum())
    _n_cant_0     = int((pd.to_numeric(df["Cantidad"], errors="coerce").fillna(0) == 0).sum())
    _n_cant_neg   = int((pd.to_numeric(df["Cantidad"], errors="coerce").fillna(0) < 0).sum())
    if _n_fecha_nan: issues.append(("⚠", f"{_n_fecha_nan} filas sin Fecha (se excluyen del análisis)"))
    if _n_sku_nan:   issues.append(("⚠", f"{_n_sku_nan} filas sin Código Producto"))
    if _n_tipo_nan:  issues.append(("⚠", f"{_n_tipo_nan} filas sin Tipo"))
    if _n_cant_0:    issues.append(("ℹ", f"{_n_cant_0} filas con Cantidad=0 (ajustes típicamente)"))
    if _n_cant_neg:  issues.append(("⚠", f"{_n_cant_neg} filas con Cantidad negativa (posible error de captura)"))
    if not issues:
        st.success("✅ No hay problemas de calidad detectados en campos clave.")
    else:
        for ico, msg in issues:
            if ico == "⚠": st.warning(f"{ico} {msg}")
            else:          st.info(f"{ico} {msg}")

    # ── Export del reporte ──
    st.markdown("#### 📥 Exportar auditoría")
    ec1, ec2 = st.columns(2)
    with ec1:
        st.download_button("📊 Categorías a Excel", to_xl(_cat_df),
            "auditoria_categorias.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch', key="aud_cat_xl")
    with ec2:
        if not no_cls.empty:
            _no_cls_export = no_cls[["Fecha","Código Producto","Nombre Producto",
                                      "Tipo","Referencia","Descripción","Cantidad",
                                      "Valor Total","Bodega Origen","Bodega Destino"]].copy()
            for c in _no_cls_export.columns:
                if _no_cls_export[c].dtype == object:
                    _no_cls_export[c] = _no_cls_export[c].fillna("").astype(str)
            st.download_button("⚠ Sin clasificar a Excel", to_xl(_no_cls_export),
                "auditoria_sin_clasificar.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch', key="aud_noc_xl")

with T_AUD:
    _render_tab_aud()

# ══ TAB 11 EXPORTAR PORTAL ══════════════════════════════════════
# Ruta canónica del CSV publicado. Streamlit sirve ./static/ en
# http://<host>:<port>/app/static/<file> cuando enableStaticServing=true.
PORTAL_STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  "static")
PORTAL_CSV_PATH = os.path.join(PORTAL_STATIC_DIR, "stock_portal.csv")
PORTAL_HTML_PATH = os.path.join(PORTAL_STATIC_DIR, "stock_ejecutivo.html")
GIST_ID_CACHE = os.path.join(PORTAL_STATIC_DIR, ".autosky_gist_id")
GIST_DESCRIPTION = "autosky-stock-ejecutivo"
GIST_FILENAME = "stock_ejecutivo.html"


def _get_github_token():
    """Lee el token de GitHub desde Streamlit secrets o env var. Prueba
    varias claves comunes para compatibilidad con setups existentes."""
    for _k in ("GITHUB_GIST_TOKEN", "GITHUB_TOKEN", "GH_TOKEN"):
        try:
            _v = st.secrets.get(_k)
            if _v: return str(_v).strip()
        except Exception:
            pass
        _v = os.environ.get(_k)
        if _v: return str(_v).strip()
    return None


def _github_api(method, path, token, body=None):
    """Llamada mínima a la API de GitHub con urllib. Retorna (status, json).
    Raise RuntimeError con mensaje claro y consejo si hay 401/403/404."""
    import urllib.request, urllib.error, json as _j
    _url = f"https://api.github.com{path}"
    _data = _j.dumps(body).encode("utf-8") if body is not None else None
    _req = urllib.request.Request(_url, data=_data, method=method)
    _req.add_header("Authorization", f"Bearer {token}")
    _req.add_header("Accept", "application/vnd.github+json")
    _req.add_header("X-GitHub-Api-Version", "2022-11-28")
    _req.add_header("User-Agent", "AutoSky-Inventario")
    if _data is not None:
        _req.add_header("Content-Type", "application/json")
    try:
        with urllib.request.urlopen(_req, timeout=20) as _rs:
            _body = _rs.read().decode("utf-8")
            return _rs.status, _j.loads(_body) if _body else {}
    except urllib.error.HTTPError as _e:
        _body = _e.read().decode("utf-8", errors="replace")
        try: _err = _j.loads(_body)
        except Exception: _err = {"message": _body[:200]}
        _msg = _err.get('message', _body[:200])
        # Consejos específicos por código de error
        if _e.code == 401:
            _hint = " → el token es inválido o expiró."
        elif _e.code == 403:
            _hint = (" → token válido pero sin permisos suficientes. "
                     "Verifica que tenga el scope **gist** (read & write).")
        elif _e.code == 404 and "/gists" in path:
            _hint = (" → típicamente significa que el token NO tiene scope "
                     "**gist**. GitHub responde 404 en lugar de 403 por "
                     "seguridad. Regenera el token y activa el scope "
                     "`gist` (classic PAT) o **Account permissions → "
                     "Gists: Read and write** (fine-grained PAT).")
        else:
            _hint = ""
        raise RuntimeError(
            f"GitHub API {_e.code} en {method} {path}: {_msg}{_hint}"
        ) from None
    except urllib.error.URLError as _e:
        raise RuntimeError(f"No se pudo contactar GitHub ({_e.reason})") from None


def _publish_html_to_gist(html_bytes, token):
    """Crea o actualiza un gist público con el HTML. Devuelve
    (gist_id, raw_url, user_login). Busca por descripción para reusar el
    mismo gist entre sesiones. Cachea gist_id en disco para publicaciones
    rápidas posteriores."""
    # 1) Intentar reusar gist_id cacheado
    _gid = None
    if os.path.exists(GIST_ID_CACHE):
        try:
            with open(GIST_ID_CACHE, "r", encoding="utf-8") as _f:
                _gid = _f.read().strip() or None
        except Exception:
            _gid = None

    # 2) Verificar que el gist cacheado sigue existiendo
    if _gid:
        try:
            _, _info = _github_api("GET", f"/gists/{_gid}", token)
            _user_login = _info.get("owner", {}).get("login") or ""
        except RuntimeError as _e:
            # Gist borrado o token cambió — invalidar cache
            _gid = None
            _user_login = ""

    # 3) Si no hay gist cacheado, buscar por descripción entre los del user
    if not _gid:
        try:
            _, _my = _github_api("GET", "/gists?per_page=100", token)
        except RuntimeError:
            _my = []
        for _g in (_my or []):
            if _g.get("description") == GIST_DESCRIPTION:
                _gid = _g.get("id")
                _user_login = _g.get("owner", {}).get("login") or ""
                break

    _content_txt = html_bytes.decode("utf-8")

    # 4) Actualizar o crear
    if _gid:
        _, _info = _github_api("PATCH", f"/gists/{_gid}", token, body={
            "description": GIST_DESCRIPTION,
            "files": {GIST_FILENAME: {"content": _content_txt}},
        })
    else:
        _, _info = _github_api("POST", "/gists", token, body={
            "description": GIST_DESCRIPTION,
            "public": True,
            "files": {GIST_FILENAME: {"content": _content_txt}},
        })
        _gid = _info.get("id")

    _user_login = _info.get("owner", {}).get("login") or ""

    # 5) Persistir gist_id para próximas publicaciones
    try:
        os.makedirs(PORTAL_STATIC_DIR, exist_ok=True)
        with open(GIST_ID_CACHE, "w", encoding="utf-8") as _f:
            _f.write(_gid)
    except Exception:
        pass

    # URLs del gist. La URL raw de gist.githubusercontent.com se sirve
    # como text/plain (GitHub no permite hosting HTML desde gists directo),
    # por eso envolvemos con htmlpreview.github.io que fetchea el raw y lo
    # renderiza client-side con el content-type correcto. Devolvemos ambas:
    # preview_url para compartir, raw_url como referencia técnica.
    #
    # Cache-buster: `?v=<timestamp>` en el raw URL dentro del preview para
    # evitar que GitHub raw (varnish) y el browser sirvan contenido viejo
    # tras una re-publicación. GitHub ignora el query string y siempre
    # devuelve el último contenido del gist; pero el query string distinto
    # fuerza cache-miss tanto en el CDN como en el browser. Cada publicación
    # genera un URL nuevo → el usuario que comparte siempre pasa fresh data.
    _cb = int(_ptime.time())
    _raw_url = f"https://gist.githubusercontent.com/{_user_login}/{_gid}/raw/{GIST_FILENAME}"
    _raw_url_cb = f"{_raw_url}?v={_cb}"
    _preview_url = f"https://htmlpreview.github.io/?{_raw_url_cb}"
    return _gid, _preview_url, _raw_url, _user_login

def _build_portal_csv_bytes(df_portal, fecha_corte_txt):
    """CSV con primera línea de metadata (Fecha Corte: DD/MM/YYYY), luego
    header SKU,Stock Disponible y filas. UTF-8 con BOM para Excel."""
    body = df_portal.to_csv(index=False)
    return (f"Fecha Corte: {fecha_corte_txt}\n" + body).encode("utf-8-sig")

def _build_portal_tsv_bytes(df_portal, fecha_corte_txt):
    body = df_portal.to_csv(index=False, sep="\t")
    return (f"Fecha Corte: {fecha_corte_txt}\n" + body).encode("utf-8")

def _portal_public_url(filename):
    """Construye URL pública del archivo publicado leyendo Host del request
    entrante. Funciona tanto en local como en Streamlit Cloud (autodetecta
    dominio) sin valores hardcodeados."""
    _host = None
    _proto = None
    try:
        _hdrs = st.context.headers  # type: ignore[attr-defined]
        _host = (_hdrs.get("Host") or _hdrs.get("host") or
                 _hdrs.get(":authority"))
        _proto = (_hdrs.get("X-Forwarded-Proto") or
                  _hdrs.get("x-forwarded-proto"))
    except Exception:
        pass
    if not _host:
        _port = os.environ.get("STREAMLIT_SERVER_PORT", "8501")
        _host = f"localhost:{_port}"
    if not _proto:
        _proto = ("https" if (".streamlit.app" in _host or
                               "streamlit.io" in _host)
                  else "http")
    return f"{_proto}://{_host}/app/static/{filename}"


@_fragment
def _render_tab_exp():
    r = st.session_state.get("result")
    eng = st.session_state.engine
    if r is None or eng.raw_df is None:
        st.info("Cargue un archivo y ejecute el análisis para exportar.")
        return

    st.markdown("### 📤 Exportar Portal")
    st.caption(
        "Exportaciones bajo demanda con datos de Stock Disponible "
        "(solo Bodega Principal). Los SKUs excluidos en el panel lateral "
        "se omiten automáticamente. El Stock Disponible incluye las "
        "importaciones ya en bodega (estado LLEGADA); las importaciones "
        "aún en tránsito se muestran en la columna **En Tránsito**."
    )

    # ── Base: sku_summary enriquecido con importaciones ──────────
    # - Stock Disponible += cantidades en estado LLEGADA (ya en bodega)
    # - Columna "En Tránsito" = cantidades en estado INGRESADA
    # - Columna "Fechas Tránsito" = fechas tentativas formateadas
    base = r.sku_summary[["Código Producto", "Nombre Producto",
                          "Stock Disponible"]].copy()
    base["Stock Disponible"] = base["Stock Disponible"].fillna(0).astype(int)
    _imp_df_exp = _get_shared_importaciones()["df"]
    base = _apply_importaciones_to_sku_base(
        base, _imp_df_exp,
        excluded_skus=st.session_state.get("excluded_skus", set()))

    # Fecha global "actualizado hasta" (último movimiento del dataset filtrado)
    _global_last = r.filtered["Fecha"].max()
    _gl_txt = (pd.to_datetime(_global_last).strftime("%d/%m/%Y")
               if pd.notna(_global_last) else "—")

    _en_transito_total = int(base["En Tránsito"].sum()) if "En Tránsito" in base.columns else 0
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("SKUs exportables", f"{len(base):,}")
    k2.metric("Con stock > 0",
              f"{int((base['Stock Disponible'] > 0).sum()):,}")
    k3.metric("En tránsito (total u)", f"{_en_transito_total:,}")
    k4.metric("Actualizado hasta", _gl_txt)

    # ── Sección A: datos para portal web ─────────────────────────────
    st.markdown("#### 🌐 Datos para portal web")
    st.caption(
        "Columnas: SKU · Stock Disponible · En Tránsito · Fechas Tránsito. "
        "La primera línea del CSV/TXT contiene la fecha de corte (último "
        "movimiento de toda la data). 'En Tránsito' suma las cantidades "
        "de importaciones en estado INGRESADA."
    )
    df_portal = base[["Código Producto", "Stock Disponible",
                        "En Tránsito", "Fechas Tránsito"]].rename(
        columns={"Código Producto": "SKU"})

    st.dataframe(df_portal, width='stretch', hide_index=True, height=320)

    _stamp = _now_ec().strftime("%Y%m%d_%H%M")
    _csv = _build_portal_csv_bytes(df_portal, _gl_txt)
    _tsv = _build_portal_tsv_bytes(df_portal, _gl_txt)
    _html = to_html(df_portal, f"Stock Disponible — Portal (corte {_gl_txt})")

    e1, e2, e3 = st.columns(3)
    with e1:
        st.download_button("📄 CSV", _csv,
            f"stock_portal_{_stamp}.csv", "text/csv",
            key="exp_portal_csv", width='stretch')
    with e2:
        st.download_button("📝 TXT", _tsv,
            f"stock_portal_{_stamp}.txt", "text/plain",
            key="exp_portal_txt", width='stretch')
    with e3:
        st.download_button("🌐 HTML", _html,
            f"stock_portal_{_stamp}.html", "text/html",
            key="exp_portal_htm", width='stretch')

    # ── Publicación al portal (URL pública para consumo HTTP GET) ────
    st.markdown("##### 📡 Publicar al portal")
    st.caption(
        "Escribe el CSV a una URL pública del servidor Streamlit para que "
        "un servicio externo lo consuma por HTTP GET. El archivo se "
        "sobrescribe en cada publicación."
    )

    try:
        os.makedirs(PORTAL_STATIC_DIR, exist_ok=True)
    except Exception as _e:
        st.error(f"No se pudo crear {PORTAL_STATIC_DIR}: {_e}")

    _exists = os.path.exists(PORTAL_CSV_PATH)
    _pub_info = ""
    if _exists:
        try:
            _mt = os.path.getmtime(PORTAL_CSV_PATH)
            _pub_info = datetime.fromtimestamp(_mt, _EC_TZ).replace(
                tzinfo=None).strftime("%d/%m/%Y %H:%M GMT-5")
        except Exception:
            _pub_info = "—"

    pc1, pc2 = st.columns([1, 2])
    with pc1:
        if st.button("📡 Publicar CSV al portal",
                     key="exp_portal_publish", type="primary",
                     width='stretch'):
            try:
                with open(PORTAL_CSV_PATH, "wb") as _f:
                    _f.write(_csv)
                log(f"Publicado stock_portal.csv ({len(df_portal)} filas, "
                    f"corte {_gl_txt})")
                st.success("✓ Publicado — URL ya disponible abajo")
                st.rerun()
            except Exception as _e:
                st.error(f"Error al publicar: {_e}")
    with pc2:
        if _exists:
            st.markdown(f"**Última publicación:** {_pub_info}")
        else:
            st.info("Aún no se ha publicado ninguna versión del CSV.")

    # URL pública — autodetectada desde Host header (funciona en local y
    # en Streamlit Cloud sin hardcodear dominio).
    if _exists:
        _public_url = _portal_public_url("stock_portal.csv")
        st.markdown("**URL pública (para el servicio externo):**")
        st.code(_public_url, language=None)
        st.caption(
            "Esta es la URL que el portal destino debe consumir vía HTTP GET. "
            "El archivo se actualiza sólo cuando pulsas **Publicar**."
        )

    st.divider()

    # ── Sección B: PDF ejecutivo vertical A4 (una sola hoja) ────────
    st.markdown("#### 📄 PDF ejecutivo")
    st.caption(
        "A4 vertical: SKU · Nombre Producto · Stock Disponible · En Tránsito · "
        "Fechas Tránsito. El encabezado muestra la fecha de corte (último "
        "registro) para indicar hasta qué punto están actualizados los datos."
    )
    df_ejec = base[["Código Producto", "Nombre Producto",
                    "Stock Disponible", "En Tránsito",
                    "Fechas Tránsito"]].rename(
        columns={"Código Producto": "SKU"})

    st.dataframe(df_ejec, width='stretch', hide_index=True, height=320)

    _pdf = to_pdf_ejecutivo(
        df_ejec,
        title="Stock Ejecutivo",
        subtitle=f"Actualizado hasta {_gl_txt}  ·  {len(df_ejec)} SKUs"
    )
    if _pdf:
        st.download_button("📄 PDF ejecutivo (A4 vertical)", _pdf,
            f"stock_ejecutivo_{_stamp}.pdf", "application/pdf",
            key="exp_ejec_pdf", width='stretch', type="primary")
    else:
        st.button("📄 PDF ejecutivo (reportlab no instalado)",
                 disabled=True, key="exp_ejec_pdf_x", width='stretch')

    st.divider()

    # ── Sección C: Vista móvil ejecutiva publicada (GitHub Gist) ────
    st.markdown("#### 📱 Vista móvil ejecutiva")
    st.caption(
        "Publica una página HTML responsive (optimizada para celular) con "
        "el stock actual de todos los SKUs a un Gist público de GitHub. "
        "La URL es permanente; cada re-publicación actualiza el mismo Gist. "
        "Compártela por WhatsApp con un clic."
    )

    # Construir HTML móvil en memoria (para descarga offline + publicación)
    _html_mobile = to_html_mobile(df_ejec, _gl_txt)

    # Estado publicación — leer gist_id cacheado si existe
    _cached_gid = None
    _cached_owner = None
    if os.path.exists(GIST_ID_CACHE):
        try:
            with open(GIST_ID_CACHE, "r", encoding="utf-8") as _f:
                _cached_gid = _f.read().strip() or None
        except Exception:
            _cached_gid = None
    # Guardamos owner en session_state para no hacer otra llamada a la API
    # solo para mostrar el URL cacheado.
    if _cached_gid:
        _cached_owner = st.session_state.get("_gist_owner_cache")

    _token = _get_github_token()
    _token_ok = bool(_token)

    if not _token_ok:
        st.warning(
            "⚠ No hay token de GitHub configurado. Agrégalo en "
            "**Settings → Secrets** de Streamlit Cloud con una de estas "
            "claves: `GITHUB_GIST_TOKEN`, `GITHUB_TOKEN` o `GH_TOKEN`. "
            "El token debe tener el scope **gist** (read & write)."
        )

    mc1, mc2 = st.columns([1, 2])
    with mc1:
        _btn = st.button("📡 Publicar vista móvil",
                          key="exp_mobile_publish", type="primary",
                          width='stretch', disabled=not _token_ok)
        if _btn and _token_ok:
            with st.spinner("Publicando al Gist…"):
                try:
                    _gid, _prev_url, _raw_url, _owner = \
                        _publish_html_to_gist(_html_mobile, _token)
                    st.session_state["_gist_owner_cache"] = _owner
                    st.session_state["_gist_last_url"]    = _prev_url
                    st.session_state["_gist_last_raw"]    = _raw_url
                    st.session_state["_gist_last_time"]   = \
                        _now_ec().strftime("%d/%m/%Y %H:%M GMT-5")
                    log(f"Publicado Gist {_gid} ({len(df_ejec)} SKUs, "
                        f"corte {_gl_txt})")
                    st.success("✓ Publicado al Gist — URL abajo")
                    st.rerun()
                except RuntimeError as _e:
                    st.error(f"Error al publicar: {_e}")
                except Exception as _e:
                    st.error(f"Error inesperado: {_e}")
    with mc2:
        _last_time = st.session_state.get("_gist_last_time")
        if _last_time:
            st.markdown(f"**Última publicación:** {_last_time}")
        elif _cached_gid:
            st.info(f"Gist existente detectado: `{_cached_gid}` — "
                    "re-publica para refrescar el contenido.")
        else:
            st.info("Aún no se ha publicado la vista móvil.")

    # Descarga directa (backup offline) — no requiere token
    st.download_button("⬇ Descargar HTML (offline)", _html_mobile,
        f"stock_ejecutivo_movil_{_stamp}.html", "text/html",
        key="exp_mobile_dl", width='stretch')

    # Mostrar URL del Gist si tenemos una publicación reciente o cacheada
    _show_url = st.session_state.get("_gist_last_url")
    _show_raw = st.session_state.get("_gist_last_raw")
    if not _show_url and _cached_gid and _cached_owner:
        _raw = (f"https://gist.githubusercontent.com/{_cached_owner}"
                f"/{_cached_gid}/raw/{GIST_FILENAME}")
        # Cache-buster para evitar que varnish/browser sirvan contenido viejo
        _cb = int(_ptime.time())
        _show_url = f"https://htmlpreview.github.io/?{_raw}?v={_cb}"
        _show_raw = _raw

    if _show_url:
        st.markdown("**URL pública (compartible) — se ve como página web:**")
        st.code(_show_url, language=None)
        st.caption(
            "Esta URL envuelve el gist con htmlpreview.github.io para que "
            "el navegador lo renderice como HTML (no como texto). Perfecta "
            "para compartir por WhatsApp / correo / chat."
        )

        import urllib.parse as _up
        _wa_url = f"https://wa.me/?text={_up.quote(_show_url)}"
        st.link_button("📲 Compartir por WhatsApp", _wa_url,
                        width='stretch', type="primary")
        st.caption(
            "Abre WhatsApp (móvil) o WhatsApp Web (desktop) con la URL "
            "lista para enviar — sólo selecciona el contacto."
        )

        if _show_raw:
            with st.expander("🔧 Ver URL raw (gist.githubusercontent.com)"):
                st.code(_show_raw, language=None)
                st.caption(
                    "Esta URL devuelve el HTML como texto plano — útil si "
                    "el consumidor es un programa que lo va a procesar, "
                    "no abrir en un navegador."
                )

with T_EXP:
    _render_tab_exp()

# ── Panel visual de performance en el sidebar ───────────────────
_perf("end")
with st.sidebar:
    # Logo corporativo — aparece en Excel/PDF ejecutivos
    with st.expander("🏢 Logo corporativo (reportes)", expanded=False):
        _logo_existente = _get_logo_path()
        if _logo_existente:
            try:
                st.image(_logo_existente, caption="Logo actual (en reportes)",
                         width='stretch')
            except Exception:
                pass
        else:
            st.caption("Aún no hay logo. Sube un PNG/JPG para incluirlo en "
                       "los reportes Excel y PDF.")
        _up_logo = st.file_uploader("Subir / reemplazar logo",
                                     type=["png","jpg","jpeg"],
                                     key="logo_upload",
                                     label_visibility="collapsed")
        _cl1, _cl2 = st.columns(2)
        with _cl1:
            if _up_logo is not None:
                if st.button("💾 Guardar", key="logo_save", width='stretch',
                             type="primary"):
                    _persist_logo(_up_logo.getvalue())
                    log(f"Logo actualizado: {_up_logo.name}")
                    st.success("✓ Logo guardado")
                    st.rerun()
        with _cl2:
            if os.path.exists(LOGO_PATH):
                if st.button("🗑 Quitar", key="logo_remove", width='stretch',
                             help="Elimina el logo subido. Si existe default, vuelve a él."):
                    try:
                        os.remove(LOGO_PATH)
                        log("Logo eliminado")
                        st.rerun()
                    except Exception as ex:
                        st.error(f"Error: {ex}")
        st.caption("💡 Usa un PNG con fondo transparente para mejores resultados. "
                   "Tamaño recomendado: 500×300 px o similar proporción.")

    with st.expander("⏱ Performance (últimos reruns)", expanded=False):
        _hist = _get_perf_history()["runs"]
        if _hist:
            _rows = []
            for _r in _hist[-15:]:
                _d = {"hora": _r["ts"], "total_ms": _r["total_ms"]}
                for _lbl, _ms in _r["ck"]:
                    _d[_lbl] = _ms
                _rows.append(_d)
            _pdf = pd.DataFrame(_rows).iloc[::-1]
            st.dataframe(_pdf, width='stretch', hide_index=True, height=300)
            st.caption(f"Archivo: `perf.log` ({_PERF_LOG_PATH})")
            if st.button("🗑 Limpiar historial", key="perf_clear"):
                _get_perf_history()["runs"] = []
                try: os.remove(_PERF_LOG_PATH)
                except: pass
                st.rerun()
        else:
            st.caption("Sin datos todavía. Interactúa con la app para ver mediciones.")

_perf_flush()
