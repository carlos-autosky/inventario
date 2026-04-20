from __future__ import annotations
import customtkinter as ctk
from tkinter import ttk, filedialog, messagebox
import tkinter as tk
import pandas as pd
import calendar as cal_module
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

from .config import APP_NAME, VERSION, BUILD_TIMESTAMP, THEME, PRIMARY_WAREHOUSE
from .engine import InventoryEngine
from .storage import load_config, save_config
from .utils import fmt_num, fmt_pct, now_str
from .toma_fisica_module import TomaFisicaWindow, DEFAULT_LOCATIONS

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

FONT_BODY      = ("Segoe UI", 11)
FONT_BOLD      = ("Segoe UI", 11, "bold")
FONT_SMALL     = ("Segoe UI", 10)
FONT_KPI_TITLE = ("Segoe UI", 9, "bold")
FONT_KPI_VALUE = ("Segoe UI", 14, "bold")
FONT_HEADER    = ("Segoe UI", 20, "bold")
FONT_TREE      = ("Segoe UI", 11)
FONT_TREE_HEAD = ("Segoe UI", 10, "bold")

MPL_STYLE = {
    "figure.facecolor": "#111827", "axes.facecolor": "#1A2436",
    "axes.edgecolor": "#243247",   "axes.labelcolor": "#9CA3AF",
    "text.color": "#F3F4F6",       "xtick.color": "#9CA3AF",
    "ytick.color": "#9CA3AF",      "grid.color": "#243247",
    "grid.linestyle": "--",        "grid.alpha": 0.5,
}

def _col_w(col: str) -> int:
    """Ancho de columna calculado al número de caracteres del label visible.
    Fórmula: len(label) * 7.5px + 16px padding. Nombre Producto siempre libre."""
    # Lookup exacto por nombre de columna interno o label visible
    _EXACT = {
        # ── Identificación ────────────────────────────────────────
        "Codigo Producto":          76,   # Cód.Prod        8c
        "Código Producto":          83,   # Cód. Prod       9c
        "Nombre Producto":          240,  # libre
        "Categoría Producto":       91,   # Categ Prod     10c
        "Bodega":                   155,
        "Bodega Origen":            155,
        "Bodega Destino":           155,
        "Fecha":                    82,
        "Coincide":                 55,
        # ── Valores financieros ───────────────────────────────────
        "Valor_Compras":            83,   # Compras $       9c
        "Valor_Ventas":             76,   # Ventas $        8c
        "Valor Compras (ING)":      158,  # Compras $ (Ingreso)  19c
        "Valor Ventas (EGR)":       136,  # Ventas $(Egreso)     16c
        "Valor Inventario ($)":     121,  # Inventario ($)       14c
        "Valor Inventario":         91,   # Inventario           10c
        "Valor Unitario Promedio":  158,  # Unitario $ Promedio  19c
        "Valor Stock":              68,   # Stock $              7c
        # ── Movimiento de unidades ────────────────────────────────
        "Compras":                  136,  # Compras(Ingreso)     16c
        "Dev_Proveedor":            76,   # Dev.Prov              8c
        "Ventas":                   121,  # Ventas(Egreso)       14c
        "Dev_Cliente":              98,   # N/C Cliente          11c
        "Muestras_Enviadas":        143,  # Muestras Enviadas    17c
        "Muestras_Devueltas":       151,  # Muestras Devueltas   18c
        "Stock Disponible":         136,  # Stock Disponible     16c
        "Stock Muestras":           151,  # Stock en  Muestras   18c
        "Stock Total":              98,   # Stock Total          11c
        "Stock en Cliente":         136,  # Stock en Cliente     16c
        # ── Otros ────────────────────────────────────────────────
        "Descripción":              240,
        "Referencia":               110,
        "Serie":                    82,
        "PVP":                      68,
    }
    if col in _EXACT:
        return _EXACT[col]
    # Fallback: calcular desde el label (puede venir con \n — tomar línea más larga)
    c = col.lower().replace("\n", " ")
    max_line = max(col.split("\n"), key=len) if "\n" in col else col
    if "nombre" in c or "descripci" in c: return 240
    if "bodega"  in c: return 155
    if "fecha"   in c: return 82
    if "coincide" in c: return 55
    # Calcular por longitud del label
    return max(60, int(len(max_line) * 7.5) + 16)

def _is_num(col: str) -> bool:
    c = col.lower()
    return any(k in c for k in ("cantidad","stock","valor","precio","pvp","compra","venta",
               "dev","muestra","entregada","devuelta","diferencia","dias","consumo",
               "margen","exactitud","total","neto","rotac","suger","rentab","unidades","costo"))


# ── Calendario propio ─────────────────────────────────────────────────────────
class DatePicker(ctk.CTkToplevel):
    """Calendario popup que devuelve fecha en formato dd/mm/yyyy."""
    def __init__(self, master, callback, initial_date: str = ""):
        super().__init__(master)
        self.callback = callback
        self.title("Seleccionar fecha")
        self.resizable(False, False)
        self.configure(fg_color=THEME["bg"])
        self.grab_set()
        import datetime
        try:
            if "/" in initial_date:
                d,m,y = initial_date.split("/")
                self.year, self.month = int(y), int(m)
            elif "-" in initial_date:
                y,m,d = initial_date.split("-")
                self.year, self.month = int(y), int(m)
            else:
                raise ValueError
        except:
            now = datetime.date.today()
            self.year, self.month = now.year, now.month
        self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        nav = ctk.CTkFrame(self, fg_color=THEME["panel_alt"])
        nav.pack(fill="x", padx=4, pady=4)
        ctk.CTkButton(nav, text="<", width=32, height=28, font=FONT_BOLD,
                      command=self._prev).pack(side="left", padx=4)
        ctk.CTkLabel(nav, text=f"{cal_module.month_name[self.month]} {self.year}",
                     font=FONT_BOLD, text_color=THEME["text"]).pack(side="left", expand=True)
        ctk.CTkButton(nav, text=">", width=32, height=28, font=FONT_BOLD,
                      command=self._next).pack(side="right", padx=4)

        grid = ctk.CTkFrame(self, fg_color=THEME["panel"])
        grid.pack(padx=4, pady=(0,4))
        days = ["Lu","Ma","Mi","Ju","Vi","Sa","Do"]
        for c,d in enumerate(days):
            ctk.CTkLabel(grid, text=d, font=FONT_KPI_TITLE,
                         text_color=THEME["muted"], width=36).grid(row=0, column=c, padx=1, pady=2)
        weeks = cal_module.monthcalendar(self.year, self.month)
        for r, week in enumerate(weeks, start=1):
            for c, day in enumerate(week):
                if day == 0:
                    ctk.CTkFrame(grid, width=36, height=30, fg_color="transparent").grid(row=r, column=c)
                else:
                    ctk.CTkButton(grid, text=str(day), width=36, height=30,
                                  font=FONT_SMALL, corner_radius=4,
                                  fg_color=THEME["card"], hover_color=THEME["accent"],
                                  command=lambda d=day: self._pick(d)).grid(row=r, column=c, padx=1, pady=1)

    def _prev(self):
        self.month -= 1
        if self.month < 1: self.month = 12; self.year -= 1
        self._build()
    def _next(self):
        self.month += 1
        if self.month > 12: self.month = 1; self.year += 1
        self._build()
    def _pick(self, day: int):
        self.callback(f"{day:02d}/{self.month:02d}/{self.year}")
        self.destroy()


# ── KPI Box ───────────────────────────────────────────────────────────────────
class KPIBox(ctk.CTkFrame):
    def __init__(self, master, title: str, color: str = "#F3F4F6"):
        super().__init__(master, fg_color=THEME["card"], corner_radius=8,
                         border_width=1, border_color=THEME["border"])
        self.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(self, text=title.upper(), font=FONT_KPI_TITLE,
                     text_color=THEME["muted"], anchor="w").grid(
            row=0, column=0, sticky="w", padx=8, pady=(5,0))
        self._val = ctk.CTkLabel(self, text="--", font=FONT_KPI_VALUE,
                                 text_color=color, anchor="w")
        self._val.grid(row=1, column=0, sticky="w", padx=8, pady=(0,5))
    def set_value(self, v: str): self._val.configure(text=v)


# ── App ───────────────────────────────────────────────────────────────────────
class InventoryApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title(f"{APP_NAME} {VERSION}")
        self.geometry("1700x1000")
        self.minsize(1400, 860)
        self.configure(fg_color=THEME["bg"])
        self.engine          = InventoryEngine()
        self.current_result  = None
        self.config_state    = load_config()
        self.wh_vars: dict[str, ctk.BooleanVar] = {}
        self.wh_checks: list[ctk.CTkCheckBox]   = []
        self.sku_excl_vars: dict[str, ctk.BooleanVar] = {}
        self._pending_recalc = False
        self._after_id: str | None = None
        self._sort_state: dict[int, tuple] = {}
        self._configure_tree_style()
        self._build_ui()
        self._apply_saved_config()
        self._rotation_df: pd.DataFrame | None = None
        self._buy_vars: dict[str, ctk.BooleanVar] = {}
        self.protocol("WM_DELETE_WINDOW", self._on_closing)
        # ── Carga automática para pruebas (solo desarrollo) ──────────────
        import os as _os
        _AUTO = r"C:\Users\carlo\Downloads\inventario_movimientos_consolidado.xlsx"
        if _os.path.exists(_AUTO):
            self.after(300, lambda: self._auto_load(_AUTO))

    def _configure_tree_style(self):
        s = ttk.Style()
        try: s.theme_use("default")
        except: pass
        s.configure("Dark.Treeview", background=THEME["panel"], foreground=THEME["text"],
            fieldbackground=THEME["panel"], bordercolor=THEME["border"],
            rowheight=24, font=FONT_TREE)
        s.map("Dark.Treeview",
            background=[("selected","#1D4ED8")], foreground=[("selected","#FFF")])
        # padding=(horizontal, vertical) — vertical amplio para que se vean 2 líneas completas
        s.configure("Dark.Treeview.Heading", background=THEME["panel_alt"],
            foreground=THEME["text"], relief="flat", font=FONT_TREE_HEAD,
            padding=(6, 16))
        # Estado de zoom global (font size + rowheight)
        self._zoom_size = 11   # tamaño de fuente base

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self._build_header()
        self._build_toolbar()
        self._build_filters()
        self._build_kpis()
        self._build_body()

    # ── Header ────────────────────────────────────────────────────────────────
    def _build_header(self):
        h = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, height=44)
        h.grid(row=0, column=0, sticky="ew"); h.grid_propagate(False)
        h.grid_columnconfigure(1, weight=1)
        ctk.CTkLabel(h, text=APP_NAME, font=FONT_HEADER, text_color=THEME["text"]).grid(
            row=0, column=0, sticky="w", padx=16, pady=10)
        ctk.CTkLabel(h, text=f"{VERSION}  |  Build {BUILD_TIMESTAMP}",
                     font=FONT_SMALL, text_color=THEME["muted"]).grid(
            row=0, column=1, sticky="e", padx=16)

    # ── Toolbar ───────────────────────────────────────────────────────────────
    def _build_toolbar(self):
        bar = ctk.CTkFrame(self, fg_color="transparent")
        bar.grid(row=1, column=0, sticky="ew", padx=14, pady=(0,0))
        # Columnas 0-4: botones (peso igual); columna 5: label BASE (expansible)
        for i in range(6): bar.grid_columnconfigure(i, weight=1)
        bar.grid_columnconfigure(6, weight=3)
        btn = {"font": FONT_BOLD, "height": 32, "corner_radius": 7}
        # Orden: Cargar Excel | Toma Fisica | Plantilla Toma | Guardar Config | Exportar Excel | BASE
        ctk.CTkButton(bar, text="Cargar Excel",    command=self.load_inventory, **btn).grid(row=0,column=0,sticky="ew",padx=3)
        ctk.CTkButton(bar, text="Toma fisica",     command=self.open_toma_fisica, **btn).grid(row=0,column=1,sticky="ew",padx=3)
        ctk.CTkButton(bar, text="Plantilla Toma",  command=self.export_toma_template,
                      fg_color=THEME["violet"], hover_color="#6D28D9", **btn).grid(row=0,column=2,sticky="ew",padx=3)
        ctk.CTkButton(bar, text="Importar Toma",   command=self.import_toma_template,
                      fg_color="#7C3AED", hover_color="#6D28D9", **btn).grid(row=0,column=3,sticky="ew",padx=3)
        ctk.CTkButton(bar, text="Guardar config",  command=self.persist_config, **btn).grid(row=0,column=4,sticky="ew",padx=3)
        ctk.CTkButton(bar, text="Exportar Excel",  command=self.export_excel,
                      fg_color=THEME["success"], hover_color="#059669", **btn).grid(row=0,column=5,sticky="ew",padx=3)
        self.base_label = ctk.CTkLabel(bar, text="Base: no cargada",
                                       text_color=THEME["muted"], anchor="w", font=FONT_SMALL)
        self.base_label.grid(row=0, column=6, sticky="ew", padx=(12,0))

    # ── Filters ───────────────────────────────────────────────────────────────
    def _build_filters(self):
        f = ctk.CTkFrame(self, fg_color=THEME["panel_alt"], corner_radius=0, height=52)
        f.grid(row=2, column=0, sticky="ew", padx=0, pady=(0,2))
        f.grid_propagate(False)

        # Barra compacta y alineada completamente a la izquierda
        left = ctk.CTkFrame(f, fg_color="transparent")
        left.pack(side="left", anchor="w", fill="x", padx=8, pady=6)

        def sep(parent):
            ctk.CTkFrame(parent, fg_color=THEME["border"], width=1, height=28).pack(side="left", padx=10, pady=0)

        # Grupo: Fecha corte
        fecha = ctk.CTkFrame(left, fg_color="transparent")
        fecha.pack(side="left", anchor="w")
        ctk.CTkLabel(fecha, text="CORTE:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.cutoff_entry = ctk.CTkEntry(fecha, placeholder_text="dd/mm/yyyy",
                                         font=FONT_BODY, height=30, width=120)
        self.cutoff_entry.pack(side="left", padx=(0,4))
        self.after(200, lambda: self._bind_date_format(self.cutoff_entry))
        ctk.CTkButton(fecha, text="📅", width=30, height=30, font=("Segoe UI",10),
                      command=lambda: DatePicker(self, self._set_cutoff,
                      self.cutoff_entry.get())).pack(side="left", padx=(0,4))
        ctk.CTkButton(fecha, text="Aplicar", width=84, height=30, font=FONT_BOLD,
                      command=self.run_analysis, fg_color=THEME["accent"],
                      hover_color="#2563EB", corner_radius=6).pack(side="left")

        sep(left)

        # Grupo: Bodegas
        bod = ctk.CTkFrame(left, fg_color="transparent")
        bod.pack(side="left", anchor="w")
        ctk.CTkLabel(bod, text="BODEGAS:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.warehouse_mode = ctk.CTkOptionMenu(
            bod, values=["Todas","Solo principal","Seleccion manual"],
            font=FONT_BODY, height=30, width=190, command=self._on_wh_mode_change)
        self.warehouse_mode.pack(side="left")
        self.warehouse_mode.set("Todas")

        sep(left)

        # Grupo: Bodegas reporte
        br = ctk.CTkFrame(left, fg_color="transparent")
        br.pack(side="left", anchor="w")
        ctk.CTkLabel(br, text="BODEGAS RPT:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.bodegas_label = ctk.CTkLabel(br, text="Todas",
                                          text_color=THEME["muted"], font=FONT_SMALL, anchor="w")
        self.bodegas_label.pack(side="left", padx=(0,6))
        ctk.CTkButton(br, text="Selec.", width=74, height=30, font=FONT_SMALL,
                      corner_radius=6, command=self.open_warehouse_selector).pack(side="left")

        sep(left)

        # Grupo: SKU excluidos
        sku = ctk.CTkFrame(left, fg_color="transparent")
        sku.pack(side="left", anchor="w")
        ctk.CTkLabel(sku, text="SKU EXCL:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.sku_label = ctk.CTkLabel(sku, text="Ninguno",
                                      text_color=THEME["muted"], font=FONT_SMALL, anchor="w")
        self.sku_label.pack(side="left", padx=(0,6))
        ctk.CTkButton(sku, text="Selec.", width=74, height=30, font=FONT_SMALL,
                      corner_radius=6, command=self.open_sku_selector).pack(side="left")

    # ── KPIs ──────────────────────────────────────────────────────────────────
    def _build_kpis(self):
        p = ctk.CTkFrame(self, fg_color="transparent")
        p.grid(row=3, column=0, sticky="ew", padx=14, pady=(1,2))
        for i in range(6): p.grid_columnconfigure(i, weight=1)
        r1 = ["Stock total","Stock disponible","Stock en muestras",
              "Valor inventario","Compras acumuladas"]
        r2 = ["Rotacion","Dias de inventario","Consumo promedio","Exactitud inventario"]
        self.kpi_boxes: dict[str,KPIBox] = {}
        colors = {"Compras acumuladas":THEME["danger"],"Valor inventario":THEME["accent"]}
        for i in range(5): p.grid_columnconfigure(i, weight=1)
        for c,name in enumerate(r1):
            b = KPIBox(p, name, color=colors.get(name,THEME["text"]))
            b.grid(row=0,column=c,sticky="ew",padx=3,pady=(0,2))
            self.kpi_boxes[name] = b
        for c,name in enumerate(r2):
            b = KPIBox(p, name)
            b.grid(row=1,column=c,sticky="ew",padx=3,pady=(0,2))
            self.kpi_boxes[name] = b

    # ── Body ──────────────────────────────────────────────────────────────────
    def _build_body(self):
        body = ctk.CTkTabview(self, fg_color=THEME["panel"],
                              segmented_button_selected_color=THEME["accent"])
        body.grid(row=4, column=0, sticky="nsew", padx=14, pady=(0,12))
        self.tab_inventory = body.add("Inventario por Bodega")
        self.tab_sku       = body.add("Detalle por SKU")
        self.tab_pivot     = body.add("SKU x Bodega")
        self.tab_samples   = body.add("Muestras por Cliente")
        self.tab_analysis  = body.add("Analisis de Periodo")
        self.tab_rotation  = body.add("Rotacion y Compras")
        self.tab_purchases = body.add("Historico de Compras")
        self.tab_kardex    = body.add("Kardex")
        self.tab_physical  = body.add("Toma Fisica")
        self.tab_log       = body.add("Log")

        self.tree_inventory = self._make_tree(self.tab_inventory)
        self._build_sku_tab()
        self._build_pivot_tab()
        self._build_samples_tab()
        self._build_analysis_tab()
        self._build_rotation_tab()
        self._build_purchases_tab()
        self._build_kardex_tab()
        self._build_physical_tab()

        self.logbox = ctk.CTkTextbox(self.tab_log, fg_color=THEME["panel"],
                                     text_color=THEME["text"], font=("Consolas",11))
        self.logbox.pack(fill="both", expand=True, padx=6, pady=6)
        self.log("Sistema listo v001.7")

    def _build_pivot_tab(self):
        """Pestaña: tabla pivote SKU (filas) × Bodega (columnas) — stock neto."""
        outer = ctk.CTkFrame(self.tab_pivot, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(1, weight=1)

        # Barra de controles — solo Generar y Exportar
        ctrl = ctk.CTkFrame(outer, fg_color=THEME["panel_alt"], corner_radius=8)
        ctrl.grid(row=0, column=0, sticky="ew", pady=(0,6))
        ctrl.grid_columnconfigure(2, weight=1)

        ctk.CTkLabel(ctrl, text="Stock neto por bodega (excluye Bodega Principal)",
                     font=FONT_KPI_TITLE, text_color=THEME["muted"]).grid(
            row=0, column=0, sticky="w", padx=12, pady=6)
        ctk.CTkButton(ctrl, text="Generar reporte", height=28, font=FONT_BOLD,
                      command=self.run_pivot_report, corner_radius=6,
                      fg_color=THEME["accent"]).grid(row=0,column=1,padx=8,pady=6)
        ctk.CTkButton(ctrl, text="Exportar Excel", height=28, font=FONT_SMALL,
                      command=self.export_pivot_excel, corner_radius=6,
                      fg_color=THEME["success"], hover_color="#059669").grid(
            row=0, column=2, sticky="w", padx=4, pady=6)

        wrap = ctk.CTkFrame(outer, fg_color="transparent")
        wrap.grid(row=1, column=0, sticky="nsew")
        wrap.grid_rowconfigure(0,weight=1); wrap.grid_columnconfigure(0,weight=1)
        self.tree_pivot = self._make_tree_in(wrap, row=0)
        self._pivot_df: pd.DataFrame | None = None

    def run_pivot_report(self):
        """Genera la tabla pivote SKU × Bodega.

        Reglas:
        - Fuente: inventory_by_warehouse (misma que KPIs) para modo stock neto.
        - Por defecto excluye la Bodega Principal — muestra solo bodegas externas.
        - Aplica el filtro BODEGAS RPT si el usuario seleccionó bodegas específicas.
        - Para modos de movimiento usa r.filtered con filtros ya aplicados.
        """
        r = self.current_result
        if r is None:
            messagebox.showinfo("Pivot","Primero ejecute el analisis principal."); return

        mode = "Cantidad neta por bodega"

        # Bodegas seleccionadas en BODEGAS RPT (vacío = todas las no-principales)
        rpt_selection = self.get_selected_warehouses()  # [] si "Todas incluidas"
        all_selected  = (len(rpt_selection) == len(self.wh_vars)) or not rpt_selection

        if mode == "Cantidad neta por bodega":
            inv = r.inventory_by_warehouse.copy()
            if inv.empty:
                messagebox.showinfo("Sin datos","No hay stock calculado."); return

            # ── Aplicar filtro de bodegas ────────────────────────────────
            if all_selected:
                # Por defecto: excluir Bodega Principal, mostrar solo externas
                inv = inv[inv["Bodega"] != PRIMARY_WAREHOUSE]
            else:
                # El usuario eligió bodegas específicas en BODEGAS RPT
                inv = inv[inv["Bodega"].isin(rpt_selection)]

            if inv.empty:
                messagebox.showinfo("Sin datos","No hay stock en las bodegas seleccionadas."); return

            pivot = inv.pivot_table(
                index=["Código Producto","Nombre Producto"],
                columns="Bodega",
                values="Stock",
                aggfunc="sum",
                fill_value=0
            ).reset_index()
            pivot.columns.name = None
            pivot = pivot.rename(columns={
                "Código Producto": "SKU",
                "Nombre Producto": "Nombre",
            })

        else:
            # Modos de movimiento — usar r.filtered (ya filtrado por engine)
            df = r.filtered.copy()
            ref = df["Referencia"].fillna("").astype(str).str.upper()
            typ = df["Tipo"].fillna("").astype(str).str.upper()

            if mode == "Solo ventas (EGR FAC)":
                df2 = df[(typ=="EGR") & ref.str.startswith("FAC")].copy()
                bodega_col, sign = "Bodega Origen", -1
            elif mode == "Solo compras (ING FAC)":
                df2 = df[(typ=="ING") & ref.str.startswith("FAC")].copy()
                bodega_col, sign = "Bodega Destino", 1
            else:
                df2 = df[typ=="TRA"].copy()
                bodega_col, sign = "Bodega Destino", 1

            df2["__SKU"] = df2["Código Producto"].fillna("").astype(str)
            df2["__BOD"] = df2[bodega_col].fillna("").astype(str).str.strip()
            df2 = df2[df2["__BOD"] != ""]

            # Aplicar filtro BODEGAS RPT
            if not all_selected:
                df2 = df2[df2["__BOD"].isin(rpt_selection)]
            else:
                # Por defecto excluir bodega principal
                df2 = df2[df2["__BOD"] != PRIMARY_WAREHOUSE]

            df2["__QTY"] = df2["Cantidad"] * sign

            if df2.empty:
                messagebox.showinfo("Sin datos","No hay movimientos para ese filtro."); return

            pivot = df2.pivot_table(
                index=["__SKU","Nombre Producto"],
                columns="__BOD", values="__QTY",
                aggfunc="sum", fill_value=0
            ).reset_index()
            pivot.columns.name = None
            pivot = pivot.rename(columns={"__SKU":"SKU","Nombre Producto":"Nombre"})

        # ── Totales ──────────────────────────────────────────────────────────
        num_cols = [c for c in pivot.columns if c not in ("SKU","Nombre")]
        if not num_cols:
            messagebox.showinfo("Sin datos","No hay columnas de bodega para mostrar."); return
        pivot["TOTAL"] = pivot[num_cols].sum(axis=1)
        # Excluir filas con todo en cero
        pivot = pivot[pivot["TOTAL"] != 0].copy()
        total_row = {"SKU":"TOTAL GENERAL","Nombre":""}
        for c in num_cols + ["TOTAL"]:
            total_row[c] = pivot[c].sum()
        pivot = pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

        self._pivot_df = pivot
        self._fill_pivot(pivot)
        bodegas_txt = ", ".join(num_cols) if len(num_cols) <= 3 else f"{len(num_cols)} bodegas"
        self.log(f"Pivot ({mode}): {len(pivot)-1} SKU | {bodegas_txt}")

    def _fill_pivot(self, df: pd.DataFrame):
        """Renderiza la tabla pivote con formato ejecutivo."""
        tree = self.tree_pivot
        tree.delete(*tree.get_children())
        if df.empty: return

        cols = list(df.columns)
        tree["columns"] = cols

        # Encabezados — quitar palabra "Bodega" de nombres de columna-bodega
        import re as _re
        for c in cols:
            is_num = (c not in ("SKU","Nombre"))
            anchor = "e" if is_num else "w"
            # Limpiar "Bodega" del label (case-insensitive) y recortar espacios
            label = _re.sub(r"(?i)\bbodega\b\s*", "", c).strip(" -_")
            if not label:
                label = c   # fallback si el nombre era solo "Bodega"
            # Partir en 2 líneas si es largo
            if len(label) > 12 and " " in label:
                mid = len(label)//2
                sp  = label.rfind(" ", 0, mid) or label.find(" ", mid)
                if sp > 0: label = label[:sp] + "\n" + label[sp+1:]
            tree.heading(c, text=label, anchor=anchor)
            w = 100 if c == "SKU" else (180 if c == "Nombre" else max(55, min(120, len(c)*7+10)))
            tree.column(c, anchor=anchor, width=w, minwidth=44, stretch=False)

        num_cols = [c for c in cols if c not in ("SKU","Nombre")]

        for i, (_, row) in enumerate(df.iterrows()):
            is_total = str(row.get("SKU","")) == "TOTAL GENERAL"
            vals = []
            for c in cols:
                v = row[c]
                if c in ("SKU","Nombre"):
                    vals.append(str(v) if str(v) not in ("nan","None") else "")
                else:
                    try:
                        fv = float(v)
                        vals.append(f"{int(fv):,}" if fv == int(fv) else f"{fv:,.1f}")
                    except:
                        vals.append(str(v))
            tag = "total" if is_total else ("even" if i%2==0 else "odd")
            tree.insert("","end",values=vals,tags=(tag,))

    def export_pivot_excel(self):
        """Exporta la tabla pivote SKU x Bodega a Excel con formato."""
        if self._pivot_df is None:
            messagebox.showinfo("Pivot","Primero genere el reporte."); return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile="reporte_sku_bodega.xlsx")
        if not path: return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            df = self._pivot_df.copy()
            # Asegurar que columnas numéricas sean float limpio
            id_cols = [c for c in ("SKU","Nombre") if c in df.columns]
            num_cols = [c for c in df.columns if c not in id_cols]
            for c in num_cols:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "SKU x Bodega"

            BLUE   = PatternFill("solid", fgColor="1E3A5F")
            YELLOW = PatternFill("solid", fgColor="FFD700")
            EVEN   = PatternFill("solid", fgColor="1A2436")
            ODD    = PatternFill("solid", fgColor="111827")
            TOTAL  = PatternFill("solid", fgColor="172A45")
            thin   = Side(style="thin", color="243247")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            nc = len(df.columns)

            # Fila 1 — título
            ws.merge_cells(f"A1:{get_column_letter(nc)}1")
            ws["A1"] = "REPORTE SKU × BODEGA — STOCK NETO POR BODEGA"
            ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
            ws["A1"].fill      = BLUE
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 26

            # Fila 2 — encabezados de columna
            for ci, col in enumerate(df.columns, 1):
                cell = ws.cell(2, ci, col)
                cell.font      = Font(bold=True, size=9, color="000000")
                cell.fill      = YELLOW
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = border
            ws.row_dimensions[2].height = 32

            # Filas de datos
            for ri, (_, row) in enumerate(df.iterrows(), 3):
                is_total = str(row.get("SKU","")) == "TOTAL GENERAL"
                fill = TOTAL if is_total else (EVEN if ri % 2 == 0 else ODD)
                for ci, col in enumerate(df.columns, 1):
                    cell        = ws.cell(ri, ci)
                    cell.border = border
                    cell.fill   = fill
                    if col in id_cols:
                        sv = str(row[col])
                        cell.value     = "" if sv in ("nan","None","NaN") else sv
                        cell.font      = Font(bold=is_total, size=9,
                                             color="FACC15" if is_total
                                             else ("93C5FD" if col == "SKU" else "E5E7EB"))
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        val = float(row[col])
                        cell.value     = int(val) if val == int(val) else round(val, 2)
                        neg = isinstance(cell.value, (int,float)) and cell.value < 0
                        cell.font      = Font(bold=is_total, size=9,
                                             color="FACC15" if is_total
                                             else ("FCA5A5" if neg else "E5E7EB"))
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                ws.row_dimensions[ri].height = 16

            # Anchos de columna
            ws.column_dimensions["A"].width = 14   # SKU
            ws.column_dimensions["B"].width = 40   # Nombre
            for ci in range(3, nc + 1):
                col_name = str(df.columns[ci - 1])
                ws.column_dimensions[get_column_letter(ci)].width = max(12, min(22, len(col_name) + 4))

            # Congelar desde C3
            ws.freeze_panes = "C3"

            wb.save(path)
            messagebox.showinfo("Exportado",
                                f"Reporte exportado correctamente:\n{path}\n"
                                f"{len(df)-1} SKU × {len(num_cols)} bodega(s)")

        except Exception as e:
            self.log(f"Error al exportar pivot: {e}")
            messagebox.showerror("Error al exportar", str(e))

    @staticmethod
    def _safe_sheet(name: str) -> str:
        """Elimina caracteres invalidos para nombres de hoja Excel."""
        result = str(name)
        for ch in ['/', '\\', '?', '*', '[', ']', ':']:
            result = result.replace(ch, '-')
        return result[:31].strip()


    def export_toma_template(self):
        """Genera plantilla Excel para toma fisica.

        RESUMEN GENERAL usa fórmulas que leen las hojas individuales —
        el usuario solo llena la columna CANTIDAD en cada hoja de ubicación.
        El resumen se actualiza automáticamente sin doble trabajo.
        """
        if self.engine.raw_df is None:
            messagebox.showinfo("Plantilla", "Primero cargue el archivo de inventario.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="plantilla_toma_fisica.xlsx")
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter, quote_sheetname

            excluded = self._get_excluded_skus()
            cols_needed = ["Código Producto", "Nombre Producto"]
            has_cat = "Categoría Producto" in self.engine.raw_df.columns
            if has_cat:
                cols_needed.append("Categoría Producto")

            skus_df = (self.engine.raw_df[cols_needed]
                       .drop_duplicates()
                       .sort_values("Nombre Producto")
                       .reset_index(drop=True))
            if excluded:
                skus_df = skus_df[~skus_df["Código Producto"].isin(excluded)]

            n_skus = len(skus_df)
            DATA_START = 4   # fila donde empiezan los datos (1-indexed)

            # ── Paleta pastel ─────────────────────────────────────────────
            HDR    = PatternFill("solid", fgColor="1E3A5F")
            YELLOW = PatternFill("solid", fgColor="FEF9C3")
            GREEN  = PatternFill("solid", fgColor="D1FAE5")
            QTY    = PatternFill("solid", fgColor="DBEAFE")
            FORMULA= PatternFill("solid", fgColor="F0FDF4")   # resumen auto: verde muy pálido
            EVEN   = PatternFill("solid", fgColor="F8FAFC")
            ODD    = PatternFill("solid", fgColor="FFFFFF")
            INSTR  = PatternFill("solid", fgColor="F1F5F9")
            thin   = Side(style="thin", color="CBD5E1")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            thin_g = Side(style="thin", color="BBF7D0")
            brd_g  = Border(left=thin_g, right=thin_g, top=thin_g, bottom=thin_g)

            wb = openpyxl.Workbook()

            # ── Hoja INSTRUCCIONES ────────────────────────────────────────
            wi = wb.active
            wi.title = "INSTRUCCIONES"
            instruc = [
                "PLANTILLA DE TOMA FISICA DE INVENTARIO",
                "",
                "INSTRUCCIONES (solo llene las hojas de ubicación):",
                "1. Vaya a la hoja de cada UBICACIÓN (ej. Bodega Principal).",
                "2. Ingrese la CANTIDAD contada en la columna D (fondo azul).",
                "3. Si un artículo no está en esa ubicación, deje la celda en blanco o en 0.",
                "4. Use la columna OBSERVACIÓN para anotar estado, serie, lote, etc.",
                "5. Llene el campo FECHA TOMA en celda B2 de cada hoja (dd/mm/yyyy).",
                "6. El RESUMEN GENERAL se actualiza AUTOMÁTICAMENTE — no lo edite.",
                "7. Guarde el archivo e impórtelo con el botón 'Importar Toma'.",
                "",
                "UBICACIONES INCLUIDAS:",
            ]
            for ri, txt in enumerate(instruc, 1):
                cell = wi.cell(ri, 1, txt)
                is_title = ri == 1
                is_hdr   = ri == 3
                is_note  = ri == 9   # instrucción importante
                color = "FFFFFF" if is_title else ("DC2626" if is_note else "1E3A5F")
                cell.font = Font(bold=(is_title or is_hdr or is_note),
                                 size=13 if is_title else 10, color=color)
                cell.fill = HDR if is_title else INSTR
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                wi.row_dimensions[ri].height = 22 if is_title else 16
            for j, loc in enumerate(DEFAULT_LOCATIONS, 1):
                ri = len(instruc) + j
                cell = wi.cell(ri, 1, f"  {j}. {loc}")
                cell.font = Font(size=10, color="065F46")
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
                wi.row_dimensions[ri].height = 16
            wi.column_dimensions["A"].width = 75

            # ── Crear primero las hojas de ubicación (se necesitan para las fórmulas) ──
            loc_sheets = {}   # loc → (sheet_name, ws_object)
            for loc in DEFAULT_LOCATIONS:
                sname = self._safe_sheet(loc)
                ws2   = wb.create_sheet(sname)
                loc_sheets[loc] = (sname, ws2)

                ws2.merge_cells("A1:E1")
                ws2["A1"] = f"TOMA FISICA — {loc.upper()}"
                ws2["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
                ws2["A1"].fill      = HDR
                ws2["A1"].alignment = Alignment(horizontal="center")
                ws2.row_dimensions[1].height = 24

                ws2["A2"] = "FECHA TOMA:"
                ws2["A2"].font = Font(bold=True, size=10, color="FFFFFF")
                ws2["A2"].fill = HDR
                ws2["B2"].fill = PatternFill("solid", fgColor="EFF6FF")
                ws2["B2"].font = Font(size=10, color="1E3A5F")
                ws2.row_dimensions[2].height = 20

                hdrs = ["Código", "Nombre", "Categoría", "Cantidad", "Observación"]
                for ci, h in enumerate(hdrs, 1):
                    cell = ws2.cell(3, ci, h)
                    cell.font      = Font(bold=True, size=9, color="1E3A5F")
                    cell.fill      = YELLOW
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                    cell.border    = border
                ws2.row_dimensions[3].height = 22

                for ri2, (_, row) in enumerate(skus_df.iterrows(), DATA_START):
                    fill = EVEN if ri2 % 2 == 0 else ODD
                    ws2.cell(ri2, 1, str(row["Código Producto"])).fill = fill
                    ws2.cell(ri2, 1).font = Font(size=9, color="1E40AF")
                    ws2.cell(ri2, 2, str(row["Nombre Producto"])).fill = fill
                    ws2.cell(ri2, 2).font = Font(size=9, color="111827")
                    cat = str(row.get("Categoría Producto","")) if has_cat else ""
                    ws2.cell(ri2, 3, cat).fill = fill
                    ws2.cell(ri2, 3).font = Font(size=9, color="6B7280")
                    # Columna D: cantidad editable — fondo azul pastel
                    qc = ws2.cell(ri2, 4, "")
                    qc.fill      = QTY
                    qc.alignment = Alignment(horizontal="right")
                    qc.font      = Font(size=10, bold=True, color="1E3A5F")
                    # Columna E: observación
                    oc = ws2.cell(ri2, 5, "")
                    oc.fill = fill
                    oc.font = Font(size=9, color="6B7280")
                    for ci in range(1, 6):
                        ws2.cell(ri2, ci).border = border
                    ws2.row_dimensions[ri2].height = 16

                tr2 = n_skus + DATA_START
                ws2.merge_cells(f"A{tr2}:C{tr2}")
                ws2.cell(tr2, 1, "TOTAL").font = Font(bold=True, size=9, color="FFFFFF")
                ws2.cell(tr2, 1).fill = HDR
                tc2 = ws2.cell(tr2, 4, f"=SUM(D{DATA_START}:D{tr2-1})")
                tc2.font      = Font(bold=True, size=9, color="065F46")
                tc2.fill      = GREEN
                tc2.alignment = Alignment(horizontal="right")
                tc2.border    = border

                ws2.column_dimensions["A"].width = 13
                ws2.column_dimensions["B"].width = 42
                ws2.column_dimensions["C"].width = 16
                ws2.column_dimensions["D"].width = 11
                ws2.column_dimensions["E"].width = 32
                ws2.freeze_panes = "D4"

            # ── Hoja RESUMEN GENERAL con fórmulas automáticas ─────────────
            # Columnas: Código | Nombre | Categoría | [loc1] | [loc2] | ... | TOTAL
            all_cols = ["Código", "Nombre", "Categoría"] + DEFAULT_LOCATIONS + ["TOTAL"]
            wr  = wb.create_sheet("RESUMEN GENERAL", 1)   # segunda hoja (después de INSTRUCCIONES)
            nc  = len(all_cols)

            wr.merge_cells(f"A1:{get_column_letter(nc)}1")
            wr["A1"] = "RESUMEN TOMA FISICA — Se actualiza automáticamente desde cada hoja"
            wr["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
            wr["A1"].fill      = HDR
            wr["A1"].alignment = Alignment(horizontal="center")
            wr.row_dimensions[1].height = 24

            # Fila 2: nota explicativa
            wr.merge_cells(f"A2:{get_column_letter(nc)}2")
            wr["A2"] = ("⚠ NO editar este resumen — los valores se calculan automáticamente "
                        "desde las hojas de cada ubicación. Solo llene las hojas individuales.")
            wr["A2"].font      = Font(bold=True, size=9, color="92400E")
            wr["A2"].fill      = PatternFill("solid", fgColor="FEF3C7")
            wr["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            wr.row_dimensions[2].height = 28

            # Fila 3: cabeceras
            for ci, col in enumerate(all_cols, 1):
                cell            = wr.cell(3, ci, col)
                cell.font       = Font(bold=True, size=9, color="1E3A5F")
                cell.fill       = YELLOW
                cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border     = border
            wr.row_dimensions[3].height = 30

            # Filas de datos: fórmulas que leen D-column de cada hoja de ubicación
            # La fila ri en el RESUMEN corresponde a la fila ri en cada hoja de ubicación
            # (misma estructura, mismo orden de SKUs)
            loc_col_start = 4   # columna D = primera ubicación

            for ri, (_, row) in enumerate(skus_df.iterrows(), DATA_START):
                fill = EVEN if ri % 2 == 0 else ODD

                # Código (columna A)
                wr.cell(ri, 1, str(row["Código Producto"])).fill = fill
                wr.cell(ri, 1).font = Font(size=9, color="1E40AF")

                # Nombre (columna B)
                wr.cell(ri, 2, str(row["Nombre Producto"])).fill = fill
                wr.cell(ri, 2).font = Font(size=9, color="111827")

                # Categoría (columna C)
                cat = str(row.get("Categoría Producto","")) if has_cat else ""
                wr.cell(ri, 3, cat).fill = fill
                wr.cell(ri, 3).font = Font(size=9, color="6B7280")

                # Una columna por ubicación — fórmula: ='NombreHoja'!D{ri}
                # Si la celda de la hoja de ubicación está vacía, la fórmula devuelve 0
                for loc_i, loc in enumerate(DEFAULT_LOCATIONS):
                    ci       = loc_col_start + loc_i
                    sname, _ = loc_sheets[loc]
                    # quote_sheetname agrega comillas simples si el nombre tiene espacios
                    formula  = f"=IFERROR({quote_sheetname(sname)}!D{ri},0)"
                    cell     = wr.cell(ri, ci, formula)
                    cell.fill       = FORMULA
                    cell.font       = Font(size=9, color="065F46")
                    cell.alignment  = Alignment(horizontal="right")
                    cell.border     = brd_g

                # Columna TOTAL = SUM de todas las ubicaciones de esta fila
                col_first = get_column_letter(loc_col_start)
                col_last  = get_column_letter(loc_col_start + len(DEFAULT_LOCATIONS) - 1)
                tc = wr.cell(ri, nc, f"=SUM({col_first}{ri}:{col_last}{ri})")
                tc.fill      = GREEN
                tc.font      = Font(bold=True, size=9, color="065F46")
                tc.alignment = Alignment(horizontal="right")
                tc.border    = border

                # Bordes en ID columns
                for ci in range(1, 4):
                    wr.cell(ri, ci).border = border
                wr.row_dimensions[ri].height = 16

            # Fila de TOTAL GENERAL
            tr = n_skus + DATA_START
            wr.merge_cells(f"A{tr}:C{tr}")
            wr.cell(tr, 1, "TOTAL GENERAL").font = Font(bold=True, size=9, color="FFFFFF")
            wr.cell(tr, 1).fill = HDR
            for ci in range(loc_col_start, nc + 1):
                cl   = get_column_letter(ci)
                cell = wr.cell(tr, ci, f"=SUM({cl}{DATA_START}:{cl}{tr-1})")
                cell.font      = Font(bold=True, size=9, color="FFFFFF")
                cell.fill      = HDR
                cell.alignment = Alignment(horizontal="right")
                cell.border    = border

            # Anchos resumen
            wr.column_dimensions["A"].width = 13
            wr.column_dimensions["B"].width = 38
            wr.column_dimensions["C"].width = 18
            for ci in range(loc_col_start, nc + 1):
                wr.column_dimensions[get_column_letter(ci)].width = 13
            wr.freeze_panes = "D4"

            wb.save(path)
            import os as _os, subprocess as _sub
            try:
                if _os.name == "nt": _os.startfile(path)
                else: _sub.Popen(["xdg-open", path])
            except Exception: pass

            n_locs = len(DEFAULT_LOCATIONS)
            self.log(f"Plantilla exportada: {path} | {n_skus} SKU | {n_locs} ubicaciones")
            messagebox.showinfo("Plantilla generada",
                f"Plantilla lista:\n{path}\n\n"
                f"SKUs incluidos: {n_skus}\n"
                f"Ubicaciones: {n_locs} hojas\n\n"
                "FLUJO DE USO:\n"
                "1. Abra la hoja de cada ubicación\n"
                "2. Llene solo la columna CANTIDAD (fondo azul)\n"
                "3. El RESUMEN GENERAL se actualiza automáticamente\n"
                "4. Guarde e importe con 'Importar Toma'")

        except Exception as e:
            import traceback
            err = traceback.format_exc()
            self.log(f"Error en plantilla: {e}")
            messagebox.showerror("Error al generar plantilla", str(e))


    def _build_sku_tab(self):
        outer = ctk.CTkFrame(self.tab_sku, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=4, pady=4)
        outer.grid_columnconfigure(0, weight=3); outer.grid_columnconfigure(2, weight=4)
        outer.grid_rowconfigure(0, weight=1)
        lv = ctk.CTkFrame(outer, fg_color=THEME["card"], corner_radius=8,
                          border_width=1, border_color=THEME["border"])
        lv.grid(row=0,column=0,sticky="nsew",padx=(0,3))
        lv.grid_rowconfigure(1,weight=1); lv.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(lv, text="VALORES FINANCIEROS", font=FONT_KPI_TITLE,
                     text_color=THEME["accent"]).grid(row=0,column=0,sticky="w",padx=8,pady=(6,2))
        self.tree_sku_val = self._make_tree_in(lv, row=1)
        ctk.CTkFrame(outer, fg_color=THEME["border"], width=2).grid(row=0,column=1,sticky="ns")
        lu = ctk.CTkFrame(outer, fg_color=THEME["card"], corner_radius=8,
                          border_width=1, border_color=THEME["border"])
        lu.grid(row=0,column=2,sticky="nsew",padx=(3,0))
        lu.grid_rowconfigure(1,weight=1); lu.grid_columnconfigure(0,weight=1)
        ctk.CTkLabel(lu, text="MOVIMIENTO DE UNIDADES", font=FONT_KPI_TITLE,
                     text_color=THEME["teal"]).grid(row=0,column=0,sticky="w",padx=8,pady=(6,2))
        self.tree_sku_unit = self._make_tree_in(lu, row=1)

    def _build_samples_tab(self):
        """Muestras: activos a la izquierda, todos a la derecha + reporte ejecutivo."""
        wrap = ctk.CTkFrame(self.tab_samples, fg_color="transparent")
        wrap.pack(fill="both", expand=True, padx=4, pady=4)
        wrap.grid_columnconfigure(0, weight=1)
        wrap.grid_rowconfigure(1, weight=1)

        # ── Barra de reporte ──────────────────────────────────────────────
        bar = ctk.CTkFrame(wrap, fg_color=THEME["panel_alt"], corner_radius=8)
        bar.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(bar, text="REPORTE CLIENTE:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).grid(row=0, column=0, padx=(10,4), pady=6)
        self.sample_client_var = ctk.StringVar(value="— seleccione cliente —")
        self.sample_client_menu = ctk.CTkOptionMenu(
            bar, variable=self.sample_client_var,
            values=["— seleccione cliente —"],
            font=FONT_BODY, height=28, width=280)
        self.sample_client_menu.grid(row=0, column=1, sticky="w", padx=4, pady=6)
        ctk.CTkButton(bar, text="📄 PDF", width=90, height=28, font=FONT_BOLD,
                      fg_color=THEME["danger"], hover_color="#B91C1C",
                      corner_radius=6,
                      command=self.export_sample_report_pdf).grid(row=0, column=2, padx=4, pady=6)
        ctk.CTkButton(bar, text="🌐 HTML", width=90, height=28, font=FONT_BOLD,
                      fg_color=THEME["accent"], hover_color="#2563EB",
                      corner_radius=6,
                      command=self.export_sample_report_html).grid(row=0, column=3, padx=(4,10), pady=6)

        # ── Tablas ────────────────────────────────────────────────────────
        outer = ctk.CTkFrame(wrap, fg_color="transparent")
        outer.grid(row=1, column=0, sticky="nsew")
        outer.grid_rowconfigure(0, weight=1)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_columnconfigure(2, weight=1)

        left = ctk.CTkFrame(outer, fg_color=THEME["card"], corner_radius=8,
                            border_width=1, border_color=THEME["border"])
        left.grid(row=0, column=0, sticky="nsew", pady=0)
        left.grid_rowconfigure(1, weight=1)
        left.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(left, text="CLIENTES CON SALDO ACTIVO  (stock > 0)",
                     font=FONT_KPI_TITLE, text_color=THEME["success"]).grid(
            row=0, column=0, sticky="w", padx=8, pady=(6, 2))
        self.tree_samples_active = self._make_tree_in(left, row=1)

        ctk.CTkFrame(outer, fg_color=THEME["border"], width=2).grid(
            row=0, column=1, sticky="ns", padx=4)

        right = ctk.CTkFrame(outer, fg_color=THEME["card"], corner_radius=8,
                             border_width=1, border_color=THEME["border"])
        right.grid(row=0, column=2, sticky="nsew", pady=0)
        right.grid_rowconfigure(1, weight=1)
        right.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(right, text="TODOS LOS CLIENTES CON MUESTRAS",
                     font=FONT_KPI_TITLE, text_color=THEME["muted"]).grid(
            row=0, column=0, sticky="w", padx=8, pady=(6, 2))
        self.tree_samples_all = self._make_tree_in(right, row=1)
    def _build_analysis_tab(self):
        outer = ctk.CTkFrame(self.tab_analysis, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(2, weight=1)

        # Selector de periodo — alineado a la izquierda
        ctrl = ctk.CTkFrame(outer, fg_color=THEME["panel_alt"], corner_radius=8)
        ctrl.grid(row=0, column=0, sticky="ew", pady=(0,4))
        # Columna expansible al final para empujar controles a la izquierda
        ctrl.grid_columnconfigure(7, weight=1)

        inner = ctk.CTkFrame(ctrl, fg_color="transparent")
        inner.pack(side="left", anchor="w", padx=8, pady=6)

        ctk.CTkLabel(inner, text="DESDE", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.anal_from = ctk.CTkEntry(inner, placeholder_text="dd/mm/yyyy",
                                      width=120, height=28, font=FONT_BODY)
        self.anal_from.pack(side="left", padx=(0,2))
        self.after(200, lambda: self._bind_date_format(self.anal_from))
        ctk.CTkButton(inner, text="📅", width=28, height=28,
                      command=lambda: DatePicker(self, lambda d: self._set_entry(self.anal_from, d),
                                                 self.anal_from.get())).pack(side="left", padx=(0,14))

        ctk.CTkLabel(inner, text="HASTA", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.anal_to = ctk.CTkEntry(inner, placeholder_text="dd/mm/yyyy",
                                    width=120, height=28, font=FONT_BODY)
        self.anal_to.pack(side="left", padx=(0,2))
        self.after(200, lambda: self._bind_date_format(self.anal_to))
        ctk.CTkButton(inner, text="📅", width=28, height=28,
                      command=lambda: DatePicker(self, lambda d: self._set_entry(self.anal_to, d),
                                                 self.anal_to.get())).pack(side="left", padx=(0,14))

        ctk.CTkButton(inner, text="Calcular Periodo", height=28, font=FONT_BOLD,
                      command=self.run_period_analysis, corner_radius=7,
                      fg_color=THEME["accent"], hover_color="#2563EB").pack(side="left")

        # KPIs periodo
        kpi_p = ctk.CTkFrame(outer, fg_color="transparent")
        kpi_p.grid(row=1, column=0, sticky="ew", pady=(0,4))
        for i in range(5): kpi_p.grid_columnconfigure(i, weight=1)
        names_p = ["Item mas rotacion","Item mas vendido","Item mas rentable"]
        colors_p = [THEME["accent"],THEME["success"],THEME["warning"]]
        self.period_kpis: dict[str,KPIBox] = {}
        for c,(name,col) in enumerate(zip(names_p,colors_p)):
            b = KPIBox(kpi_p, name, color=col)
            b.grid(row=0,column=c,sticky="ew",padx=3)
            self.period_kpis[name] = b

        # Sub-tabs análisis
        tabs2 = ctk.CTkTabview(outer, fg_color=THEME["panel"])
        tabs2.grid(row=2, column=0, sticky="nsew")
        self.ptab_top10v  = tabs2.add("Top 10 Vendidos")
        self.ptab_top10r  = tabs2.add("Top 10 Rentabilidad")

        self.fig_top10v, self.ax_top10v, self.canvas_top10v, self.tree_top10v = \
            self._make_analysis_panel(self.ptab_top10v)
        self.fig_top10r, self.ax_top10r, self.canvas_top10r, self.tree_top10r = \
            self._make_analysis_panel(self.ptab_top10r)
        # Alias para compatibilidad con run_period_analysis
        self.ptab_monthly = self.ptab_top10v
        self.fig_monthly = self.fig_top10v; self.ax_monthly = self.ax_top10v
        self.canvas_monthly = self.canvas_top10v; self.tree_monthly = self.tree_top10v
        self.ptab_margin = self.ptab_top10r
        self.fig_margin = self.fig_top10r; self.ax_margin = self.ax_top10r
        self.canvas_margin = self.canvas_top10r; self.tree_margin = self.tree_top10r

    def _make_analysis_panel(self, parent):
        fr = ctk.CTkFrame(parent, fg_color="transparent")
        fr.pack(fill="both", expand=True, padx=4, pady=4)
        fr.grid_columnconfigure(0, weight=1)
        fr.grid_columnconfigure(1, weight=2)
        fr.grid_rowconfigure(0, weight=1)

        # Contenedor propio para el tree — evita que su scrollbar
        # se superponga al canvas del gráfico
        tree_wrap = ctk.CTkFrame(fr, fg_color="transparent")
        tree_wrap.grid(row=0, column=0, sticky="nsew")
        tree_wrap.grid_rowconfigure(0, weight=1)
        tree_wrap.grid_columnconfigure(0, weight=1)
        tree = self._make_tree_in(tree_wrap, row=0)

        with plt.rc_context(MPL_STYLE):
            fig, ax = plt.subplots(figsize=(6,4))
            fig.tight_layout(pad=1.5)
        canvas = FigureCanvasTkAgg(fig, master=fr)
        canvas.get_tk_widget().grid(row=0, column=1, sticky="nsew", padx=(6,0))
        return fig, ax, canvas, tree

    def _build_rotation_tab(self):
        outer = ctk.CTkFrame(self.tab_rotation, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(3, weight=1)

        # ── Panel de contexto metodológico ────────────────────────────────
        ctx = ctk.CTkFrame(outer, fg_color=THEME["card"], corner_radius=8,
                           border_width=1, border_color=THEME["border"])
        ctx.grid(row=0, column=0, sticky="ew", pady=(0,3))
        ctx.grid_columnconfigure((0,1,2,3), weight=1)
        formulas = [
            ("Consumo/dia",
             "Ventas (u) / Dias del periodo"
             "  |  Promedio diario de unidades vendidas"
             "  |  Fuente: EGR+FAC del periodo de corte",
             THEME["muted"]),
            ("Rotacion (x)",
             "Ventas (u) / Stock Disponible"
             "  |  Veces que el inventario se renueva en el periodo"
             "  |  Alta=activo  Baja=sobrestock",
             THEME["accent"]),
            ("Dias de Inventario",
             "Stock Disponible / Consumo/dia"
             "  |  Dias de cobertura con el stock actual"
             "  |  CRITICO<LT  BAJO<LT+Seg  OK>=LT+Seg",
             THEME["teal"]),
            ("Sugerido de Compra",
             "max(0, Consumo/dia x (Lead Time + Stock Seg.) - Stock Disp.)"
             "  |  Unidades para cubrir LT + colchon de seguridad"
             "  |  P.Reorden = Consumo/dia x Lead Time",
             THEME["warning"]),
        ]
        for c, (name, formula, color) in enumerate(formulas):
            ff = ctk.CTkFrame(ctx, fg_color="transparent")
            ff.grid(row=0, column=c, sticky="nsew", padx=8, pady=6)
            ctk.CTkLabel(ff, text=name, font=FONT_KPI_TITLE, text_color=color).pack(anchor="w")
            ctk.CTkLabel(ff, text=formula, font=("Segoe UI",8),
                         text_color=THEME["muted"], wraplength=240, justify="left").pack(anchor="w")

        # ── Controles ─────────────────────────────────────────────────────
        ctrl = ctk.CTkFrame(outer, fg_color=THEME["panel_alt"], corner_radius=8)
        ctrl.grid(row=1, column=0, sticky="ew", pady=(0,3))
        for i in range(11): ctrl.grid_columnconfigure(i, weight=1 if i==10 else 0)

        # Escenario Marítimo
        ctk.CTkLabel(ctrl, text="🚢 MARÍTIMO:", font=FONT_KPI_TITLE,
                     text_color=THEME["teal"]).grid(row=0,column=0,sticky="w",padx=(10,2),pady=6)
        ctk.CTkLabel(ctrl, text="Lead(d):", font=FONT_SMALL,
                     text_color=THEME["muted"]).grid(row=0,column=1,padx=(0,2),pady=6)
        self.lt_mar_days = ctk.CTkEntry(ctrl, width=48, height=28, font=FONT_BODY)
        self.lt_mar_days.insert(0,"45")
        self.lt_mar_days.grid(row=0,column=2,padx=(0,8),pady=6)

        # Escenario Aéreo
        ctk.CTkLabel(ctrl, text="✈ AÉREO:", font=FONT_KPI_TITLE,
                     text_color=THEME["accent"]).grid(row=0,column=3,sticky="w",padx=(8,2),pady=6)
        ctk.CTkLabel(ctrl, text="Lead(d):", font=FONT_SMALL,
                     text_color=THEME["muted"]).grid(row=0,column=4,padx=(0,2),pady=6)
        self.lt_air_days = ctk.CTkEntry(ctrl, width=48, height=28, font=FONT_BODY)
        self.lt_air_days.insert(0,"15")
        self.lt_air_days.grid(row=0,column=5,padx=(0,12),pady=6)

        # Stock de seguridad
        ctk.CTkLabel(ctrl, text="Stock Seg.(d):", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).grid(row=0,column=6,padx=(4,2),pady=6)
        self.safety_days = ctk.CTkEntry(ctrl, width=48, height=28, font=FONT_BODY)
        self.safety_days.insert(0,"15")
        self.safety_days.grid(row=0,column=7,padx=(0,8),pady=6)

        ctk.CTkButton(ctrl, text="Calcular", height=28, font=FONT_BOLD,
                      command=self.run_rotation_analysis, corner_radius=7,
                      fg_color=THEME["success"], hover_color="#059669").grid(
            row=0,column=8,padx=4,pady=6)
        ctk.CTkButton(ctrl, text="✓ Sel. CRITICOS", height=28,
                      font=FONT_SMALL, fg_color=THEME["danger"], hover_color="#B91C1C",
                      command=self._select_all_critical).grid(row=0,column=9,padx=4,pady=6)
        ctk.CTkButton(ctrl, text="📄 PDF Solicitud", height=28,
                      font=FONT_BOLD, fg_color=THEME["violet"], hover_color="#6D28D9",
                      command=self._generate_purchase_pdf).grid(
            row=0,column=10,sticky="e",padx=8,pady=6)

        # ── Dos paneles: Marítimo | Aéreo ─────────────────────────────────
        panels = ctk.CTkFrame(outer, fg_color="transparent")
        panels.grid(row=3, column=0, sticky="nsew")
        panels.grid_rowconfigure(0, weight=1)   # fila 0 expandible
        panels.grid_columnconfigure(0, weight=1)
        panels.grid_columnconfigure(1, weight=1)

        # Panel Marítimo
        lm = ctk.CTkFrame(panels, fg_color=THEME["card"], corner_radius=8,
                          border_width=1, border_color=THEME["teal"])
        lm.grid(row=0, column=0, sticky="nsew", padx=(0,3))
        lm.grid_rowconfigure(1, weight=1); lm.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(lm, text="🚢  ESCENARIO MARÍTIMO", font=FONT_KPI_TITLE,
                     text_color=THEME["teal"]).grid(row=0,column=0,sticky="w",padx=8,pady=(6,2))
        wm = ctk.CTkFrame(lm, fg_color="transparent")
        wm.grid(row=1, column=0, sticky="nsew")
        wm.grid_rowconfigure(0, weight=1); wm.grid_columnconfigure(0, weight=1)
        self.tree_rotation_mar = self._make_tree_in(wm, row=0)

        # Panel Aéreo
        la = ctk.CTkFrame(panels, fg_color=THEME["card"], corner_radius=8,
                          border_width=1, border_color=THEME["accent"])
        la.grid(row=0, column=1, sticky="nsew", padx=(3,0))
        la.grid_rowconfigure(1, weight=1); la.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(la, text="✈  ESCENARIO AÉREO", font=FONT_KPI_TITLE,
                     text_color=THEME["accent"]).grid(row=0,column=0,sticky="w",padx=8,pady=(6,2))
        wa = ctk.CTkFrame(la, fg_color="transparent")
        wa.grid(row=1, column=0, sticky="nsew")
        wa.grid_rowconfigure(0, weight=1); wa.grid_columnconfigure(0, weight=1)
        self.tree_rotation_air = self._make_tree_in(wa, row=0)

        # Mantener referencia legacy para _fill_rotation y PDF
        self.tree_rotation = self.tree_rotation_mar
    def _build_purchases_tab(self):
        """Pestaña Histórico de Compras: filtro por SKU + tabla de facturas."""
        outer = ctk.CTkFrame(self.tab_purchases, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(1, weight=1)

        # Barra de filtros
        bar = ctk.CTkFrame(outer, fg_color=THEME["panel_alt"], corner_radius=8)
        bar.grid(row=0, column=0, sticky="ew", pady=(0,6))
        bar.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(bar, text="SKU / PRODUCTO:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).grid(row=0, column=0, padx=(10,4), pady=6)
        self.purch_filter = ctk.CTkEntry(bar, placeholder_text="Filtrar por código o nombre...",
                                          font=FONT_BODY, height=28, width=280)
        self.purch_filter.grid(row=0, column=1, sticky="ew", padx=4, pady=6)
        ctk.CTkButton(bar, text="Buscar", width=80, height=28, font=FONT_BOLD,
                      fg_color=THEME["accent"], hover_color="#2563EB", corner_radius=6,
                      command=self.render_purchases).grid(row=0, column=2, padx=4, pady=6)
        ctk.CTkButton(bar, text="Todos", width=70, height=28, font=FONT_SMALL,
                      fg_color=THEME["card"], hover_color=THEME["border"], corner_radius=6,
                      command=lambda: [self.purch_filter.delete(0,"end"),
                                       self.render_purchases()]).grid(row=0, column=3, padx=(0,10), pady=6)

        wrap = ctk.CTkFrame(outer, fg_color="transparent")
        wrap.grid(row=1, column=0, sticky="nsew")
        wrap.grid_rowconfigure(0, weight=1)
        wrap.grid_columnconfigure(0, weight=1)
        self.tree_purchases = self._make_tree_in(wrap, row=0)

    def render_purchases(self):
        """Llena la tabla Histórico de Compras: SKU+Fecha, subtotal por SKU con costo promedio."""
        r = self.current_result
        if r is None:
            self.fill_tree(self.tree_purchases, pd.DataFrame(), empty_msg="Ejecute el análisis primero.")
            return

        df = self.engine.raw_df.copy()
        ref = df["Referencia"].fillna("").astype(str).str.upper()
        typ = df["Tipo"].fillna("").astype(str).str.upper()
        cdf = df[(typ=="ING") & ref.str.startswith("FAC")].copy()

        if cdf.empty:
            self.fill_tree(self.tree_purchases, pd.DataFrame(),
                           empty_msg="No hay compras registradas.")
            return

        # Filtro de texto
        filt = self.purch_filter.get().strip().lower()
        if filt:
            mask = (cdf["Código Producto"].fillna("").astype(str).str.lower().str.contains(filt) |
                    cdf["Nombre Producto"].fillna("").astype(str).str.lower().str.contains(filt))
            cdf = cdf[mask]

        if cdf.empty:
            self.fill_tree(self.tree_purchases, pd.DataFrame(),
                           empty_msg="Sin resultados para el filtro aplicado.")
            return

        cdf = cdf.sort_values(["Código Producto","Fecha"]).reset_index(drop=True)

        # Construir filas de detalle + subtotales por SKU con costo promedio ponderado
        tree = self.tree_purchases
        tree.delete(*tree.get_children())

        cols = ["Fecha","Factura","Código Producto","Nombre Producto",
                "Descripción","Cantidad","Valor Total","V. Unitario","Costo Prom."]
        tree["columns"] = cols

        # Anchos
        ws = {"Fecha":82,"Factura":100,"Código Producto":95,"Nombre Producto":210,
              "Descripción":200,"Cantidad":70,"Valor Total":100,"V. Unitario":100,"Costo Prom.":100}
        for c in cols:
            anchor = "e" if c in ("Cantidad","Valor Total","V. Unitario","Costo Prom.") else "w"
            tree.heading(c, text=c, anchor=anchor)
            tree.column(c, anchor=anchor, width=ws.get(c,90), minwidth=44, stretch=False)

        self._apply_tags(tree)
        tree.tag_configure("subtotal", background="#1E3A5F", foreground="#FACC15",
                           font=("Segoe UI",11,"bold"))
        tree.tag_configure("grand",    background="#172A45", foreground="#F59E0B",
                           font=("Segoe UI",11,"bold"))

        grand_qty = grand_val = 0.0
        row_i = 0

        for sku, grp in cdf.groupby("Código Producto", sort=False):
            costo_prom = None   # costo promedio acumulado para este SKU
            sku_qty = sku_val = 0.0

            for _, row in grp.iterrows():
                qty   = float(row["Cantidad"]) if float(row["Cantidad"]) > 0 else 1
                v_tot = float(row["Valor Total"])
                v_unit = v_tot / qty

                # Costo promedio ponderado
                if costo_prom is None:
                    costo_prom = v_unit          # primera compra
                else:
                    costo_prom = (v_unit + costo_prom) / 2.0

                fecha = row["Fecha"].strftime("%d/%m/%Y") if pd.notna(row["Fecha"]) else ""
                vals = [
                    fecha,
                    str(row.get("Referencia","")).strip(),
                    str(row.get("Código Producto","")).strip(),
                    str(row.get("Nombre Producto","")).strip(),
                    str(row.get("Descripción","")).strip(),
                    f"{int(qty):,}",
                    f"{v_tot:,.2f}",
                    f"{v_unit:,.4f}",
                    "",   # costo prom solo en subtotal
                ]
                tag = "even" if row_i % 2 == 0 else "odd"
                tree.insert("", "end", values=vals, tags=(tag,))
                row_i += 1
                sku_qty += qty; sku_val += v_tot

            # Subtotal del SKU
            nom = str(grp["Nombre Producto"].iloc[0]) if not grp.empty else sku
            tree.insert("", "end", values=[
                "SUBTOTAL", "", sku, nom, "",
                f"{int(sku_qty):,}",
                f"{sku_val:,.2f}",
                "",
                f"{costo_prom:,.4f}" if costo_prom is not None else "",
            ], tags=("subtotal",))

            grand_qty += sku_qty; grand_val += sku_val

        # Total general
        tree.insert("", "end", values=[
            "TOTAL GENERAL", "", "", "", "",
            f"{int(grand_qty):,}",
            f"{grand_val:,.2f}",
            "", "",
        ], tags=("grand",))

        self.log(f"Histórico compras: {int(grand_qty):,} uds | ${grand_val:,.2f}" +
                 (f" | filtro: '{filt}'" if filt else ""))

    # ══════════════════════════════════════════════════════════════════
    # KARDEX
    # ══════════════════════════════════════════════════════════════════
    # ── Helper: entrada de fecha con auto "/" ─────────────────────────────────
    def _make_date_entry(self, parent, width=110):
        """Entry que inserta '/' automáticamente al tipear una fecha ddmmyyyy."""
        entry = ctk.CTkEntry(parent, placeholder_text="dd/mm/yyyy",
                             width=width, height=28, font=FONT_BODY)
        def _on_key(event, e=entry):
            cur = e.get().replace("/","")
            digits = "".join(c for c in cur if c.isdigit())
            if len(digits) >= 2:
                digits = digits[:2] + "/" + digits[2:]
            if len(digits) >= 5:
                digits = digits[:5] + "/" + digits[5:9]
            pos = e.index("insert")
            e.delete(0,"end")
            e.insert(0, digits)
            # Mantener cursor al final si se está escribiendo
            try: e.icursor(min(pos+1, len(digits)))
            except: pass
        entry._entry.bind("<KeyRelease>", _on_key)
        return entry

    def _build_kardex_tab(self):
        """Pestaña Kardex: movimientos cronológicos con costo promedio ponderado."""
        outer = ctk.CTkFrame(self.tab_kardex, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1)
        outer.grid_rowconfigure(1, weight=1)

        # ── Barra de controles ────────────────────────────────────────────
        bar = ctk.CTkFrame(outer, fg_color=THEME["panel_alt"], corner_radius=8)
        bar.grid(row=0, column=0, sticky="ew", pady=(0,6))

        inner = ctk.CTkFrame(bar, fg_color="transparent")
        inner.pack(side="left", anchor="w", padx=8, pady=6)

        ctk.CTkLabel(inner, text="DESDE:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.kdx_from = self._make_date_entry(inner)
        self.kdx_from.pack(side="left", padx=(0,2))
        ctk.CTkButton(inner, text="📅", width=28, height=28,
                      command=lambda: DatePicker(self, lambda d: self._set_entry(self.kdx_from, d),
                                                 self.kdx_from.get())).pack(side="left", padx=(0,10))

        ctk.CTkLabel(inner, text="HASTA:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))
        self.kdx_to = self._make_date_entry(inner)
        self.kdx_to.pack(side="left", padx=(0,2))
        ctk.CTkButton(inner, text="📅", width=28, height=28,
                      command=lambda: DatePicker(self, lambda d: self._set_entry(self.kdx_to, d),
                                                 self.kdx_to.get())).pack(side="left", padx=(0,10))

        ctk.CTkLabel(inner, text="SKU:", font=FONT_KPI_TITLE,
                     text_color=THEME["muted"]).pack(side="left", padx=(0,4))

        # Autocomplete SKU — Listbox flotante debajo del Entry
        sku_frame = ctk.CTkFrame(inner, fg_color="transparent")
        sku_frame.pack(side="left", padx=(0,8))
        self.kdx_sku = ctk.CTkEntry(sku_frame, placeholder_text="Código o nombre (vacío=todos)",
                                    width=220, height=28, font=FONT_BODY)
        self.kdx_sku.pack()
        # Listbox de sugerencias (tk nativo — se superpone)
        import tkinter as _tk
        self._kdx_listbox_win = None

        def _show_suggestions(event=None):
            txt = self.kdx_sku.get().strip().lower()
            if not txt or self.engine.raw_df is None:
                _hide_suggestions(); return
            raw = self.engine.raw_df
            matches = []
            for sku in raw["Código Producto"].dropna().unique():
                s = str(sku)
                if txt in s.lower():
                    matches.append(s)
            # También buscar por nombre
            for nom in raw["Nombre Producto"].dropna().unique():
                s = str(nom)
                if txt in s.lower() and s not in matches:
                    matches.append(s)
            matches = matches[:12]
            if not matches:
                _hide_suggestions(); return
            _hide_suggestions()
            # Crear ventana Toplevel transparente con Listbox
            x = self.kdx_sku.winfo_rootx()
            y = self.kdx_sku.winfo_rooty() + self.kdx_sku.winfo_height()
            win = _tk.Toplevel(self)
            win.wm_overrideredirect(True)
            win.geometry(f"220x{min(len(matches),8)*20+4}+{x}+{y}")
            win.configure(bg=THEME["panel_alt"])
            lb = _tk.Listbox(win, font=("Segoe UI",10), bg=THEME["card"],
                             fg=THEME["text"], selectbackground=THEME["accent"],
                             bd=0, highlightthickness=0, activestyle="none")
            lb.pack(fill="both", expand=True)
            for m in matches:
                lb.insert("end", m)
            # Índice de navegación por teclado (sin cambiar foco al Listbox)
            _nav_idx = [-1]   # mutable para closures

            def _pick(val):
                """Selecciona un valor, cierra el listbox y ejecuta el Kardex."""
                if val:
                    self.kdx_sku.delete(0, "end")
                    self.kdx_sku.insert(0, val)
                _hide_suggestions()
                if val and self.kdx_from.get().strip() and self.kdx_to.get().strip():
                    self.after(80, self.run_kardex)

            def _pick_current():
                """Selecciona el ítem actualmente resaltado."""
                idx = _nav_idx[0]
                if 0 <= idx < lb.size():
                    _pick(lb.get(idx))
                elif lb.size() > 0:
                    _pick(lb.get(0))

            def _nav(direction):
                """Mueve la selección visual sin cambiar el foco."""
                size = lb.size()
                if size == 0: return
                cur = _nav_idx[0]
                if direction == "down":
                    _nav_idx[0] = min(cur + 1, size - 1) if cur >= 0 else 0
                else:
                    _nav_idx[0] = max(cur - 1, 0) if cur >= 0 else 0
                lb.selection_clear(0, "end")
                lb.selection_set(_nav_idx[0])
                lb.see(_nav_idx[0])
                # Mostrar en el entry como preview
                lb_val = lb.get(_nav_idx[0])
                self.kdx_sku.delete(0, "end")
                self.kdx_sku.insert(0, lb_val)
                self.kdx_sku._entry.icursor("end")

            # Clic con mouse — without changing focus
            def _on_mouse(evt):
                idx = lb.nearest(evt.y)
                if idx >= 0:
                    _nav_idx[0] = idx
                    _pick(lb.get(idx))
            lb.bind("<ButtonRelease-1>", _on_mouse)
            lb.bind("<Button-1>",        _on_mouse)

            # Teclado desde el Entry — foco permanece en el Entry
            def _on_entry_key(evt):
                key = evt.keysym
                if key in ("Return", "Tab"):
                    _pick_current()
                    return "break"
                elif key == "Down":
                    _nav("down")
                    return "break"
                elif key == "Up":
                    _nav("up")
                    return "break"
                elif key == "Escape":
                    _hide_suggestions()
                    return "break"

            self.kdx_sku._entry.bind("<Return>",  _on_entry_key, add="+")
            self.kdx_sku._entry.bind("<Tab>",     _on_entry_key, add="+")
            self.kdx_sku._entry.bind("<Down>",    _on_entry_key, add="+")
            self.kdx_sku._entry.bind("<Up>",      _on_entry_key, add="+")
            self.kdx_sku._entry.bind("<Escape>",  _on_entry_key, add="+")
            self._kdx_listbox_win = win

        def _hide_suggestions(event=None):
            if self._kdx_listbox_win:
                try: self._kdx_listbox_win.destroy()
                except: pass
                self._kdx_listbox_win = None

        self.kdx_sku._entry.bind("<KeyRelease>", _show_suggestions)
        self.kdx_sku._entry.bind("<FocusOut>",
            lambda e: self.after(150, _hide_suggestions))

        ctk.CTkButton(inner, text="Generar", height=28, font=FONT_BOLD,
                      fg_color=THEME["accent"], hover_color="#2563EB", corner_radius=6,
                      command=self.run_kardex).pack(side="left", padx=(0,4))
        ctk.CTkButton(inner, text="🌐 HTML", height=28, font=FONT_BOLD,
                      fg_color=THEME["teal"], hover_color="#0D9488", corner_radius=6,
                      command=self.export_kardex_html).pack(side="left", padx=(0,4))
        ctk.CTkButton(inner, text="📄 PDF", height=28, font=FONT_BOLD,
                      fg_color=THEME["danger"], hover_color="#B91C1C", corner_radius=6,
                      command=self.export_kardex_pdf).pack(side="left", padx=(0,4))
        ctk.CTkButton(inner, text="📊 Excel", height=28, font=FONT_BOLD,
                      fg_color=THEME["success"], hover_color="#059669", corner_radius=6,
                      command=self.export_kardex_excel).pack(side="left")

        # ── Tabla ─────────────────────────────────────────────────────────
        wrap = ctk.CTkFrame(outer, fg_color="transparent")
        wrap.grid(row=1, column=0, sticky="nsew")
        wrap.grid_rowconfigure(0, weight=1); wrap.grid_columnconfigure(0, weight=1)
        self.tree_kardex = self._make_tree_in(wrap, row=0)
        self._kardex_df: "pd.DataFrame | None" = None

    def _calc_kardex(self, d_from, d_to, sku_filter: str = ""):
        """Construye el DataFrame del Kardex completo.

        Lógica:
        - Toma TODOS los movimientos históricos hasta d_to.
        - Calcula el costo promedio ponderado de forma incremental
          solo con las COMPRAS (ING+FAC). TRA no afecta al costo.
        - Genera una fila de SALDO INICIAL al cierre del día anterior
          a d_from, con el costo promedio vigente en ese momento.
        - Muestra todos los movimientos del período d_from→d_to.
        """
        raw = self.engine.raw_df.copy()

        # Filtro de SKU
        if sku_filter:
            fl = sku_filter.lower()
            mask = (raw["Código Producto"].fillna("").astype(str).str.lower().str.contains(fl) |
                    raw["Nombre Producto"].fillna("").astype(str).str.lower().str.contains(fl))
            raw = raw[mask]

        if raw.empty:
            return pd.DataFrame()

        ref = raw["Referencia"].fillna("").astype(str).str.upper()
        typ = raw["Tipo"].fillna("").astype(str).str.upper()
        raw["_is_purchase"] = (typ == "ING") & ref.str.startswith("FAC")
        raw["_is_sale"]     = (typ == "EGR") & ref.str.startswith("FAC")
        raw["_is_sup_ret"]  = (typ == "EGR") & ref.str.startswith("NCT")   # dev proveedor
        raw["_is_cust_ret"] = (typ == "ING") & ref.str.startswith("NCT")   # dev cliente
        raw["_is_tra"]      = typ == "TRA"

        raw = raw.sort_values(["Código Producto", "Fecha", "Código"]).reset_index(drop=True)

        # ── Paso 1: calcular costo promedio por SKU hasta d_from-1 ────────
        hist = raw[raw["Fecha"] < d_from].copy()

        # Costo promedio acumulado y stock antes del período
        cp_map    = {}   # SKU → costo promedio vigente
        cpqty_map = {}   # SKU → cantidad acumulada (para promedio ponderado)
        cpval_map = {}   # SKU → valor acumulado
        stock_map = {}   # SKU → stock neto antes del período

        for _, row in hist.iterrows():
            sku = row["Código Producto"]
            qty = float(row["Cantidad"])
            if stock_map.get(sku) is None:
                stock_map[sku] = 0.0

            # Stock neto: ING suma, EGR resta, TRA neto cero consolidado
            if row["_is_purchase"] or row["_is_cust_ret"]:
                stock_map[sku] += qty
            elif row["_is_sale"] or row["_is_sup_ret"]:
                stock_map[sku] -= qty
            # TRA no afecta stock consolidado (sale de una bodega y entra a otra)

            # Costo promedio solo con compras
            if row["_is_purchase"] and qty > 0:
                vtot = float(row["Valor Total"])
                cunit = vtot / qty if qty > 0 else 0
                if cunit == 0:
                    cunit = float(row.get("Valor Unitario", 0) or 0)
                if cunit > 0:
                    old_qty = cpqty_map.get(sku, 0.0)
                    old_val = cpval_map.get(sku, 0.0)
                    new_qty = old_qty + qty
                    new_val = old_val + vtot
                    cpqty_map[sku] = new_qty
                    cpval_map[sku] = new_val
                    cp_map[sku]    = new_val / new_qty

        # ── Paso 2: recorrer movimientos del período y generar filas ──────
        period = raw[(raw["Fecha"] >= d_from) & (raw["Fecha"] <= d_to)].copy()

        rows = []

        for sku, grp in raw.groupby("Código Producto"):
            nom = str(grp["Nombre Producto"].iloc[0])

            # Fila saldo inicial
            s0   = stock_map.get(sku, 0.0)
            cp0  = cp_map.get(sku, 0.0)
            val0 = s0 * cp0
            rows.append({
                "Fecha":        d_from.strftime("%d/%m/%Y"),
                "Código":       sku,
                "Nombre":       nom,
                "Referencia":   "—",
                "Descripción":  "SALDO INICIAL",
                "Tipo Mov.":    "INICIO",
                "Cantidad":     round(s0, 2),
                "V.Unit":       round(cp0, 4),
                "Costo Prom.":  round(cp0, 4),
                "Saldo Uds":    round(s0, 2),
                "Valor Inv.":   round(val0, 2),
                "_sku":         sku,
                "_ord":         0,
            })

            cp_current  = cp_map.get(sku, 0.0)
            qty_current = cpqty_map.get(sku, 0.0)
            val_current = cpval_map.get(sku, 0.0)
            saldo       = s0
            ord_i       = 1

            for _, row in grp[(grp["Fecha"] >= d_from) & (grp["Fecha"] <= d_to)].iterrows():
                qty  = float(row["Cantidad"])
                vtot = float(row["Valor Total"])
                vunit= float(row.get("Valor Unitario", 0) or 0)
                if vunit == 0 and qty > 0:
                    vunit = vtot / qty

                fecha  = row["Fecha"].strftime("%d/%m/%Y") if pd.notna(row["Fecha"]) else ""
                ref    = str(row.get("Referencia", "")).strip()
                desc   = str(row.get("Descripción", "")).strip()

                # Determinar tipo de movimiento y efecto en saldo/costo
                if row["_is_purchase"]:
                    tipo_mov = "INGRESO"
                    efecto   = +qty
                    # Actualizar costo promedio ponderado solo en compras
                    if qty > 0 and vunit > 0:
                        nqty = qty_current + qty
                        nval = val_current + vtot
                        cp_current  = nval / nqty
                        qty_current = nqty
                        val_current = nval

                elif row["_is_cust_ret"]:
                    tipo_mov = "ING DEV.CLI"
                    efecto   = +qty

                elif row["_is_sale"]:
                    tipo_mov = "EGRESO"
                    efecto   = -qty

                elif row["_is_sup_ret"]:
                    tipo_mov = "EGR DEV.PROV"
                    efecto   = -qty

                elif row["_is_tra"]:
                    tipo_mov = "TRANSFERENCIA"
                    efecto   = 0   # consolidado: sale de una bodega, entra a otra → neto 0

                else:
                    tipo_mov = "OTRO"
                    efecto   = 0

                saldo += efecto

                rows.append({
                    "Fecha":       fecha,
                    "Código":      sku,
                    "Nombre":      nom,
                    "Referencia":  ref,
                    "Descripción": desc,
                    "Tipo Mov.":   tipo_mov,
                    "Cantidad":    round(abs(qty) if tipo_mov != "TRANSFERENCIA" else qty, 2),
                    "V.Unit":      round(vunit, 4),
                    "Costo Prom.": round(cp_current, 4),
                    "Saldo Uds":   round(saldo, 2),
                    "Valor Inv.":  round(saldo * cp_current, 2),
                    "_sku":        sku,
                    "_ord":        ord_i,
                })
                ord_i += 1

        if not rows:
            return pd.DataFrame()

        df = pd.DataFrame(rows).sort_values(["_sku", "_ord"]).drop(columns=["_sku","_ord"])
        return df.reset_index(drop=True)

    def run_kardex(self):
        """Genera el Kardex en pantalla."""
        if self.engine.raw_df is None:
            messagebox.showinfo("Kardex", "Cargue primero el archivo Excel."); return
        try:
            d_from = self._parse_date(self.kdx_from.get())
            d_to   = self._parse_date(self.kdx_to.get())
        except:
            messagebox.showerror("Fechas", "Use el formato dd/mm/yyyy."); return
        if d_from > d_to:
            messagebox.showerror("Fechas", "La fecha inicial debe ser menor a la final."); return

        sku_filter = self.kdx_sku.get().strip()
        self.log(f"Kardex: generando {d_from.date()} → {d_to.date()}" +
                 (f" | SKU: {sku_filter}" if sku_filter else " | todos los SKU"))

        df = self._calc_kardex(d_from, d_to, sku_filter)
        self._kardex_df = df

        if df.empty:
            self.fill_tree(self.tree_kardex, pd.DataFrame(),
                           empty_msg="Sin movimientos para el filtro aplicado.")
            return

        # Renderizar con tags por tipo de movimiento
        tree = self.tree_kardex
        tree.delete(*tree.get_children())
        display_cols = [c for c in df.columns]
        tree["columns"] = display_cols

        col_w_kdx = {
            "Fecha":10, "Código":90, "Nombre":210, "Referencia":100,
            "Descripción":160, "Tipo Mov.":110, "Cantidad":70,
            "V.Unit":90, "Costo Prom.":90, "Saldo Uds":80, "Valor Inv.":100,
        }
        for c in display_cols:
            anchor = "e" if c in ("Cantidad","V.Unit","Costo Prom.","Saldo Uds","Valor Inv.") else "w"
            tree.heading(c, text=c, anchor=anchor)
            tree.column(c, anchor=anchor, width=col_w_kdx.get(c,90), minwidth=44, stretch=False)

        self._apply_tags(tree)
        # Paleta azul/celeste pastel — sin colores saturados
        tree.tag_configure("kdx_inicio", background="#1E3A5F", foreground="#BFDBFE",
                           font=("Segoe UI",11,"bold"))
        tree.tag_configure("kdx_subtot", background="#172A45", foreground="#93C5FD",
                           font=("Segoe UI",10,"bold"))
        tree.tag_configure("kdx_ing",    background="#0C1E35", foreground="#7DD3FC")   # azul claro
        tree.tag_configure("kdx_egr",    background="#162032", foreground="#BAE6FD")   # celeste más claro
        tree.tag_configure("kdx_tra",    background="#0F1A2E", foreground="#93C5FD")   # azul medio
        tree.tag_configure("kdx_dev",    background="#1A2D44", foreground="#E0F2FE")   # casi blanco azul
        tree.tag_configure("kdx_even",   background="#141E2E", foreground="#CBD5E1")
        tree.tag_configure("kdx_odd",    background=THEME["panel"], foreground="#CBD5E1")

        tag_map = {
            "INICIO":       "kdx_inicio",
            "INGRESO":      "kdx_ing",
            "ING DEV.CLI":  "kdx_dev",
            "EGRESO":       "kdx_egr",
            "EGR DEV.PROV": "kdx_dev",
            "TRANSFERENCIA":"kdx_tra",
        }

        NUM_COLS = {"Cantidad","Saldo Uds","Valor Inv.","V.Unit","Costo Prom."}

        def fmt_val(c, v):
            try:
                fv = float(v)
                return f"{fv:,.4f}" if c in ("V.Unit","Costo Prom.") else f"{fv:,.2f}"
            except:
                return str(v) if str(v) not in ("nan","None","NaN") else ""

        prev_sku = None
        row_i    = 0
        # Acumuladores para subtotal
        sub_qty_in = sub_qty_out = sub_qty_tra = 0.0

        for _, row in df.iterrows():
            tipo = str(row.get("Tipo Mov.", ""))
            sku  = str(row.get("Código", ""))

            # Insertar fila subtotal cuando cambia el SKU (excepto en el primero)
            if prev_sku is not None and sku != prev_sku:
                # Subtotal del SKU anterior
                tree.insert("", "end", values=["", prev_sku, "", "", "── SUBTOTAL SKU ──",
                    "SUBTOTAL",
                    f"{sub_qty_in:,.2f}", "", "", "", ""],
                    tags=("kdx_subtot",))
                sub_qty_in = sub_qty_out = sub_qty_tra = 0.0

            # Acumular por tipo
            qty = float(row.get("Cantidad", 0))
            if tipo in ("INGRESO","ING DEV.CLI"): sub_qty_in  += qty
            elif tipo in ("EGRESO","EGR DEV.PROV"): sub_qty_out += qty
            elif tipo == "TRANSFERENCIA": sub_qty_tra += qty

            tag = tag_map.get(tipo, "kdx_even" if row_i % 2 == 0 else "kdx_odd")
            vals = [fmt_val(c, row[c]) if c in NUM_COLS
                    else (str(row[c]) if str(row[c]) not in ("nan","None","NaN") else "")
                    for c in display_cols]
            tree.insert("", "end", values=vals, tags=(tag,))
            prev_sku = sku
            row_i   += 1

        # Subtotal del último SKU
        if prev_sku is not None:
            tree.insert("", "end", values=["", prev_sku, "", "", "── SUBTOTAL SKU ──",
                "SUBTOTAL", f"{sub_qty_in:,.2f}", "", "", "", ""],
                tags=("kdx_subtot",))

        self.log(f"Kardex: {len(df):,} filas generadas")

    def export_kardex_html(self):
        """Exporta el Kardex a HTML con tema claro profesional."""
        if self._kardex_df is None or self._kardex_df.empty:
            messagebox.showinfo("Kardex", "Primero genere el Kardex."); return

        sku_t = self.kdx_sku.get().strip().replace(" ","_").replace("/","_")
        fname = f"kardex_{sku_t}.html" if sku_t else "kardex.html"
        path = filedialog.asksaveasfilename(
            defaultextension=".html", filetypes=[("HTML","*.html")],
            initialfile=fname)
        if not path: return

        try:
            from datetime import datetime as _dt
            df   = self._kardex_df.copy()
            cols = list(df.columns)
            NUM  = {"Cantidad","Saldo Uds","Valor Inv.","V.Unit","Costo Prom."}

            COLOR_MAP = {
                "INICIO":       ("#DBEAFE","#1E40AF"),  # azul
                "INGRESO":      ("#D1FAE5","#065F46"),  # verde
                "ING DEV.CLI":  ("#FEF9C3","#854D0E"),  # amarillo
                "EGRESO":       ("#F3F4F6","#374151"),  # gris
                "EGR DEV.PROV": ("#F3F4F6","#374151"),
                "TRANSFERENCIA":("#EDE9FE","#5B21B6"),  # violeta
                "SUBTOTAL":     ("#EFF6FF","#1E3A8A"),  # azul pastel fuerte
            }

            rows_html = ""
            prev_sku  = None
            sub_ing = sub_egr = 0.0

            for _, row in df.iterrows():
                tipo = str(row.get("Tipo Mov.",""))
                sku  = str(row.get("Código",""))

                if prev_sku is not None and sku != prev_sku:
                    # Fila subtotal SKU anterior
                    bg,fg = COLOR_MAP["SUBTOTAL"]
                    rows_html += (
                        f'<tr style="background:{bg};color:{fg};font-weight:bold">' +
                        f'<td></td><td>{prev_sku}</td>' +
                        '<td colspan="3" style="text-align:center">── SUBTOTAL ──</td>' +
                        f'<td style="text-align:right">{sub_ing:,.2f}</td>' +
                        f'<td style="text-align:right">{sub_egr:,.2f}</td>' +
                        '<td colspan="4"></td></tr>'
                    )
                    sub_ing = sub_egr = 0.0

                qty = float(row.get("Cantidad",0))
                if tipo in ("INGRESO","ING DEV.CLI"): sub_ing += qty
                elif tipo in ("EGRESO","EGR DEV.PROV"): sub_egr += qty
                prev_sku = sku

                bg, fg = COLOR_MAP.get(tipo, ("#FFFFFF","#111827"))
                cells = ""
                for c in cols:
                    v = row[c]
                    align = "right" if c in NUM else "left"
                    if c in NUM:
                        try:
                            fv = float(v)
                            txt = f"{fv:,.4f}" if c in ("V.Unit","Costo Prom.") else f"{fv:,.2f}"
                        except: txt = str(v)
                    else:
                        txt = str(v) if str(v) not in ("nan","None","NaN") else ""
                    cells += f'<td style="text-align:{align}">{txt}</td>'
                rows_html += f'<tr style="background:{bg};color:{fg}">{cells}</tr>'

            # Último subtotal
            if prev_sku:
                bg,fg = COLOR_MAP["SUBTOTAL"]
                rows_html += (
                    f'<tr style="background:{bg};color:{fg};font-weight:bold">' +
                    f'<td></td><td>{prev_sku}</td>' +
                    '<td colspan="3" style="text-align:center">── SUBTOTAL ──</td>' +
                    f'<td style="text-align:right">{sub_ing:,.2f}</td>' +
                    f'<td style="text-align:right">{sub_egr:,.2f}</td>' +
                    '<td colspan="4"></td></tr>'
                )

            hdr_cells = "".join(f"<th>{c}</th>" for c in cols)
            sku_txt   = self.kdx_sku.get().strip()

            html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Kardex de Inventario</title>
<style>
  body  {{ font-family:Segoe UI,Arial,sans-serif; background:#fff; color:#111827;
           margin:0; padding:20px; font-size:11px }}
  h1    {{ color:#1E3A5F; font-size:18px; margin-bottom:2px;
           border-bottom:2px solid #1E3A5F; padding-bottom:6px }}
  .meta {{ color:#6B7280; font-size:10px; margin-bottom:14px }}
  table {{ width:100%; border-collapse:collapse; font-size:10px }}
  th    {{ background:#1E3A5F; color:#fff; padding:6px 8px;
           text-align:left; position:sticky; top:0 }}
  td    {{ padding:4px 8px; border-bottom:1px solid #E5E7EB }}
  .leg  {{ display:flex; gap:12px; margin-bottom:10px; flex-wrap:wrap }}
  .lb   {{ padding:3px 10px; border-radius:4px; font-size:10px }}
  @media print {{ th {{ position:static }} }}
</style>
</head>
<body>
<h1>KARDEX DE INVENTARIO</h1>
<div class="meta">
  Período: <b>{self.kdx_from.get()} → {self.kdx_to.get()}</b>
  {"&nbsp;|&nbsp; SKU: <b>" + sku_txt + "</b>" if sku_txt else ""}
  &nbsp;|&nbsp; Generado: {_dt.now().strftime("%d/%m/%Y %H:%M")}
</div>
<div class="leg">
  <span class="lb" style="background:#DBEAFE;color:#1E40AF">■ INICIO</span>
  <span class="lb" style="background:#D1FAE5;color:#065F46">■ INGRESO</span>
  <span class="lb" style="background:#F3F4F6;color:#374151;border:1px solid #D1D5DB">■ EGRESO</span>
  <span class="lb" style="background:#EDE9FE;color:#5B21B6">■ TRANSFERENCIA</span>
  <span class="lb" style="background:#FEF9C3;color:#854D0E">■ DEV.</span>
  <span class="lb" style="background:#EFF6FF;color:#1E3A8A;font-weight:bold">■ SUBTOTAL</span>
</div>
<table>
  <thead><tr>{hdr_cells}</tr></thead>
  <tbody>{rows_html}</tbody>
</table>
</body></html>"""

            with open(path,"w",encoding="utf-8") as f: f.write(html)
            import webbrowser, os
            webbrowser.open(f"file:///{os.path.abspath(path)}")
            messagebox.showinfo("HTML", f"Kardex HTML exportado:\n{path}")
            self.log(f"Kardex HTML → {path}")

        except Exception as e:
            self.log(f"Error Kardex HTML: {e}")
            messagebox.showerror("Error", str(e))

    def export_kardex_pdf(self):
        """Exporta el Kardex a PDF (landscape A4, tabla compacta)."""
        if self._kardex_df is None or self._kardex_df.empty:
            messagebox.showinfo("Kardex", "Primero genere el Kardex."); return

        sku_t = self.kdx_sku.get().strip().replace(" ","_").replace("/","_")
        fname = f"kardex_{sku_t}.pdf" if sku_t else "kardex.pdf"
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=fname)
        if not path: return

        try:
            from reportlab.lib.pagesizes import landscape, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, HRFlowable)
            from datetime import datetime as _dt

            PAGE  = landscape(A4)
            doc   = SimpleDocTemplate(path, pagesize=PAGE,
                                      topMargin=1.0*cm, bottomMargin=1.0*cm,
                                      leftMargin=1.2*cm, rightMargin=1.2*cm)
            PW    = PAGE[0] - 2.4*cm

            styles = getSampleStyleSheet()
            C_BLUE  = colors.HexColor("#1E3A5F")
            C_TEXT  = colors.HexColor("#111827")
            C_MUTED = colors.HexColor("#6B7280")
            C_GREEN = colors.HexColor("#059669")
            C_RED   = colors.HexColor("#DC2626")
            C_EVEN  = colors.HexColor("#F9FAFB")
            C_INIT  = colors.HexColor("#EFF6FF")
            C_ING   = colors.HexColor("#ECFDF5")
            C_EGR   = colors.HexColor("#F3F4F6")  # gris claro
            C_TRA   = colors.HexColor("#F5F3FF")

            title_s = ParagraphStyle("t", fontSize=13, textColor=C_BLUE,
                                     fontName="Helvetica-Bold", spaceAfter=2)
            sub_s   = ParagraphStyle("s", fontSize=9, textColor=C_MUTED,
                                     fontName="Helvetica", spaceAfter=4)
            cell_s  = ParagraphStyle("c", fontSize=7, textColor=C_TEXT,
                                     fontName="Helvetica", leading=9)

            elems = []
            elems.append(Paragraph("KARDEX DE INVENTARIO", title_s))
            elems.append(Paragraph(
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}  |  "
                f"Período: {self.kdx_from.get()} → {self.kdx_to.get()}"
                + (f"  |  SKU: {self.kdx_sku.get()}" if self.kdx_sku.get().strip() else ""),
                sub_s))
            elems.append(HRFlowable(width="100%", thickness=1, color=C_BLUE, spaceAfter=6))

            df = self._kardex_df.copy()
            cols     = list(df.columns)
            num_cols = {"Cantidad","Saldo Uds","Valor Inv.","V.Unit","Costo Prom."}

            # Anchos proporcionales al ancho útil
            col_ratios = {
                "Fecha":0.065,"Código":0.07,"Nombre":0.16,"Referencia":0.075,
                "Descripción":0.12,"Tipo Mov.":0.08,"Cantidad":0.055,
                "V.Unit":0.075,"Costo Prom.":0.075,"Saldo Uds":0.06,"Valor Inv.":0.075,
            }
            cws = [PW * col_ratios.get(c, 0.07) for c in cols]

            # Header
            hdr = [Paragraph(f"<b>{c}</b>", ParagraphStyle("h", fontSize=7,
                   textColor=colors.white, fontName="Helvetica-Bold", leading=9))
                   for c in cols]
            tdata = [hdr]

            color_map = {
                "INICIO":"INIT","INGRESO":"ING","ING DEV.CLI":"ING",
                "EGRESO":"EGR","EGR DEV.PROV":"EGR","TRANSFERENCIA":"TRA",
            }
            bg_map = {"INIT":C_INIT,"ING":C_ING,"EGR":C_EGR,"TRA":C_TRA}

            table_styles = [
                ("BACKGROUND",    (0,0), (-1,0), C_BLUE),
                ("FONTSIZE",      (0,0), (-1,-1), 7),
                ("TOPPADDING",    (0,0), (-1,-1), 2),
                ("BOTTOMPADDING", (0,0), (-1,-1), 2),
                ("BOX",           (0,0), (-1,-1), 0.5, colors.HexColor("#D1D5DB")),
                ("INNERGRID",     (0,0), (-1,-1), 0.3, colors.HexColor("#E5E7EB")),
            ]

            for ri, (_, row) in enumerate(df.iterrows(), 1):
                tipo  = str(row.get("Tipo Mov.", ""))
                grp   = color_map.get(tipo, "")
                bg    = bg_map.get(grp, (C_EVEN if ri%2==0 else colors.white))
                table_styles.append(("BACKGROUND", (0,ri), (-1,ri), bg))

                vals = []
                for c in cols:
                    v = row[c]
                    if c in num_cols:
                        fmt = f"{float(v):,.4f}" if c in ("V.Unit","Costo Prom.") else f"{float(v):,.2f}"
                        vals.append(Paragraph(fmt, ParagraphStyle("nr", fontSize=7,
                            fontName="Helvetica", alignment=2, leading=9)))
                    else:
                        s = str(v) if str(v) not in ("nan","None","NaN") else ""
                        vals.append(Paragraph(s, cell_s))
                tdata.append(vals)

            t = Table(tdata, colWidths=cws, repeatRows=1)
            t.setStyle(TableStyle(table_styles))
            elems.append(t)

            doc.build(elems)
            import os as _os, subprocess as _sub
            try:
                if _os.name == "nt":
                    _os.startfile(path)
                else:
                    _sub.Popen(["xdg-open", path])
            except Exception:
                pass
            messagebox.showinfo("PDF", f"Kardex exportado:\n{path}")
            self.log(f"Kardex PDF → {path}")

        except Exception as e:
            self.log(f"Error Kardex PDF: {e}")
            messagebox.showerror("Error", str(e))

    def export_kardex_excel(self):
        """Exporta el Kardex a Excel con formato."""
        if self._kardex_df is None or self._kardex_df.empty:
            messagebox.showinfo("Kardex", "Primero genere el Kardex."); return

        sku_t = self.kdx_sku.get().strip().replace(" ","_").replace("/","_")
        fname = f"kardex_{sku_t}.xlsx" if sku_t else "kardex.xlsx"
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=fname)
        if not path: return

        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from datetime import datetime as _dt

            df  = self._kardex_df.copy()
            wb  = openpyxl.Workbook()
            ws  = wb.active; ws.title = "Kardex"

            HDR   = PatternFill("solid", fgColor="1E3A5F")
            INIT  = PatternFill("solid", fgColor="DBEAFE")
            ING   = PatternFill("solid", fgColor="D1FAE5")
            EGR   = PatternFill("solid", fgColor="E5E7EB")  # gris
            TRA   = PatternFill("solid", fgColor="EDE9FE")
            EVEN  = PatternFill("solid", fgColor="F9FAFB")
            ODD   = PatternFill("solid", fgColor="FFFFFF")
            thin  = Side(style="thin", color="E5E7EB")
            brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

            fill_map = {
                "INICIO":"INIT","INGRESO":"ING","ING DEV.CLI":"ING",
                "EGRESO":"EGR","EGR DEV.PROV":"EGR","TRANSFERENCIA":"TRA",
            }
            fills = {"INIT":INIT,"ING":ING,"EGR":EGR,"TRA":TRA}

            cols     = list(df.columns)
            num_cols = {"Cantidad","Saldo Uds","Valor Inv.","V.Unit","Costo Prom."}

            # Título
            nc = len(cols)
            ws.merge_cells(f"A1:{get_column_letter(nc)}1")
            ws["A1"] = f"KARDEX DE INVENTARIO  —  {self.kdx_from.get()} → {self.kdx_to.get()}"
            ws["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
            ws["A1"].fill      = HDR
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            ws.merge_cells(f"A2:{get_column_letter(nc)}2")
            ws["A2"] = f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
            ws["A2"].font = Font(size=9, color="6B7280")
            ws.row_dimensions[2].height = 16

            # Headers
            for ci, c in enumerate(cols, 1):
                cell = ws.cell(3, ci, c)
                cell.font      = Font(bold=True, size=8, color="FFFFFF")
                cell.fill      = HDR
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = brd
            ws.row_dimensions[3].height = 28

            # Datos
            for ri, (_, row) in enumerate(df.iterrows(), 4):
                tipo  = str(row.get("Tipo Mov.", ""))
                grp   = fill_map.get(tipo, "")
                fill  = fills.get(grp, EVEN if ri%2==0 else ODD)
                for ci, c in enumerate(cols, 1):
                    cell        = ws.cell(ri, ci)
                    cell.border = brd
                    cell.fill   = fill
                    v = row[c]
                    if c in num_cols:
                        try:
                            fv = float(v)
                            cell.value = round(fv, 4 if c in ("V.Unit","Costo Prom.") else 2)
                            fmt = '#,##0.0000' if c in ("V.Unit","Costo Prom.") else '#,##0.00'
                            cell.number_format = fmt
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        except:
                            cell.value = str(v)
                    else:
                        sv = str(v) if str(v) not in ("nan","None","NaN") else ""
                        cell.value     = sv
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.font = Font(size=8, color="111827",
                                     bold=(tipo in ("INICIO",)))
                ws.row_dimensions[ri].height = 14

            # Anchos
            col_widths = {
                "Fecha":11,"Código":12,"Nombre":32,"Referencia":14,
                "Descripción":24,"Tipo Mov.":14,"Cantidad":10,
                "V.Unit":12,"Costo Prom.":12,"Saldo Uds":11,"Valor Inv.":13,
            }
            for ci, c in enumerate(cols, 1):
                ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(c, 12)

            ws.freeze_panes = "A4"

            wb.save(path)
            import os as _os, subprocess as _sub
            try:
                if _os.name == "nt": _os.startfile(path)
                else: _sub.Popen(["xdg-open", path])
            except Exception: pass
            messagebox.showinfo("Excel", f"Kardex exportado:\n{path}\n{len(df):,} filas")
            self.log(f"Kardex Excel → {path}")

        except Exception as e:
            self.log(f"Error Kardex Excel: {e}")
            messagebox.showerror("Error", str(e))

    def _build_physical_tab(self):
        outer = ctk.CTkFrame(self.tab_physical, fg_color="transparent")
        outer.pack(fill="both", expand=True, padx=6, pady=6)
        outer.grid_columnconfigure(0, weight=1); outer.grid_rowconfigure(1, weight=1)

        ctk.CTkLabel(outer, text="Toma Fisica — Comparacion con Inventario Calculado",
                     font=FONT_BOLD, text_color=THEME["muted"]).grid(
            row=0, column=0, sticky="w", pady=(0,4))

        wrap = ctk.CTkFrame(outer, fg_color="transparent")
        wrap.grid(row=1, column=0, sticky="nsew")
        wrap.grid_rowconfigure(0, weight=1); wrap.grid_columnconfigure(0, weight=1)
        self.tree_physical = self._make_tree_in(wrap, row=0)

    # ── Tree helpers ──────────────────────────────────────────────────────────
    def _make_tree(self, parent) -> ttk.Treeview:
        wrap = ctk.CTkFrame(parent, fg_color="transparent")
        wrap.pack(fill="both", expand=True, padx=6, pady=6)
        wrap.grid_rowconfigure(0,weight=1); wrap.grid_columnconfigure(0,weight=1)
        return self._make_tree_in(wrap, row=0)

    def _make_tree_in(self, parent, row: int) -> ttk.Treeview:
        # Fila de zoom encima del tree (row - 1 reservado para zoom bar)
        zoom_row = row
        tree_row = row + 1
        # Reconfigurar pesos
        parent.grid_rowconfigure(zoom_row, weight=0)
        parent.grid_rowconfigure(tree_row, weight=1)
        parent.grid_columnconfigure(0, weight=1)

        # ── Zoom bar ──────────────────────────────────────────────────────
        zbar = ctk.CTkFrame(parent, fg_color=THEME["panel_alt"], height=26, corner_radius=0)
        zbar.grid(row=zoom_row, column=0, columnspan=2, sticky="ew")
        zbar.grid_propagate(False)

        ctk.CTkLabel(zbar, text="Zoom:", font=("Segoe UI", 9),
                     text_color=THEME["muted"]).pack(side="left", padx=(8, 2))

        def zoom_out(t=None):
            self._zoom_size = max(7, self._zoom_size - 1)
            self._apply_zoom()
        def zoom_in(t=None):
            self._zoom_size = min(18, self._zoom_size + 1)
            self._apply_zoom()
        def zoom_reset(t=None):
            self._zoom_size = 11
            self._apply_zoom()

        ctk.CTkButton(zbar, text="−", width=26, height=20, font=("Segoe UI", 12, "bold"),
                      corner_radius=4, fg_color=THEME["card"], hover_color=THEME["border"],
                      command=zoom_out).pack(side="left", padx=1)
        ctk.CTkButton(zbar, text="+", width=26, height=20, font=("Segoe UI", 12, "bold"),
                      corner_radius=4, fg_color=THEME["card"], hover_color=THEME["border"],
                      command=zoom_in).pack(side="left", padx=1)
        ctk.CTkButton(zbar, text="↺", width=26, height=20, font=("Segoe UI", 11),
                      corner_radius=4, fg_color=THEME["card"], hover_color=THEME["border"],
                      command=zoom_reset).pack(side="left", padx=1)

        # ── Treeview ──────────────────────────────────────────────────────
        tree = ttk.Treeview(parent, style="Dark.Treeview", show="headings")
        ys = ttk.Scrollbar(parent, orient="vertical",   command=tree.yview)
        xs = ttk.Scrollbar(parent, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        tree.grid(row=tree_row, column=0, sticky="nsew")
        ys.grid(row=tree_row, column=1, sticky="ns")
        xs.grid(row=tree_row+1, column=0, sticky="ew")
        self._apply_tags(tree)
        # Registrar tree en lista global para zoom
        if not hasattr(self, "_all_trees"):
            self._all_trees = []
        self._all_trees.append(tree)

        # ── Auto-fit al doble clic en separador de heading (como Excel) ───
        def _autofit_col(event, t=tree):
            """Detecta doble clic en el borde derecho de un heading y ajusta el ancho."""
            region = t.identify_region(event.x, event.y)
            if region != "separator":
                return
            col = t.identify_column(event.x)
            if not col:
                return
            col_id = t["columns"][int(col[1:]) - 1]
            # Medir ancho del heading
            heading_text = t.heading(col_id, "text").replace("\n", " ")
            head_w = int(len(heading_text) * 7.5) + 20
            # Medir ancho máximo de las celdas de datos
            max_data_w = 0
            for iid in t.get_children():
                val = t.set(iid, col_id)
                max_data_w = max(max_data_w, int(len(str(val)) * 7.2) + 12)
            new_w = max(head_w, max_data_w, 44)
            t.column(col_id, width=new_w)

        tree.bind("<Double-Button-1>", _autofit_col)
        return tree

    def _apply_zoom(self):
        """Aplica el tamaño de fuente actual a todos los trees registrados."""
        sz = self._zoom_size
        rh = max(16, sz + 13)   # rowheight proporcional al font size
        s  = ttk.Style()
        s.configure("Dark.Treeview",
                    font=("Segoe UI", sz),
                    rowheight=rh)
        s.configure("Dark.Treeview.Heading",
                    font=("Segoe UI", max(8, sz - 1), "bold"),
                    padding=(6, max(6, sz - 3)))
        # Refrescar tags en todos los trees (total/bold mantiene relación)
        for tree in getattr(self, "_all_trees", []):
            try:
                tree.tag_configure("total",
                    font=("Segoe UI", sz, "bold"))
                tree.tag_configure("wh_hdr",
                    font=("Segoe UI", sz, "bold"))
                tree.tag_configure("wh_sub",
                    font=("Segoe UI", max(8, sz-1), "bold"))
            except Exception:
                pass

    def _apply_tags(self, tree):
        tree.tag_configure("avail",    background="#0F2744", foreground="#E5E7EB")
        tree.tag_configure("sample",   background="#0F2D20", foreground="#D1FAE5")
        tree.tag_configure("wh_hdr",   background="#1E3A5F", foreground="#93C5FD",
                           font=("Segoe UI",11,"bold"))
        tree.tag_configure("wh_sub",   background="#172A45", foreground="#FACC15",
                           font=("Segoe UI",10,"bold"))
        tree.tag_configure("diff_ok",  background="#0F2D20", foreground="#D1FAE5")
        tree.tag_configure("diff_bad", background="#3B1010", foreground="#FEE2E2")
        tree.tag_configure("even",     background="#141E2E", foreground="#E5E7EB")
        tree.tag_configure("odd",      background=THEME["panel"],  foreground="#E5E7EB")
        tree.tag_configure("total",    background="#1a3055", foreground="#FACC15",
                           font=("Segoe UI",11,"bold"))
        tree.tag_configure("alert",    background="#3B1010", foreground="#FCA5A5")
        tree.tag_configure("warn",     background="#3B2A0A", foreground="#FCD34D")
        tree.tag_configure("ok",       background="#0F2D20", foreground="#6EE7B7")

    # ── Date helpers ──────────────────────────────────────────────────────────
    def _set_cutoff(self, date_str: str):
        self.cutoff_entry.delete(0, "end")
        self.cutoff_entry.insert(0, date_str)

    def _set_entry(self, entry: ctk.CTkEntry, date_str: str):
        entry.delete(0, "end")
        entry.insert(0, date_str)

    def _parse_date(self, s: str) -> pd.Timestamp:
        s = s.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try: return pd.to_datetime(s, format=fmt)
            except: pass
        raise ValueError(f"Formato de fecha no reconocido: {s}")

    # ── Config ────────────────────────────────────────────────────────────────
    def _apply_saved_config(self):
        self.warehouse_mode.set(self.config_state.get("warehouse_mode","Todas"))
        self.cutoff_entry.delete(0,"end")
        lc = self.config_state.get("last_cutoff","")
        if lc:
            # Convertir a dd/mm/yyyy si viene en formato antiguo
            try:
                dt = pd.to_datetime(lc)
                lc = dt.strftime("%d/%m/%Y")
            except: pass
            self.cutoff_entry.insert(0, lc)

    def log(self, msg: str):
        self.logbox.insert("end", f"[{now_str()}] {msg}\n")
        self.logbox.see("end")

    def persist_config(self):
        data = {
            "warehouse_mode":      self.warehouse_mode.get(),
            "selected_warehouses": self.get_selected_warehouses(),
            "excluded_skus":       list(self._get_excluded_skus()),
            "last_cutoff":         self.cutoff_entry.get().strip(),
        }
        save_config(data)
        self.config_state = data
        self.log("Configuracion guardada.")

    def _get_excluded_skus(self) -> set:
        return {sku for sku,var in self.sku_excl_vars.items() if var.get()}

    # ── SKU Selector ──────────────────────────────────────────────────────────
    def open_sku_selector(self):
        if self.engine.raw_df is None:
            messagebox.showinfo("SKU Excluidos","Primero cargue el archivo de inventario."); return
        self._open_checklist_window(
            title="Seleccionar SKU a excluir",
            instruction="Marque los SKU que desea EXCLUIR del analisis:",
            vars_dict=self.sku_excl_vars,
            items=self._get_sku_items(),
            label_widget=self.sku_label,
            label_fn=lambda n: f"{n} SKU excluido(s)" if n else "Ninguno seleccionado",
            on_apply=lambda: self._on_sku_apply())

    def _get_sku_items(self) -> list[tuple[str,str]]:
        df = self.engine.raw_df
        if df is None: return []
        return [(str(r["Código Producto"]), str(r["Nombre Producto"]))
                for _,r in df[["Código Producto","Nombre Producto"]]
                .drop_duplicates().sort_values("Nombre Producto").iterrows()]

    def _on_sku_apply(self):
        self.engine.excluded_skus = self._get_excluded_skus()
        n = len(self.engine.excluded_skus)
        self.sku_label.configure(text=f"{n} SKU excluido(s)" if n else "Ninguno seleccionado")
        if self.current_result is not None: self.run_analysis(silent=True)

    # ── Warehouse Selector ────────────────────────────────────────────────────
    def open_warehouse_selector(self):
        if not self.wh_vars:
            messagebox.showinfo("Bodegas","Primero cargue el archivo de inventario."); return
        self._open_checklist_window(
            title="Bodegas en el Reporte",
            instruction="Seleccione las bodegas a incluir en el reporte:",
            vars_dict=self.wh_vars,
            items=[(wh, "") for wh in sorted(self.wh_vars.keys())],
            label_widget=self.bodegas_label,
            label_fn=lambda n: f"{n} bodega(s) seleccionada(s)" if n < len(self.wh_vars) else "Todas incluidas",
            on_apply=lambda: self._schedule_recalc())

    def _open_checklist_window(self, title, instruction, vars_dict, items,
                                label_widget, label_fn, on_apply):
        win = ctk.CTkToplevel(self)
        win.title(title)
        win.geometry("500x580")
        win.configure(fg_color=THEME["bg"])
        win.grab_set()

        ctk.CTkLabel(win, text=instruction, font=FONT_BOLD,
                     text_color=THEME["muted"]).pack(anchor="w", padx=14, pady=(12,4))

        search_var = tk.StringVar()
        ctk.CTkEntry(win, textvariable=search_var,
                     placeholder_text="Buscar...", height=30, font=FONT_BODY).pack(
            fill="x", padx=14, pady=(0,4))

        scroll = ctk.CTkScrollableFrame(win, fg_color=THEME["panel"], corner_radius=8)
        scroll.pack(fill="both", expand=True, padx=14, pady=(0,4))

        check_refs: list[ctk.CTkCheckBox] = []

        def render(ft=""):
            for w in scroll.winfo_children(): w.destroy()
            check_refs.clear()
            fl = ft.lower()
            for key, label in items:
                if fl and fl not in key.lower() and fl not in label.lower(): continue
                if key not in vars_dict:
                    vars_dict[key] = ctk.BooleanVar(value=True)
                display = f"{key}  —  {label[:40]}" if label else key
                chk = ctk.CTkCheckBox(scroll, text=display, variable=vars_dict[key],
                                      font=FONT_SMALL, height=22)
                chk.pack(anchor="w", padx=6, pady=1)
                check_refs.append(chk)

        render()
        search_var.trace_add("write", lambda *_: render(search_var.get()))

        def select_all():
            for key,_ in items: vars_dict.get(key) and vars_dict[key].set(True)
        def clear_all():
            for v in vars_dict.values(): v.set(False)

        br = ctk.CTkFrame(win, fg_color="transparent")
        br.pack(fill="x", padx=14, pady=(0,10))
        ctk.CTkButton(br, text="Marcar todos",  width=120, height=28,
                      command=select_all, font=FONT_SMALL).pack(side="left", padx=(0,6))
        ctk.CTkButton(br, text="Limpiar",       width=100, height=28,
                      command=clear_all, font=FONT_SMALL,
                      fg_color=THEME["danger"], hover_color="#B91C1C").pack(side="left")

        def apply_close():
            n = sum(1 for v in vars_dict.values() if v.get())
            label_widget.configure(text=label_fn(n))
            on_apply()
            win.destroy()

        ctk.CTkButton(br, text="Aplicar y cerrar", height=28,
                      command=apply_close,
                      fg_color=THEME["accent"], hover_color="#2563EB",
                      font=FONT_BOLD).pack(side="right")

    # ── Warehouses state ──────────────────────────────────────────────────────
    def refresh_warehouses(self, warehouses: list[str]):
        selected = set(self.config_state.get("selected_warehouses", warehouses))
        for wh in warehouses:
            if wh not in self.wh_vars:
                self.wh_vars[wh] = ctk.BooleanVar(value=(wh in selected))
        # Limpiar vars de bodegas que ya no existen
        for wh in list(self.wh_vars.keys()):
            if wh not in warehouses: del self.wh_vars[wh]
        n = sum(1 for v in self.wh_vars.values() if v.get())
        total = len(self.wh_vars)
        self.bodegas_label.configure(
            text="Todas incluidas" if n == total else f"{n} bodega(s) seleccionada(s)")

    def get_selected_warehouses(self) -> list[str]:
        return [wh for wh,var in self.wh_vars.items() if var.get()]

    def _on_wh_mode_change(self, _=None): self._schedule_recalc()

    def _bind_date_format(self, entry):
        """Agrega auto-inserción de '/' al escribir fechas en un CTkEntry."""
        def _fmt(event):
            cur = entry.get()
            # Solo procesar teclas que produzcan caracteres numéricos
            if event.keysym in ("BackSpace","Delete","Left","Right","Home","End","Tab"):
                return
            digits = "".join(c for c in cur if c.isdigit())
            if len(digits) > 8:
                digits = digits[:8]
            formatted = digits
            if len(digits) >= 3:
                formatted = digits[:2] + "/" + digits[2:]
            if len(digits) >= 5:
                formatted = digits[:2] + "/" + digits[2:4] + "/" + digits[4:]
            if cur != formatted:
                pos = entry.index("insert")
                entry.delete(0, "end")
                entry.insert(0, formatted)
                # Mantener cursor al final si está escribiendo
                try: entry.icursor(len(formatted))
                except: pass
        try:
            entry._entry.bind("<KeyRelease>", _fmt, add="+")
        except Exception:
            pass

    def _on_closing(self):
        """Cierre limpio: cancela callbacks pendientes antes de destruir la ventana."""
        if self._after_id is not None:
            try:
                self.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None
        self._pending_recalc = False
        self.quit()
        self.destroy()

    def _schedule_recalc(self):
        if self.current_result is not None and not self._pending_recalc:
            self._pending_recalc = True
            self._after_id = self.after(350, self._do_recalc)

    def _do_recalc(self):
        self._after_id = None
        self._pending_recalc = False
        if self.engine.raw_df is not None: self.run_analysis(silent=True)

    # ── Load ──────────────────────────────────────────────────────────────────
    def import_toma_template(self):
        """Importa conteos desde la plantilla Excel generada por Plantilla Toma.

        Lee cada hoja de ubicación, extrae Código y Cantidad.
        Abre la ventana de Toma Física y carga los datos directamente.
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel","*.xlsx")],
            title="Seleccione la plantilla de toma física completada")
        if not path:
            return
        try:
            import openpyxl as _opx
            wb  = _opx.load_workbook(path, data_only=True)
            SKIP = {"INSTRUCCIONES","RESUMEN GENERAL"}
            imported_data = {}   # {loc: {cod: {qty, obs}}}
            toma_date = ""

            for sname in wb.sheetnames:
                if sname.strip().upper() in SKIP:
                    continue
                ws2 = wb[sname]
                loc = sname.strip()
                loc_data = {}

                # Detectar fecha de la hoja (celda B2)
                fecha_cell = ws2["B2"].value
                if fecha_cell and not toma_date:
                    toma_date = str(fecha_cell).strip()

                # Cabeceras en fila 3
                headers = [str(ws2.cell(3, c).value or "").strip()
                           for c in range(1, ws2.max_column + 1)]
                try:
                    col_cod  = next(i+1 for i,h in enumerate(headers)
                                    if h.lower() in ("código","codigo","code","sku"))
                    col_cant = next(i+1 for i,h in enumerate(headers)
                                    if h.lower() in ("cantidad","qty","conteo"))
                except StopIteration:
                    continue

                col_obs = None
                for i, h in enumerate(headers):
                    if "obs" in h.lower(): col_obs = i+1; break

                for r in range(4, ws2.max_row + 1):
                    cod = str(ws2.cell(r, col_cod).value or "").strip()
                    if not cod or cod in ("None","TOTAL",""):
                        continue
                    try:
                        qty = int(float(str(ws2.cell(r, col_cant).value or 0)))
                    except:
                        qty = 0
                    obs = str(ws2.cell(r, col_obs).value or "") if col_obs else ""
                    if qty > 0 or obs:   # solo importar filas con datos
                        loc_data[cod] = {"qty": qty, "obs": obs}

                if loc_data:
                    imported_data[loc] = loc_data

            if not imported_data:
                messagebox.showinfo("Sin datos",
                    "No se encontraron conteos en la plantilla.\n"
                    "Asegúrese de llenar la columna CANTIDAD en cada hoja.")
                return

            # Abrir ventana de toma física con los datos precargados
            skus_df = self.engine.raw_df if self.engine.raw_df is not None else None
            excluded = self._get_excluded_skus()
            from .toma_fisica_module import TomaFisicaWindow
            win = TomaFisicaWindow(self, skus_df, excluded_skus=excluded)

            # Cargar datos importados en la ventana
            total_items = 0
            for loc, loc_data in imported_data.items():
                if loc not in win.locations:
                    win.locations.append(loc)
                    win.data[loc] = {}
                win.data[loc].update(loc_data)
                total_items += len(loc_data)

            if toma_date:
                try:
                    win.toma_date_var.set(toma_date)
                except Exception:
                    pass

            win._refresh_location_list()
            win._load_location(win.locations[0])

            n_locs = len(imported_data)
            messagebox.showinfo("Importado",
                f"Plantilla importada correctamente:\n"
                f"• {n_locs} ubicaciones cargadas\n"
                f"• {total_items} ítems con conteo\n"
                f"• Fecha: {toma_date or 'no especificada'}\n\n"
                f"Revise y guarde la toma desde la ventana de Toma Física.")
            self.log(f"Plantilla importada: {n_locs} ubicaciones, {total_items} ítems")

        except Exception as e:
            self.log(f"Error importar plantilla: {e}")
            messagebox.showerror("Error al importar", str(e))

    def _auto_load(self, path: str):
        """Carga automática del archivo de prueba al iniciar."""
        try:
            df = self.engine.load_inventory_file(path)
            import os as _os
            short = _os.path.basename(path)
            self.base_label.configure(text=f"Base: {short}")
            self.refresh_warehouses(self.engine.get_warehouses())
            excluded_saved = set(self.config_state.get("excluded_skus", []))
            for sku in df["Código Producto"].unique():
                if sku not in self.sku_excl_vars:
                    self.sku_excl_vars[sku] = ctk.BooleanVar(value=(sku in excluded_saved))
            n_excl = sum(1 for v in self.sku_excl_vars.values() if v.get())
            self.sku_label.configure(text=f"{n_excl} SKU excluido(s)" if n_excl else "Ninguno seleccionado")
            if df["Fecha"].notna().any():
                if not self.cutoff_entry.get().strip():
                    self.cutoff_entry.insert(0, df["Fecha"].max().strftime("%d/%m/%Y"))
            self.log(f"[AUTO] Cargado: {short} | {len(df):,} registros")
            if self.cutoff_entry.get().strip():
                self.run_analysis(silent=True)
        except Exception as e:
            self.log(f"[AUTO] No se pudo cargar: {e}")

    def load_inventory(self):
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx *.xls")])
        if not path: return
        try:
            df = self.engine.load_inventory_file(path)
            short = path.replace("\\","/").split("/")[-1]
            self.base_label.configure(text=f"Base: {short}")
            self.refresh_warehouses(self.engine.get_warehouses())
            excluded_saved = set(self.config_state.get("excluded_skus",[]))
            for sku in df["Código Producto"].unique():
                if sku not in self.sku_excl_vars:
                    self.sku_excl_vars[sku] = ctk.BooleanVar(value=(sku in excluded_saved))
            n_excl = sum(1 for v in self.sku_excl_vars.values() if v.get())
            self.sku_label.configure(text=f"{n_excl} SKU excluido(s)" if n_excl else "Ninguno seleccionado")
            if df["Fecha"].notna().any():
                if not self.cutoff_entry.get().strip():
                    self.cutoff_entry.insert(0, df["Fecha"].max().strftime("%d/%m/%Y"))
            self.log(f"Cargado: {short} | {len(df):,} registros")
            messagebox.showinfo("Importacion", f"Archivo importado correctamente.\n{len(df):,} registros cargados.")
            if self.cutoff_entry.get().strip():
                self.run_analysis(silent=True)
                self.log("Analisis ejecutado automaticamente tras la carga del Excel.")
        except Exception as e:
            self.log(f"Error de carga: {e}")
            messagebox.showerror("Error al cargar", str(e))

    def open_toma_fisica(self):
        skus_df = self.engine.raw_df
        excluded = self._get_excluded_skus()
        TomaFisicaWindow(self, skus_df, excluded_skus=excluded)

    # ── Analysis ──────────────────────────────────────────────────────────────
    def run_analysis(self, silent: bool = False):
        try:
            cutoff_raw = self.cutoff_entry.get().strip()
            if not cutoff_raw:
                if not silent: raise ValueError("Ingrese una fecha de corte.")
                return
            cutoff_dt = self._parse_date(cutoff_raw)
            cutoff_str = cutoff_dt.strftime("%Y-%m-%d")
            self.engine.excluded_skus = self._get_excluded_skus()
            result = self.engine.analyze(
                cutoff_date=cutoff_str,
                warehouse_mode=self.warehouse_mode.get(),
                selected_warehouses=self.get_selected_warehouses())
            self.current_result = result
            self.render_result()
            self.persist_config()
            if not silent:
                self.log(f"Analisis OK | corte {cutoff_str} | {len(result.filtered):,} mov.")
        except Exception as e:
            self.log(f"Error: {e}")
            if not silent: messagebox.showerror("Error en analisis", str(e))

    # ── Render ────────────────────────────────────────────────────────────────
    def render_result(self):
        r = self.current_result
        if r is None: return
        kpi_map = {"Rotacion":"Rotación","Dias de inventario":"Días de inventario",
                   "Exactitud inventario":"Exactitud inventario"}
        pct_keys = {"Margen","Exactitud inventario"}
        for k,box in self.kpi_boxes.items():
            val = r.kpis.get(kpi_map.get(k,k), r.kpis.get(k,0))
            box.set_value(fmt_pct(val) if k in pct_keys else fmt_num(val))

        self._fill_inventory_grouped(self.tree_inventory, self._enrich_inventory(r))
        self._fill_sku_split(r.sku_summary)

        samples = r.samples_by_client
        # Actualizar selector de clientes para reporte
        clients = sorted(samples["Cliente"].dropna().astype(str).tolist())
        self.sample_client_menu.configure(values=["— seleccione cliente —"] + clients)
        self.sample_client_var.set("— seleccione cliente —")
        active = samples[samples["Stock en Cliente"]>0].sort_values(
            ["Stock en Cliente","Cliente"], ascending=[False,True])
        self.fill_tree(self.tree_samples_active, active,  tag_mode="zebra", decimals=False)
        self.fill_tree(self.tree_samples_all,    samples, tag_mode="zebra", decimals=False)

        if r.physical_compare is not None:
            self.fill_tree(self.tree_physical, r.physical_compare, tag_mode="physical")
        else:
            self.fill_tree(self.tree_physical, pd.DataFrame(columns=["Estado"]),
                           empty_msg="Sin toma fisica cargada")
        # Actualizar histórico de compras automáticamente
        self.render_purchases()

    def _enrich_inventory(self, r) -> pd.DataFrame:
        inv = r.inventory_by_warehouse.copy()
        if inv.empty: return inv
        df = r.filtered; SKU = "Código Producto"
        out_df = (df[df["is_sample_out"]].groupby([SKU,"Bodega Destino"],as_index=False)["Cantidad"].sum()
                  .rename(columns={"Bodega Destino":"Bodega","Cantidad":"Entregadas"}))
        in_df  = (df[df["is_sample_in"]].groupby([SKU,"Bodega Origen"],as_index=False)["Cantidad"].sum()
                  .rename(columns={"Bodega Origen":"Bodega","Cantidad":"Devueltas"}))
        inv = inv.merge(out_df,on=[SKU,"Bodega"],how="left")
        inv = inv.merge(in_df, on=[SKU,"Bodega"],how="left")
        inv["Entregadas"] = inv["Entregadas"].fillna(0)
        inv["Devueltas"]  = inv["Devueltas"].fillna(0)
        ordered = [SKU,"Nombre Producto",
                   "Valor Unitario Promedio","Valor Stock","Entregadas","Devueltas","Stock"]
        present = [c for c in ordered if c in inv.columns]
        return inv[present + ["Bodega"]]

    def _fill_inventory_grouped(self, tree, df):
        tree.delete(*tree.get_children())
        if df.empty:
            tree["columns"]=["Info"]; tree.heading("Info",text="Info")
            tree.column("Info",anchor="w",width=400)
            tree.insert("","end",values=["Sin datos"]); return

        display_cols = [c for c in df.columns if c != "Bodega"]
        two_line = {
            "Código Producto":"Código\nProducto",
            "Nombre Producto":"Nombre\nProducto",
            "Valor Unitario Promedio":"Valor Unit.\nPromedio",
            "Valor Stock":"Valor\nStock",
            "Entregadas":"Entregadas\n(Muestras)",
            "Devueltas":"Devueltas\n(Muestras)",
            "Stock":"Stock\nActual",
        }

        def _fmt_inventory_value(col, value):
            if pd.isna(value):
                return ""
            if col in {"Entregadas", "Devueltas", "Stock"}:
                try:
                    return f"{int(float(value)):,}"
                except Exception:
                    return str(value)
            if _is_num(col):
                try:
                    return f"{float(value):,.2f}"
                except Exception:
                    return str(value)
            s = str(value)
            return "" if s in ("nan", "None", "NaN") else s

        def _inventory_col_width(col):
            header_lines = two_line.get(col, col).split("\n")
            max_len = max(len(x) for x in header_lines)
            if col == "Código Producto":
                max_len = max(max_len, max((len(str(v)) for v in df[col].fillna("")), default=0), len("TOTAL GENERAL"))
                return max(110, min(165, int(max_len * 10 + 26)))
            if col == "Nombre Producto":
                max_len = max(max_len, max((len(str(v)) for v in df[col].fillna("")), default=0))
                return max(260, min(560, int(max_len * 8 + 24)))
            if col in {"Valor Unitario Promedio", "Valor Stock", "Entregadas", "Devueltas", "Stock"}:
                formatted = [_fmt_inventory_value(col, v) for v in df[col].fillna(0)]
                max_len = max(max_len, max((len(x) for x in formatted), default=0))
                if col in {"Valor Unitario Promedio", "Valor Stock"}:
                    return max(125, min(190, int(max_len * 9 + 26)))
                return max(100, min(135, int(max_len * 9 + 22)))
            return _col_w(col)

        tree["columns"] = display_cols
        for c in display_cols:
            anchor = "e" if _is_num(c) else "w"
            tree.heading(c, text=two_line.get(c,c), anchor=anchor,
                         command=lambda col=c, t=tree, d=df: self._sort_grouped(t,col,d))
            tree.column(c, anchor=anchor, width=_inventory_col_width(c), minwidth=60, stretch=False)

        num_cols = [c for c in display_cols if _is_num(c)]
        grand = {c:0.0 for c in num_cols}
        for bodega in df["Bodega"].unique():
            sub = df[df["Bodega"]==bodega]
            hv = [""]*len(display_cols); hv[0] = f"  ▼  {bodega.upper()}"
            tree.insert("","end",values=hv,tags=("wh_hdr",))
            INT_INV = {"Entregadas","Devueltas","Stock"}
            for i,(_,row) in enumerate(sub.iterrows()):
                vals = []
                for c in display_cols:
                    vals.append(_fmt_inventory_value(c, row[c]))
                tree.insert("","end",values=vals,tags=("even" if i%2==0 else "odd",))
            sub_t = {}
            for c in num_cols:
                v = pd.to_numeric(sub[c],errors="coerce").sum()
                sub_t[c]=v; grand[c]+=v
            INT_INV2 = {"Entregadas","Devueltas","Stock"}
            sv = []
            for i2,c in enumerate(display_cols):
                if i2==0:
                    sv.append(f"  Subtotal — {bodega}")
                elif c in sub_t:
                    sv.append(f"{int(sub_t[c]):,}" if c in INT_INV2 else f"{sub_t[c]:,.2f}")
                else:
                    sv.append("")
            tree.insert("","end",values=sv,tags=("wh_sub",))
        INT_INV3 = {"Entregadas","Devueltas","Stock"}
        gv = []
        for i3,c in enumerate(display_cols):
            if i3==0:
                gv.append("TOTAL GENERAL")
            elif c in grand:
                gv.append(f"{int(grand[c]):,}" if c in INT_INV3 else f"{grand[c]:,.2f}")
            else:
                gv.append("")
        tree.insert("","end",values=gv,tags=("total",))

    def _sort_grouped(self, tree, col, df):
        prev_col,prev_rev = self._sort_state.get(id(tree),(None,False))
        rev = not prev_rev if col==prev_col else False
        self._sort_state[id(tree)] = (col,rev)
        try:
            sdf = df.sort_values(col, ascending=not rev,
                key=lambda s: pd.to_numeric(s,errors="coerce").fillna(0) if _is_num(col)
                              else s.astype(str).str.lower())
        except: sdf = df.sort_values(col, ascending=not rev)
        self._fill_inventory_grouped(tree, sdf)

    def _fill_sku_split(self, df):
        if df.empty:
            self.fill_tree(self.tree_sku_val, df)
            self.fill_tree(self.tree_sku_unit, df)
            return

        id_c = ["Código Producto", "Nombre Producto"]
        id_p = [c for c in id_c if c in df.columns]

        # ── Valores Financieros ─────────────────────────────────────────────
        val_c  = ["Valor_Compras", "Valor_Ventas", "Valor Inventario"]
        df_val = df[[*id_p, *[c for c in val_c if c in df.columns]]].copy()
        df_val = df_val.rename(columns={
            "Valor_Compras":    "Valor Compras (ING)",
            "Valor_Ventas":     "Valor Ventas (EGR)",
            "Valor Inventario": "Valor Inventario ($)",
        })
        self.fill_tree(self.tree_sku_val, df_val, tag_mode="zebra", sortable=True, decimals=True)

        # ── Movimiento de Unidades ──────────────────────────────────────────
        unit_c = ["Compras", "Dev_Proveedor",
                  "Ventas", "Dev_Cliente",
                  "Muestras_Enviadas", "Muestras_Devueltas",
                  "Stock Disponible", "Stock Muestras", "Stock Total"]
        df_unit = df[[*id_p, *[c for c in unit_c if c in df.columns]]].copy()
        self.fill_tree(self.tree_sku_unit, df_unit, tag_mode="zebra", sortable=True, decimals=False)

    # ── Period analysis ───────────────────────────────────────────────────────
    def run_period_analysis(self):
        if self.engine.raw_df is None:
            messagebox.showinfo("Analisis","Cargue primero el archivo."); return
        try:
            d_from = self._parse_date(self.anal_from.get())
            d_to   = self._parse_date(self.anal_to.get())
        except:
            messagebox.showerror("Fechas","Use el formato dd/mm/yyyy."); return
        if d_from > d_to:
            messagebox.showerror("Fechas","La fecha inicial debe ser menor a la final."); return

        df = self.engine.raw_df.copy()
        df = df[(df["Fecha"]>=d_from)&(df["Fecha"]<=d_to)]
        if df.empty: messagebox.showinfo("Sin datos","No hay movimientos en ese periodo."); return

        ref = df["Referencia"].fillna("").astype(str).str.upper()
        typ = df["Tipo"].fillna("").astype(str).str.upper()
        df["is_sale"]     = (typ=="EGR") & ref.str.startswith("FAC")
        df["is_purchase"] = (typ=="ING") & ref.str.startswith("FAC")

        vdf = df[df["is_sale"]].copy()
        cdf = df[df["is_purchase"]].copy()

        tv = vdf["Valor Total"].sum()
        tc = cdf["Valor Total"].sum()
        margen_p = (tv-tc)/tv*100 if tv else 0

        days = max((d_to-d_from).days+1, 1)

        if not vdf.empty:
            rot_by = (vdf.groupby("Código Producto")["Cantidad"].sum()/days)
            item_rot  = rot_by.idxmax()
            item_vend = vdf.groupby("Código Producto")["Cantidad"].sum().idxmax()
            vt = vdf.groupby("Código Producto")["Valor Total"].sum()
            ct = cdf.groupby("Código Producto")["Valor Total"].sum()
            rentab = (vt-ct).dropna()
            item_rent = rentab.idxmax() if not rentab.empty else "—"
        else:
            item_rot=item_vend=item_rent="—"

        self.period_kpis["Item mas rotacion"].set_value(str(item_rot)[:28])
        self.period_kpis["Item mas vendido"].set_value(str(item_vend)[:28])
        self.period_kpis["Item mas rentable"].set_value(str(item_rent)[:28])
        # Ventas periodo y Margen periodo eliminados de KPIs (mostrar solo en tabla)

        # ── Costo Promedio Ponderado — cálculo cronológico por SKU ──────
        # Algoritmo correcto (PEPS/Promedio móvil):
        #   1. Se usa TODO el histórico de compras (no solo el periodo) ordenado por fecha
        #   2. Por cada compra: costo_unit = Valor Total / Cantidad
        #   3. Primera compra del SKU: costo_prom = costo_unit
        #   4. Compras siguientes: costo_prom = (qty_nueva*costo_nueva + qty_ant*costo_ant)
        #                                        / (qty_nueva + qty_ant)   ← promedio ponderado REAL
        #   5. Fallback si no hay compra: usa Valor Unitario del propio movimiento de venta
        #   6. Costo 0 imposible: si Valor Unitario = 0 y hay ventas, se toma
        #      el PVP del movimiento × 0.6 como estimado de costo

        df_hist_all = self.engine.raw_df.copy()
        h_ref = df_hist_all["Referencia"].fillna("").astype(str).str.upper()
        h_typ = df_hist_all["Tipo"].fillna("").astype(str).str.upper()
        df_hist_all["is_purchase"] = (h_typ=="ING") & h_ref.str.startswith("FAC")
        purchases_all = (df_hist_all[df_hist_all["is_purchase"]]
                         .sort_values(["Código Producto","Fecha"])
                         .copy())

        # Construir mapa costo_prom acumulado hasta cada fecha
        # costo_prom_timeline[sku] = [(fecha, costo_prom)] — cronológico
        from collections import defaultdict
        cpm_qty   = defaultdict(float)   # cantidad acumulada por SKU
        cpm_valor = defaultdict(float)   # valor acumulado por SKU
        costo_prom_map = {}              # SKU → costo promedio final (hasta d_to)

        # También guardamos costo promedio por (SKU, mes) para el gráfico mensual
        # Procesamos compra a compra y anotamos el costo vigente al cierre de cada mes
        cp_by_date = defaultdict(dict)   # {fecha: {sku: costo_prom}}

        for _, row in purchases_all.iterrows():
            sku  = row["Código Producto"]
            qty  = float(row["Cantidad"]) if float(row["Cantidad"]) > 0 else 1
            vtot = float(row["Valor Total"])
            # Ignorar compras sin valor (solo si también sin cantidad)
            if vtot == 0 and qty == 0:
                continue
            costo_u = vtot / qty if qty > 0 else 0
            if costo_u == 0:
                # Intentar recuperar del campo Valor Unitario
                costo_u = float(row.get("Valor Unitario", 0) or 0)
            if costo_u == 0:
                continue   # no contaminar con costo 0

            # Promedio ponderado: (qty_ant * cp_ant + qty_nueva * c_nueva) / (qty_ant + qty_nueva)
            old_qty   = cpm_qty[sku]
            old_valor = cpm_valor[sku]
            new_qty   = old_qty + qty
            new_valor = old_valor + vtot
            cpm_qty[sku]   = new_qty
            cpm_valor[sku] = new_valor
            cp_current     = new_valor / new_qty
            costo_prom_map[sku] = cp_current

        def get_costo(sku, row_sale=None):
            """Devuelve costo promedio; nunca 0."""
            cp = costo_prom_map.get(sku, 0.0)
            if cp > 0:
                return cp
            # Fallback 1: Valor Unitario del propio movimiento
            if row_sale is not None:
                vu = float(row_sale.get("Valor Unitario", 0) or 0)
                if vu > 0:
                    return vu
                # Fallback 2: PVP * 0.6
                pvp = float(row_sale.get("PVP", 0) or 0)
                if pvp > 0:
                    return pvp * 0.6
            return 0.0  # realmente sin datos

        # Ventas mensuales con costo correcto por fila de venta
        vdf["Mes"] = vdf["Fecha"].dt.to_period("M").astype(str)
        monthly_rows = []
        for mes, grp in vdf.groupby("Mes"):
            unidades = int(grp["Cantidad"].sum())
            ventas   = float(grp["Valor Total"].sum())
            costo_m  = 0.0
            for _, rv in grp.iterrows():
                sku = rv["Código Producto"]
                qty = float(rv["Cantidad"])
                cp  = get_costo(sku, rv)
                costo_m += qty * cp
            margen = (ventas - costo_m) / ventas * 100 if ventas else 0.0
            monthly_rows.append({
                "Mes": mes, "Unidades": unidades,
                "Ventas(Egreso)": round(ventas, 2),
                "Costo Prom.": round(costo_m, 2),
                "Margen %": round(margen, 2),
            })
        monthly = pd.DataFrame(monthly_rows)
        self.fill_tree(self.tree_monthly, monthly, tag_mode="zebra")
        self._plot_line_monthly(monthly.rename(columns={"Ventas(Egreso)":"Ventas","Costo Prom.":"Costo"}))

        # Top 10 vendidos
        top10v = (vdf.groupby(["Código Producto","Nombre Producto"])
            .agg(Unidades=("Cantidad","sum"),Ventas=("Valor Total","sum"))
            .reset_index().nlargest(10,"Ventas"))
        self.fill_tree(self.tree_top10v, top10v, tag_mode="zebra")
        self._plot_bar_h(self.ax_top10v, self.canvas_top10v,
                         top10v["Código Producto"], top10v["Ventas"],
                         "Top 10 por Ventas ($)", THEME["accent"])

        # Top 10 rentabilidad con costo promedio ponderado por fila
        t10r_rows = []
        for (sku, nom), grp in vdf.groupby(["Código Producto","Nombre Producto"]):
            ventas_sku = float(grp["Valor Total"].sum())
            costo_sku  = sum(float(rv["Cantidad"]) * get_costo(sku, rv)
                             for _, rv in grp.iterrows())
            rent       = ventas_sku - costo_sku
            margen_sku = rent / ventas_sku * 100 if ventas_sku else 0.0
            t10r_rows.append({
                "Código Producto": sku, "Nombre Producto": nom,
                "Ventas": round(ventas_sku,2), "Costo Prom.": round(costo_sku,2),
                "Rentabilidad": round(rent,2), "Margen %": round(margen_sku,2),
            })
        t10r = (pd.DataFrame(t10r_rows)
                .nlargest(10,"Rentabilidad")
                .reset_index(drop=True))
        self.fill_tree(self.tree_top10r, t10r, tag_mode="zebra")
        self._plot_bar_h(self.ax_top10r, self.canvas_top10r,
                         t10r["Código Producto"], t10r["Rentabilidad"],
                         "Top 10 por Rentabilidad ($)", THEME["success"])

        mg = t10r[["Código Producto","Nombre Producto","Ventas","Costo Prom.","Rentabilidad","Margen %"]].copy()
        self.fill_tree(self.tree_margin, mg, tag_mode="zebra")
        self._plot_bar_h(self.ax_margin, self.canvas_margin,
                         mg["Código Producto"], mg["Margen %"],
                         "Margen % (Top 10)", THEME["warning"])

        self.log(f"Periodo {d_from.date()} → {d_to.date()} | ventas ${tv:,.0f} | margen {margen_p:.1f}%")

    def _plot_line_monthly(self, df):
        ax = self.ax_monthly; fig = self.fig_monthly
        ax.clear()
        with plt.rc_context(MPL_STYLE):
            x = range(len(df))
            ax.plot(list(x), df["Ventas"].tolist(), color=THEME["accent"],
                    marker="o", linewidth=2, label="Ventas", zorder=3)
            ax.plot(list(x), df["Costo"].tolist(), color=THEME["danger"],
                    marker="s", linewidth=2, label="Costo", zorder=3)
            for xi, (v, c) in enumerate(zip(df["Ventas"], df["Costo"])):
                ax.annotate(f"${v:,.0f}", (xi,v), textcoords="offset points",
                            xytext=(0,6), ha="center", fontsize=7, color=THEME["accent"])
                ax.annotate(f"${c:,.0f}", (xi,c), textcoords="offset points",
                            xytext=(0,-12), ha="center", fontsize=7, color=THEME["danger"])
            ax.set_xticks(list(x))
            ax.set_xticklabels(df["Mes"].tolist(), rotation=45, ha="right", fontsize=8)
            ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"${v:,.0f}"))
            ax.legend(fontsize=8); ax.set_title("Ventas vs Costo Mensual", fontsize=10)
            ax.grid(axis="y", zorder=0); fig.tight_layout(pad=1.5)
        self.canvas_monthly.draw()

    def _plot_bar_h(self, ax, canvas, labels, values, title, color):
        ax.clear()
        with plt.rc_context(MPL_STYLE):
            y = range(len(labels))
            ax.barh(list(y), list(values), color=color, alpha=0.85, zorder=3)
            ax.set_yticks(list(y))
            ax.set_yticklabels([str(l) for l in labels], fontsize=9)
            ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda v,_: f"${v:,.0f}"))
            ax.set_title(title, fontsize=10); ax.grid(axis="x", zorder=0)
            ax.figure.tight_layout(pad=1.5)
        canvas.draw()

    # ── Rotation ──────────────────────────────────────────────────────────────
    def run_rotation_analysis(self):
        """Calcula rotación e inventario sugerido para dos escenarios de Lead Time.

        Fórmulas exactas:
          consumo_dia  = ventas_unidades / días_período
          rotación     = ventas_unidades / stock_disponible
          días_inv     = stock_disponible / consumo_dia
          punto_reorden= consumo_dia × lead_time          ← cuándo ordenar
          stock_seg_u  = consumo_dia × safety_days        ← colchón de seguridad
          sugerido     = (consumo_dia × (lead_time + safety_days)) - stock_disponible
                       = máx(0, lo anterior)              ← nunca negativo
        """
        r = self.current_result
        if r is None:
            messagebox.showinfo("Rotacion","Ejecute primero el analisis."); return
        try: safety = max(0, int(self.safety_days.get().strip() or "15"))
        except: safety = 15
        try: lt_mar = max(1, int(self.lt_mar_days.get().strip() or "45"))
        except: lt_mar = 45
        try: lt_air = max(1, int(self.lt_air_days.get().strip() or "15"))
        except: lt_air = 15

        df  = r.filtered
        sku = r.sku_summary.copy()
        if sku.empty: return

        # Días reales del período analizado (fecha mín a máx del filtered)
        days = max(int((df["Fecha"].max() - df["Fecha"].min()).days) + 1, 1)

        rows_mar, rows_air = [], []

        for _, row in sku.iterrows():
            cod      = str(row["Código Producto"])
            name     = str(row["Nombre Producto"])
            stock_d  = max(0.0, float(row.get("Stock Disponible", 0)))
            ventas_u = float(row.get("Ventas", 0))

            # consumo/día = ventas totales del período / días del período
            consumo = ventas_u / days

            # rotación = veces que el stock se renueva en el período
            rot = ventas_u / stock_d if stock_d > 0 else (float("inf") if ventas_u > 0 else 0.0)

            # días de inventario = días que dura el stock actual al ritmo actual
            dias_inv = stock_d / consumo if consumo > 0 else (0.0 if stock_d == 0 else float("inf"))

            def alert_level(lt):
                if consumo > 0 and stock_d == 0:  return "SIN STOCK"
                if consumo > 0 and dias_inv < lt:  return "CRÍTICO"
                if consumo > 0 and dias_inv < lt + safety: return "BAJO"
                if consumo == 0:                   return "SIN VENTA"
                return "OK"

            def sugerido(lt):
                """Unidades para cubrir lead_time + stock_seguridad menos lo que hay."""
                return max(0.0, consumo * (lt + safety) - stock_d)

            for rows, lt, scenario in [(rows_mar, lt_mar, "mar"), (rows_air, lt_air, "air")]:
                sug  = sugerido(lt)
                alrt = alert_level(lt)
                pr   = consumo * lt   # punto de reorden en unidades

                rows.append({
                    "Sel":           "",
                    "Código":        cod,
                    "Nombre":        name,
                    "Stock Disp.":   int(stock_d),
                    "Ventas (u)":    int(ventas_u),
                    "Consumo/día":   round(consumo, 3),
                    "P.Reorden(u)":  round(pr, 1),
                    "Días Inv.":     round(dias_inv, 1) if dias_inv != float("inf") else "∞",
                    "Rotación(x)":   round(rot, 2) if rot != float("inf") else "∞",
                    "Sug.Compra":    int(round(sug)),
                    "Estado":        alrt,
                })

        order = {"SIN STOCK":0,"CRÍTICO":1,"BAJO":2,"OK":3,"SIN VENTA":4}
        for rows in (rows_mar, rows_air):
            rows.sort(key=lambda r: (order.get(r["Estado"], 5), -r["Ventas (u)"]))

        df_mar = pd.DataFrame(rows_mar)
        df_air = pd.DataFrame(rows_air)

        self._fill_rotation_table(self.tree_rotation_mar, df_mar, scenario="mar")
        self._fill_rotation_table(self.tree_rotation_air, df_air, scenario="air")
        self._rotation_df = df_mar   # para PDF usa el marítimo por defecto
        self.log(
            f"Rotación | {len(df_mar)} SKU | período {days}d | "
            f"LT mar={lt_mar}d air={lt_air}d | seg={safety}d"
        )

    def _fill_rotation_table(self, tree, df: pd.DataFrame, scenario: str = "mar"):
        """Renderiza una tabla de rotación (marítimo o aéreo)."""
        tree.delete(*tree.get_children())
        if df.empty: return

        data_cols = [c for c in df.columns if c != "Sel"]
        cols = ["Sel"] + data_cols
        tree["columns"] = cols

        labels = {
            "Sel":          "✓",
            "Código":       "Código",
            "Nombre":       "Nombre",
            "Stock Disp.":  "Stock\nDisp.",
            "Ventas (u)":   "Ventas\n(u)",
            "Consumo/día":  "Consumo\n/día",
            "P.Reorden(u)": "P.Reorden\n(u)",
            "Días Inv.":    "Días\nInv.",
            "Rotación(x)":  "Rotación\n(x)",
            "Sug.Compra":   "Sug.\nCompra",
            "Estado":       "Estado",
        }
        widths = {
            "Sel":10, "Código":90, "Nombre":200, "Stock Disp.":70,
            "Ventas (u)":70, "Consumo/día":78, "P.Reorden(u)":80,
            "Días Inv.":65, "Rotación(x)":72, "Sug.Compra":78, "Estado":80,
        }
        for c in cols:
            anchor = "center" if c == "Sel" else ("w" if c in ("Código","Nombre","Estado") else "e")
            tree.heading(c, text=labels.get(c, c), anchor=anchor)
            tree.column(c, anchor=anchor, width=widths.get(c, 80), minwidth=36, stretch=False)

        if scenario == "mar":
            self._buy_vars.clear()

        for _, row in df.iterrows():
            cod    = str(row.get("Código", ""))
            estado = str(row.get("Estado", ""))
            pre_sel = estado in ("SIN STOCK", "CRÍTICO")
            if scenario == "mar":
                self._buy_vars[cod] = ctk.BooleanVar(value=pre_sel)

            vals = ["☑" if pre_sel else "☐"]
            for c in data_cols:
                v = row[c]
                vals.append(str(v))

            tag = {"SIN STOCK":"alert","CRÍTICO":"alert","BAJO":"warn",
                   "OK":"ok","SIN VENTA":"even"}.get(estado, "even")
            tree.insert("", "end", values=vals, tags=(tag,))

        tree.bind("<ButtonRelease-1>", self._on_rotation_click)

    def _fill_rotation(self, df):
        """Alias legacy."""
        self._fill_rotation_table(self.tree_rotation_mar, df)

    def _on_rotation_click(self, event):
        """Toggle checkbox al hacer clic en la columna Sel."""
        tree = self.tree_rotation
        region = tree.identify("region", event.x, event.y)
        if region != "cell": return
        col = tree.identify_column(event.x)
        if col != "#1": return  # Solo columna Sel
        iid = tree.identify_row(event.y)
        if not iid: return
        vals = list(tree.item(iid, "values"))
        # El código está en la columna índice 1 (después de Sel)
        if len(vals) < 2: return
        cod = vals[1]  # columna Código
        if cod in self._buy_vars:
            new_val = not self._buy_vars[cod].get()
            self._buy_vars[cod].set(new_val)
            vals[0] = "☑" if new_val else "☐"
            tree.item(iid, values=vals)

    def _select_all_critical(self):
        """Selecciona todos los items CRITICO y SIN STOCK."""
        if self._rotation_df is None: return
        tree = self.tree_rotation
        for iid in tree.get_children():
            vals = list(tree.item(iid, "values"))
            if len(vals) < 2: continue
            cod    = vals[1]
            estado = vals[-1] if vals else ""
            # Buscar estado en df
            row = self._rotation_df[self._rotation_df["Código"]==cod]
            if not row.empty:
                estado = str(row.iloc[0].get("Estado",""))
            if estado in ("CRITICO","SIN STOCK","BAJO"):
                if cod in self._buy_vars:
                    self._buy_vars[cod].set(True)
                    vals[0] = "☑"
                    tree.item(iid, values=vals)

    def _generate_purchase_pdf(self):
        """Genera PDF de solicitud de compra con los items seleccionados."""
        if self._rotation_df is None or not self._buy_vars:
            messagebox.showinfo("Solicitud","Primero calcule la rotacion."); return

        selected = [cod for cod,var in self._buy_vars.items() if var.get()]
        if not selected:
            messagebox.showinfo("Sin seleccion","Seleccione al menos un item."); return

        df_sel = self._rotation_df[self._rotation_df["Código"].isin(selected)].copy()

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                            Paragraph, Spacer, HRFlowable)
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER
            HAS_PDF = True
        except ImportError:
            HAS_PDF = False

        if not HAS_PDF:
            # Fallback: exportar Excel
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
                initialfile="solicitud_compra.xlsx")
            if not path: return
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active; ws.title = "Solicitud Compra"
            for ci,col in enumerate(df_sel.columns,1):
                ws.cell(1,ci,col)
            for ri,(_,row) in enumerate(df_sel.iterrows(),2):
                for ci,v in enumerate(row,1):
                    ws.cell(ri,ci,str(v))
            wb.save(path)
            messagebox.showinfo("Exportado", f"Solicitud exportada:\n{path}"); return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile="solicitud_compra.pdf")
        if not path: return

        import datetime
        today = datetime.date.today().strftime("%d/%m/%Y")
        use_mar = self.lt_maritime.get()
        use_air = self.lt_air.get()

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=2*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("title", parent=styles["Title"],
                                     fontSize=16, textColor=colors.HexColor("#1E3A5F"))
        sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                                     fontSize=10, textColor=colors.HexColor("#6B7280"))
        elems = []

        elems.append(Paragraph("SOLICITUD DE COMPRA — REPOSICIÓN DE INVENTARIO", title_style))
        elems.append(Paragraph(f"Fecha: {today}  |  Items seleccionados: {len(df_sel)}  |  "
                               f"Modalidad: {'Marítimo (45d) ' if use_mar else ''}"
                               f"{'/ Aéreo (15d)' if use_air else ''}", sub_style))
        elems.append(Spacer(1, 0.5*cm))
        elems.append(HRFlowable(width="100%", thickness=1,
                                color=colors.HexColor("#1E3A5F")))
        elems.append(Spacer(1, 0.3*cm))

        # Columnas del PDF
        pdf_cols = ["Codigo","Nombre","Estado","Stock Disp.","Ventas (u)",
                    "Consumo/dia","Dias Inv."]
        if use_mar: pdf_cols.append("Sug. Maritimo")
        if use_air:  pdf_cols.append("Sug. Aereo")

        col_map = {
            "Codigo":"Codigo","Nombre":"Nombre","Estado":"Estado",
            "Stock Disp.":"Stock Disp.","Ventas (u)":"Ventas (u)",
            "Consumo/dia":"Consumo/dia","Dias Inv.":"Dias Inv.",
            "Sug. Maritimo":"Sug. Maritimo","Sug. Aereo":"Sug. Aereo",
        }

        data_rows = [pdf_cols]
        for _, row in df_sel.iterrows():
            r = [str(row.get(col_map.get(c,c),"")) for c in pdf_cols]
            data_rows.append(r)

        # Fila de totales para sugeridos
        if len(df_sel) > 0:
            total_row = ["","","TOTAL","","","",""]
            if use_mar:
                try:
                    total_m = pd.to_numeric(df_sel["Sug. Maritimo"],errors="coerce").sum()
                    total_row.append(f"{int(total_m):,}")
                except: total_row.append("")
            if use_air:
                try:
                    total_a = pd.to_numeric(df_sel["Sug. Aereo"],errors="coerce").sum()
                    total_row.append(f"{int(total_a):,}")
                except: total_row.append("")
            data_rows.append(total_row)

        n = len(pdf_cols)
        col_widths = [2.5*cm, 7*cm, 2*cm] + [2*cm]*(n-3)

        tbl = Table(data_rows, colWidths=col_widths, repeatRows=1)
        estado_colors = {
            "SIN STOCK": colors.HexColor("#3B1010"),
            "CRITICO":   colors.HexColor("#3B1010"),
            "BAJO":      colors.HexColor("#3B2A0A"),
            "OK":        colors.HexColor("#0F2D20"),
        }
        style_cmds = [
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("ALIGN",      (0,1), (1,-1),  "LEFT"),
            ("ROWBACKGROUNDS", (0,1),(-2,-1),
             [colors.HexColor("#F8FAFC"), colors.HexColor("#EFF6FF")]),
            ("GRID",       (0,0),(-1,-1), 0.5, colors.HexColor("#CBD5E1")),
            ("FONTNAME",   (0,-1),(-1,-1), "Helvetica-Bold"),
            ("BACKGROUND", (0,-1),(-1,-1), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR",  (0,-1),(-1,-1), colors.HexColor("#FACC15")),
            ("ROWHEIGHT",  (0,0), (-1,0), 30),
        ]
        # Colorear filas por estado
        for ri, (_, row) in enumerate(df_sel.iterrows(), 1):
            estado = str(row.get("Estado",""))
            bg = estado_colors.get(estado, colors.HexColor("#F8FAFC"))
            tc = colors.HexColor("#FEE2E2") if estado in ("SIN STOCK","CRITICO") else                  colors.HexColor("#FEF3C7") if estado=="BAJO" else colors.black
            style_cmds.append(("BACKGROUND",(0,ri),(-1,ri),bg))
            style_cmds.append(("TEXTCOLOR",  (0,ri),(-1,ri),tc))

        tbl.setStyle(TableStyle(style_cmds))
        elems.append(tbl)

        # Leyenda de fórmulas al final
        elems.append(Spacer(1, 0.8*cm))
        elems.append(HRFlowable(width="100%", thickness=0.5,
                                color=colors.HexColor("#CBD5E1")))
        legend_style = ParagraphStyle("legend", parent=styles["Normal"],
                                      fontSize=7, textColor=colors.HexColor("#6B7280"))
        formulas_text = (
            "<b>Fórmulas de cálculo:</b>  "
            "Consumo/día = Ventas (u) ÷ Días del período  |  "
            "Rotación = Ventas (u) ÷ Stock Disponible  |  "
            "Días Inventario = Stock Disp. ÷ Consumo/día  |  "
            "Sugerido = (Lead Time + Stock Seg.) × Consumo/día − Stock Disp.  |  "
            "Lead Time Marítimo: 45d · Aéreo: 15d"
        )
        elems.append(Spacer(1, 0.2*cm))
        elems.append(Paragraph(formulas_text, legend_style))

        doc.build(elems)
        n_items = len(df_sel)
        messagebox.showinfo("PDF generado",
                            f"Solicitud de compra guardada:\n{path}\n{n_items} items seleccionados")

    def _get_sample_report_data(self, cliente: str):
        """Construye el detalle completo de muestras para un cliente desde r.filtered."""
        r = self.current_result
        if r is None:
            messagebox.showinfo("Reporte", "Primero ejecute el análisis."); return None
        if not cliente or cliente == "— seleccione cliente —":
            messagebox.showinfo("Reporte", "Seleccione un cliente en el selector."); return None

        df = r.filtered.copy()

        # Columnas disponibles — Descripción puede no estar en todos los df
        base_cols = ["Fecha","Código Producto","Nombre Producto","Cantidad"]
        if "Descripción" in df.columns:
            base_cols.append("Descripción")

        # Movimientos de envío para este cliente
        enviados = df[
            df["is_sample_out"] & (df["Bodega Destino"] == cliente)
        ][base_cols].copy()
        enviados["Movimiento"] = "Enviado"

        # Movimientos de devolución de este cliente
        devueltos = df[
            df["is_sample_in"] & (df["Bodega Origen"] == cliente)
        ][base_cols].copy()
        devueltos["Movimiento"] = "Devuelto"

        detalle = pd.concat([enviados, devueltos], ignore_index=True)
        detalle = detalle.sort_values(["Código Producto","Fecha"]).reset_index(drop=True)
        detalle["Fecha"] = detalle["Fecha"].dt.strftime("%d/%m/%Y")

        # Resumen por SKU
        env_grp = enviados.groupby(["Código Producto","Nombre Producto"])["Cantidad"].sum()
        dev_grp = devueltos.groupby(["Código Producto","Nombre Producto"])["Cantidad"].sum()
        resumen = env_grp.rename("Enviadas").to_frame()
        resumen["Devueltas"] = dev_grp.reindex(resumen.index).fillna(0).astype(int)
        resumen["Saldo"] = resumen["Enviadas"] - resumen["Devueltas"]
        resumen = resumen.reset_index()

        # Totales generales
        total_env = int(resumen["Enviadas"].sum())
        total_dev = int(resumen["Devueltas"].sum())
        saldo_total = total_env - total_dev

        # Fecha última devolución por SKU
        if not devueltos.empty:
            ult_dev = (devueltos.groupby("Código Producto")["Fecha"]
                       .max().rename("Última Devolución"))
            resumen = resumen.merge(ult_dev, on="Código Producto", how="left")
        else:
            resumen["Última Devolución"] = ""

        cutoff = self.cutoff_entry.get().strip()
        return {
            "cliente":      cliente,
            "cutoff":       cutoff,
            "resumen":      resumen,
            "detalle":      detalle,
            "total_env":    total_env,
            "total_dev":    total_dev,
            "saldo_total":  saldo_total,
        }

    def export_sample_report_pdf(self):
        """Genera reporte ejecutivo PDF de muestras por cliente."""
        cliente = self.sample_client_var.get()
        data = self._get_sample_report_data(cliente)
        if data is None: return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF","*.pdf")],
            initialfile=f"muestras_{cliente.replace(' ','_')}.pdf")
        if not path: return

        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                            Table, TableStyle, HRFlowable)
            from datetime import datetime as _dt

            from reportlab.lib.pagesizes import landscape as _landscape
            PAGE = _landscape(A4)   # horizontal: 29.7 × 21 cm
            doc  = SimpleDocTemplate(path, pagesize=PAGE,
                                     topMargin=1.2*cm, bottomMargin=1.2*cm,
                                     leftMargin=1.5*cm, rightMargin=1.5*cm)
            PAGE_W = PAGE[0] - 3.0*cm   # ancho útil
            styles = getSampleStyleSheet()

            C_DARK  = colors.white
            C_BLUE  = colors.HexColor("#1E3A5F")
            C_TEAL  = colors.HexColor("#1E3A5F")
            C_GREEN = colors.HexColor("#059669")
            C_WARN  = colors.HexColor("#D97706")
            C_RED   = colors.HexColor("#DC2626")
            C_TEXT  = colors.HexColor("#111827")
            C_MUTED = colors.HexColor("#6B7280")
            C_EVEN  = colors.HexColor("#F9FAFB")
            C_ODD   = colors.white

            title_style = ParagraphStyle("title", parent=styles["Title"],
                fontSize=16, textColor=C_BLUE, spaceAfter=4,
                fontName="Helvetica-Bold")
            sub_style   = ParagraphStyle("sub", parent=styles["Normal"],
                fontSize=10, textColor=C_MUTED, spaceAfter=2,
                fontName="Helvetica")
            sect_style  = ParagraphStyle("sect", parent=styles["Normal"],
                fontSize=11, textColor=C_BLUE, spaceBefore=10, spaceAfter=4,
                fontName="Helvetica-Bold")
            body_style  = ParagraphStyle("body", parent=styles["Normal"],
                fontSize=9, textColor=C_TEXT, fontName="Helvetica")

            elems = []

            # ── Encabezado ────────────────────────────────────────────────
            elems.append(Paragraph(f"REPORTE EJECUTIVO DE MUESTRAS", title_style))
            elems.append(Paragraph(f"Cliente: {data['cliente']}", sub_style))
            elems.append(Paragraph(
                f"Fecha de corte: {data['cutoff']}  |  "
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}",
                sub_style))
            elems.append(HRFlowable(width="100%", thickness=1,
                                    color=C_TEAL, spaceAfter=8))

            # ── KPIs resumen ──────────────────────────────────────────────
            kpi_data = [
                ["TOTAL ENVIADAS", "TOTAL DEVUELTAS", "SALDO EN CLIENTE"],
                [str(data["total_env"]), str(data["total_dev"]),
                 str(data["saldo_total"])],
            ]
            saldo_color = C_RED if data["saldo_total"] > 0 else C_GREEN
            kpi_table = Table(kpi_data, colWidths=[PAGE_W/3, PAGE_W/3, PAGE_W/3])
            kpi_table.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,0), C_BLUE),
                ("TEXTCOLOR",   (0,0), (-1,0), C_MUTED),
                ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",    (0,0), (-1,0), 8),
                ("ALIGN",       (0,0), (-1,-1), "CENTER"),
                ("BACKGROUND",  (0,1), (1,1), C_EVEN),
                ("BACKGROUND",  (2,1), (2,1), C_BLUE),
                ("TEXTCOLOR",   (0,1), (1,1), C_TEXT),
                ("TEXTCOLOR",   (2,1), (2,1), saldo_color),
                ("FONTNAME",    (0,1), (-1,1), "Helvetica-Bold"),
                ("FONTSIZE",    (0,1), (-1,1), 18),
                ("TOPPADDING",  (0,0), (-1,-1), 6),
                ("BOTTOMPADDING",(0,0),(-1,-1), 6),
                ("BOX",         (0,0), (-1,-1), 1, C_TEAL),
                ("INNERGRID",   (0,0), (-1,-1), 0.5, C_BLUE),
            ]))
            elems.append(kpi_table)
            elems.append(Spacer(1, 0.4*cm))

            # ── Resumen por SKU ───────────────────────────────────────────
            elems.append(Paragraph("RESUMEN POR PRODUCTO", sect_style))
            res = data["resumen"]
            res_hdr = ["Código", "Nombre Producto", "Enviadas", "Devueltas",
                       "Saldo", "Últ. Devolución"]
            res_rows = [res_hdr]
            for _, row in res.iterrows():
                saldo = int(row["Saldo"])
                res_rows.append([
                    str(row["Código Producto"]),
                    str(row["Nombre Producto"]),
                    str(int(row["Enviadas"])),
                    str(int(row["Devueltas"])),
                    str(saldo),
                    str(row.get("Última Devolución","")) or "—",
                ])
            # Fila de total
            res_rows.append([
                "TOTAL", "",
                str(data["total_env"]),
                str(data["total_dev"]),
                str(data["saldo_total"]), ""
            ])

            col_ws = [2.5*cm, PAGE_W-2.5*cm-1.8*cm-2.0*cm-1.8*cm-2.8*cm, 1.8*cm, 2.0*cm, 1.8*cm, 2.8*cm]
            from reportlab.platypus import Paragraph as _P
            # Wrap en Nombre Producto para evitar solapamiento
            res_rows_wrap = []
            for ri, row_r in enumerate(res_rows):
                if ri == 0:
                    res_rows_wrap.append(row_r)
                else:
                    wrapped = list(row_r)
                    wrapped[1] = _P(str(row_r[1]), ParagraphStyle("cell", fontSize=7,
                        textColor=C_TEXT if ri < len(res_rows)-1 else C_WARN,
                        fontName="Helvetica"))
                    res_rows_wrap.append(wrapped)
            res_table = Table(res_rows_wrap, colWidths=col_ws, repeatRows=1)
            ts = [
                ("BACKGROUND",    (0,0), (-1,0), C_BLUE),
                ("TEXTCOLOR",     (0,0), (-1,0), C_TEXT),
                ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0), (-1,0), 8),
                ("ALIGN",         (0,0), (-1,0), "CENTER"),
                ("FONTSIZE",      (0,1), (-1,-1), 8),
                ("FONTNAME",      (0,-1),(-1,-1), "Helvetica-Bold"),
                ("BACKGROUND",    (0,-1),(-1,-1), C_BLUE),
                ("TEXTCOLOR",     (0,-1),(-1,-1), C_WARN),
                ("TOPPADDING",    (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("BOX",           (0,0), (-1,-1), 1, C_TEAL),
                ("INNERGRID",     (0,0), (-1,-1), 0.3, C_BLUE),
            ]
            for i in range(1, len(res_rows)-1):
                bg = C_EVEN if i % 2 == 0 else C_ODD
                ts.append(("BACKGROUND", (0,i), (-1,i), bg))
                ts.append(("TEXTCOLOR",  (0,i), (-1,i), C_TEXT))
                # Colorear saldo
                saldo_val = res_rows[i][4]
                try:
                    sv = int(saldo_val)
                    col_s = C_RED if sv > 0 else C_GREEN
                    ts.append(("TEXTCOLOR", (4,i), (4,i), col_s))
                except: pass
            ts.append(("ALIGN", (2,1), (-1,-1), "CENTER"))
            res_table.setStyle(TableStyle(ts))
            elems.append(res_table)
            elems.append(Spacer(1, 0.4*cm))

            # ── Detalle de movimientos ────────────────────────────────────
            elems.append(Paragraph("DETALLE DE MOVIMIENTOS", sect_style))
            det = data["detalle"]
            if det.empty:
                elems.append(Paragraph("Sin movimientos registrados.", body_style))
            else:
                has_desc = "Descripción" in det.columns
                det_hdr = ["Fecha","Código","Nombre Producto","Descripción","Cant.","Movimiento"]
                det_rows = [det_hdr]
                for _, row in det.iterrows():
                    det_rows.append([
                        str(row["Fecha"]),
                        str(row["Código Producto"]),
                        str(row["Nombre Producto"]),
                        str(row["Descripción"]) if has_desc and pd.notna(row.get("Descripción")) else "",
                        str(int(row["Cantidad"])),
                        str(row["Movimiento"]),
                    ])
                det_cws = [2.0*cm, 2.0*cm,
                           PAGE_W*0.30, PAGE_W*0.25,
                           1.2*cm, 2.2*cm]
                # Wrap en Nombre y Descripción
                det_rows_wrap = []
                for ri2, row_d in enumerate(det_rows):
                    if ri2 == 0:
                        det_rows_wrap.append(row_d)
                    else:
                        wd = list(row_d)
                        txt_c = C_GREEN if str(row_d[5]) == "Devuelto" else C_TEXT
                        wd[2] = _P(str(row_d[2]), ParagraphStyle("cn", fontSize=7,
                            textColor=txt_c, fontName="Helvetica"))
                        wd[3] = _P(str(row_d[3]), ParagraphStyle("cd", fontSize=7,
                            textColor=C_MUTED, fontName="Helvetica"))
                        det_rows_wrap.append(wd)
                det_table = Table(det_rows_wrap, colWidths=det_cws, repeatRows=1)
                det_ts = [
                    ("BACKGROUND",    (0,0), (-1,0), C_BLUE),
                    ("TEXTCOLOR",     (0,0), (-1,0), C_TEXT),
                    ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
                    ("FONTSIZE",      (0,0), (-1,-1), 8),
                    ("TOPPADDING",    (0,0), (-1,-1), 3),
                    ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                    ("BOX",           (0,0), (-1,-1), 1, C_TEAL),
                    ("INNERGRID",     (0,0), (-1,-1), 0.3, C_BLUE),
                    ("ALIGN",         (3,0), (3,-1), "CENTER"),
                ]
                for i in range(1, len(det_rows)):
                    bg = C_EVEN if i % 2 == 0 else C_ODD
                    mov = det_rows[i][5]   # índice 5 porque agregamos Descripción
                    txt = C_GREEN if mov == "Devuelto" else C_TEXT
                    det_ts.append(("BACKGROUND", (0,i), (-1,i), bg))
                    det_ts.append(("TEXTCOLOR",  (0,i), (-1,i), txt))
                det_table.setStyle(TableStyle(det_ts))
                elems.append(det_table)

            doc.build(elems)
            messagebox.showinfo("PDF generado", f"Reporte guardado:\n{path}")
            self.log(f"PDF muestras '{cliente}' → {path}")

        except Exception as e:
            self.log(f"Error PDF muestras: {e}")
            messagebox.showerror("Error al generar PDF", str(e))

    def export_sample_report_html(self):
        """Genera reporte ejecutivo HTML de muestras por cliente."""
        cliente = self.sample_client_var.get()
        data = self._get_sample_report_data(cliente)
        if data is None: return

        path = filedialog.asksaveasfilename(
            defaultextension=".html", filetypes=[("HTML","*.html")],
            initialfile=f"muestras_{cliente.replace(' ','_')}.html")
        if not path: return

        try:
            from datetime import datetime as _dt

            def tbl_rows(df, cols):
                rows = ""
                for i, (_, row) in enumerate(df.iterrows()):
                    bg = "#1A2436" if i % 2 == 0 else "#111827"
                    cells = "".join(f"<td>{row[c]}</td>" for c in cols)
                    rows += f'<tr style="background:{bg}">{cells}</tr>'
                return rows

            res = data["resumen"]
            det = data["detalle"]
            saldo_color = "#EF4444" if data["saldo_total"] > 0 else "#10B981"

            res_hdr = ["Código Producto","Nombre Producto","Enviadas",
                       "Devueltas","Saldo","Última Devolución"]
            res_rows_html = ""
            for i, (_, row) in enumerate(res.iterrows()):
                sv = int(row["Saldo"])
                sc = "#DC2626" if sv > 0 else "#059669"
                res_rows_html += (
                    f'<tr>'
                    f'<td>{row["Código Producto"]}</td>'
                    f'<td>{row["Nombre Producto"]}</td>'
                    f'<td class="num">{int(row["Enviadas"])}</td>'
                    f'<td class="num">{int(row["Devueltas"])}</td>'
                    f'<td class="num" style="color:{sc};font-weight:bold">{sv}</td>'
                    f'<td class="num">{row.get("Última Devolución","") or "—"}</td>'
                    f'</tr>'
                )
            res_rows_html += (
                f'<tr style="background:#1E3A5F;color:#fff;font-weight:bold">'
                f'<td>TOTAL</td><td></td>'
                f'<td class="num">{data["total_env"]}</td>'
                f'<td class="num">{data["total_dev"]}</td>'
                f'<td class="num" style="color:{"#DC2626" if data["saldo_total"]>0 else "#059669"}">{data["saldo_total"]}</td>'
                f'<td></td></tr>'
            )

            has_desc = "Descripción" in det.columns
            det_rows_html = ""
            for i, (_, row) in enumerate(det.iterrows()):
                color = "#059669" if row["Movimiento"] == "Devuelto" else "#111827"
                desc  = str(row["Descripción"]) if has_desc and pd.notna(row.get("Descripción")) else ""
                det_rows_html += (
                    f'<tr>'
                    f'<td style="color:{color}">{row["Fecha"]}</td>'
                    f'<td>{row["Código Producto"]}</td>'
                    f'<td>{row["Nombre Producto"]}</td>'
                    f'<td>{desc}</td>'
                    f'<td class="num">{int(row["Cantidad"])}</td>'
                    f'<td style="color:{color};font-weight:bold">{row["Movimiento"]}</td>'
                    f'</tr>'
                )

            html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Muestras — {cliente}</title>
<style>
  body      {{ background:#ffffff; color:#111827; font-family:Segoe UI,Arial,sans-serif; margin:0; padding:24px }}
  h1        {{ color:#1E3A5F; font-size:20px; margin-bottom:2px; border-bottom:2px solid #1E3A5F; padding-bottom:6px }}
  h2        {{ color:#1E3A5F; font-size:13px; margin:18px 0 6px; text-transform:uppercase; letter-spacing:1px }}
  .meta     {{ color:#6B7280; font-size:11px; margin-bottom:16px }}
  .kpi-row  {{ display:flex; gap:12px; margin-bottom:20px }}
  .kpi      {{ background:#F9FAFB; border:1px solid #D1D5DB; border-radius:6px;
               padding:12px 20px; text-align:center; flex:1 }}
  .kpi-lbl  {{ color:#6B7280; font-size:10px; font-weight:bold; letter-spacing:1px }}
  .kpi-val  {{ font-size:26px; font-weight:bold; margin-top:4px; color:#6B7280 }}
  .kpi-val.saldo {{ color: inherit }}
  table     {{ width:100%; border-collapse:collapse; font-size:11px; margin-bottom:20px }}
  th        {{ background:#1E3A5F; color:#ffffff; padding:7px 10px; text-align:left }}
  td        {{ padding:5px 10px; border-bottom:1px solid #E5E7EB; color:#374151 }}
  td.num    {{ text-align:right; color:#6B7280; font-variant-numeric:tabular-nums }}
  tr:nth-child(even) td {{ background:#F9FAFB }}
  .num      {{ text-align:right; color:#6B7280 }}
  hr        {{ border:none; border-top:1px solid #D1D5DB; margin:12px 0 }}
  @media print {{ body {{ padding:10px }} .kpi {{ border:1px solid #ccc }} }}
</style>
</head>
<body>
<h1>REPORTE EJECUTIVO DE MUESTRAS</h1>
<div class="meta">
  Cliente: <strong>{cliente}</strong> &nbsp;|&nbsp;
  Corte: <strong>{data["cutoff"]}</strong> &nbsp;|&nbsp;
  Generado: {_dt.now().strftime("%d/%m/%Y %H:%M")}
</div>
<hr>
<div class="kpi-row">
  <div class="kpi">
    <div class="kpi-lbl">TOTAL ENVIADAS</div>
    <div class="kpi-val" style="color:#F3F4F6">{data["total_env"]}</div>
  </div>
  <div class="kpi">
    <div class="kpi-lbl">TOTAL DEVUELTAS</div>
    <div class="kpi-val" style="color:#F3F4F6">{data["total_dev"]}</div>
  </div>
  <div class="kpi">
    <div class="kpi-lbl">SALDO EN CLIENTE</div>
    <div class="kpi-val saldo" style="color:{saldo_color}">{data["saldo_total"]}</div>
  </div>
</div>

<h2>RESUMEN POR PRODUCTO</h2>
<table>
  <thead><tr>
    <th>Código</th><th>Nombre Producto</th>
    <th class="num">Enviadas</th><th class="num">Devueltas</th>
    <th class="num">Saldo</th><th class="num">Últ. Devolución</th>
  </tr></thead>
  <tbody>{res_rows_html}</tbody>
</table>

<h2>DETALLE DE MOVIMIENTOS</h2>
<table>
  <thead><tr>
    <th>Fecha</th><th>Código</th><th>Nombre Producto</th><th>Descripción</th>
    <th class="num">Cant.</th><th>Movimiento</th>
  </tr></thead>
  <tbody>{det_rows_html}</tbody>
</table>
</body>
</html>"""

            with open(path, "w", encoding="utf-8") as f:
                f.write(html)

            import webbrowser, os
            webbrowser.open(f"file:///{os.path.abspath(path)}")
            messagebox.showinfo("HTML generado", f"Reporte guardado y abierto:\n{path}")
            self.log(f"HTML muestras '{cliente}' → {path}")

        except Exception as e:
            self.log(f"Error HTML muestras: {e}")
            messagebox.showerror("Error al generar HTML", str(e))


    # ── Fill genérico ────────────────────────────────────────────────────────
    def fill_tree(self, tree, df, tag_mode=None, empty_msg=None, sortable=False, decimals=True):
        tree.delete(*tree.get_children())
        cols = list(df.columns)
        if not cols or df.empty:
            tree["columns"]=["Info"]; tree.heading("Info",text="Info")
            tree.column("Info",anchor="w",width=400)
            tree.insert("","end",values=[empty_msg or "Sin datos"]); return

        two_line = {
            "Codigo Producto":        "Cód.Prod",
            "Nombre Producto":        "Nombre Producto",
            "Código Producto":        "Cód. Prod",
            "Categoría Producto":     "Categ Prod",
            "Valor_Compras":          "Compras $",
            "Valor_Ventas":           "Ventas $",
            "Valor Compras (ING)":    "Compras $ (Ingreso)",
            "Valor Ventas (EGR)":     "Ventas $(Egreso)",
            "Valor Inventario ($)":   "Inventario ($)",
            "Valor Inventario":       "Inventario",
            "Valor Unitario Promedio":"Unitario $ Promedio",
            "Valor Stock":            "Stock $",
            "Compras":                "Compras(Ingreso)",
            "Dev_Proveedor":          "Dev.Prov",
            "Ventas":                 "Ventas(Egreso)",
            "Dev_Cliente":            "N/C Cliente",
            "Muestras_Enviadas":      "Muestras Enviadas",
            "Muestras_Devueltas":     "Muestras Devueltas",
            "Stock Disponible":       "Stock Disponible",
            "Stock Muestras":         "Stock en  Muestras",
            "Stock Total":            "Stock Total",
            "Stock en Cliente":       "Stock en Cliente",
        }
        tree["columns"] = cols
        for c in cols:
            anchor = "e" if _is_num(c) else "w"
            label  = two_line.get(c, c)
            if sortable:
                tree.heading(c, text=label, anchor=anchor,
                             command=lambda col=c, t=tree, d=df: self._sort_col(t,col,d,decimals))
            else:
                tree.heading(c, text=label, anchor=anchor)
            tree.column(c, anchor=anchor, width=max(_col_w(c), _col_w(label)), minwidth=50, stretch=False)

        num_cols = [c for c in cols if _is_num(c)]
        totals = {}
        for c in num_cols:
            try: totals[c] = pd.to_numeric(df[c],errors="coerce").sum()
            except: pass

        for i,(_,row) in enumerate(df.iterrows()):
            vals = []
            for c in cols:
                v = row[c]
                if isinstance(v,float):
                    vals.append(f"{v:,.0f}" if not decimals else f"{v:,.2f}")
                elif isinstance(v,bool): vals.append("Si" if v else "No")
                else:
                    s=str(v); vals.append("" if s in ("nan","None","NaN") else s)
            tree.insert("","end",values=vals,tags=(self._row_tag(tag_mode,i,row,cols),))

        if totals:
            tr=["TOTAL" if i==0 else (
                    f"{totals[c]:,.0f}" if not decimals and c in totals else
                    f"{totals[c]:,.2f}" if c in totals else "")
                for i,c in enumerate(cols)]
            tree.insert("","end",values=tr,tags=("total",))

    def _sort_col(self, tree, col, df, decimals):
        prev_col,prev_rev = self._sort_state.get(id(tree),(None,False))
        rev = not prev_rev if col==prev_col else False
        self._sort_state[id(tree)] = (col,rev)
        try:
            sdf = df.sort_values(col, ascending=not rev,
                key=lambda s: pd.to_numeric(s,errors="coerce").fillna(0) if _is_num(col)
                              else s.astype(str).str.lower())
        except: sdf = df.sort_values(col, ascending=not rev)
        self.fill_tree(tree, sdf, tag_mode="zebra", sortable=True, decimals=decimals)

    def _row_tag(self, tag_mode, idx, row, cols):
        if tag_mode == "physical":
            try:
                ci = cols.index("Coincide")
                v  = row.iloc[ci] if hasattr(row,"iloc") else row[ci]
                return "diff_ok" if str(v) in ("True","Si","1") else "diff_bad"
            except: pass
        return "even" if idx%2==0 else "odd"

    # ── Export ────────────────────────────────────────────────────────────────
    def export_excel(self):
        if self.current_result is None:
            messagebox.showwarning("Exportacion","Primero ejecute el analisis."); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel","*.xlsx")])
        if not path: return
        try:
            self.engine.export_result(self.current_result, path)
            self.log(f"Exportado: {path}")
            messagebox.showinfo("Exportacion","Archivo exportado correctamente.")
        except Exception as e:
            self.log(f"Error al exportar: {e}")
            messagebox.showerror("Error",str(e))
