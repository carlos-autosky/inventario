"""
Módulo de Toma Física v001.9
- Ingreso directo en tabla (sin doble clic)
- Enter avanza a la siguiente fila
- Cantidad a la derecha de la descripción
- Columna Observación
- Persistencia entre tomas (columna anterior en gris)
- Fecha de toma
- Opción de duplicar toma anterior
"""
from __future__ import annotations
try:
    import customtkinter as ctk
    from tkinter import ttk, filedialog, messagebox
    import tkinter as tk
    HAS_GUI = True
except ImportError:
    HAS_GUI = False
    class _GuiStub:
        CTkToplevel = object
        def __getattr__(self, _): return object
    ctk = _GuiStub()
    tk = _GuiStub()
    ttk = _GuiStub()
    filedialog = _GuiStub()
    messagebox = _GuiStub()
import pandas as pd
import json
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

from .config import THEME
from .utils import now_str

FONT_BODY  = ("Segoe UI", 11)
FONT_BOLD  = ("Segoe UI", 11, "bold")
FONT_SMALL = ("Segoe UI", 10)
FONT_KPI   = ("Segoe UI", 9, "bold")

HISTORY_FILE = Path.home() / "inventario_toma_fisica_history.json"

DEFAULT_LOCATIONS = [
    "Bodega Principal",
    "Servicio Tecnico",
    "Reparados",
    "Sospechosos / Dañados",
    "En Cajas Master",
    "Percha",
    "Reempaque",
    "Por Facturar",
    "En Revision",
    "Muestras",
]


def load_history() -> dict:
    """Carga el historial de tomas anteriores desde disco."""
    if HISTORY_FILE.exists():
        try:
            return json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_history(data: dict) -> None:
    HISTORY_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


class TomaFisicaWindow(ctk.CTkToplevel):
    def __init__(self, master, skus_df: pd.DataFrame | None = None, excluded_skus: set | None = None):
        super().__init__(master)
        self.title("Toma Física — Ingreso de Conteo")
        self.geometry("1400x860")
        self.configure(fg_color=THEME["bg"])
        self.grab_set()

        excluded_skus = excluded_skus or set()
        if skus_df is not None and not skus_df.empty:
            all_skus = (skus_df[["Código Producto", "Nombre Producto"]]
                        .drop_duplicates()
                        .sort_values("Nombre Producto")
                        .reset_index(drop=True))
            # Excluir los SKUs marcados como excluidos del analisis
            if excluded_skus:
                all_skus = all_skus[~all_skus["Código Producto"].isin(excluded_skus)]
            self.skus = all_skus.reset_index(drop=True)
        else:
            self.skus = pd.DataFrame(columns=["Código Producto", "Nombre Producto"])

        self.locations: list[str] = list(DEFAULT_LOCATIONS)
        # data[loc][sku] = {"qty": int, "obs": str}
        self.data: dict[str, dict[str, dict]] = {loc: {} for loc in self.locations}
        self.history: dict = load_history()   # {loc: {date_str: {sku: {qty,obs}}}}
        self.current_location: str = self.locations[0]

        # Fecha de la toma actual
        self.toma_date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))

        # Referencia al entry activo en la tabla
        self._active_entry: tk.Entry | None = None
        self._entry_widgets: list[tuple[str, tk.Entry, tk.Entry]] = []  # (sku, qty_entry, obs_entry)

        self._build_ui()
        self._refresh_location_list()
        self._load_location(self.current_location)

    # ── UI ────────────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        # ── Panel izquierdo ────────────────────────────────────────────────
        left = ctk.CTkFrame(self, fg_color=THEME["panel"], corner_radius=0, width=230)
        left.grid(row=0, column=0, sticky="nsew")
        left.grid_propagate(False)
        left.grid_rowconfigure(2, weight=1)
        left.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(left, text="UBICACIONES", font=FONT_KPI,
                     text_color=THEME["muted"]).grid(row=0, column=0, sticky="w", padx=10, pady=(10, 2))

        add_row = ctk.CTkFrame(left, fg_color="transparent")
        add_row.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 4))
        add_row.grid_columnconfigure(0, weight=1)
        self.new_loc_entry = ctk.CTkEntry(add_row, placeholder_text="Nueva ubicacion...",
                                          font=FONT_SMALL, height=26)
        self.new_loc_entry.grid(row=0, column=0, sticky="ew")
        ctk.CTkButton(add_row, text="+", width=28, height=26, font=FONT_BOLD,
                      command=self._add_location,
                      fg_color=THEME["accent"], hover_color="#2563EB").grid(row=0, column=1, padx=(3, 0))

        self.loc_frame = ctk.CTkScrollableFrame(left, fg_color="transparent")
        self.loc_frame.grid(row=2, column=0, sticky="nsew", padx=6, pady=4)
        self.loc_buttons: list[ctk.CTkButton] = []

        exp_frame = ctk.CTkFrame(left, fg_color="transparent")
        exp_frame.grid(row=3, column=0, sticky="ew", padx=8, pady=8)
        exp_frame.grid_columnconfigure((0,1,2), weight=1)
        ctk.CTkButton(exp_frame, text="📥 Importar", height=36, font=FONT_SMALL,
                      fg_color="#7C3AED", hover_color="#6D28D9",
                      command=self._import_excel).grid(row=0, column=0, sticky="ew", padx=(0,2))
        ctk.CTkButton(exp_frame, text="📊 Excel", height=36, font=FONT_SMALL,
                      fg_color=THEME["success"], hover_color="#059669",
                      command=self._export_excel).grid(row=0, column=1, sticky="ew", padx=(2,2))
        if HAS_PDF:
            ctk.CTkButton(exp_frame, text="📄 PDF", height=36, font=FONT_SMALL,
                          fg_color=THEME["violet"], hover_color="#6D28D9",
                          command=self._export_pdf).grid(row=0, column=2, sticky="ew", padx=(2,0))

        # ── Panel derecho ──────────────────────────────────────────────────
        right = ctk.CTkFrame(self, fg_color=THEME["bg"], corner_radius=0)
        right.grid(row=0, column=1, sticky="nsew")
        right.grid_rowconfigure(2, weight=1)
        right.grid_columnconfigure(0, weight=1)

        # Header con fecha y controles
        hdr = ctk.CTkFrame(right, fg_color=THEME["panel_alt"], corner_radius=8)
        hdr.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 4))
        hdr.grid_columnconfigure(1, weight=1)

        self.loc_title = ctk.CTkLabel(hdr, text="", font=("Segoe UI", 14, "bold"),
                                       text_color=THEME["accent"])
        self.loc_title.grid(row=0, column=0, sticky="w", padx=12, pady=6)

        # Fecha de toma
        date_f = ctk.CTkFrame(hdr, fg_color="transparent")
        date_f.grid(row=0, column=1, sticky="e", padx=8, pady=6)
        ctk.CTkLabel(date_f, text="Fecha toma:", font=FONT_KPI,
                     text_color=THEME["muted"]).pack(side="left", padx=(0, 4))
        self.date_entry = ctk.CTkEntry(date_f, textvariable=self.toma_date_var,
                                       width=110, height=28, font=FONT_BODY)
        self.date_entry.pack(side="left")

        # Botones de acción
        btn_f = ctk.CTkFrame(hdr, fg_color="transparent")
        btn_f.grid(row=0, column=2, sticky="e", padx=8, pady=6)
        ctk.CTkButton(btn_f, text="Duplicar toma anterior", width=160, height=28,
                      font=FONT_SMALL, fg_color=THEME["teal"], hover_color="#0D9488",
                      command=self._duplicate_previous).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_f, text="Limpiar", width=80, height=28,
                      font=FONT_SMALL, fg_color=THEME["danger"], hover_color="#B91C1C",
                      command=self._clear_location).pack(side="left", padx=(0, 4))
        ctk.CTkButton(btn_f, text="Guardar toma", width=110, height=28,
                      font=FONT_BOLD, fg_color=THEME["accent"], hover_color="#2563EB",
                      command=self._save_toma).pack(side="left")

        # Barra de búsqueda
        search_f = ctk.CTkFrame(right, fg_color="transparent")
        search_f.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 4))
        search_f.grid_columnconfigure(0, weight=1)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._render_table())
        ctk.CTkEntry(search_f, textvariable=self.search_var,
                     placeholder_text="Buscar SKU o nombre...",
                     height=28, font=FONT_BODY).grid(row=0, column=0, sticky="ew")

        self.summary_label = ctk.CTkLabel(search_f, text="",
                                           font=FONT_SMALL, text_color=THEME["muted"])
        self.summary_label.grid(row=0, column=1, padx=(12, 0))

        # Canvas + Scrollbar para la tabla editable
        table_outer = ctk.CTkFrame(right, fg_color=THEME["panel"], corner_radius=8)
        table_outer.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0, 8))
        table_outer.grid_rowconfigure(0, weight=1)
        table_outer.grid_columnconfigure(0, weight=1)

        self.canvas = tk.Canvas(table_outer, bg=THEME["panel"],
                                highlightthickness=0, bd=0)
        self.canvas.grid(row=0, column=0, sticky="nsew")

        ys = ttk.Scrollbar(table_outer, orient="vertical", command=self.canvas.yview)
        ys.grid(row=0, column=1, sticky="ns")
        self.canvas.configure(yscrollcommand=ys.set)

        self.table_frame = tk.Frame(self.canvas, bg=THEME["panel"])
        self.canvas_window = self.canvas.create_window(
            (0, 0), window=self.table_frame, anchor="nw")

        self.table_frame.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_frame_configure(self, _=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    # ── Locations ─────────────────────────────────────────────────────────────
    def _refresh_location_list(self):
        for btn in self.loc_buttons:
            btn.destroy()
        self.loc_buttons.clear()
        for loc in self.locations:
            total = sum(v.get("qty", 0) for v in self.data.get(loc, {}).values())
            label = f"{loc}  [{total}]" if total else loc
            color = THEME["accent"] if loc == self.current_location else THEME["card"]
            btn = ctk.CTkButton(
                self.loc_frame, text=label, anchor="w",
                fg_color=color, hover_color="#1D4ED8",
                font=FONT_SMALL, height=28, corner_radius=5,
                command=lambda l=loc: self._switch_location(l))
            btn.pack(fill="x", padx=2, pady=1)
            self.loc_buttons.append(btn)

    def _add_location(self):
        name = self.new_loc_entry.get().strip()
        if not name:
            return
        if name in self.locations:
            messagebox.showwarning("Duplicado", f"La ubicacion '{name}' ya existe.")
            return
        self.locations.append(name)
        self.data[name] = {}
        self.new_loc_entry.delete(0, "end")
        self._refresh_location_list()

    def _switch_location(self, loc: str):
        self.current_location = loc
        self._refresh_location_list()
        self._load_location(loc)

    def _load_location(self, loc: str):
        self.loc_title.configure(text=f"  {loc}")
        self.search_var.set("")
        self._render_table()

    # ── Obtener toma anterior ──────────────────────────────────────────────────
    def _get_previous_toma(self, loc: str) -> tuple[str, dict]:
        """Retorna (fecha_str, {sku: {qty, obs}}) de la última toma guardada."""
        loc_history = self.history.get(loc, {})
        if not loc_history:
            return ("", {})
        # La última fecha guardada
        last_date = sorted(loc_history.keys())[-1]
        return (last_date, loc_history[last_date])

    def _duplicate_previous(self):
        prev_date, prev_data = self._get_previous_toma(self.current_location)
        if not prev_data:
            messagebox.showinfo("Sin historial",
                                "No hay toma anterior guardada para esta ubicación.")
            return
        if not messagebox.askyesno("Duplicar",
                f"¿Copiar el conteo de la toma del {prev_date} a la toma actual?\n"
                "Se sobrescribirán los valores actuales."):
            return
        self.data[self.current_location] = {
            sku: dict(v) for sku, v in prev_data.items()
        }
        self._render_table()
        self._refresh_location_list()

    # ── Render tabla editable ─────────────────────────────────────────────────
    def _render_table(self):
        # Limpiar tabla
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        self._entry_widgets.clear()

        loc = self.current_location
        current_data = self.data.get(loc, {})
        prev_date, prev_data = self._get_previous_toma(loc)

        ft = self.search_var.get().lower()

        # ── Colores ────────────────────────────────────────────────────────
        BG_HDR   = THEME["panel_alt"]
        BG_EVEN  = "#141E2E"
        BG_ODD   = THEME["panel"]
        BG_PREV  = "#1A2436"   # gris azulado para toma anterior
        FG_TEXT  = THEME["text"]
        FG_MUTED = "#6B7280"
        FG_PREV  = "#4B5563"   # gris para columna anterior
        FG_QTY   = "#10B981"   # verde para cantidad ingresada

        # ── Anchos de columna (px) ─────────────────────────────────────────
        W_NUM  = 40
        W_SKU  = 100
        W_NAME = 260
        W_PREV = 80  if prev_data else 0
        W_QTY  = 90
        W_OBS  = 200

        def make_header(parent, text, width, bg=BG_HDR, fg=FG_TEXT):
            lbl = tk.Label(parent, text=text, font=("Segoe UI", 10, "bold"),
                           bg=bg, fg=fg, width=0, anchor="center",
                           relief="flat", padx=4, pady=5)
            lbl.pack(side="left", ipadx=0)
            lbl.configure(width=width // 8)

        # ── Fila de cabecera ───────────────────────────────────────────────
        hdr_row = tk.Frame(self.table_frame, bg=BG_HDR)
        hdr_row.pack(fill="x", side="top")

        # Usar grid dentro de hdr_row para control exacto de anchos
        col_defs = [
            ("#",      W_NUM,  BG_HDR, FG_MUTED),
            ("Código", W_SKU,  BG_HDR, FG_TEXT),
            ("Nombre / Descripción", W_NAME, BG_HDR, FG_TEXT),
        ]
        if prev_data:
            col_defs.append((f"Anterior\n{prev_date}", W_PREV, BG_PREV, FG_PREV))
        col_defs += [
            ("Cantidad\n(nueva toma)", W_QTY, BG_HDR, FG_QTY),
            ("Observación",           W_OBS, BG_HDR, FG_TEXT),
        ]

        for ci, (text, w, bg, fg) in enumerate(col_defs):
            lbl = tk.Label(hdr_row, text=text, font=("Segoe UI", 9, "bold"),
                           bg=bg, fg=fg, width=w // 7, anchor="center",
                           relief="flat", padx=2, pady=6, justify="center")
            lbl.grid(row=0, column=ci, sticky="nsew", padx=1, pady=0)

        # Separador
        tk.Frame(self.table_frame, bg=THEME["border"], height=1).pack(fill="x")

        # ── Filas de datos ─────────────────────────────────────────────────
        filtered_skus = []
        for _, row in self.skus.iterrows():
            cod  = str(row["Código Producto"])
            name = str(row["Nombre Producto"])
            if ft and ft not in cod.lower() and ft not in name.lower():
                continue
            filtered_skus.append((cod, name))

        qty_entries: list[tk.Entry] = []
        obs_entries: list[tk.Entry] = []

        for idx, (cod, name) in enumerate(filtered_skus):
            bg = BG_EVEN if idx % 2 == 0 else BG_ODD
            cur = current_data.get(cod, {})
            cur_qty = cur.get("qty", 0)
            cur_obs = cur.get("obs", "")
            prev_qty = prev_data.get(cod, {}).get("qty", 0) if prev_data else None
            has_value = cur_qty > 0 or (prev_qty and prev_qty > 0)

            row_frame = tk.Frame(self.table_frame, bg=bg)
            row_frame.pack(fill="x", side="top")

            ci = 0
            # Número
            tk.Label(row_frame, text=str(idx+1), font=("Segoe UI", 10),
                     bg=bg, fg=FG_MUTED, width=W_NUM//8, anchor="center",
                     padx=2, pady=3).grid(row=0, column=ci, sticky="nsew"); ci+=1

            # Código
            tk.Label(row_frame, text=cod, font=("Segoe UI", 10, "bold"),
                     bg=bg, fg=THEME["accent"], width=W_SKU//8, anchor="w",
                     padx=4, pady=3).grid(row=0, column=ci, sticky="nsew"); ci+=1

            # Nombre (truncado)
            short_name = name[:100] + "…" if len(name) > 100 else name
            tk.Label(row_frame, text=short_name, font=("Segoe UI", 10),
                     bg=bg, fg=FG_TEXT, width=W_NAME//8, anchor="w",
                     padx=4, pady=3).grid(row=0, column=ci, sticky="nsew"); ci+=1

            # Anterior (gris, no editable)
            if prev_data is not None and prev_date:
                prev_val = str(prev_qty) if prev_qty else ""
                tk.Label(row_frame, text=prev_val, font=("Segoe UI", 10),
                         bg=BG_PREV, fg=FG_PREV, width=W_PREV//8, anchor="e",
                         padx=6, pady=3).grid(row=0, column=ci, sticky="nsew"); ci+=1

            # Cantidad (editable, verde si tiene valor)
            qty_var = tk.StringVar(value=str(cur_qty) if cur_qty else "")
            qty_entry = tk.Entry(row_frame, textvariable=qty_var,
                                 font=("Segoe UI", 11, "bold"),
                                 bg="#0F2D20" if cur_qty else bg,
                                 fg=FG_QTY, insertbackground=FG_QTY,
                                 relief="flat", bd=0, width=W_QTY//8,
                                 justify="right")
            qty_entry.grid(row=0, column=ci, sticky="nsew", padx=2, pady=2); ci+=1

            # Observación (editable)
            obs_var = tk.StringVar(value=cur_obs)
            obs_entry = tk.Entry(row_frame, textvariable=obs_var,
                                 font=("Segoe UI", 10),
                                 bg=bg, fg="#9CA3AF",
                                 insertbackground=FG_TEXT,
                                 relief="flat", bd=0, width=W_OBS//8)
            obs_entry.grid(row=0, column=ci, sticky="nsew", padx=4, pady=2)

            # Guardar referencias
            qty_entries.append(qty_entry)
            obs_entries.append(obs_entry)
            self._entry_widgets.append((cod, qty_var, obs_var))

            # Eventos en qty_entry
            qty_entry.bind("<FocusOut>",
                           lambda e, c=cod, qv=qty_var, ov=obs_var: self._save_row(c, qv, ov))
            qty_entry.bind("<FocusIn>",
                           lambda e, ent=qty_entry: ent.configure(bg="#0F2D20"))

            # Separador entre filas
            tk.Frame(self.table_frame, bg=THEME["border"], height=1).pack(fill="x")

        # ── Navegación con Enter ───────────────────────────────────────────
        all_qty = qty_entries
        for i, qe in enumerate(all_qty):
            def make_enter(idx_=i):
                def on_enter(e):
                    # Guardar fila actual
                    cod_, qty_v_, obs_v_ = self._entry_widgets[idx_]
                    self._save_row(cod_, qty_v_, obs_v_)
                    # Mover al siguiente
                    next_i = idx_ + 1
                    if next_i < len(all_qty):
                        all_qty[next_i].focus_set()
                        all_qty[next_i].select_range(0, "end")
                    return "break"
                return on_enter
            qe.bind("<Return>", make_enter(i))
            qe.bind("<Tab>",    make_enter(i))

        # Resumen
        total_items = sum(1 for _,qv,_ in self._entry_widgets if self._parse_qty(qv.get()) > 0)
        total_units = sum(self._parse_qty(qv.get()) for _,qv,_ in self._entry_widgets)
        self.summary_label.configure(
            text=f"Items con conteo: {total_items}  |  Total unidades: {total_units:,}  |  "
                 f"Enter / Tab → siguiente fila")

    def _parse_qty(self, s: str) -> int:
        try: return max(0, int(str(s).strip()))
        except: return 0

    def _save_row(self, cod: str, qty_var: tk.StringVar, obs_var: tk.StringVar):
        qty = self._parse_qty(qty_var.get())
        obs = obs_var.get().strip()
        loc_data = self.data.setdefault(self.current_location, {})
        if qty > 0 or obs:
            loc_data[cod] = {"qty": qty, "obs": obs}
        else:
            loc_data.pop(cod, None)
        self._refresh_location_list()

    def _clear_location(self):
        if not messagebox.askyesno("Limpiar",
                f"¿Borrar todos los conteos de '{self.current_location}'?"):
            return
        self.data[self.current_location] = {}
        self._render_table()
        self._refresh_location_list()

    def _save_toma(self):
        """Guarda la toma actual en el historial persistente."""
        toma_date = self.toma_date_var.get().strip()
        if not toma_date:
            messagebox.showerror("Fecha", "Ingrese la fecha de la toma.")
            return

        # Recoger todos los valores actuales de los entries antes de guardar
        for cod, qty_var, obs_var in self._entry_widgets:
            self._save_row(cod, qty_var, obs_var)

        loc = self.current_location
        loc_data = self.data.get(loc, {})
        if not loc_data:
            messagebox.showinfo("Sin datos", "No hay conteos en esta ubicación.")
            return

        if loc not in self.history:
            self.history[loc] = {}
        self.history[loc][toma_date] = {
            sku: {"qty": v["qty"], "obs": v.get("obs", "")}
            for sku, v in loc_data.items()
        }
        save_history(self.history)
        messagebox.showinfo("Guardado",
                            f"Toma del {toma_date} guardada correctamente\n"
                            f"Ubicación: {loc}\n"
                            f"Items: {len(loc_data)}")
        # Recargar para mostrar la nueva toma como "anterior"
        self._render_table()

    # ── Summary DataFrame ─────────────────────────────────────────────────────
    def _build_summary_df(self) -> pd.DataFrame:
        if self.skus.empty:
            return pd.DataFrame()

        # Recoger valores actuales
        for cod, qty_var, obs_var in self._entry_widgets:
            self._save_row(cod, qty_var, obs_var)

        prev_date, prev_data = self._get_previous_toma(self.current_location)
        rows = []
        for _, row in self.skus.iterrows():
            cod  = str(row["Código Producto"])
            name = str(row["Nombre Producto"])
            cur  = self.data.get(self.current_location, {}).get(cod, {})
            cur_qty = cur.get("qty", 0)
            cur_obs = cur.get("obs", "")
            prev_qty = prev_data.get(cod, {}).get("qty", 0) if prev_data else 0
            if cur_qty == 0 and prev_qty == 0:
                continue
            entry = {"Código": cod, "Nombre": name}
            if prev_date:
                entry[f"Anterior ({prev_date})"] = prev_qty or ""
            entry["Cantidad"] = cur_qty or ""
            entry["Observación"] = cur_obs
            rows.append(entry)

        return pd.DataFrame(rows).reset_index(drop=True)

    def _build_full_df(self) -> pd.DataFrame:
        """Resumen de todas las ubicaciones."""
        if self.skus.empty:
            return pd.DataFrame()
        rows = []
        for _, row in self.skus.iterrows():
            cod  = str(row["Código Producto"])
            name = str(row["Nombre Producto"])
            entry = {"Código": cod, "Nombre": name}
            total = 0
            for loc in self.locations:
                qty = self.data.get(loc, {}).get(cod, {}).get("qty", 0)
                entry[loc] = qty if qty else ""
                total += qty
            entry["TOTAL"] = total if total else ""
            rows.append(entry)
        df = pd.DataFrame(rows)
        loc_cols = self.locations
        df = df[df[loc_cols].apply(
            lambda r: any(v != "" for v in r), axis=1)].copy()
        return df.reset_index(drop=True)

    # ── Import Excel ──────────────────────────────────────────────────────────
    def _import_excel(self):
        """Importa conteos desde un Excel exportado previamente.

        Reglas:
        - Cada hoja del Excel = una ubicación.
        - Hoja nueva en el Excel → se crea la ubicación automáticamente.
        - Hoja presente en el sistema pero AUSENTE del Excel →
          la ubicación se mantiene pero todos sus ítems pasan a qty=0
          (significa que esa ubicación fue inspeccionada y no tiene unidades).
        - Columnas esperadas por hoja: Código, [Anterior (fecha)], Cantidad, Observación.
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel","*.xlsx")],
            title="Seleccione el archivo de toma física exportado")
        if not path:
            return
        try:
            import openpyxl as _opx
            wb = _opx.load_workbook(path, data_only=True)
            imported = 0
            created  = 0
            zeroed   = 0
            skipped  = {"Resumen General"}   # hojas que no son ubicaciones

            # Ubicaciones presentes en el Excel
            excel_locs = {s.strip() for s in wb.sheetnames if s.strip() not in skipped}

            # ── Ubicaciones en el sistema que NO están en el Excel → qty=0 ──
            for loc in self.locations:
                if loc not in excel_locs:
                    # Poner todos los ítems en cero
                    for cod in self.data.get(loc, {}):
                        self.data[loc][cod]["qty"] = 0
                        zeroed += 1

            # ── Procesar cada hoja del Excel ─────────────────────────────────
            for sheet_name in wb.sheetnames:
                if sheet_name.strip() in skipped:
                    continue
                ws2 = wb[sheet_name]
                loc = sheet_name.strip()

                # Crear ubicación si no existe
                if loc not in self.locations:
                    self.locations.append(loc)
                    self.data[loc] = {}
                    created += 1

                # Leer cabeceras
                headers = [str(ws2.cell(1, c).value or "").strip()
                           for c in range(1, ws2.max_column + 1)]
                try:
                    col_cod  = headers.index("Código") + 1
                    col_cant = next(i+1 for i,h in enumerate(headers)
                                    if h.lower() in ("cantidad","qty","conteo"))
                except (ValueError, StopIteration):
                    continue   # hoja sin columnas reconocibles

                col_obs = None
                for i, h in enumerate(headers):
                    if "obs" in h.lower():
                        col_obs = i + 1
                        break

                # Leer filas
                for r in range(2, ws2.max_row + 1):
                    cod = str(ws2.cell(r, col_cod).value or "").strip()
                    if not cod or cod == "None":
                        continue
                    try:
                        qty = int(float(str(ws2.cell(r, col_cant).value or 0)))
                    except:
                        qty = 0
                    obs = str(ws2.cell(r, col_obs).value or "") if col_obs else ""
                    self.data[loc][cod] = {"qty": qty, "obs": obs}
                    imported += 1

            self._refresh_location_list()
            self._load_location(self.current_location)

            zeroed_msg = f"\n• {zeroed} ítems puestos en 0 (ubicaciones no incluidas en el Excel)" if zeroed else ""
            msg = (f"Importación completada:\n"
                   f"• {imported} conteos actualizados\n"
                   f"• {created} ubicaciones nuevas creadas"
                   f"{zeroed_msg}")
            messagebox.showinfo("Importado", msg)

        except Exception as e:
            messagebox.showerror("Error al importar", str(e))

    # ── Export Excel ──────────────────────────────────────────────────────────
    def _export_excel(self):
        df_full = self._build_full_df()
        if df_full.empty:
            messagebox.showinfo("Sin datos", "No hay conteos ingresados."); return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"toma_fisica_{self.toma_date_var.get().replace('/','_')}.xlsx")
        if not path: return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen General"

        # Paleta pastel/clara — apta para impresión
        HDR    = PatternFill("solid", fgColor="1E3A5F")   # título: azul marino (único oscuro)
        YELLOW = PatternFill("solid", fgColor="FEF9C3")   # cabeceras: amarillo pastel
        GREEN  = PatternFill("solid", fgColor="D1FAE5")   # total: verde pastel
        GRAY   = PatternFill("solid", fgColor="F1F5F9")   # datos numéricos: gris muy claro
        PREV   = PatternFill("solid", fgColor="E0E7FF")   # anterior: violeta pastel
        EVEN   = PatternFill("solid", fgColor="FFFFFF")   # filas pares: blanco
        ODD    = PatternFill("solid", fgColor="F8FAFC")   # filas impares: gris muy suave
        bold14 = Font(bold=True, size=13, color="FFFFFF")
        bold11 = Font(bold=True, size=10, color="1E3A5F")
        bold_y = Font(bold=True, size=10, color="1E3A5F")
        norm   = Font(size=10, color="111827")
        norm_n = Font(size=10, color="1E3A5F")            # números: azul marino
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_a = Alignment(horizontal="left",   vertical="center")
        right_a= Alignment(horizontal="right",  vertical="center")
        thin   = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ncols = len(df_full.columns)

        # Título
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        ws["A1"] = f"TOMA FÍSICA — {self.toma_date_var.get()} — {self.current_location}"
        ws["A1"].font = bold14
        ws["A1"].fill = HDR
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 28

        # Cabeceras
        for ci, col_name in enumerate(df_full.columns, 1):
            cell = ws.cell(2, ci, col_name)
            cell.font = bold_y; cell.fill = YELLOW
            cell.alignment = center if ci > 2 else left_a
            cell.border = border
        ws.row_dimensions[2].height = 20

        # Datos con zebra pastel
        for ri, (_, row) in enumerate(df_full.iterrows(), 3):
            row_fill = EVEN if ri % 2 == 0 else ODD
            for ci, col_name in enumerate(df_full.columns, 1):
                val  = row[col_name]
                cell = ws.cell(ri, ci, val if val != 0 else "")
                cell.border = border
                if col_name == "TOTAL":
                    cell.fill = GREEN
                    cell.font = Font(bold=True, size=10, color="065F46")
                    cell.alignment = right_a
                elif ci > 2:
                    cell.fill = row_fill; cell.font = norm_n; cell.alignment = right_a
                else:
                    cell.fill = row_fill; cell.alignment = left_a; cell.font = norm
            ws.row_dimensions[ri].height = 18

        # Anchos
        ws.column_dimensions["A"].width = 13
        ws.column_dimensions["B"].width = 36
        for ci in range(3, ncols+1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

        # Total row
        tr = df_full.shape[0] + 3
        ws.cell(tr, 1, "TOTAL GENERAL").font = bold11
        ws.cell(tr, 1).fill = HDR
        ws.cell(tr, 2, "").fill = HDR
        for ci in range(3, ncols+1):
            cl = get_column_letter(ci)
            cell = ws.cell(tr, ci, f"=SUM({cl}3:{cl}{tr-1})")
            cell.font = Font(bold=True, color="FEF9C3")
            cell.fill = HDR; cell.alignment = right_a; cell.border = border

        # Hoja por ubicación
        for loc in self.locations:
            loc_data = self.data.get(loc, {})
            if not loc_data:
                continue
            prev_date, prev_data = self._get_previous_toma(loc)
            ws2 = wb.create_sheet(title=loc[:28])
            headers = ["Código", "Nombre"]
            if prev_date:
                headers.append(f"Anterior ({prev_date})")
            headers += ["Cantidad", "Observación"]
            for ci, h in enumerate(headers, 1):
                cell = ws2.cell(1, ci, h)
                cell.font = Font(bold=True, size=10, color="1E3A5F")
                cell.fill = YELLOW
                cell.alignment = center; cell.border = border

            for ri, sku in enumerate(sorted(loc_data.keys()), 2):
                v    = loc_data[sku]
                name = self.skus[self.skus["Código Producto"]==sku]["Nombre Producto"]
                name = name.iloc[0] if not name.empty else ""
                ci_  = 1
                ws2.cell(ri, ci_, sku).border = border; ci_ += 1
                ws2.cell(ri, ci_, name).border = border; ci_ += 1
                if prev_date:
                    prev_q = prev_data.get(sku, {}).get("qty", "")
                    cell = ws2.cell(ri, ci_, prev_q)
                    cell.fill = PREV; cell.font = Font(size=10, color="3730A3")
                    cell.border = border; ci_ += 1
                ws2.cell(ri, ci_, v.get("qty","")).border = border; ci_ += 1
                ws2.cell(ri, ci_, v.get("obs","")).border = border

            ws2.column_dimensions["A"].width = 13
            ws2.column_dimensions["B"].width = 36

        wb.save(path)
        import os as _os, subprocess as _sub
        try:
            if _os.name == "nt":
                _os.startfile(path)
            else:
                _sub.Popen(["xdg-open", path])
        except Exception:
            pass
        messagebox.showinfo("Exportado", f"Archivo guardado:\n{path}")

    # ── Export PDF ────────────────────────────────────────────────────────────
    def _export_pdf(self):
        if not HAS_PDF:
            messagebox.showerror("PDF", "reportlab no disponible."); return
        df = self._build_full_df()
        if df.empty:
            messagebox.showinfo("Sin datos", "No hay conteos."); return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf", filetypes=[("PDF", "*.pdf")],
            initialfile=f"toma_fisica_{self.toma_date_var.get().replace('/','_')}.pdf")
        if not path: return

        doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                leftMargin=1*cm, rightMargin=1*cm,
                                topMargin=1.5*cm, bottomMargin=1*cm)
        styles = getSampleStyleSheet()
        elems = [
            Paragraph(f"<b>TOMA FÍSICA — {self.toma_date_var.get()}</b>", styles["Title"]),
            Spacer(1, 0.3*cm)
        ]
        headers = list(df.columns)
        data_rows = [headers]
        for _, row in df.iterrows():
            data_rows.append([str(v) if v != "" else "" for v in row])

        col_w = [2.5*cm, 7*cm] + [2*cm]*(len(headers)-2)
        tbl = Table(data_rows, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 8),
            ("ALIGN",      (2,0), (-1,-1), "RIGHT"),
            ("ALIGN",      (0,0), (1,-1),  "LEFT"),
            ("ROWBACKGROUNDS", (0,1),(-1,-1),
             [colors.HexColor("#1A2436"), colors.HexColor("#111827")]),
            ("TEXTCOLOR",  (0,1),(-1,-1), colors.HexColor("#E5E7EB")),
            ("GRID",       (0,0),(-1,-1), 0.4, colors.HexColor("#243247")),
            ("BACKGROUND", (-1,1),(-1,-1), colors.HexColor("#0F2D20")),
            ("TEXTCOLOR",  (-1,1),(-1,-1), colors.HexColor("#D1FAE5")),
            ("FONTNAME",   (-1,0),(-1,-1), "Helvetica-Bold"),
        ]))
        elems.append(tbl)
        doc.build(elems)
        messagebox.showinfo("Exportado", f"PDF guardado:\n{path}")
